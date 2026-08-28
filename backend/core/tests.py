from decimal import Decimal
from datetime import timedelta
import json
import importlib
from pathlib import Path
import tempfile
import zipfile
from types import SimpleNamespace
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import F
from django.test import TestCase, override_settings
from django.utils import timezone
import requests
from rest_framework.test import APIClient
from .models import AICatalogItem, AISettings, AuditLog, Debt, Expense, BusinessSettings, CatalogComposition, CatalogHistory, CatalogItem, CatalogMaterialUsage, Conversation, Customer, FloristAttendance, FloristProfile, FloristSalaryEntry, FloristVolumeRate, Flower, FlowerVariant, IntegrationSettings, Lead, LeadCatalogUsage, LeadStatus, Message, Notification, Packaging, PackagingMovement, PagePermission, Reservation, ReservationPayment, SocialPost, StockDelivery, Branch, CatalogTransfer, FloristDayOff, FloristStockBalance, FloristStockIssue, StockBatch, StockMovement, Supplier, SupplierPayment, UserProfile
from .serializers import CatalogItemSerializer, ConversationSerializer, FloristProfileSerializer, FloristSalaryEntrySerializer, FloristVolumeRateSerializer, PackagingSerializer, StockBatchSerializer, permission_matrix
from .inventory_services import catalog_remaining, close_selected_florist_issues, create_catalog_rework, issue_stock_to_florist, deduct_catalog_stock, mark_catalog_sold, sync_catalog_financials
from .services import available_catalog_queryset, catalog_composition_summary, AI_FOLLOW_UP_DELAY_SECONDS, ai_allowed_for_conversation, ai_catalog_rows, ai_flower_variant_rows, ai_reply, ai_stock_rows, ai_tool_definitions, apply_media_match_safeguard, calculate_custom_arrangement_price, create_ai_reply_for_conversation, customer_attachment_rows, execute_ai_tool, mini_app_custom_quote_ai, mini_app_quote_note, normalize_phone, process_pending_customer_reply, process_stalled_conversation_follow_up, recent_customer_orders, send_stock_batch_image, stock_batch_ai_row
from .tasks import process_conversation_follow_up, process_delayed_instagram_reply, process_delayed_telegram_reply
from .webhook_services import resolve_instagram_event, resolve_telegram_update, social_post_from_ai_catalog_item
from .backup_services import backup_command_matches, backup_caption, create_media_backup
from . import services, vision_services


def vision_fingerprint(**overrides):
    """Test uchun tayyor fingerprint. Model qaytaradigan shakl bilan bir xil."""
    row = {
        "flower_form": "peony_rose",
        "flower_variety_guess": "",
        "dominant_colors": ["cream", "pink"],
        "color_pattern": "two_tone",
        "container": "basket",
        "wrap_colors": [],
        "size": "large",
        "count_bucket": "50_to_100",
        "distinctive_features": [],
        "summary": "savatdagi kompozitsiya",
    }
    row.update(overrides)
    return row


def catalog_fingerprint_fields(image_url, **overrides):
    """AICatalogItem.objects.create ga qo'shiladigan tayyor fingerprint maydonlari."""
    return {
        "visual_fingerprint": vision_services.clean_fingerprint(vision_fingerprint(**overrides)),
        "fingerprint_source_url": image_url,
        "fingerprint_updated_at": timezone.now(),
    }


def verdict_payload(verdict="same_product", flower_form_match=True, color_match=True, container_match=True, differences=""):
    return {"verdict": verdict, "flower_form_match": flower_form_match, "color_match": color_match, "container_match": container_match, "differences": differences}


def patch_vision(source_fingerprint, verdicts):
    """Rasm tahlili va nomzod tekshiruvini bevosita almashtiradi.

    Tekshiruv har nomzod uchun alohida va parallel ketadi, shuning uchun OpenAI
    javoblarini ketma-ketlik bilan berish ishonchsiz bo'lardi.
    """
    from unittest.mock import patch

    return patch.multiple(
        "core.vision_services",
        analyze_image=lambda *args, **kwargs: vision_services.clean_fingerprint(source_fingerprint),
        verify_candidates=lambda source_url, source, rows, **kwargs: {row["item"].id: verdicts.get(row["item"].id, verdict_payload(verdict="different")) for row in rows},
    )


def media_conversation(external_id, image_url="https://cdn.example.com/customer.jpg", kind="photo"):
    customer = Customer.objects.create(instagram_user_id=external_id)
    conversation = Conversation.objects.create(customer=customer)
    conversation.messages.create(sender="customer", text="shu nechpul", metadata={"attachments": [{"kind": kind, "url": image_url}]})
    return conversation


class BusinessRulesTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("admin", password="password")
        flower = Flower.objects.create(name_uz="Atirgul", slug="rose")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Mondial", color_uz="Oq")
        self.batch = StockBatch.objects.create(variant=variant, batch_number="T-1", height_cm=60, stems_per_bunch=20, received_stems=100, remaining_stems=100, cost_per_stem=20000, sale_price_per_stem=30000, sale_price_per_bunch=580000)
        self.item = CatalogItem.objects.create(name_uz="Oq buket", arrangement_type="bouquet", price=500000)
        CatalogComposition.objects.create(catalog_item=self.item, stock_batch=self.batch, quantity_stems=15)

    @override_settings(AI_TEST_INSTAGRAM_USERNAMES=["extra_teest"], AI_TEST_INSTAGRAM_USER_IDS=[])
    def test_global_ai_off_allows_only_configured_test_instagram_user(self):
        AISettings.objects.update_or_create(pk=1, defaults={"is_active": False})
        test_customer = Customer.objects.create(instagram_user_id="ig-extra", instagram_username="extra_teest")
        regular_customer = Customer.objects.create(instagram_user_id="ig-real", instagram_username="real_user")
        self.assertTrue(ai_allowed_for_conversation(Conversation.objects.create(customer=test_customer)))
        self.assertFalse(ai_allowed_for_conversation(Conversation.objects.create(customer=regular_customer)))

    def test_instagram_webhook_saves_username_without_display_name(self):
        from unittest.mock import patch

        IntegrationSettings.objects.update_or_create(pk=1, defaults={"instagram_access_token": "token", "instagram_account_id": "biz-account"})
        response = SimpleNamespace(raise_for_status=lambda: None, json=lambda: {"username": "@extra_teest"})
        payload = {"entry": [{"messaging": [{"sender": {"id": "ig-extra-id"}, "recipient": {"id": "biz-account"}, "message": {"mid": "mid-profile-1", "text": "salom"}}]}]}
        with patch("core.webhook_services.requests.get", return_value=response):
            resolve_instagram_event(payload)
        customer = Customer.objects.get(instagram_user_id="ig-extra-id")
        self.assertEqual(customer.instagram_username, "extra_teest")
        self.assertEqual(str(customer), "extra_teest")

    def test_instagram_webhook_ignores_ai_sent_image_echo(self):
        IntegrationSettings.objects.update_or_create(pk=1, defaults={"instagram_access_token": "token", "instagram_account_id": "biz-account"})
        customer = Customer.objects.create(instagram_user_id="ig-customer", instagram_username="extra_teest")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="system", text="", metadata={"image_tool_result": {"sent": {"message_id": "image-echo-mid"}}})
        payload = {"entry": [{"messaging": [{"sender": {"id": "biz-account"}, "recipient": {"id": "ig-customer"}, "message": {"mid": "image-echo-mid", "attachments": [{"type": "image", "payload": {"url": "https://cdn.example.com/sent.jpg"}}]}}]}]}
        self.assertEqual(resolve_instagram_event(payload), [])
        conversation.refresh_from_db()
        self.assertEqual(conversation.status, "ai")
        self.assertFalse(conversation.messages.filter(sender="operator").exists())

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_context_exposes_already_known_lead_fields(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-known", name="Ahmad", phone="+998901112233")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="borib olaman")
        Lead.objects.create(customer=customer, conversation=conversation, request_uz="2 pochka", fulfillment="pickup", desired_date="2026-07-30", desired_time="15:00")
        payload = {"reply": "Yaxshi", "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="r1")
            ai_reply(conversation)
        context = json.loads(client.responses.create.call_args.kwargs["input"][0]["content"].split("REAL_CONTEXT_JSON:\n", 1)[1])
        known = context["conversation"]["already_known"]
        self.assertTrue(known["name"])
        self.assertTrue(known["phone"])
        self.assertEqual(known["fulfillment"], "pickup")
        self.assertEqual(known["payment_type"], "")
        self.assertTrue(known["desired_date"])
        self.assertTrue(known["desired_time"])
        self.assertEqual(context["conversation"]["open_lead"]["fulfillment"], "pickup")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_context_marks_returning_customer(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-return", name="Ahmad", phone="+998901112233")
        conversation = Conversation.objects.create(customer=customer)
        Lead.objects.create(customer=customer, request_uz="oldingi buyurtma")
        conversation.messages.create(sender="customer", text="salom")
        payload = {"reply": "Assalomu alaykum, Ahmad", "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="r2")
            ai_reply(conversation)
        context = json.loads(client.responses.create.call_args.kwargs["input"][0]["content"].split("REAL_CONTEXT_JSON:\n", 1)[1])
        self.assertTrue(context["customer"]["is_returning"])
        self.assertEqual(context["customer"]["previous_orders_count"], 1)
        self.assertEqual(context["customer"]["name"], "Ahmad")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_context_exposes_last_delivery_address(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-addr", name="Ahmad", phone="+998901112233")
        Lead.objects.create(customer=customer, request_uz="oldingi", delivery_address="Xadra 9")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="salom")
        payload = {"reply": "ok", "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="r3")
            ai_reply(conversation)
        context = json.loads(client.responses.create.call_args.kwargs["input"][0]["content"].split("REAL_CONTEXT_JSON:\n", 1)[1])
        self.assertEqual(context["customer"]["last_delivery_address"], "Xadra 9")

    def test_cyrillic_latin_roundtrip(self):
        from .services import cyrillic_to_latin, latin_to_cyrillic, detect_text_script
        self.assertEqual(cyrillic_to_latin("Ассалому Алайкум"), "Assalomu Alaykum")
        self.assertEqual(cyrillic_to_latin("Етказиб бериш керакми"), "Yetkazib berish kerakmi")
        self.assertEqual(cyrillic_to_latin("силада нечпул"), "silada nechpul")
        self.assertEqual(latin_to_cyrillic("Yetkazib berish kerakmi"), "Етказиб бериш керакми")
        self.assertEqual(latin_to_cyrillic("Rahmat, kuningiz xayrli o'tsin"), "Раҳмат, кунингиз хайрли ўтсин")
        self.assertEqual(latin_to_cyrillic("Sizga qanday gul kerak edi?"), "Сизга қандай гул керак эди?")
        self.assertEqual(latin_to_cyrillic("Florist haqi taxminan 50 000 so'm"), "Флорист ҳақи тахминан 50 000 сўм")

    def test_transliteration_protects_links_and_brands(self):
        from .services import latin_to_cyrillic
        self.assertIn("https://yandex.uz/maps/-/CTfQ6TMD", latin_to_cyrillic("Manzil https://yandex.uz/maps/-/CTfQ6TMD"))
        self.assertIn("Next Mall", latin_to_cyrillic("Next Mall dan keyin"))
        self.assertIn("EuroFlowers", latin_to_cyrillic("EuroFlowers Premium gul do'koni"))

    def test_detect_text_script_separates_uzbek_cyrillic_from_russian(self):
        from .services import detect_text_script
        self.assertEqual(detect_text_script("qanaqa gullar bor"), "latin")
        self.assertEqual(detect_text_script("канака гулла бор"), "uz_cyril")
        self.assertEqual(detect_text_script("Ассалому алайкум"), "uz_cyril")
        self.assertEqual(detect_text_script("какие цветы есть"), "ru")
        self.assertEqual(detect_text_script("сколько стоит букет"), "ru")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_cyrillic_customer_gets_cyrillic_reply_from_latin_model_output(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-cyr")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="канака гулла бор")
        payload = {"reply": "Skladimizda hozir quyidagi gullar bor", "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="rc")
            result = ai_reply(conversation)
        sent = client.responses.create.call_args.kwargs["input"][-1]["content"]
        self.assertIn("kanaka gulla bor", sent)
        self.assertEqual(result["reply"], "Складимизда ҳозир қуйидаги гуллар бор")
        self.assertEqual(result["reply_latin"], "Skladimizda hozir quyidagi gullar bor")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_russian_customer_reply_is_not_transliterated(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-ru2")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="какие цветы есть")
        payload = {"reply": "На складе сейчас есть розы", "detected_language": "ru", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="rr")
            result = ai_reply(conversation)
        sent = client.responses.create.call_args.kwargs["input"][-1]["content"]
        self.assertIn("какие цветы есть", sent)
        self.assertEqual(result["reply"], "На складе сейчас есть розы")

    def test_stock_row_exposes_pochka_fields(self):
        row = stock_batch_ai_row(self.batch)
        self.assertEqual(row["stems_per_pochka"], self.batch.stems_per_bunch)
        self.assertEqual(row["price_per_pochka"], str(self.batch.sale_price_per_bunch))
        self.assertEqual(row["price_per_stem"], str(self.batch.sale_price_per_stem))

    def test_stock_search_does_not_match_term_inside_another_word(self):
        from .services import haystack_has_term
        self.assertFalse(haystack_has_term("atirgul jumila podgallan pushti", "all"))
        self.assertTrue(haystack_has_term("atirgul jumila podgallan pushti", "jumila"))
        self.assertTrue(haystack_has_term("atirgul prut oq", "oq"))
        flower = Flower.objects.create(name_uz="Atirgul2", slug="rose2")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="podgallan", color_uz="Pushti")
        StockBatch.objects.create(variant=variant, batch_number="T-2", height_cm=50, stems_per_bunch=25, received_stems=100, remaining_stems=100, cost_per_stem=10000, sale_price_per_stem=15000, sale_price_per_bunch=375000)
        self.assertEqual(len(ai_stock_rows("", limit=50)), 2)

    def test_lead_tool_refuses_without_name_or_phone(self):
        customer = Customer.objects.create(instagram_user_id="ig-lead-guard")
        conversation = Conversation.objects.create(customer=customer)
        payload = {"customer_name": None, "phone": None, "request_text": "50 ta atirgul buket", "arrangement_type": "bouquet", "estimated_price": 800000, "fulfillment": None, "delivery_address": None, "desired_date": None, "desired_time": None, "catalog_items": [], "stock_items": [], "note": None}
        self.assertEqual(execute_ai_tool("client_lead_create", payload, conversation)["detail"], "customer_name_required")
        payload["customer_name"] = "Ahmad"
        self.assertEqual(execute_ai_tool("client_lead_create", payload, conversation)["detail"], "phone_required")
        self.assertFalse(Lead.objects.filter(customer=customer).exists())

    def test_lead_tool_persists_fulfillment_address_and_date(self):
        customer = Customer.objects.create(instagram_user_id="ig-lead-full")
        conversation = Conversation.objects.create(customer=customer)
        payload = {"customer_name": "Ahmad", "phone": "901112233", "request_text": "50 ta Atirgul Mondial oq buket", "arrangement_type": "bouquet", "estimated_price": 800000, "fulfillment": "delivery", "delivery_address": "Xadra 9", "desired_date": "2026-07-30", "desired_time": "15:00", "catalog_items": [], "stock_items": [{"batch_id": self.batch.id, "quantity_stems": 50, "quantity_bunches": 0}], "note": None}
        result = execute_ai_tool("client_lead_create", payload, conversation)
        self.assertTrue(result["ok"])
        lead = Lead.objects.get(id=result["lead_id"])
        self.assertEqual(lead.fulfillment, "delivery")
        self.assertEqual(lead.delivery_address, "Xadra 9")
        self.assertEqual(lead.desired_date.isoformat(), "2026-07-30")
        self.assertEqual(lead.desired_time, "15:00")
        # Florist haqini AI yozmaydi — u CRM da operator qo'yadi.
        self.assertEqual(lead.florist_fee, Decimal("0"))
        self.assertNotIn("custom", lead.request_uz.lower())
        edited = execute_ai_tool("client_lead_edit", {"lead_id": lead.id, "customer_name": None, "phone": None, "request_text": None, "status": None, "arrangement_type": None, "estimated_price": None, "fulfillment": "pickup", "delivery_address": None, "desired_date": None, "desired_time": None, "catalog_items": None, "stock_items": None, "note": None}, conversation)
        self.assertTrue(edited["ok"])
        lead.refresh_from_db()
        self.assertEqual(lead.fulfillment, "pickup")

    def test_stock_image_helper_reports_failure_instead_of_claiming_sent(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-image-fail")
        conversation = Conversation.objects.create(customer=customer)
        self.batch.image_url = "https://example.com/rose.jpg"
        self.batch.save(update_fields=["image_url", "updated_at"])
        with patch("core.services.instagram_send_image", side_effect=requests.HTTPError("400 Bad Request")):
            result = send_stock_batch_image(conversation, self.batch)
        self.assertFalse(result["ok"])
        self.assertFalse(result["image_sent"])
        self.assertEqual(result["detail"], "send_failed")

    def test_stock_image_helper_confirms_delivery_on_success(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-image-ok")
        conversation = Conversation.objects.create(customer=customer)
        self.batch.image_url = "https://example.com/rose.jpg"
        self.batch.save(update_fields=["image_url", "updated_at"])
        with patch("core.services.instagram_send_image", return_value={"message_id": "mid-1"}):
            result = send_stock_batch_image(conversation, self.batch)
        self.assertTrue(result["ok"])
        self.assertTrue(result["image_sent"])

    def test_catalog_image_result_carries_operator_note(self):
        """Bitta mahsulot rasmi ketgach AI izohdan javob bera olsin."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Izohli buket", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://example.com/note.jpg", note="100 ta guldan yasalgan, bo'yi 60 sm\n\nnarxi:1000000 kelishtirilgan narxi 800000")
        customer = Customer.objects.create(instagram_user_id="telegram:5010")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_image", return_value={"ok": True}):
            result = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation)
        self.assertTrue(result["ok"])
        self.assertEqual(result["note_uz"], item.note)

    def test_catalog_album_items_carry_operator_note(self):
        """Albomdan keyin mijoz raqam bilan tanlaydi — izoh o'sha natijada turishi kerak."""
        from unittest.mock import patch
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        items = self.make_album_catalog(1)
        items[0].note = "50 ta guli bor, bo'yi 45 sm\n\nnarxi:100000 kelishtirilgan narxi 80000"
        items[0].save(update_fields=["note", "updated_at"])
        customer = Customer.objects.create(instagram_user_id="telegram:5011")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_image", return_value={"ok": True}):
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertTrue(result["ok"])
        self.assertEqual(result["items"][0]["note_uz"], items[0].note)

    def make_album_catalog(self, count):
        items = []
        for index in range(count):
            items.append(AICatalogItem.objects.create(name=f"Albom buket {index + 1}", arrangement_type="bouquet", price=100000 * (index + 1), quantity=1, image_url=f"https://example.com/album-{index + 1}.jpg"))
        return items

    def test_catalog_album_sends_every_item_in_one_message(self):
        from unittest.mock import patch
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        self.make_album_catalog(3)
        customer = Customer.objects.create(instagram_user_id="telegram:5001")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_media_group", return_value={"ok": True}) as album_mock:
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertTrue(result["ok"])
        self.assertEqual(result["sent_as"], "album")
        self.assertEqual(result["messages_sent"], 1)
        self.assertTrue(result["numbering_visible"])
        self.assertEqual([row["position"] for row in result["items"]], [1, 2, 3])
        media = album_mock.call_args.args[1]
        self.assertEqual(len(media), 3)
        self.assertTrue(media[0]["caption"].startswith("1. "))
        self.assertIn("so'm", media[0]["caption"])
        self.assertEqual({row["catalog_id"] for row in result["items"]}, {row["catalog_id"] for row in ai_catalog_rows("", limit=80)})
        self.assertNotIn("image_url", result["items"][0])
        stored = conversation.messages.filter(sender="system").order_by("-id").first().metadata["catalog_album_result"]
        self.assertEqual([row["image_url"] for row in stored["items"]], ["https://example.com/album-3.jpg", "https://example.com/album-2.jpg", "https://example.com/album-1.jpg"])
        self.assertEqual([row["position"] for row in stored["items"]], [1, 2, 3])

    def test_catalog_album_keeps_numbering_across_chunks(self):
        from unittest.mock import patch
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        self.make_album_catalog(12)
        customer = Customer.objects.create(instagram_user_id="telegram:5002")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_media_group", return_value={"ok": True}) as album_mock:
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertEqual(album_mock.call_count, 2)
        self.assertEqual(result["messages_sent"], 2)
        self.assertEqual([row["position"] for row in result["items"]], list(range(1, 13)))
        self.assertEqual(len(album_mock.call_args_list[0].args[1]), 10)
        self.assertEqual(len(album_mock.call_args_list[1].args[1]), 2)
        self.assertTrue(album_mock.call_args_list[1].args[1][0]["caption"].startswith("11. "))

    def test_catalog_album_falls_back_to_single_images_when_album_fails(self):
        from unittest.mock import patch
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        self.make_album_catalog(2)
        customer = Customer.objects.create(instagram_user_id="telegram:5003")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_media_group", side_effect=requests.HTTPError("400 Bad Request")):
            with patch("core.services.telegram_send_image", return_value={"ok": True}) as single_mock:
                result = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertTrue(result["ok"])
        self.assertEqual(result["sent_as"], "one_by_one")
        self.assertFalse(result["numbering_visible"])
        self.assertEqual(single_mock.call_count, 2)

    def test_catalog_album_reports_failure_instead_of_claiming_sent(self):
        from unittest.mock import patch
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        self.make_album_catalog(1)
        customer = Customer.objects.create(instagram_user_id="telegram:5004")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_image", side_effect=requests.HTTPError("400 Bad Request")):
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertFalse(result["ok"])
        self.assertEqual(result["items"], [])
        self.assertEqual(len(result["not_sent"]), 1)

    def test_catalog_album_uses_instagram_carousel(self):
        from unittest.mock import patch
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        items = self.make_album_catalog(2)
        customer = Customer.objects.create(instagram_user_id="ig-album")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.instagram_send_carousel", return_value={"message_id": "mid-1"}) as carousel_mock:
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": [items[1].id, items[0].id]}, conversation)
        self.assertTrue(result["ok"])
        self.assertEqual(result["sent_as"], "album")
        elements = carousel_mock.call_args.args[1]
        self.assertEqual(len(elements), 2)
        self.assertTrue(elements[0]["title"].startswith("1. "))
        self.assertEqual([row["catalog_id"] for row in result["items"]], [items[1].id, items[0].id])

    def test_catalog_album_returns_empty_detail_when_nothing_on_sale(self):
        self.item.status = "archived"
        self.item.save(update_fields=["status", "updated_at"])
        customer = Customer.objects.create(instagram_user_id="telegram:5005")
        conversation = Conversation.objects.create(customer=customer)
        result = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertFalse(result["ok"])
        self.assertEqual(result["detail"], "catalog_empty")

    def test_send_catalog_image_accepts_catalog_id(self):
        from unittest.mock import patch
        item = self.make_album_catalog(1)[0]
        customer = Customer.objects.create(instagram_user_id="telegram:5006")
        conversation = Conversation.objects.create(customer=customer)
        with patch("core.services.telegram_send_image", return_value={"ok": True}):
            result = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation)
        self.assertTrue(result["ok"])
        self.assertEqual(result["catalog_id"], item.id)
        self.assertEqual(result["price_text"], "100 000 so'm")
        self.assertIn("Narxi 100 000 so'm", result["reply_instruction"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_context_exposes_operator_contact(self):
        from unittest.mock import patch
        settings_row, _ = BusinessSettings.objects.get_or_create(pk=1)
        settings_row.operator_phone = "+998 88 111 22 33"
        settings_row.operator_hours = "08:00 dan 00:00 gacha"
        settings_row.operator_hours_ru = "с 08:00 до 00:00"
        settings_row.save()
        customer = Customer.objects.create(instagram_user_id="ig-operator")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="operator bilan gaplashsam boladimi")
        payload = {"reply": "Aloqa raqamimiz", "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="r-op")
            ai_reply(conversation)
        context = json.loads(client.responses.create.call_args.kwargs["input"][0]["content"].split("REAL_CONTEXT_JSON:\n", 1)[1])
        self.assertEqual(context["business"]["operator_phone"], "+998 88 111 22 33")
        self.assertEqual(context["business"]["operator_hours_uz"], "08:00 dan 00:00 gacha")
        self.assertEqual(context["business"]["operator_hours_ru"], "с 08:00 до 00:00")
        # Mijozga username berilmaydi — javob shunchaki kutishga chorlaydi.
        self.assertEqual(context["business"]["operator_telegram_text"],
                         "Operatorlarimiz sizga tez orada yozib yuborishadi")
        self.assertNotIn("@", context["business"]["operator_telegram_text"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_context_uses_business_settings_not_hardcoded_values(self):
        from unittest.mock import patch
        settings_row, _ = BusinessSettings.objects.get_or_create(pk=1)
        settings_row.working_hours = {"uz": "Har kuni 09:00-22:00", "ru": "Ежедневно 09:00-22:00"}
        settings_row.shop_phone = "+998 90 000 00 00"
        settings_row.save()
        customer = Customer.objects.create(instagram_user_id="ig-ctx")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="ish vaqti qanaqa")
        payload = {"reply": "Har kuni 09:00 dan 22:00 gacha", "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="resp_ctx")
            ai_reply(conversation)
        context = client.responses.create.call_args.kwargs["input"][0]["content"]
        self.assertIn("Har kuni 09:00-22:00", context)
        self.assertIn("+998 90 000 00 00", context)
        self.assertNotIn("24/7", context)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_reply_is_stored_without_post_processing(self):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-nopostproc")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="manzil qayerda")
        original = "Manzilimiz Bobur ko'chasi 10\nNext Mall dan keyin o'ng qo'lda"
        with patch("core.services.ai_reply", return_value={"reply": original, "detected_language": "uz", "customer_name": None, "phone": None, "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None, "handoff": False, "catalog_items": [], "stock_items": []}):
            reply = create_ai_reply_for_conversation(conversation)
        self.assertEqual(reply.text, original)

    def test_selling_does_not_automatically_deduct_stock(self):
        mark_catalog_sold(self.item, self.user)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 100)
        self.assertFalse(Notification.objects.filter(notification_type="stock_pending", reference_id=self.item.id).exists())

    def test_manual_deduction_is_atomic_and_once_only(self):
        mark_catalog_sold(self.item, self.user)
        deduct_catalog_stock(self.item, self.user)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 85)
        with self.assertRaises(ValueError):
            deduct_catalog_stock(self.item, self.user)

    def test_catalog_partial_sales_deduct_composition_per_quantity(self):
        item = CatalogItem.objects.create(name_uz="Qizil set", arrangement_type="bouquet", price=900000, quantity_total=10, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=3)
        mark_catalog_sold(item, self.user, quantity=3)
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 3)
        deduct_catalog_stock(item, self.user, quantity=3)
        self.batch.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 91)
        self.assertEqual(item.quantity_stock_deducted, 3)
        self.assertEqual(item.status, "available")

    def test_catalog_discounted_sale_requires_reason_and_creates_history(self):
        item = CatalogItem.objects.create(name_uz="Skidka buket", arrangement_type="bouquet", price=500000, quantity_total=2, status="available")
        with self.assertRaises(ValueError):
            mark_catalog_sold(item, self.user, quantity=1, sale_price=450000)
        mark_catalog_sold(item, self.user, quantity=1, sale_price=450000, discount_reason="Doimiy mijoz")
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 1)
        history = CatalogHistory.objects.get(catalog_item=item, action="sold")
        self.assertEqual(history.listed_unit_price, Decimal("500000.00"))
        self.assertEqual(history.sold_unit_price, Decimal("450000.00"))
        self.assertEqual(history.discount_amount, Decimal("50000.00"))
        self.assertEqual(history.discount_percent, Decimal("10.00"))
        self.assertEqual(history.discount_reason, "Doimiy mijoz")

    def test_ai_stock_rows_return_every_distinct_offer(self):
        """Bo'yi yoki narxi boshqa partiya alohida taklif — AI ikkalasini ham ko'radi."""
        other_price = StockBatch.objects.create(variant=self.batch.variant, batch_number="T-2", height_cm=60, stems_per_bunch=20, received_stems=200, remaining_stems=200, cost_per_stem=21000, sale_price_per_stem=35000, sale_price_per_bunch=680000, received_at=timezone.localdate() + timedelta(days=1))
        batch_ids = [row["batch_id"] for row in ai_stock_rows("atirgul mondial", limit=10)]
        self.assertIn(self.batch.id, batch_ids)
        self.assertIn(other_price.id, batch_ids)

    def test_ai_stock_rows_skip_a_repeated_offer(self):
        """Bo'yi ham, narxi ham bir xil partiya takrorlanmaydi — eng eskisi qoladi."""
        same_offer = StockBatch.objects.create(variant=self.batch.variant, batch_number="T-3", height_cm=60, stems_per_bunch=20, received_stems=200, remaining_stems=200, cost_per_stem=20000, sale_price_per_stem=30000, sale_price_per_bunch=580000, received_at=timezone.localdate() + timedelta(days=1))
        batch_ids = [row["batch_id"] for row in ai_stock_rows("atirgul mondial", limit=10)]
        self.assertIn(self.batch.id, batch_ids)
        self.assertNotIn(same_offer.id, batch_ids)

    def test_ai_stock_rows_treats_whitespace_query_as_all_stock(self):
        rows = ai_stock_rows(" ", limit=10)
        self.assertTrue(rows)
        self.assertEqual(rows[0]["batch_id"], self.batch.id)

    def test_ai_stock_rows_include_cyrillic_display_names(self):
        rows = ai_stock_rows("atirgul", limit=10)
        self.assertTrue(rows)
        self.assertIn("display_name_uz_cyril", rows[0])
        self.assertIn("Атиргул", rows[0]["display_name_uz_cyril"])

    def test_ai_catalog_rows_treats_whitespace_query_as_all_catalog(self):
        item = AICatalogItem.objects.create(name="AI oq buket", arrangement_type="bouquet", price=500000, quantity=2, volume="M", note="Oq premium buket")
        rows = ai_catalog_rows(" ", limit=10)
        self.assertTrue(rows)
        self.assertEqual(rows[0]["catalog_id"], item.id)
        self.assertEqual(rows[0]["name_uz"], item.name)
        self.assertEqual(rows[0]["quantity"], 2)
        self.assertEqual(rows[0]["volume"], "M")

    def test_ai_catalog_rows_match_words_from_a_whole_sentence(self):
        """Mijoz butun jumla yozadi — nom bo'yicha so'zma-so'z moslik ishlashi kerak."""
        item = AICatalogItem.objects.create(name="Shoxli bambastik gulidan kompozitsiya", arrangement_type="bouquet", price=900000, quantity=1)
        AICatalogItem.objects.create(name="Oq atirgul buketi", arrangement_type="bouquet", price=400000, quantity=1)
        rows = ai_catalog_rows("buket kotta shoxli bambastik gulidan yasalgan qancha turadi", limit=20)
        self.assertEqual([row["catalog_id"] for row in rows], [item.id])

    def test_ai_catalog_rows_stay_empty_when_nothing_matches(self):
        AICatalogItem.objects.create(name="Oq atirgul buketi", arrangement_type="bouquet", price=400000, quantity=1)
        self.assertEqual(ai_catalog_rows("kaktus kerak edi", limit=20), [])

    def test_ai_catalog_rows_carry_operator_note(self):
        """Izoh AI ga note_uz bo'lib boradi. Ichida mahsulot tafsiloti ham, kelishilgan narx ham bo'ladi."""
        item = AICatalogItem.objects.create(name="Izohli kompozitsiya", arrangement_type="bouquet", price=1000000, quantity=1, note="100 ta guldan yasalgan, bo'yi 60 sm\n\nnarxi:1000000 kelishtirilgan narxi 800000")
        row = [candidate for candidate in ai_catalog_rows(" ", limit=20) if candidate["catalog_id"] == item.id][0]
        self.assertEqual(row["note_uz"], item.note)
        self.assertEqual(row["price"], "1000000.00")
        # Izoh mijozga tayyor matn emas, shuning uchun description nomi bilan berilmaydi.
        self.assertNotIn("description_uz", row)
        self.assertNotIn("description_ru", row)

    def test_ai_catalog_instagram_link_syncs_to_social_post(self):
        item = AICatalogItem.objects.create(name="Reel buket", arrangement_type="bouquet", price=750000, quantity=1, image_url="https://cdn.example.com/reel.jpg", instagram_link="https://www.instagram.com/reel/ABC123/?igsh=test")
        post = social_post_from_ai_catalog_item(item)
        self.assertEqual(post.post_type, "reel")
        self.assertEqual(post.title_uz, item.name)
        self.assertEqual(post.price, item.price)
        item.name = "Yangilangan reel buket"
        item.price = Decimal("800000")
        item.save()
        updated = social_post_from_ai_catalog_item(item)
        self.assertEqual(updated.id, post.id)
        updated.refresh_from_db()
        self.assertEqual(updated.title_uz, "Yangilangan reel buket")
        self.assertEqual(updated.price, Decimal("800000"))
        second = AICatalogItem.objects.create(name="Ikkinchi reel buket", arrangement_type="bouquet", price=650000, quantity=1, image_url="https://cdn.example.com/reel-2.jpg", instagram_link=item.instagram_link)
        second_post = social_post_from_ai_catalog_item(second)
        self.assertNotEqual(second_post.id, post.id)
        self.assertEqual(SocialPost.objects.filter(permalink=item.instagram_link).count(), 2)

    def test_custom_catalog_deducts_inventory_and_creates_salary_from_volume_rate(self):
        florist_user = User.objects.create_user("florist", password="password", first_name="Ali")
        florist = FloristProfile.objects.create(user=florist_user, staff_type="florist")
        # Tarif aniq floristga biriktiriladi, umumiy tarif endi ishlatilmaydi
        FloristVolumeRate.objects.create(florist=florist, arrangement_type="bouquet", volume="small", default_stems=10, florist_fee=70000)
        # Florist tanlangan katalog gulni floristning qo'lidagi qoldiqdan oladi
        from .inventory_services import issue_stock_to_florist
        issue_stock_to_florist(florist, self.batch, 10, "test", self.user)
        packaging = Packaging.objects.create(packaging_type="wrap", name_uz="Test qogoz", cost_price=10000, sale_price=20000, quantity=5)
        serializer = CatalogItemSerializer(data={
            "name_uz": "Custom buket",
            "arrangement_type": "bouquet",
            "catalog_kind": "custom",
            "volume": "small",
            "florist": florist.id,
            "price": "250000.00",
            "discount_reason": "Doimiy mijozga chegirma",
            "quantity_total": 1,
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10, "quantity_bunches": "0.50"}],
            "materials": [{"packaging": packaging.id, "quantity": 1}],
        }, context={"request": SimpleNamespace(user=self.user)})
        self.assertTrue(serializer.is_valid(), serializer.errors)
        item = serializer.save()
        self.batch.refresh_from_db()
        packaging.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(item.status, "draft")
        self.assertEqual(item.quantity_sold, 0)
        self.assertIsNone(item.sold_at)
        self.assertEqual(self.batch.remaining_stems, 90)
        self.assertEqual(packaging.quantity, 4)
        self.assertEqual(item.florist_fee, Decimal("50000.00"))
        self.assertEqual(item.florist_salary_amount, Decimal("70000.00"))
        self.assertEqual(item.calculated_component_price, Decimal("370000.00"))
        self.assertEqual(item.calculated_cost_price, Decimal("260000.00"))
        self.assertEqual(item.discount_amount, Decimal("120000.00"))
        self.assertEqual(item.discount_percent, Decimal("32.43"))
        salary = FloristSalaryEntry.objects.get(catalog_item=item)
        self.assertEqual(salary.amount, Decimal("70000.00"))
        self.assertEqual(salary.florist, florist)
        self.assertFalse(CatalogHistory.objects.filter(catalog_item=item, action="sold").exists())

    def test_catalog_for_apprentice_does_not_create_salary_entry(self):
        apprentice_user = User.objects.create_user("catalog-apprentice", password="password", first_name="Vali")
        apprentice = FloristProfile.objects.create(user=apprentice_user, staff_type="apprentice", daily_pay=100000)
        FloristVolumeRate.objects.create(florist=apprentice, arrangement_type="bouquet", volume="small", default_stems=10, florist_fee=70000)
        from .inventory_services import issue_stock_to_florist
        issue_stock_to_florist(apprentice, self.batch, 10, "test", self.user)
        serializer = CatalogItemSerializer(data={
            "name_uz": "Shogird buketi",
            "arrangement_type": "bouquet",
            "catalog_kind": "standard",
            "volume": "small",
            "florist": apprentice.id,
            "price": "250000.00",
            "quantity_total": 1,
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10, "quantity_bunches": "0.50"}],
        }, context={"request": SimpleNamespace(user=self.user)})
        self.assertTrue(serializer.is_valid(), serializer.errors)
        item = serializer.save()
        item.refresh_from_db()
        self.assertEqual(item.florist_salary_amount, Decimal("0.00"))
        self.assertFalse(FloristSalaryEntry.objects.filter(catalog_item=item).exists())

    def test_apprentice_issue_close_does_not_require_volume_rate(self):
        apprentice_user = User.objects.create_user("close-apprentice", password="password", first_name="Vali")
        apprentice = FloristProfile.objects.create(user=apprentice_user, staff_type="apprentice", daily_pay=100000)
        issue_stock_to_florist(apprentice, self.batch, 10, "test", self.user)
        item = CatalogItem.objects.create(
            name_uz="Shogird yopish buketi",
            arrangement_type="bouquet",
            catalog_kind="standard",
            volume="small",
            florist=apprentice,
            price=Decimal("250000.00"),
            quantity_total=1,
        )
        row = CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=0)
        from .inventory_services import close_florist_issue
        result = close_florist_issue(apprentice, self.batch, user=self.user)
        row.refresh_from_db()
        self.assertEqual(result["weight_source"], "apprentice_equal")
        self.assertEqual(row.quantity_stems, 10)
        self.assertFalse(FloristSalaryEntry.objects.filter(catalog_item=item).exists())

    def test_custom_catalog_accepts_custom_volume_and_manual_salary_amount(self):
        florist_user = User.objects.create_user("custom-florist", password="password", first_name="Ali")
        florist = FloristProfile.objects.create(user=florist_user, staff_type="florist")
        from .inventory_services import issue_stock_to_florist
        issue_stock_to_florist(florist, self.batch, 12, "test", self.user)
        serializer = CatalogItemSerializer(data={
            "name_uz": "Juda katta custom buket",
            "arrangement_type": "bouquet",
            "catalog_kind": "custom",
            "volume": "extra katta 120 dona",
            "florist": florist.id,
            "price": "450000.00",
            "florist_fee": "80000.00",
            "florist_salary_amount": "125000.00",
            "quantity_total": 1,
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10, "quantity_bunches": "0.50"}],
        }, context={"request": SimpleNamespace(user=self.user)})
        self.assertTrue(serializer.is_valid(), serializer.errors)
        item = serializer.save()
        item.refresh_from_db()
        self.assertEqual(item.volume, "extra katta 120 dona")
        self.assertEqual(item.florist_fee, Decimal("80000.00"))
        self.assertEqual(item.florist_salary_amount, Decimal("125000.00"))
        self.assertEqual(item.calculated_component_price, Decimal("380000.00"))
        salary = FloristSalaryEntry.objects.get(catalog_item=item)
        self.assertEqual(salary.amount, Decimal("125000.00"))

    def test_stock_batch_serializer_returns_fractional_remaining_bunches(self):
        self.batch.remaining_stems = 85
        self.batch.save(update_fields=["remaining_stems", "updated_at"])
        data = StockBatchSerializer(self.batch).data
        self.assertEqual(data["remaining_bunches"], "4.25")
        self.assertEqual(data["remaining_bunches_label"], "4.25 pochka")

    def test_florist_profile_accepts_precise_coordinates_and_nested_volume_rates(self):
        florist_user = User.objects.create_user("precise-florist", password="password", first_name="Ali")
        serializer = FloristProfileSerializer(data={
            "user": florist_user.id,
            "staff_type": "florist",
            "daily_pay": "150000.00",
            "shop_latitude": "41.31108123456789",
            "shop_longitude": "69.24056234567891",
            "volume_rates": [
                {"arrangement_type": "bouquet", "volume": "small", "default_stems": 15, "florist_fee": "50000.00"},
                {"arrangement_type": "basket", "volume": "large", "default_stems": 45, "florist_fee": "120000.00"},
            ],
        })
        self.assertTrue(serializer.is_valid(), serializer.errors)
        profile = serializer.save()
        self.assertEqual(profile.daily_pay, Decimal("0"))
        self.assertEqual(profile.shop_latitude, Decimal("41.3110812346"))
        self.assertEqual(profile.shop_longitude, Decimal("69.2405623457"))
        self.assertEqual(profile.volume_rates.count(), 2)
        self.assertTrue(profile.volume_rates.filter(arrangement_type="basket", volume="large", florist_fee=Decimal("120000.00")).exists())
        self.assertNotIn("branch", FloristProfileSerializer(profile).data)
        rate = profile.volume_rates.first()
        self.assertNotIn("branch", FloristVolumeRateSerializer(rate).data)

    def test_apprentice_daily_salary_update_requires_reason(self):
        apprentice_user = User.objects.create_user("apprentice", password="password", first_name="Vali")
        apprentice = FloristProfile.objects.create(user=apprentice_user, staff_type="apprentice", daily_pay=100000)
        entry = FloristSalaryEntry.objects.create(florist=apprentice, source="daily", amount=100000, work_date=timezone.localdate(), note="Kunlik")
        serializer = FloristSalaryEntrySerializer(entry, data={"amount": "120000.00"}, partial=True)
        self.assertFalse(serializer.is_valid())
        serializer = FloristSalaryEntrySerializer(entry, data={"amount": "120000.00", "reason": "Qo‘shimcha smena"}, partial=True)
        self.assertTrue(serializer.is_valid(), serializer.errors)
        updated = serializer.save()
        self.assertEqual(updated.amount, Decimal("120000.00"))
        self.assertIn("Qo‘shimcha smena", updated.note)

    def test_packaging_serializer_accepts_image_and_returns_quantity_label(self):
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                serializer = PackagingSerializer(data={
                    "packaging_type": "other",
                    "name_uz": "Shokolad",
                    "cost_price": "10000.00",
                    "sale_price": "20000.00",
                    "quantity": 12,
                    "image": SimpleUploadedFile("shokolad.jpg", b"image-bytes", content_type="image/jpeg"),
                })
                self.assertTrue(serializer.is_valid(), serializer.errors)
                packaging = serializer.save()
                data = PackagingSerializer(packaging).data
                self.assertEqual(data["quantity_label"], "12 dona")
                self.assertTrue(data["image_url"].endswith("shokolad.jpg"))

    def test_phone_normalization(self):
        self.assertEqual(normalize_phone("90 123-45-67"), "+998901234567")
        self.assertEqual(normalize_phone("+998 90 123 45 67"), "+998901234567")
        self.assertEqual(normalize_phone("998901234567"), "+998901234567")
        self.assertEqual(normalize_phone("+998 ** *** ** 67"), "")
        self.assertEqual(normalize_phone("+99867"), "")
        self.assertEqual(normalize_phone("67"), "")

    def test_ai_catalog_generic_query_returns_available_items(self):
        from .services import ai_catalog_rows
        AICatalogItem.objects.create(name="AI oq buket", arrangement_type="bouquet", price=500000, quantity=1)
        AICatalogItem.objects.create(name="AI tugagan buket", arrangement_type="bouquet", price=500000, quantity=0)
        AICatalogItem.objects.create(name="AI arxiv buket", arrangement_type="bouquet", price=500000, quantity=1, is_active=False)
        CatalogItem.objects.create(name_uz="Ichki CRM buket", arrangement_type="bouquet", price=500000, status="available")
        for query in ["vitrina", "vitrinada qanaqa gulla bor", "katalogdagi tayyor mahsulotlar"]:
            rows = ai_catalog_rows(query)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["name_uz"], "AI oq buket")

    @override_settings(BACKUP_TELEGRAM_GROUP_ID="-1003718639311", BACKUP_TELEGRAM_THREAD_ID="1542", BACKUP_TELEGRAM_COMMAND="/eurodan_backup_tashachi")
    def test_backup_command_matches_only_configured_group_thread(self):
        payload = {"message": {"text": "/eurodan_backup_tashachi", "chat": {"id": -1003718639311}, "message_thread_id": 1542}}
        self.assertTrue(backup_command_matches(payload))
        self.assertFalse(backup_command_matches({"message": {"text": "/eurodan_backup_tashachi", "chat": {"id": -1003718639311}, "message_thread_id": 999}}))
        self.assertFalse(backup_command_matches({"message": {"text": "/start", "chat": {"id": -1003718639311}, "message_thread_id": 1542}}))

    def test_backup_media_zip_is_created_separately(self):
        with tempfile.TemporaryDirectory() as media_root, tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(media_root) / "catalog" / "test.jpg"
            image_path.parent.mkdir(parents=True)
            image_path.write_bytes(b"image-bytes")
            with override_settings(MEDIA_ROOT=media_root):
                zip_path = create_media_backup(temp_dir)
            self.assertEqual(zip_path.name, "euroflowers_media_images.zip")
            with zipfile.ZipFile(zip_path) as archive:
                self.assertEqual(archive.namelist(), ["catalog/test.jpg"])

    def test_backup_caption_uses_markdown_and_escapes_filename(self):
        caption = backup_caption(Path("euroflowers_media_images.zip"), "manual:test_user", 3, 3)
        self.assertIn("📦 *EuroFlowers backup*", caption)
        self.assertIn("Media rasmlar ZIP", caption)
        self.assertIn("euroflowers\\_media\\_images\\.zip", caption)
        self.assertIn("manual:test\\_user", caption)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_ai_reply_sends_context_conversation_and_allowed_tools(self):
        customer = Customer.objects.create(instagram_user_id="ig-tools", name="Ahmad")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="qanaqa gullar bor")
        conversation.messages.create(sender="ai", text="Skladimizda bor", metadata={"tool_results": [{"name": "get_stock", "output": {"stock": [{"price_per_stem": "105000.00"}]}}]})
        payload = {
            "reply": "Katalogdagi gullarni ko‘rib beraman.",
            "detected_language": "uz",
            "customer_name": None,
            "phone": None,
            "lead_ready": False,
            "lead_request": None,
            "arrangement_type": None,
            "estimated_price": None,
            "handoff": False,
            "catalog_items": [],
            "stock_items": [],
        }
        from unittest.mock import patch
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="resp_1")
            result = ai_reply(conversation)
        kwargs = client.responses.create.call_args.kwargs
        self.assertEqual({tool["name"] for tool in kwargs["tools"]}, {"client_leads_get", "client_lead_create", "client_lead_edit", "client_payment_update", "call_operator", "delivery_location_link", "match_ai_catalog_by_media", "get_catalog", "send_catalog_image", "send_catalog_album", "send_post_image"})
        self.assertTrue(kwargs["parallel_tool_calls"] is False)
        self.assertEqual(result["reply"], payload["reply"])
        self.assertIn(AISettings.objects.get(pk=1).system_prompt, kwargs["instructions"])
        self.assertTrue(kwargs["instructions"].startswith("MEDIA MATCHING FIRST"))
        self.assertTrue(kwargs["input"][0]["content"].startswith("REAL_CONTEXT_JSON:"))
        self.assertIn("shop_phone", kwargs["input"][0]["content"])
        self.assertIn("working_hours_uz", kwargs["input"][0]["content"])
        self.assertIn("qanaqa gullar bor", kwargs["input"][1]["content"])
        self.assertIn("105000.00", kwargs["input"][-1]["content"])

    def test_prompt_shows_free_container_colors_first(self):
        """Idish rangi savolida avval tekinlari aytiladi, pullisi keyin.

        Promptni keyingi migratsiya qayta yozganda shu qoida tushib qolmasin.
        """
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("7A. IDISH RANGI", prompt)
        section = prompt.split("7A. IDISH RANGI", 1)[1].split("8. BUYURTMA", 1)[0]
        for color in ["Havo rang", "Malla", "Oq", "Pushti", "Ko'k"]:
            self.assertIn(color, section)
        self.assertIn("Qizil va Tilla", section)
        self.assertIn("100 000 so'm", section)
        # Mijoz rang nomini aytsa yashirilmaydi — bu alohida holat.
        self.assertIn("3-HOLAT. MIJOZ QIZIL YOKI TILLANI NOMMA-NOM SO'RADI", section)
        # Pulli ranglar faqat shu bo'limda tilga olinadi, boshqa joyda emas.
        self.assertEqual(prompt.count("Tilla"), section.count("Tilla"))
        # Eski qoida narxni oldindan aytishga undardi, u olib tashlandi.
        self.assertNotIn("savat idishi narxi qo'shilishini ayt", prompt)

    def test_ai_tool_definitions_are_whitelisted(self):
        self.assertEqual({tool["name"] for tool in ai_tool_definitions()}, {"client_leads_get", "client_lead_create", "client_lead_edit", "client_payment_update", "call_operator", "delivery_location_link", "match_ai_catalog_by_media", "get_catalog", "send_catalog_image", "send_catalog_album", "send_post_image"})

    def test_get_catalog_tool_filters_baskets(self):
        basket = AICatalogItem.objects.create(name="Oq savat", arrangement_type="basket", price=700000, quantity=1)
        bouquet = AICatalogItem.objects.create(name="Oq buket", arrangement_type="bouquet", price=500000, quantity=1)
        customer = Customer.objects.create(instagram_user_id="telegram:11")
        conversation = Conversation.objects.create(customer=customer)
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": "basket"}, conversation)
        names = {row["name_uz"] for row in result["catalog"]}
        self.assertIn(basket.name, names)
        self.assertNotIn(bouquet.name, names)

    def test_stock_rows_do_not_return_baskets_when_flower_is_missing(self):
        self.assertEqual(ai_stock_rows("gortenziya", limit=10), [])

    def test_stock_tools_are_hidden_from_the_ai(self):
        """Sklad AI ga ko'rsatilmaydi — tool ro'yxatida ham, javobda ham."""
        customer = Customer.objects.create(instagram_user_id="telegram:13")
        conversation = Conversation.objects.create(customer=customer)
        exposed = {tool["name"] for tool in ai_tool_definitions()}
        for name in ["get_stock", "get_flower_variant_info", "calculate_custom_arrangement_price", "send_stock_image", "send_stock_images"]:
            self.assertNotIn(name, exposed)
            result = execute_ai_tool(name, {"query": "atirgul"}, conversation)
            self.assertFalse(result["ok"])
            self.assertEqual(result["detail"], "stock_not_available_to_ai")
            self.assertIn("client_lead_create", result["instruction"])

    def test_calculate_custom_arrangement_price_is_deterministic(self):
        BusinessSettings.objects.update_or_create(pk=1, defaults={"default_florist_fee": Decimal("50000")})
        self.batch.sale_price_per_stem = Decimal("15000")
        self.batch.remaining_stems = 100
        self.batch.save(update_fields=["sale_price_per_stem", "remaining_stems", "updated_at"])
        result = calculate_custom_arrangement_price([{"batch_id": self.batch.id, "quantity_stems": 50, "quantity_bunches": 0}])
        self.assertTrue(result["ok"])
        self.assertEqual(result["flower_subtotal"], "750000")
        self.assertEqual(result["florist_fee"], "50000")
        self.assertEqual(result["total"], "800000")
        self.assertIn("Jami taxminan 800 000 so'm", result["display_summary_uz"]["total"])

    def test_calculate_custom_arrangement_price_handles_multiple_flowers(self):
        BusinessSettings.objects.update_or_create(pk=1, defaults={"default_florist_fee": Decimal("50000")})
        self.batch.sale_price_per_stem = Decimal("15000")
        self.batch.remaining_stems = 100
        self.batch.save(update_fields=["sale_price_per_stem", "remaining_stems", "updated_at"])
        second = StockBatch.objects.create(variant=self.batch.variant, batch_number="T-PRUT", height_cm=60, stems_per_bunch=20, received_stems=100, remaining_stems=100, cost_per_stem=10000, sale_price_per_stem=15000, sale_price_per_bunch=300000)
        customer = Customer.objects.create(instagram_user_id="telegram:calc")
        conversation = Conversation.objects.create(customer=customer)
        result = calculate_custom_arrangement_price([
            {"batch_id": self.batch.id, "quantity_stems": 10, "quantity_bunches": 0},
            {"batch_id": second.id, "quantity_stems": 10, "quantity_bunches": 0},
        ])
        self.assertTrue(result["ok"])
        self.assertEqual(result["flower_subtotal"], "300000")
        self.assertEqual(result["total"], "350000")
        self.assertEqual(len(result["lines"]), 2)

    def test_send_stock_batch_image_sends_flower_image(self):
        self.batch.image_url = "https://example.com/freedom.jpg"
        self.batch.save(update_fields=["image_url", "updated_at"])
        customer = Customer.objects.create(instagram_user_id="telegram:13")
        conversation = Conversation.objects.create(customer=customer)
        from unittest.mock import patch
        with patch("core.services.telegram_send_image", return_value={"ok": True}) as image_mock:
            result = send_stock_batch_image(conversation, self.batch)
        self.assertTrue(result["ok"])
        self.assertEqual(result["image_url"], "https://example.com/freedom.jpg")
        image_mock.assert_called_once_with("13", "https://example.com/freedom.jpg")

    def _ai_catalog_lead(self, catalog_name="Oq buket"):
        AICatalogItem.objects.create(name="Oq buket", arrangement_type="bouquet", quantity=3, price=Decimal("500000"))
        customer = Customer.objects.create(instagram_user_id="telegram:10")
        conversation = Conversation.objects.create(customer=customer)
        result = execute_ai_tool("client_lead_create", {
            "customer_name": "Ahmad",
            "phone": "901234567",
            "request_text": "Oq buket 1 dona, kelib olish",
            "arrangement_type": "catalog",
            "estimated_price": 500000,
            "catalog_items": [{"catalog_name": catalog_name, "quantity": 1}],
            "stock_items": [],
            "note": None,
        }, conversation)
        return customer, result

    def test_client_lead_create_tool_creates_customer_and_lead(self):
        customer, result = self._ai_catalog_lead()
        self.assertTrue(result["ok"])
        customer.refresh_from_db()
        self.assertEqual(customer.name, "Ahmad")
        self.assertEqual(customer.phone, "+998901234567")
        lead = Lead.objects.get(id=result["lead_id"])
        self.assertEqual(lead.request_uz, "Oq buket 1 dona, kelib olish")
        self.assertEqual(lead.source, "telegram")

    def test_client_lead_create_writes_ai_catalog_choice_into_details(self):
        item = AICatalogItem.objects.filter(name="Oq buket").first()
        _, result = self._ai_catalog_lead()
        lead = Lead.objects.get(id=result["lead_id"])
        row = lead.details["catalog_items"][0]
        self.assertEqual(row["catalog_name"], "Oq buket")
        self.assertEqual(row["quantity"], 1)
        self.assertEqual(row["price"], "500000.00")
        self.assertEqual(row["ai_catalog_item"], AICatalogItem.objects.get(name="Oq buket").id)
        # AI katalogi sklad katalogi emas — bog'lanish ochilmaydi, operator o'zi tanlaydi
        self.assertFalse(LeadCatalogUsage.objects.filter(lead=lead).exists())

    def test_client_lead_create_keeps_unmatched_catalog_name(self):
        _, result = self._ai_catalog_lead(catalog_name="Nomalum kompozitsiya")
        lead = Lead.objects.get(id=result["lead_id"])
        row = lead.details["catalog_items"][0]
        self.assertEqual(row["catalog_name"], "Nomalum kompozitsiya")
        self.assertIsNone(row["ai_catalog_item"])

    def test_recent_orders_show_ai_catalog_choice(self):
        customer, result = self._ai_catalog_lead()
        orders = recent_customer_orders(customer)
        self.assertEqual(orders[0]["catalog_items"][0]["name_uz"], "Oq buket")
        self.assertEqual(orders[0]["catalog_items"][0]["quantity"], 1)

    def test_ai_stock_rows_matches_long_price_query(self):
        rows = ai_stock_rows("Mondial oq atirgul narxi va mavjudlik 10 dona", limit=10)
        self.assertTrue(any(row["batch_id"] == self.batch.id for row in rows))

    def test_ai_stock_rows_excludes_variants_without_stock_from_general_list(self):
        flower = Flower.objects.create(name_uz="Gortenziya", slug="gortenziya")
        FlowerVariant.objects.create(flower=flower, name_uz="Snowball", color_uz="Oq")
        FlowerVariant.objects.create(flower=flower, name_uz="Limelight", color_uz="Yashil")
        rows = ai_stock_rows("gortenziya", limit=10)
        self.assertFalse(any(row["variant_uz"] in {"Snowball", "Limelight"} for row in rows))

    def test_ai_stock_rows_include_full_variant_display_names(self):
        flower = Flower.objects.create(name_uz="Gortenziya", slug="gortenziya-display")
        golland = FlowerVariant.objects.create(flower=flower, name_uz="Gortenziya Golland", color_uz="Moviy")
        kolumbiya = FlowerVariant.objects.create(flower=flower, name_uz="Gortenziya Kolumbiya", color_uz="Moviy")
        StockBatch.objects.create(variant=golland, batch_number="GOL-1", height_cm=50, stems_per_bunch=5, received_stems=20, remaining_stems=20, cost_per_stem=70000, sale_price_per_stem=105000, sale_price_per_bunch=500000)
        StockBatch.objects.create(variant=kolumbiya, batch_number="KOL-1", height_cm=50, stems_per_bunch=5, received_stems=20, remaining_stems=20, cost_per_stem=35000, sale_price_per_stem=60000, sale_price_per_bunch=285000)
        rows = ai_stock_rows("gortenziya kok", limit=10)
        rows_by_name = {row["display_name_uz"]: row for row in rows}
        self.assertIn("Gortenziya Golland Moviy", rows_by_name)
        self.assertIn("Gortenziya Kolumbiya Moviy", rows_by_name)
        self.assertEqual(rows_by_name["Gortenziya Golland Moviy"]["price_per_stem"], "105000.00")
        self.assertEqual(rows_by_name["Gortenziya Kolumbiya Moviy"]["price_per_stem"], "60000.00")
        self.assertTrue(all(row["color_uz"] == "Moviy" for row in rows if row["flower_uz"] == "Gortenziya"))

    def test_ai_flower_variant_rows_can_show_specific_variant_without_stock(self):
        flower = Flower.objects.create(name_uz="Gortenziya", slug="gortenziya")
        FlowerVariant.objects.create(flower=flower, name_uz="Snowball", color_uz="Oq")
        rows = ai_flower_variant_rows("gortenziya snow ball", limit=10)
        self.assertTrue(any(row["variant_uz"] == "Snowball" and row["active_stock"] == [] for row in rows))

    def test_pending_customer_reply_debounces_to_latest_message(self):
        customer = Customer.objects.create(instagram_user_id="ig-debounce", name="Ahmad", phone="+998901234567")
        conversation = Conversation.objects.create(customer=customer)
        first = conversation.messages.create(sender="customer", text="katalog")
        second = conversation.messages.create(sender="customer", text="rasmlari bormi")
        from unittest.mock import patch
        with patch("core.services.create_ai_reply_for_conversation", side_effect=lambda conv: Message.objects.create(conversation=conv, sender="ai", text="Javob")) as mocked:
            self.assertIsNone(process_pending_customer_reply(conversation.id, first.id))
            reply = process_pending_customer_reply(conversation.id, second.id)
            self.assertIsNotNone(reply)
            self.assertIsNone(process_pending_customer_reply(conversation.id, second.id))
        conversation.refresh_from_db()
        self.assertEqual(mocked.call_count, 1)
        self.assertEqual(conversation.ai_replied_to_message_id, second.id)

    def test_delayed_reply_waits_until_latest_message_is_old_enough(self):
        customer = Customer.objects.create(instagram_user_id="telegram:444", name="Ahmad", phone="+998901234567")
        conversation = Conversation.objects.create(customer=customer)
        message = conversation.messages.create(sender="customer", text="salom")
        from unittest.mock import patch
        with patch("core.tasks.process_delayed_telegram_reply.apply_async") as schedule_mock, patch("core.tasks.process_pending_customer_reply") as reply_mock:
            result = process_delayed_telegram_reply(conversation.id, message.id, "444")
        self.assertIsNone(result)
        schedule_mock.assert_called_once()
        reply_mock.assert_not_called()

    def test_pending_customer_reply_handles_empty_social_post(self):
        customer = Customer.objects.create(instagram_user_id="ig-no-post", name="Ahmad", phone="+998901234567")
        conversation = Conversation.objects.create(customer=customer, social_post=None)
        message = conversation.messages.create(sender="customer", text="salom")
        from unittest.mock import patch
        with patch("core.services.create_ai_reply_for_conversation", side_effect=lambda conv: Message.objects.create(conversation=conv, sender="ai", text="Javob")):
            reply = process_pending_customer_reply(conversation.id, message.id)
        self.assertIsNotNone(reply)
        conversation.refresh_from_db()
        self.assertEqual(conversation.ai_replied_to_message_id, message.id)

    def test_conversation_serializer_exposes_ai_active_and_clears_expired_pause(self):
        customer = Customer.objects.create(instagram_user_id="ig-pause")
        conversation = Conversation.objects.create(customer=customer, ai_paused_until=timezone.now() - timedelta(minutes=1), ai_pause_reason="operator_message")
        data = ConversationSerializer(conversation).data
        conversation.refresh_from_db()
        self.assertTrue(data["ai_is_active"])
        self.assertIsNone(data["ai_paused_until"])
        self.assertEqual(conversation.ai_pause_reason, "")

    def test_delayed_instagram_reply_does_not_show_typing_when_ai_paused(self):
        customer = Customer.objects.create(instagram_user_id="ig-paused")
        conversation = Conversation.objects.create(customer=customer, ai_paused_until=timezone.now() + timedelta(minutes=15), ai_pause_reason="operator_message")
        message = conversation.messages.create(sender="customer", text="salom")
        from unittest.mock import patch
        with patch("core.tasks.instagram_sender_action") as typing_mock, patch("core.tasks.process_pending_customer_reply") as reply_mock:
            result = process_delayed_instagram_reply(conversation.id, message.id, "ig-paused")
        self.assertIsNone(result)
        typing_mock.assert_not_called()
        reply_mock.assert_not_called()

    def test_delayed_telegram_reply_does_not_auto_send_catalog_image(self):
        self.item.status = "available"
        self.item.image_url = "https://example.com/oq-buket.jpg"
        self.item.save(update_fields=["status", "image_url", "updated_at"])
        customer = Customer.objects.create(instagram_user_id="telegram:123", name="Ahmad", phone="+998901234567")
        conversation = Conversation.objects.create(customer=customer)
        customer_message = conversation.messages.create(sender="customer", text="oq buket rasmi")
        Message.objects.filter(id=customer_message.id).update(created_at=timezone.now() - timedelta(seconds=8))
        reply_message = Message.objects.create(conversation=conversation, sender="ai", text="Mana rasmi", metadata={"catalog_items": [{"catalog_id": self.item.id, "quantity": 1}]})
        from unittest.mock import patch
        with patch("core.tasks.process_pending_customer_reply", return_value=reply_message), patch("core.tasks.telegram_sender_action", return_value={"ok": True}), patch("core.tasks.telegram_send", return_value={"ok": True}) as text_mock, patch("core.tasks.process_conversation_follow_up.apply_async") as follow_up_schedule:
            result = process_delayed_telegram_reply(conversation.id, customer_message.id, "555")
        self.assertEqual(result, reply_message.id)
        text_mock.assert_called_once_with("555", "Mana rasmi")
        follow_up_schedule.assert_called_once_with(args=[conversation.id, reply_message.id], countdown=AI_FOLLOW_UP_DELAY_SECONDS)

    def test_delayed_telegram_reply_skips_context_image_after_tool_image(self):
        customer = Customer.objects.create(instagram_user_id="telegram:123", name="Ahmad", phone="+998901234567")
        conversation = Conversation.objects.create(customer=customer)
        customer_message = conversation.messages.create(sender="customer", text="oq buket rasmi")
        Message.objects.filter(id=customer_message.id).update(created_at=timezone.now() - timedelta(seconds=8))
        reply_message = Message.objects.create(conversation=conversation, sender="ai", text="Rasmini yubordim.", metadata={"catalog_items": [{"catalog_id": self.item.id, "quantity": 1}], "image_tool_results": [{"image_sent": True, "catalog_id": self.item.id}]})
        from unittest.mock import patch
        with patch("core.tasks.process_pending_customer_reply", return_value=reply_message), patch("core.tasks.telegram_sender_action", return_value={"ok": True}), patch("core.tasks.telegram_send", return_value={"ok": True}), patch("core.tasks.process_conversation_follow_up.apply_async") as follow_up_schedule:
            result = process_delayed_telegram_reply(conversation.id, customer_message.id, "555")
        self.assertEqual(result, reply_message.id)
        follow_up_schedule.assert_called_once_with(args=[conversation.id, reply_message.id], countdown=AI_FOLLOW_UP_DELAY_SECONDS)

    def test_follow_up_waits_thirty_minutes_after_last_ai_message(self):
        customer = Customer.objects.create(instagram_user_id="ig-follow-up")
        conversation = Conversation.objects.create(customer=customer)
        customer_message = conversation.messages.create(sender="customer", text="katalog narxlari")
        ai_message = Message.objects.create(conversation=conversation, sender="ai", text="Oq buket\nNarxi 500 000 so'm")
        Message.objects.filter(id=customer_message.id).update(created_at=timezone.now() - timedelta(minutes=30))
        Message.objects.filter(id=ai_message.id).update(created_at=timezone.now() - timedelta(minutes=29))
        from unittest.mock import patch
        with patch("core.services.ai_follow_up_decision") as decision_mock:
            result = process_stalled_conversation_follow_up(conversation.id, ai_message.id)
        self.assertIsNone(result)
        decision_mock.assert_not_called()

    def test_follow_up_skips_when_lead_exists(self):
        customer = Customer.objects.create(instagram_user_id="ig-follow-up-lead", name="Ahmad", phone="+998901234567")
        conversation = Conversation.objects.create(customer=customer)
        customer_message = conversation.messages.create(sender="customer", text="oq buket olaman Ahmad 901234567")
        ai_message = Message.objects.create(conversation=conversation, sender="ai", text="Buyurtmangiz qabul qilindi", metadata={"lead_created_id": 10})
        Lead.objects.create(customer=customer, conversation=conversation, request_uz="Oq buket", arrangement_type="catalog")
        Message.objects.filter(id=customer_message.id).update(created_at=timezone.now() - timedelta(minutes=32))
        Message.objects.filter(id=ai_message.id).update(created_at=timezone.now() - timedelta(minutes=31))
        from unittest.mock import patch
        with patch("core.services.ai_follow_up_decision") as decision_mock:
            result = process_stalled_conversation_follow_up(conversation.id, ai_message.id)
        self.assertIsNone(result)
        decision_mock.assert_not_called()

    def test_follow_up_creates_ai_message_from_decision(self):
        customer = Customer.objects.create(instagram_user_id="ig-follow-up-create")
        conversation = Conversation.objects.create(customer=customer)
        customer_message = conversation.messages.create(sender="customer", text="shu savat nechpul")
        ai_message = Message.objects.create(conversation=conversation, sender="ai", text="Gortenziya Mix savat\nNarxi 900 000 so'm", metadata={"catalog_items": [{"catalog_name": "Gortenziya Mix", "quantity": 1}]})
        Message.objects.filter(id=customer_message.id).update(created_at=timezone.now() - timedelta(minutes=32))
        Message.objects.filter(id=ai_message.id).update(created_at=timezone.now() - timedelta(minutes=31))
        from unittest.mock import patch
        with patch("core.services.ai_follow_up_decision", return_value={"send_follow_up": True, "message": "Hurmatli mijoz, budjetingiz qancha edi?", "reason": "price_shown"}):
            follow_up = process_stalled_conversation_follow_up(conversation.id, ai_message.id)
        self.assertIsNotNone(follow_up)
        self.assertEqual(follow_up.sender, "ai")
        self.assertTrue(follow_up.metadata["follow_up"])
        self.assertIn("budjetingiz", follow_up.text)

    def test_follow_up_task_sends_instagram_message(self):
        customer = Customer.objects.create(instagram_user_id="ig-follow-up-send")
        conversation = Conversation.objects.create(customer=customer)
        customer_message = conversation.messages.create(sender="customer", text="katalog")
        ai_message = Message.objects.create(conversation=conversation, sender="ai", text="Narxi 500 000 so'm")
        Message.objects.filter(id=customer_message.id).update(created_at=timezone.now() - timedelta(minutes=32))
        Message.objects.filter(id=ai_message.id).update(created_at=timezone.now() - timedelta(minutes=31))
        from unittest.mock import patch
        with patch("core.services.ai_follow_up_decision", return_value={"send_follow_up": True, "message": "Budjetingiz qancha edi?", "reason": "stalled"}), patch("core.tasks.instagram_send", return_value={"ok": True}) as send_mock:
            result = process_conversation_follow_up(conversation.id, ai_message.id)
        self.assertIsNotNone(result)
        send_mock.assert_called_once_with("ig-follow-up-send", "Budjetingiz qancha edi?")

    def test_discount_negotiation_prompt_rule_requires_lead_creation(self):
        migration = importlib.import_module("core.migrations.0038_ai_prompt_discount_negotiation")
        rule = migration.DISCOUNT_NEGOTIATION_PROMPT_RULE
        self.assertIn("Arzonlashtirish va budjetga mos variant qoidasi", rule)
        self.assertIn("client_lead_create", rule)
        self.assertIn("Mijoz 200 000 so'mlik katalog buketni 150 000 so'mga so'radi", rule)

    def test_stock_image_prompt_rule_asks_quantity_before_date(self):
        migration = importlib.import_module("core.migrations.0041_ai_prompt_stock_image_and_discount_flow")
        rule = migration.STOCK_IMAGE_AND_DISCOUNT_FLOW_RULE
        self.assertIn("send_stock_image", rule)
        self.assertIn("sizga qachonga kerak edi", rule)
        self.assertIn("Shu guldan nechta dona qilib buket yoki savat yasaymiz", rule)
        self.assertIn("Gullarimiz hamyonbop narxlarda", rule)
        self.assertIn("client_lead_create", rule)

    def test_stock_list_and_calculation_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0042_ai_prompt_stock_list_and_calculation_flow")
        rule = migration.STOCK_LIST_AND_CALCULATION_FLOW_RULE
        self.assertIn("Qaysi turini ko'rgingiz keladi", rule)
        self.assertIn("Qaysi biridan buket yoki savat yasaymiz", rule)
        self.assertIn("javob juda qisqa bo'lsin", rule)
        self.assertIn("quantity_stems x sale_price_per_stem", rule)
        self.assertIn("150 000 + 150 000 + 50 000 = 350 000", rule)
        self.assertIn("Hech qachon shu holatni 550 000", rule)

    def test_public_reply_boundaries_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0043_ai_prompt_public_reply_boundaries")
        rule = migration.PUBLIC_REPLY_BOUNDARIES_RULE
        self.assertIn("lead, CRM, tizimga yozish", rule)
        self.assertIn("leadga qo'shsam bo'ladimi", rule)
        self.assertIn("Manzil javobining oxiriga", rule)
        self.assertIn("Rahmat, tez orada operatorlarimiz", rule)
        self.assertIn("Ko'pi bilan 3-5 qator", rule)
        self.assertIn("Jami taxminan 350 000 so'm", rule)

    def test_deterministic_price_tool_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0044_ai_prompt_deterministic_price_tool")
        rule = migration.DETERMINISTIC_PRICE_TOOL_RULE
        self.assertIn("Custom buket yoki savat narxini hech qachon o'zing hisoblama", rule)
        self.assertIn("calculate_custom_arrangement_price", rule)
        self.assertIn("quantity_stems 50", rule)
        self.assertIn("775 000", rule)
        self.assertIn("errors qaytarsa, narx aytma", rule)

    def test_natural_sales_flow_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0045_ai_prompt_natural_sales_flow")
        rule = migration.NATURAL_SALES_FLOW_RULE
        self.assertIn("qattiq shablon emas", rule)
        self.assertIn("50 dona guldan bitta buket", rule)
        self.assertIn("50 dona bitta buketmi yoki 50 ta buket kerakmi", rule)
        self.assertIn("darhol get_stock va calculate_custom_arrangement_price", rule)
        self.assertIn("Sizga qachonga kerak edi?", rule)
        self.assertIn("Yetkazib berish kerakmi yoki kelib olib ketasizmi", rule)
        self.assertIn("Tushunarli", rule)

    def test_clean_ai_prompt_replacement_contains_core_rules(self):
        migration = importlib.import_module("core.migrations.0046_replace_ai_prompt_clean_version")
        prompt = migration.CLEAN_EUROFLOWERS_AI_PROMPT
        self.assertIn("EUROFLOWERS PREMIUM gul do'konining Instagram va Telegramdagi AI sotuv menejerisan", prompt)
        self.assertIn("Custom narxni hech qachon o'zing hisoblama", prompt)
        self.assertIn("calculate_custom_arrangement_price", prompt)
        self.assertIn("lead_ready doim false", prompt)
        self.assertIn("REAL_CONTEXT_JSON.business", prompt)
        self.assertIn("shop_phone", prompt)
        self.assertNotIn("# EUROFLOWERS PREMIUM", prompt)

    def test_legacy_ai_prompt_restore_reverts_clean_prompt(self):
        migration = importlib.import_module("core.migrations.0047_restore_legacy_ai_prompt")
        prompt = migration.LEGACY_EUROFLOWERS_AI_PROMPT
        self.assertIn("Sen EuroFlowers Premium gul do‘konining Instagram va Telegramdagi AI sotuvchisian", prompt)
        self.assertIn("Deterministik custom narx hisoblash qoidasi", prompt)
        self.assertIn("Tabiiy sotuv muloqoti qoidasi", prompt)
        self.assertIn("calculate_custom_arrangement_price", prompt)
        self.assertNotIn("EUROFLOWERS PREMIUM gul do'konining Instagram va Telegramdagi AI sotuv menejerisan", prompt)

    def test_quality_reassurance_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0048_ai_prompt_quality_reassurance")
        rule = migration.QUALITY_REASSURANCE_PROMPT_RULE
        self.assertIn("Sifat, yangi gul va obyom bo'yicha aniq javob qoidasi", rule)
        self.assertIn("Ko'nglingiz xotirjam bo'lsin", rule)
        self.assertIn("so'lib qolgan gullar bilan hech qachon buket yoki savat yasalmaydi", rule)
        self.assertIn("o'zingdan generate qilma", rule)

    def test_stock_image_tool_required_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0049_ai_prompt_stock_image_tool_required")
        rule = migration.STOCK_IMAGE_TOOL_REQUIRED_RULE
        self.assertIn("Sklad rasmi bo'yicha qat'iy qoida", rule)
        self.assertIn("rasmni ko'rmoqchimisiz", rule)
        self.assertIn("send_stock_image", rule)
        self.assertIn("Tool chaqirmasdan hech qachon", rule)

    def test_stock_language_pickup_prompt_rule(self):
        migration = importlib.import_module("core.migrations.0050_ai_prompt_stock_language_pickup_rules")
        rule = migration.STOCK_LANGUAGE_PICKUP_PROMPT_RULE
        self.assertIn("Stock mavjudlik, til va pickup aniqligi", rule)
        self.assertIn("majburiy get_stock chaqir", rule)
        self.assertIn("display_name_uz_cyril", rule)
        self.assertIn("vitrinada yo'q", rule)
        self.assertIn("qayd etildi", rule)

    def test_telegram_update_creates_conversation_message_once(self):
        SocialPost.objects.create(post_type="post", title_uz="Test post", title_ru="Test post", is_active=True)
        payload = {
            "update_id": 1001,
            "message": {
                "message_id": 77,
                "text": "Assalomu alaykum",
                "chat": {"id": 555},
                "from": {"id": 999, "first_name": "Ali", "last_name": "Valiyev"},
            },
        }
        jobs = resolve_telegram_update(payload)
        self.assertEqual(len(jobs), 1)
        customer = Customer.objects.get(instagram_user_id="telegram:999")
        self.assertEqual(customer.name, "")
        self.assertTrue(Conversation.objects.filter(customer=customer).exists())
        self.assertTrue(Message.objects.filter(conversation__customer=customer, text="Assalomu alaykum", instagram_message_id="telegram:555:77").exists())
        self.assertEqual(resolve_telegram_update(payload), [])

    def test_instagram_media_links_are_saved_in_chat_message(self):
        SocialPost.objects.create(post_type="story", title_uz="Story", title_ru="Story", is_active=True)
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-user-1"},
                    "recipient": {"id": "ig-business"},
                    "message": {
                        "mid": "mid-link-1",
                        "text": "shu qancha",
                        "attachments": [{
                            "type": "ig_story",
                            "payload": {
                                "story_media_id": "story-1",
                                "story_media_url": "https://lookaside.fbsbx.com/ig_messaging_cdn/story.jpg",
                            },
                        }],
                    },
                }],
            }],
        }
        jobs = resolve_instagram_event(payload)
        self.assertEqual(len(jobs), 1)
        message = Message.objects.get(instagram_message_id="mid-link-1")
        self.assertIn("Story link: https://lookaside.fbsbx.com/ig_messaging_cdn/story.jpg", message.text)
        self.assertEqual(message.metadata["attachments"][0]["kind"], "story")
        self.assertEqual(message.metadata["attachments"][0]["url"], "https://lookaside.fbsbx.com/ig_messaging_cdn/story.jpg")

    def test_instagram_echo_message_is_saved_as_operator_message(self):
        IntegrationSettings.objects.update_or_create(pk=1, defaults={"instagram_account_id": "ig-business", "instagram_business_id": "ig-business"})
        customer = Customer.objects.create(instagram_user_id="ig-user-echo")
        Conversation.objects.create(customer=customer)
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-business"},
                    "recipient": {"id": "ig-user-echo"},
                    "message": {"mid": "mid-echo-1", "text": "Operator javobi", "is_echo": True},
                }],
            }],
        }
        jobs = resolve_instagram_event(payload)
        self.assertEqual(jobs, [])
        conversation = Conversation.objects.get(customer=customer)
        message = Message.objects.get(instagram_message_id="mid-echo-1")
        self.assertEqual(message.sender, "operator")
        self.assertEqual(message.text, "Operator javobi")
        self.assertEqual(conversation.status, "operator")
        self.assertEqual(conversation.ai_pause_reason, "instagram_operator_message")
        self.assertGreater(conversation.ai_paused_until, timezone.now())

    def test_instagram_echo_for_backend_sent_message_is_ignored(self):
        IntegrationSettings.objects.update_or_create(pk=1, defaults={"instagram_account_id": "ig-business", "instagram_business_id": "ig-business"})
        customer = Customer.objects.create(instagram_user_id="ig-user-backend-echo")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="ai", text="AI javobi", instagram_message_id="mid-backend-1")
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-business"},
                    "recipient": {"id": "ig-user-backend-echo"},
                    "message": {"mid": "mid-backend-1", "text": "AI javobi", "is_echo": True},
                }],
            }],
        }
        jobs = resolve_instagram_event(payload)
        self.assertEqual(jobs, [])
        self.assertEqual(conversation.messages.count(), 1)
        conversation.refresh_from_db()
        self.assertEqual(conversation.status, "ai")
        self.assertIsNone(conversation.ai_paused_until)

    def test_global_ai_inactive_blocks_delayed_reply(self):
        from core.services import should_start_ai_reply
        AISettings.objects.update_or_create(pk=1, defaults={"is_active": False})
        customer = Customer.objects.create(instagram_user_id="ig-global-off")
        conversation = Conversation.objects.create(customer=customer)
        message = conversation.messages.create(sender="customer", text="salom")
        self.assertFalse(should_start_ai_reply(conversation.id, message.id))

    def test_instagram_story_reply_links_catalog_by_active_story_asset_url(self):
        post = SocialPost.objects.create(
            post_type="story",
            media_id="story-share-3946136376066774555",
            permalink="https://www.instagram.com/stories/extra_teest/3946136376066774555/",
            story_share_id="3946136376066774555",
            webhook_story_id="17900000000000000",
            title_uz="Pion buketi",
            title_ru="Pion bouquet",
            price=800000,
            is_active=True,
        )
        CatalogItem.objects.create(
            social_post=post,
            name_uz="Pion buketi",
            arrangement_type="bouquet",
            price=800000,
            status="available",
        )
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-user-story-asset"},
                    "recipient": {"id": "ig-business"},
                    "message": {
                        "mid": "mid-story-asset-1",
                        "text": "shu nechpul",
                        "reply_to": {
                            "story": {
                                "url": "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=18151925590500461&signature=test",
                            }
                        },
                    },
                }],
            }],
        }
        from unittest.mock import patch

        with patch("core.webhook_services.find_active_story_by_media_url", return_value={
            "id": "18151925590500461",
            "media_url": "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=18151925590500461&signature=live",
            "permalink": "https://www.instagram.com/stories/extra_teest/3946136376066774555/",
        }):
            jobs = resolve_instagram_event(payload)
        self.assertEqual(len(jobs), 1)
        conversation = Conversation.objects.get(customer__instagram_user_id="ig-user-story-asset")
        message = Message.objects.get(instagram_message_id="mid-story-asset-1")
        post.refresh_from_db()
        self.assertEqual(conversation.social_post_id, post.id)
        self.assertIn("18151925590500461", post.webhook_story_id)
        self.assertNotIn("bazadagi story/post/reel katalogiga bog‘lanmagan", message.text)

    def test_instagram_story_reply_creates_social_post_from_catalog_story_url(self):
        item = CatalogItem.objects.create(
            name_uz="Pushti atirgul buketi",
            arrangement_type="bouquet",
            price=500000,
            status="available",
            instagram_story_url="https://www.instagram.com/stories/extra_teest/3948457236253594433/",
        )
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-user-catalog-story"},
                    "recipient": {"id": "ig-business"},
                    "message": {
                        "mid": "mid-catalog-story-1",
                        "text": "narxi qancha",
                        "attachments": [{
                            "type": "ig_story",
                            "payload": {
                                "story_media_id": "18151925590500461",
                                "story_media_url": "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=18151925590500461&signature=test",
                            },
                        }],
                    },
                }],
            }],
        }
        from unittest.mock import patch

        with patch("core.webhook_services.find_active_story_by_media_url", return_value={
            "id": "18151925590500461",
            "media_url": "https://scontent.example/story.jpg",
            "permalink": "https://www.instagram.com/stories/extra_teest/3948457236253594433/",
        }):
            jobs = resolve_instagram_event(payload)
        self.assertEqual(len(jobs), 1)
        item.refresh_from_db()
        conversation = Conversation.objects.get(customer__instagram_user_id="ig-user-catalog-story")
        message = Message.objects.get(instagram_message_id="mid-catalog-story-1")
        self.assertIsNotNone(item.social_post_id)
        self.assertEqual(conversation.social_post_id, item.social_post_id)
        self.assertNotIn("bazadagi story/post/reel katalogiga bog‘lanmagan", message.text)

    def test_instagram_lookaside_base_url_does_not_match_other_story(self):
        SocialPost.objects.create(
            post_type="story",
            media_id="18151925590500461",
            permalink="https://www.instagram.com/stories/extra_teest/3948457236253594433/",
            story_share_id="3948457236253594433",
            webhook_story_id="18151925590500461",
            webhook_story_url="https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=18151925590500461&signature=old",
            title_uz="Gortenziya Mix",
            title_ru="Hydrangea Mix",
            price=1500000,
            is_active=True,
        )
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-user-other-story"},
                    "recipient": {"id": "ig-business"},
                    "message": {
                        "mid": "mid-other-story-1",
                        "text": "buni narxi qancha",
                        "attachments": [{
                            "type": "ig_story",
                            "payload": {
                                "story_media_id": "18090487133179534",
                                "story_media_url": "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=18090487133179534&signature=new",
                            },
                        }],
                    },
                }],
            }],
        }
        from unittest.mock import patch

        with patch("core.webhook_services.find_active_story_by_media_url", return_value=None):
            jobs = resolve_instagram_event(payload)
        self.assertEqual(len(jobs), 1)
        conversation = Conversation.objects.get(customer__instagram_user_id="ig-user-other-story")
        message = Message.objects.get(instagram_message_id="mid-other-story-1")
        self.assertIsNone(conversation.social_post_id)
        self.assertIn("bazadagi story/post/reel katalogiga bog‘lanmagan", message.text)

    def test_unknown_instagram_media_clears_previous_post_context(self):
        old_post = SocialPost.objects.create(post_type="reel", media_id="old-pion", title_uz="Pion buket", title_ru="Pion", is_active=True)
        customer = Customer.objects.create(instagram_user_id="ig-user-2")
        conversation = Conversation.objects.create(customer=customer, social_post=old_post)
        payload = {
            "entry": [{
                "messaging": [{
                    "sender": {"id": "ig-user-2"},
                    "recipient": {"id": "ig-business"},
                    "message": {
                        "mid": "mid-link-2",
                        "text": "shu qancha",
                        "attachments": [{
                            "type": "share",
                            "payload": {"url": "https://www.instagram.com/reel/UNKNOWN123/"},
                        }],
                    },
                }],
            }],
        }
        jobs = resolve_instagram_event(payload)
        self.assertEqual(len(jobs), 1)
        conversation.refresh_from_db()
        message = Message.objects.get(instagram_message_id="mid-link-2")
        self.assertIsNone(conversation.social_post)
        self.assertIn("bazadagi story/post/reel katalogiga bog‘lanmagan", message.text)

    def test_telegram_voice_link_is_saved_in_chat_message(self):
        SocialPost.objects.create(post_type="post", title_uz="Test post", title_ru="Test post", is_active=True)
        IntegrationSettings.objects.update_or_create(pk=1, defaults={"telegram_bot_token": "test-token"})
        payload = {
            "update_id": 1002,
            "message": {
                "message_id": 78,
                "chat": {"id": 555},
                "from": {"id": 1000, "first_name": "Ali"},
                "voice": {"file_id": "voice-file-id"},
            },
        }
        from unittest.mock import patch

        class MockResponse:
            def raise_for_status(self):
                return None

            def json(self):
                return {"ok": True, "result": {"file_path": "voice/file.ogg"}}

        with patch("core.platform_services.requests.post", return_value=MockResponse()):
            jobs = resolve_telegram_update(payload)
        self.assertEqual(len(jobs), 1)
        message = Message.objects.get(instagram_message_id="telegram:555:78")
        self.assertIn("Voice message: https://api.telegram.org/file/bottest-token/voice/file.ogg", message.text)
        self.assertEqual(message.metadata["attachments"][0]["kind"], "voice")


class CatalogReworkTests(TestCase):
    """Restavratsiya: katalogni buzib, undan yangi mahsulot yasash."""

    def setUp(self):
        self.user = User.objects.create_user("rework-admin", password="password")
        florist_user = User.objects.create_user("florist-dilnoza", password="password", first_name="Dilnoza")
        self.florist = FloristProfile.objects.create(user=florist_user)
        flower = Flower.objects.create(name_uz="Atirgul", slug="rose-rework")
        self.variant = FlowerVariant.objects.create(flower=flower, name_uz="Mondial", color_uz="Oq")
        self.batch = StockBatch.objects.create(
            variant=self.variant, batch_number="RW-1", height_cm=60, stems_per_bunch=25,
            received_stems=500, remaining_stems=500, cost_per_stem=10000,
            sale_price_per_stem=15000, sale_price_per_bunch=375000,
        )

    def make_item(self, name, stems, quantity=1, price=500000):
        item = CatalogItem.objects.create(
            name_uz=name, arrangement_type="bouquet", price=price, status="available",
            quantity_total=quantity, quantity_stock_deducted=quantity, florist_fee=0,
        )
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=stems)
        return sync_catalog_financials(item)

    def test_small_plus_stock_becomes_one_big_bouquet(self):
        """1-holat: kichkina buket + skladdan gul = 1 ta katta buket."""
        small = self.make_item("Kichkina buket", 25)
        before = self.batch.remaining_stems
        rework = create_catalog_rework(
            florist=self.florist, florist_amount=Decimal("60000"),
            sources=[{"catalog_item": small, "quantity": 1}],
            stock_inputs=[{"stock_batch": self.batch, "quantity_stems": 25}],
            outputs=[{
                "name_uz": "Katta buket", "arrangement_type": "bouquet", "quantity": 1,
                "price": Decimal("900000"),
                "composition": [{"stock_batch": self.batch, "quantity_stems": 50}],
            }],
            user=self.user,
        )
        self.batch.refresh_from_db()
        small.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 25)
        self.assertEqual(small.quantity_reworked, 1)
        self.assertEqual(small.status, "archived")
        self.assertEqual(rework.input_stems, 50)
        self.assertEqual(rework.output_stems, 50)
        self.assertEqual(rework.waste_stems, 0)
        output = rework.outputs.get()
        self.assertEqual(output.catalog_item.quantity_total, 1)
        self.assertEqual(output.catalog_item.composition.get().quantity_stems, 50)

    def test_one_big_becomes_two_medium_and_three_small(self):
        """2-holat: 1 ta katta + skladdan gul = 2 ta o'rtancha + 3 ta kichkina."""
        big = self.make_item("Katta buket", 60, price=1000000)
        before = self.batch.remaining_stems
        rework = create_catalog_rework(
            florist=self.florist, florist_amount=Decimal("150000"),
            sources=[{"catalog_item": big, "quantity": 1}],
            stock_inputs=[{"stock_batch": self.batch, "quantity_stems": 40}],
            outputs=[
                {"name_uz": "O'rtancha", "arrangement_type": "bouquet", "quantity": 2,
                 "price": Decimal("450000"),
                 "composition": [{"stock_batch": self.batch, "quantity_stems": 25}]},
                {"name_uz": "Kichkina", "arrangement_type": "bouquet", "quantity": 3,
                 "price": Decimal("280000"),
                 "composition": [{"stock_batch": self.batch, "quantity_stems": 15}]},
            ],
            user=self.user,
        )
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 40)
        self.assertEqual(rework.input_stems, 100)
        self.assertEqual(rework.output_stems, 95)
        self.assertEqual(rework.waste_stems, 5)
        self.assertEqual(rework.outputs.count(), 2)
        self.assertEqual(sum(row.catalog_item.quantity_total for row in rework.outputs.all()), 5)
        # florist haqi gul soniga proporsional taqsimlanadi va yig'indisi to'liq mos keladi
        self.assertEqual(sum(row.allocated_florist_amount for row in rework.outputs.all()), Decimal("150000"))

    def test_florist_salary_entry_is_created_manually(self):
        item = self.make_item("Buket", 20)
        rework = create_catalog_rework(
            florist=self.florist, florist_amount=Decimal("75000"),
            sources=[{"catalog_item": item, "quantity": 1}], stock_inputs=[],
            outputs=[{"name_uz": "Yangi", "arrangement_type": "bouquet", "quantity": 1,
                      "price": Decimal("400000"),
                      "composition": [{"stock_batch": self.batch, "quantity_stems": 20}]}],
            user=self.user,
        )
        entry = FloristSalaryEntry.objects.get(rework=rework)
        self.assertEqual(entry.amount, Decimal("75000"))
        self.assertEqual(entry.source, "rework")
        self.assertEqual(entry.florist, self.florist)

    def test_output_flowers_cannot_exceed_available_pool(self):
        item = self.make_item("Buket", 20)
        with self.assertRaises(ValueError):
            create_catalog_rework(
                florist=self.florist, florist_amount=0,
                sources=[{"catalog_item": item, "quantity": 1}], stock_inputs=[],
                outputs=[{"name_uz": "Katta", "arrangement_type": "bouquet", "quantity": 1,
                          "price": Decimal("400000"),
                          "composition": [{"stock_batch": self.batch, "quantity_stems": 50}]}],
                user=self.user,
            )

    def test_cannot_break_more_units_than_remaining(self):
        item = self.make_item("Buket", 10, quantity=2)
        with self.assertRaises(ValueError):
            create_catalog_rework(
                florist=self.florist, florist_amount=0,
                sources=[{"catalog_item": item, "quantity": 3}], stock_inputs=[],
                outputs=[{"name_uz": "Yangi", "arrangement_type": "bouquet", "quantity": 1,
                          "price": Decimal("100000"),
                          "composition": [{"stock_batch": self.batch, "quantity_stems": 10}]}],
                user=self.user,
            )

    def test_partial_break_keeps_item_available(self):
        item = self.make_item("Buket", 10, quantity=3)
        create_catalog_rework(
            florist=self.florist, florist_amount=Decimal("20000"),
            sources=[{"catalog_item": item, "quantity": 1}], stock_inputs=[],
            outputs=[{"name_uz": "Yangi", "arrangement_type": "bouquet", "quantity": 1,
                      "price": Decimal("150000"),
                      "composition": [{"stock_batch": self.batch, "quantity_stems": 10}]}],
            user=self.user,
        )
        item.refresh_from_db()
        self.assertEqual(item.quantity_reworked, 1)
        self.assertEqual(item.status, "available")
        self.assertEqual(catalog_remaining(item), 2)

    def test_multiple_sources_merge_into_one_output(self):
        a = self.make_item("Kichkina A", 20)
        b = self.make_item("Kichkina B", 20)
        rework = create_catalog_rework(
            florist=self.florist, florist_amount=Decimal("50000"),
            sources=[{"catalog_item": a, "quantity": 1}, {"catalog_item": b, "quantity": 1}],
            stock_inputs=[],
            outputs=[{"name_uz": "Katta", "arrangement_type": "bouquet", "quantity": 1,
                      "price": Decimal("800000"),
                      "composition": [{"stock_batch": self.batch, "quantity_stems": 40}]}],
            user=self.user,
        )
        self.assertEqual(rework.sources.count(), 2)
        self.assertEqual(rework.input_stems, 40)
        self.assertEqual(rework.waste_stems, 0)

    def test_reworked_item_is_hidden_from_sale_but_kept_in_history(self):
        item = self.make_item("Buket", 20)
        rework = create_catalog_rework(
            florist=self.florist, florist_amount=0,
            sources=[{"catalog_item": item, "quantity": 1}], stock_inputs=[],
            outputs=[{"name_uz": "Yangi", "arrangement_type": "bouquet", "quantity": 1,
                      "price": Decimal("300000"),
                      "composition": [{"stock_batch": self.batch, "quantity_stems": 20}]}],
            user=self.user,
        )
        item.refresh_from_db()
        self.assertNotIn(item.id, list(available_catalog_queryset().values_list("id", flat=True)))
        self.assertIn(item.id, list(rework.sources.values_list("catalog_item_id", flat=True)))
        self.assertTrue(item.history.filter(action="reworked").exists())

    def test_output_stock_is_not_deducted_twice(self):
        item = self.make_item("Buket", 20)
        before = self.batch.remaining_stems
        rework = create_catalog_rework(
            florist=self.florist, florist_amount=0,
            sources=[{"catalog_item": item, "quantity": 1}], stock_inputs=[],
            outputs=[{"name_uz": "Yangi", "arrangement_type": "bouquet", "quantity": 1,
                      "price": Decimal("300000"),
                      "composition": [{"stock_batch": self.batch, "quantity_stems": 20}]}],
            user=self.user,
        )
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before)
        new_item = rework.outputs.get().catalog_item
        self.assertEqual(new_item.quantity_stock_deducted, new_item.quantity_total)
        with self.assertRaises(ValueError):
            deduct_catalog_stock(new_item, self.user)


    def test_api_creates_rework_and_lists_history(self):
        UserProfile.objects.update_or_create(user=self.user, defaults={"role": "admin"})
        for page, _ in PagePermission.PAGE_CHOICES:
            PagePermission.objects.update_or_create(user=self.user, page=page, defaults={"can_view": True, "can_control": True})
        item = self.make_item("Katta buket", 60, price=1000000)
        client = APIClient()
        client.force_authenticate(self.user)
        payload = {
            "florist": self.florist.id,
            "florist_amount": "150000",
            "note": "Vitrinadagi buket buzildi",
            "sources": [{"catalog_item": item.id, "quantity": 1}],
            "stock_inputs": [{"stock_batch": self.batch.id, "quantity_stems": 40}],
            "outputs": [
                {"name_uz": "O'rtancha", "arrangement_type": "bouquet", "quantity": 2, "price": "450000",
                 "composition": [{"stock_batch": self.batch.id, "quantity_stems": 25}]},
                {"name_uz": "Kichkina", "arrangement_type": "bouquet", "quantity": 3, "price": "280000",
                 "composition": [{"stock_batch": self.batch.id, "quantity_stems": 15}]},
            ],
        }
        response = client.post("/api/catalog-reworks/", payload, format="json")
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["waste_stems"], 5)
        self.assertEqual(len(response.data["outputs"]), 2)
        listing = client.get("/api/catalog-reworks/")
        self.assertEqual(listing.status_code, 200)
        self.assertEqual(listing.data["count"], 1)

    def test_api_rejects_missing_flowers(self):
        UserProfile.objects.update_or_create(user=self.user, defaults={"role": "admin"})
        for page, _ in PagePermission.PAGE_CHOICES:
            PagePermission.objects.update_or_create(user=self.user, page=page, defaults={"can_view": True, "can_control": True})
        item = self.make_item("Buket", 20)
        client = APIClient()
        client.force_authenticate(self.user)
        response = client.post("/api/catalog-reworks/", {
            "florist": self.florist.id, "florist_amount": "0",
            "sources": [{"catalog_item": item.id, "quantity": 1}],
            "outputs": [{"name_uz": "Katta", "arrangement_type": "bouquet", "quantity": 1, "price": "400000",
                         "composition": [{"stock_batch": self.batch.id, "quantity_stems": 50}]}],
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("yetmayapti", response.data["detail"])


class SupplierDateFilterTests(TestCase):
    """Postavshik hisobini sana oralig'i bo'yicha filtrlash."""

    def setUp(self):
        self.user = User.objects.create_user("supplier-admin", password="password")
        UserProfile.objects.update_or_create(user=self.user, defaults={"role": "admin"})
        for page, _ in PagePermission.PAGE_CHOICES:
            PagePermission.objects.update_or_create(user=self.user, page=page, defaults={"can_view": True, "can_control": True})
        self.supplier = Supplier.objects.create(name="Davron Aka")
        flower = Flower.objects.create(name_uz="Atirgul", slug="rose-supplier")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Prut", color_uz="Oq")
        # 04.08 — 100 dona × 1000 = 100 000
        StockBatch.objects.create(
            variant=variant, supplier=self.supplier, batch_number="S-1", received_at="2026-08-04",
            height_cm=50, stems_per_bunch=25, received_stems=100, remaining_stems=100,
            cost_per_stem=1000, sale_price_per_stem=3000, sale_price_per_bunch=75000,
        )
        # 05.08 — 50 dona × 2000 = 100 000
        StockBatch.objects.create(
            variant=variant, supplier=self.supplier, batch_number="S-2", received_at="2026-08-05",
            height_cm=50, stems_per_bunch=25, received_stems=50, remaining_stems=50,
            cost_per_stem=2000, sale_price_per_stem=4000, sale_price_per_bunch=100000,
        )
        SupplierPayment.objects.create(supplier=self.supplier, amount=Decimal("60000"), paid_at="2026-08-04")
        SupplierPayment.objects.create(supplier=self.supplier, amount=Decimal("40000"), paid_at="2026-08-05")
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def row(self, params=""):
        response = self.client.get(f"/api/suppliers/{params}")
        self.assertEqual(response.status_code, 200, response.data)
        return next(r for r in response.data["results"] if r["id"] == self.supplier.id)

    def test_without_filter_counts_everything(self):
        row = self.row()
        self.assertEqual(Decimal(row["purchase_total"]), Decimal("200000.00"))
        self.assertEqual(row["batches_count"], 2)
        self.assertEqual(row["total_received_stems"], 150)
        self.assertEqual(Decimal(row["paid_total"]), Decimal("100000.00"))

    def test_single_day_filter(self):
        row = self.row("?date_from=2026-08-04&date_to=2026-08-04")
        self.assertEqual(Decimal(row["purchase_total"]), Decimal("100000.00"))
        self.assertEqual(row["batches_count"], 1)
        self.assertEqual(row["total_received_stems"], 100)
        self.assertEqual(Decimal(row["paid_total"]), Decimal("60000.00"))

    def test_date_from_only(self):
        row = self.row("?date_from=2026-08-05")
        self.assertEqual(Decimal(row["purchase_total"]), Decimal("100000.00"))
        self.assertEqual(row["total_received_stems"], 50)

    def test_date_to_only(self):
        row = self.row("?date_to=2026-08-04")
        self.assertEqual(Decimal(row["purchase_total"]), Decimal("100000.00"))
        self.assertEqual(row["total_received_stems"], 100)

    def test_range_with_no_data_returns_zero(self):
        row = self.row("?date_from=2026-09-01&date_to=2026-09-30")
        self.assertEqual(Decimal(row["purchase_total"]), Decimal("0.00"))
        self.assertEqual(row["batches_count"], 0)
        self.assertEqual(row["total_received_stems"], 0)

    def test_last_payment_stays_global(self):
        row = self.row("?date_from=2026-08-04&date_to=2026-08-04")
        self.assertEqual(str(row["last_payment_at"]), "2026-08-05")

    def test_bad_date_returns_400(self):
        response = self.client.get("/api/suppliers/?date_from=04.08.2026")
        self.assertEqual(response.status_code, 400)
        self.assertIn("date_from", response.data)

    def test_reversed_range_returns_400(self):
        response = self.client.get("/api/suppliers/?date_from=2026-08-05&date_to=2026-08-04")
        self.assertEqual(response.status_code, 400)

    def test_detail_endpoint_respects_filter(self):
        response = self.client.get(f"/api/suppliers/{self.supplier.id}/?date_from=2026-08-04&date_to=2026-08-04")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Decimal(response.data["purchase_total"]), Decimal("100000.00"))


class FloristBulkCloseTests(TestCase):
    """Tanlangan chiqimlarni birga yopish."""

    def setUp(self):
        self.user = User.objects.create_user("bulk-admin", password="password")
        UserProfile.objects.update_or_create(user=self.user, defaults={"role": "admin"})
        for page, _ in PagePermission.PAGE_CHOICES:
            PagePermission.objects.update_or_create(user=self.user, page=page, defaults={"can_view": True, "can_control": True})
        flower = Flower.objects.create(name_uz="Atirgul", slug="rose-bulk")
        self.variant = FlowerVariant.objects.create(flower=flower, name_uz="Prut", color_uz="Oq")
        self.b1 = self.make_batch("B-1")
        self.b2 = self.make_batch("B-2")
        self.b3 = self.make_batch("B-3")
        self.f1 = self.make_florist("bulk-f1", "Abror")
        self.f2 = self.make_florist("bulk-f2", "Bekzod")
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def make_batch(self, number):
        return StockBatch.objects.create(
            variant=self.variant, batch_number=number, height_cm=50, stems_per_bunch=25,
            received_stems=300, remaining_stems=300, cost_per_stem=1000,
            sale_price_per_stem=3000, sale_price_per_bunch=75000,
        )

    def make_florist(self, username, name):
        user = User.objects.create_user(username, password="password", first_name=name)
        return FloristProfile.objects.create(user=user, staff_type="florist")

    def issue(self, florist, batch, stems):
        issue_stock_to_florist(florist, batch, stems, "test", self.user)

    def balance(self, florist, batch):
        row = FloristStockBalance.objects.filter(florist=florist, batch=batch).first()
        return row.remaining_stems if row else 0

    def test_closes_only_selected_issues(self):
        self.issue(self.f1, self.b1, 40)
        self.issue(self.f1, self.b2, 30)
        self.issue(self.f1, self.b3, 20)
        result = close_selected_florist_issues(
            [{"florist": self.f1, "batch": self.b1, "return_stems": 40},
             {"florist": self.f1, "batch": self.b2, "return_stems": 30}],
            self.user,
        )
        self.assertEqual(result["closed_batches"], 2)
        self.assertEqual(self.balance(self.f1, self.b1), 0)
        self.assertEqual(self.balance(self.f1, self.b2), 0)
        # tanlanmagani tegilmaydi
        self.assertEqual(self.balance(self.f1, self.b3), 20)

    def test_closes_across_multiple_florists(self):
        self.issue(self.f1, self.b1, 25)
        self.issue(self.f2, self.b2, 35)
        result = close_selected_florist_issues(
            [{"florist": self.f1, "batch": self.b1, "return_stems": 25},
             {"florist": self.f2, "batch": self.b2, "return_stems": 35}],
            self.user,
        )
        self.assertEqual(result["closed_batches"], 2)
        self.assertEqual(len(result["florists"]), 2)
        self.assertEqual(result["returned_stems"], 60)
        self.assertEqual(self.balance(self.f1, self.b1), 0)
        self.assertEqual(self.balance(self.f2, self.b2), 0)

    def test_returned_stems_go_back_to_stock(self):
        self.issue(self.f1, self.b1, 40)
        before = StockBatch.objects.get(pk=self.b1.pk).remaining_stems
        close_selected_florist_issues(
            [{"florist": self.f1, "batch": self.b1, "return_stems": 40}], self.user,
        )
        self.assertEqual(StockBatch.objects.get(pk=self.b1.pk).remaining_stems, before + 40)

    def test_one_bad_row_rolls_back_everything(self):
        self.issue(self.f1, self.b1, 30)
        # b2 chiqarilmagan — yopib bo'lmaydi
        with self.assertRaises(ValueError):
            close_selected_florist_issues(
                [{"florist": self.f1, "batch": self.b1, "return_stems": 30},
                 {"florist": self.f1, "batch": self.b2, "return_stems": 10}],
                self.user,
            )
        # birinchisi ham yopilmagan bo'lishi kerak
        self.assertEqual(self.balance(self.f1, self.b1), 30)

    def test_duplicate_selection_is_rejected(self):
        self.issue(self.f1, self.b1, 20)
        with self.assertRaises(ValueError):
            close_selected_florist_issues(
                [{"florist": self.f1, "batch": self.b1, "return_stems": 10},
                 {"florist": self.f1, "batch": self.b1, "return_stems": 10}],
                self.user,
            )

    def test_empty_selection_is_rejected(self):
        with self.assertRaises(ValueError):
            close_selected_florist_issues([], self.user)

    def test_api_closes_selected(self):
        self.issue(self.f1, self.b1, 40)
        self.issue(self.f2, self.b2, 20)
        self.issue(self.f1, self.b3, 15)
        response = self.client.post("/api/florist-stock-balances/close-issues/", {
            "items": [
                {"florist": self.f1.id, "batch": self.b1.id, "return_stems": 40},
                {"florist": self.f2.id, "batch": self.b2.id, "return_stems": 20},
            ],
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["closed_batches"], 2)
        self.assertEqual(self.balance(self.f1, self.b3), 15)

    def test_api_preview_changes_nothing(self):
        self.issue(self.f1, self.b1, 40)
        response = self.client.post("/api/florist-stock-balances/close-issues-preview/", {
            "items": [{"florist": self.f1.id, "batch": self.b1.id, "return_stems": 40}],
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["selected"], 1)
        self.assertEqual(self.balance(self.f1, self.b1), 40)

    def test_api_rejects_duplicate_rows(self):
        self.issue(self.f1, self.b1, 20)
        response = self.client.post("/api/florist-stock-balances/close-issues/", {
            "items": [
                {"florist": self.f1.id, "batch": self.b1.id, "return_stems": 10},
                {"florist": self.f1.id, "batch": self.b1.id, "return_stems": 10},
            ],
        }, format="json")
        self.assertEqual(response.status_code, 400)

    def test_api_rejects_empty_items(self):
        response = self.client.post("/api/florist-stock-balances/close-issues/", {"items": []}, format="json")
        self.assertEqual(response.status_code, 400)


class ApiTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("admin", password="password", is_superuser=True, is_staff=True)
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        flower = Flower.objects.create(name_uz="Atirgul API", slug="rose-api")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Freedom", color_uz="Qizil")
        self.batch = StockBatch.objects.create(variant=variant, batch_number="API-1", height_cm=60, stems_per_bunch=20, received_stems=100, remaining_stems=100, cost_per_stem=10000, sale_price_per_stem=20000, sale_price_per_bunch=400000)

    def _florist_with_history(self):
        user = User.objects.create_user("florist-stats", password="password", first_name="Dilnoza", last_name="F")
        profile = FloristProfile.objects.create(user=user, staff_type="florist", phone="+998901112233")
        bouquet = CatalogItem.objects.create(name_uz="Katta buket", arrangement_type="bouquet", volume="Katta", catalog_kind="standard", price=Decimal("900000"), quantity_total=1, status="available", florist=profile, florist_fee=Decimal("60000"))
        basket = CatalogItem.objects.create(name_uz="Mini savat", arrangement_type="basket", volume="Kichik", catalog_kind="custom", price=Decimal("400000"), quantity_total=1, status="available", florist=profile, florist_fee=Decimal("30000"))
        CatalogComposition.objects.create(catalog_item=bouquet, stock_batch=self.batch, quantity_stems=20)
        CatalogComposition.objects.create(catalog_item=basket, stock_batch=self.batch, quantity_stems=10)
        FloristSalaryEntry.objects.create(florist=profile, amount=Decimal("60000"), source="catalog", work_date="2026-07-20", catalog_item=bouquet)
        FloristSalaryEntry.objects.create(florist=profile, amount=Decimal("30000"), source="custom_catalog", work_date="2026-07-21", catalog_item=basket)
        FloristSalaryEntry.objects.create(florist=profile, amount=Decimal("15000"), source="manual", work_date="2026-07-21", note="Qo‘shimcha")
        mark_catalog_sold(bouquet, self.user)
        FloristAttendance.objects.create(florist=profile, work_date="2026-07-20")
        return profile

    def _catalog_with_composition(self, stems=15):
        response = self.client.post("/api/catalog/", {"name_uz": "Tarkibli buket", "arrangement_type": "bouquet", "price": "300000", "quantity_total": 1, "status": "available", "composition": [{"stock_batch": self.batch.id, "quantity_stems": stems}]}, format="json")
        self.assertEqual(response.status_code, 201)
        listed = self.client.get("/api/catalog/")
        return [row for row in listed.data["results"] if row["id"] == response.data["id"]][0]

    def test_catalog_list_shows_flower_composition(self):
        row = self._catalog_with_composition()
        self.assertEqual(len(row["composition"]), 1)
        composition = row["composition"][0]
        self.assertEqual(composition["quantity_stems"], 15)
        self.assertEqual(composition["batch_detail"]["flower_name"], "Atirgul API")
        self.assertEqual(composition["batch_detail"]["variant_detail"]["color_uz"], "Qizil")

    def test_catalog_list_composition_stays_light(self):
        batch_detail = self._catalog_with_composition()["composition"][0]["batch_detail"]
        for key in ["supplier_detail", "delivery_detail", "cost_per_stem", "sale_price_per_stem", "remaining_stems"]:
            self.assertNotIn(key, batch_detail)

    def test_catalog_list_hides_composition_row_without_stems(self):
        item = CatalogItem.objects.create(name_uz="Soni yozilmagan", arrangement_type="bouquet", price=Decimal("100000"), quantity_total=1, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=0)
        listed = self.client.get("/api/catalog/")
        row = [r for r in listed.data["results"] if r["id"] == item.id][0]
        self.assertEqual(row["composition"], [])

    def test_deleting_transferred_catalog_does_not_restore_flowers_to_florist(self):
        profile = FloristProfile.objects.create(user=User.objects.create_user("branch-florist", password="p"), staff_type="florist")
        source = CatalogItem.objects.create(name_uz="Asosiy buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=1, quantity_stock_deducted=1, status="available", florist=profile)
        branch = Branch.objects.create(name="Test filial")
        UserProfile.objects.create(user=self.user, role="admin", branch=branch)
        item = CatalogItem.objects.create(name_uz="Filial buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=1, quantity_stock_deducted=1, status="available", florist=profile, source_item=source, branch=branch)
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=12)
        response = self.client.delete(f"/api/catalog/{item.id}/")
        self.assertEqual(response.status_code, 204)
        self.assertFalse(FloristStockBalance.objects.filter(florist=profile, batch=self.batch, remaining_stems__gt=0).exists())

    def test_catalog_item_can_be_created_without_customer(self):
        response = self.client.post("/api/catalog/", {"name_uz": "Mijozsiz buket", "arrangement_type": "bouquet", "price": "300000", "quantity_total": 1, "status": "available"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertIsNone(response.data["customer"])
        self.assertIsNone(response.data["customer_detail"])

    def test_catalog_item_can_attach_existing_customer(self):
        customer = Customer.objects.create(name="Ahmad", phone="+998901112233", instagram_user_id="ig-cat-1")
        response = self.client.post("/api/catalog/", {"name_uz": "Ahmad buketi", "arrangement_type": "bouquet", "price": "300000", "quantity_total": 1, "status": "available", "customer": customer.id}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["customer"], customer.id)
        self.assertEqual(response.data["customer_detail"]["name"], "Ahmad")

    def test_reservation_payment_is_linked_when_catalog_is_sold(self):
        reservation_response = self.client.post("/api/reservations/", {
            "customer_name": "Bron Mijoz",
            "customer_phone": "901112233",
            "request_uz": "Qizil atirgul bron",
            "arrangement_type": "bouquet",
            "estimated_price": "500000",
        }, format="json")
        self.assertEqual(reservation_response.status_code, 201, reservation_response.json())
        reservation_id = reservation_response.data["id"]
        payment_response = self.client.post(f"/api/reservations/{reservation_id}/add-payment/", {
            "amount": "200000",
            "method": "cash",
        }, format="json")
        self.assertEqual(payment_response.status_code, 201, payment_response.json())
        item = CatalogItem.objects.create(name_uz="Bron buket", arrangement_type="bouquet", price=Decimal("500000"), quantity_total=1, status="available")
        sell_response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1,
            "sale_price": "500000",
            "reservation": reservation_id,
            "payment_type": "cash",
        }, format="json")
        self.assertEqual(sell_response.status_code, 200, sell_response.json())
        history = CatalogHistory.objects.get(catalog_item=item, action="sold")
        reservation = Reservation.objects.get(pk=reservation_id)
        self.assertEqual(history.reservation_id, reservation_id)
        self.assertEqual(reservation.status, "fulfilled")
        self.assertEqual(reservation.payment_status, "deposit")
        self.assertEqual(history.snapshot["reservation"]["paid_amount"], "200000.00")
        self.assertEqual(history.snapshot["reservation"]["remaining_due"], "300000.00")
        self.assertEqual(Customer.objects.count(), 1)

    def test_catalog_item_creates_new_customer_from_name_and_phone(self):
        response = self.client.post("/api/catalog/", {"name_uz": "Yangi mijoz buketi", "arrangement_type": "basket", "catalog_kind": "custom", "price": "450000", "quantity_total": 1, "customer_name": "Dilnoza", "customer_phone": "901119988"}, format="json")
        self.assertEqual(response.status_code, 201)
        customer = Customer.objects.get(name="Dilnoza")
        self.assertEqual(customer.phone, "+998901119988")
        self.assertTrue(customer.instagram_user_id.startswith("manual:"))
        self.assertEqual(response.data["customer_detail"]["phone"], "+998901119988")

    def test_catalog_item_reuses_customer_by_phone(self):
        existing = Customer.objects.create(name="", phone="+998901119988", instagram_user_id="ig-cat-2")
        response = self.client.post("/api/catalog/", {"name_uz": "Takror mijoz", "arrangement_type": "bouquet", "price": "200000", "quantity_total": 1, "status": "available", "customer_name": "Dilnoza", "customer_phone": "+998901119988"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["customer"], existing.id)
        existing.refresh_from_db()
        self.assertEqual(existing.name, "Dilnoza")
        self.assertEqual(Customer.objects.count(), 1)

    def test_catalog_item_customer_can_be_changed_and_cleared(self):
        first = Customer.objects.create(name="Birinchi", phone="+998901110001", instagram_user_id="ig-cat-3")
        item = CatalogItem.objects.create(name_uz="O‘zgaruvchi", arrangement_type="bouquet", price=Decimal("200000"), quantity_total=1, status="available", customer=first)
        changed = self.client.patch(f"/api/catalog/{item.id}/", {"customer_name": "Ikkinchi", "customer_phone": "901110002"}, format="json")
        self.assertEqual(changed.status_code, 200)
        self.assertEqual(changed.data["customer_detail"]["name"], "Ikkinchi")
        cleared = self.client.patch(f"/api/catalog/{item.id}/", {"customer": None}, format="json")
        self.assertEqual(cleared.status_code, 200)
        self.assertIsNone(cleared.data["customer"])

    def test_catalog_can_be_filtered_by_customer(self):
        customer = Customer.objects.create(name="Filtr", phone="+998901110003", instagram_user_id="ig-cat-4")
        CatalogItem.objects.create(name_uz="Filtr buketi", arrangement_type="bouquet", price=Decimal("200000"), quantity_total=1, status="available", customer=customer)
        CatalogItem.objects.create(name_uz="Boshqa", arrangement_type="bouquet", price=Decimal("200000"), quantity_total=1, status="available")
        response = self.client.get(f"/api/catalog/?customer={customer.id}")
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["name_uz"], "Filtr buketi")

    def test_florist_stats_endpoint_returns_full_breakdown(self):
        profile = self._florist_with_history()
        response = self.client.get(f"/api/florists/{profile.id}/stats/")
        self.assertEqual(response.status_code, 200)
        data = response.data
        summary = data["summary"]
        self.assertEqual(Decimal(summary["salary_total"]), Decimal("105000.00"))
        self.assertEqual(Decimal(summary["catalog_salary_total"]), Decimal("90000.00"))
        self.assertEqual(Decimal(summary["manual_salary_total"]), Decimal("15000.00"))
        self.assertEqual(summary["catalog_count"], 2)
        self.assertEqual(summary["bouquet_count"], 1)
        self.assertEqual(summary["basket_count"], 1)
        self.assertEqual(summary["standard_count"], 1)
        self.assertEqual(summary["custom_count"], 1)
        self.assertEqual(summary["attendance_days"], 1)
        self.assertEqual(summary["sold_quantity"], 1)
        self.assertEqual(Decimal(summary["sale_revenue"]), Decimal("900000.00"))
        self.assertEqual(Decimal(summary["avg_fee_per_item"]), Decimal("45000.00"))
        self.assertEqual({row["arrangement_type"] for row in data["by_arrangement"]}, {"bouquet", "basket"})
        self.assertEqual({(row["arrangement_type"], row["volume"]) for row in data["by_volume"]}, {("bouquet", "Katta"), ("basket", "Kichik")})
        self.assertEqual({row["source"] for row in data["by_source"]}, {"catalog", "custom_catalog", "manual"})
        self.assertEqual(len(data["by_day"]), 2)
        sold_row = next(row for row in data["salary_entries"] if row["catalog_name"] == "Katta buket")
        self.assertEqual(sold_row["arrangement_label"], "Buket")
        self.assertEqual(sold_row["volume"], "Katta")
        self.assertEqual(Decimal(sold_row["amount"]), Decimal("60000.00"))
        self.assertEqual(Decimal(sold_row["sale_revenue"]), Decimal("900000.00"))
        self.assertTrue(sold_row["is_sold"])
        unsold_row = next(row for row in data["salary_entries"] if row["catalog_name"] == "Mini savat")
        self.assertFalse(unsold_row["is_sold"])
        self.assertEqual(Decimal(unsold_row["sale_revenue"]), Decimal("0"))

    def test_florist_stats_counts_catalog_quantity_total(self):
        user = User.objects.create_user("quantity-florist", password="password", first_name="Quantity")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        item = CatalogItem.objects.create(name_uz="10 dona buket", arrangement_type="bouquet", volume="L", catalog_kind="standard", price=Decimal("500000"), quantity_total=10, status="available", florist=profile)
        FloristSalaryEntry.objects.create(florist=profile, amount=Decimal("500000"), source="catalog", work_date="2026-07-25", catalog_item=item)
        mark_catalog_sold(item, self.user, quantity=3)
        response = self.client.get(f"/api/florists/{profile.id}/stats/")
        self.assertEqual(response.status_code, 200)
        summary = response.data["summary"]
        self.assertEqual(summary["catalog_count"], 10)
        self.assertEqual(summary["bouquet_count"], 10)
        self.assertEqual(summary["standard_count"], 10)
        self.assertEqual(summary["sold_quantity"], 3)
        self.assertEqual(response.data["by_arrangement"][0]["count"], 10)
        self.assertEqual(response.data["by_volume"][0]["count"], 10)
        self.assertEqual(response.data["by_day"][0]["count"], 10)
        self.assertEqual(Decimal(summary["avg_fee_per_item"]), Decimal("50000.00"))

    def test_florist_stats_respects_date_range(self):
        profile = self._florist_with_history()
        response = self.client.get(f"/api/florists/{profile.id}/stats/?date_from=2026-07-21&date_to=2026-07-21")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["salary_entries_count"], 2)
        self.assertEqual(Decimal(response.data["summary"]["salary_total"]), Decimal("45000.00"))
        self.assertEqual(response.data["period"]["date_from"], "2026-07-21")

    def test_florist_me_dashboard_returns_own_stats(self):
        profile = self._florist_with_history()
        PagePermission.objects.create(user=profile.user, page="florists", can_view=True, can_control=False)
        UserProfile.objects.update_or_create(user=profile.user, defaults={"role": "florist"})
        client = APIClient()
        client.force_authenticate(profile.user)
        response = client.get("/api/florists/me/dashboard/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["florist"]["id"], profile.id)
        self.assertEqual(Decimal(response.data["summary"]["salary_total"]), Decimal("105000.00"))
        self.assertTrue(response.data["salary_entries"])

    def test_florist_excel_export_has_all_sheets(self):
        import io
        from openpyxl import load_workbook
        profile = self._florist_with_history()
        response = self.client.get(f"/api/exports/florist/?florist={profile.id}&date_from=2026-07-01&date_to=2026-07-31")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Ish haqi tarixi", "Umumiy", "Kunlar bo‘yicha", "Hajm bo‘yicha", "Manba bo‘yicha", "Keldi-ketdi"])
        header = [cell.value for cell in workbook["Ish haqi tarixi"][1]]
        self.assertIn("Sotuvdan tushgan", header)
        self.assertIn("Floristga qo‘shilgan", header)
        self.assertIn("Buket yoki savat", header)

    def test_supplier_with_payments_is_archived_not_deleted(self):
        supplier = Supplier.objects.create(name="To‘lovli postavshik")
        SupplierPayment.objects.create(supplier=supplier, amount=Decimal("100000"), paid_at="2026-07-29", method="cash")
        SupplierPayment.objects.create(supplier=supplier, amount=Decimal("50000"), paid_at="2026-07-30", method="card")
        response = self.client.delete(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["archived"])
        self.assertFalse(response.data["deleted"])
        self.assertFalse(response.data["is_active"])
        self.assertEqual(response.data["id"], supplier.id)
        self.assertEqual(response.data["blocked_by"], [{"model": "supplierpayment", "label": "To‘lovlar", "count": 2}])
        self.assertIn("To‘lovlar (2 ta)", response.data["detail"])
        supplier.refresh_from_db()
        self.assertFalse(supplier.is_active)
        self.assertTrue(AuditLog.objects.filter(action="supplier_archived").exists())

    def test_supplier_without_relations_is_deleted_with_204(self):
        supplier = Supplier.objects.create(name="Bo‘sh postavshik")
        response = self.client.delete(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Supplier.objects.filter(id=supplier.id).exists())
        self.assertTrue(AuditLog.objects.filter(action="supplier_deleted").exists())

    def test_supplier_with_only_batches_is_deleted_and_batches_lose_supplier(self):
        """StockBatch.supplier SET_NULL, shuning uchun faqat partiyasi bor postavshik
        o'chib ketadi va partiyalar egasiz qoladi. Hozirgi xatti-harakat shunday."""
        supplier = Supplier.objects.create(name="Partiyali postavshik")
        batch = StockBatch.objects.create(variant=self.batch.variant, supplier=supplier, batch_number="ARCH-1", height_cm=50, stems_per_bunch=25, received_stems=10, remaining_stems=10, cost_per_stem=1000, sale_price_per_stem=2000, sale_price_per_bunch=50000)
        response = self.client.delete(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Supplier.objects.filter(id=supplier.id).exists())
        batch.refresh_from_db()
        self.assertIsNone(batch.supplier_id)

    def test_supplier_with_payment_and_batches_is_archived_and_keeps_batches(self):
        supplier = Supplier.objects.create(name="To‘lovli va partiyali")
        batch = StockBatch.objects.create(variant=self.batch.variant, supplier=supplier, batch_number="ARCH-2", height_cm=50, stems_per_bunch=25, received_stems=10, remaining_stems=10, cost_per_stem=1000, sale_price_per_stem=2000, sale_price_per_bunch=50000)
        SupplierPayment.objects.create(supplier=supplier, amount=Decimal("100000"), paid_at="2026-07-29", method="cash")
        response = self.client.delete(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["archived"])
        batch.refresh_from_db()
        self.assertEqual(batch.supplier_id, supplier.id)

    def _sell_catalog(self, price="6700000", quantity=1, sold_at=None):
        item = CatalogItem.objects.create(name_uz="Sotuv buketi", arrangement_type="bouquet", price=Decimal(price), quantity_total=quantity, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=5)
        payload = {"quantity": quantity, "sale_image_url": "https://example.com/sale.jpg"}
        if sold_at:
            payload["sold_at"] = sold_at
        response = self.client.post(f"/api/catalog/{item.id}/sell/", payload, format="json")
        self.assertEqual(response.status_code, 200)
        return item

    def test_dashboard_revenue_includes_catalog_sales(self):
        self._sell_catalog()
        response = self.client.get("/api/dashboard/")
        self.assertEqual(response.status_code, 200)
        data = response.data
        self.assertEqual(Decimal(data["revenue_today"]), Decimal("6700000.00"))
        self.assertEqual(Decimal(data["catalog_sales_revenue_today"]), Decimal("6700000.00"))
        self.assertEqual(Decimal(data["lead_revenue_today"]), Decimal("0"))
        self.assertEqual(data["catalog_sales_orders_today"], 1)
        self.assertEqual(data["orders_today"], 1)

    def test_dashboard_revenue_matches_accounting_total_sales(self):
        self._sell_catalog()
        dashboard = self.client.get("/api/dashboard/").data
        accounting = self.client.get("/api/accounting/").data
        self.assertEqual(Decimal(dashboard["catalog_sales_revenue_today"]), Decimal(accounting["summary"]["total_sales"]))

    def test_analytics_summary_and_daily_stats_include_catalog(self):
        self._sell_catalog()
        response = self.client.get("/api/analytics/")
        self.assertEqual(response.status_code, 200)
        summary = response.data["summary"]
        self.assertEqual(Decimal(summary["revenue"]), Decimal("6700000.00"))
        self.assertEqual(Decimal(summary["catalog_sales_revenue"]), Decimal("6700000.00"))
        self.assertEqual(Decimal(summary["lead_revenue"]), Decimal("0"))
        self.assertEqual(summary["catalog_sales_orders"], 1)
        today = timezone.localdate().isoformat()
        row = next(day for day in response.data["daily_stats"] if day["date"] == today)
        self.assertEqual(Decimal(row["catalog_revenue"]), Decimal("6700000.00"))
        self.assertEqual(row["catalog_orders"], 1)
        self.assertEqual(row["catalog_quantity"], 1)
        self.assertEqual(Decimal(row["revenue"]), Decimal("6700000.00"))

    def test_dashboard_daily_stats_include_catalog_series(self):
        self._sell_catalog()
        response = self.client.get("/api/dashboard/")
        today = timezone.localdate().isoformat()
        row = next(day for day in response.data["daily_stats"] if day["date"] == today)
        self.assertEqual(Decimal(row["catalog_revenue"]), Decimal("6700000.00"))
        self.assertEqual(row["catalog_orders"], 1)

    def test_top_catalog_items_and_revenue_by_source_use_catalog_sales(self):
        item = self._sell_catalog()
        response = self.client.get("/api/analytics/")
        top = response.data["top_catalog_items"]
        self.assertTrue(top)
        self.assertEqual(top[0]["catalog_item_id"], item.id)
        self.assertEqual(Decimal(top[0]["revenue"]), Decimal("6700000.00"))
        self.assertTrue(response.data["recent_top_catalog_items"])
        sources = {row["source"]: row for row in response.data["revenue_by_source"]}
        self.assertIn("catalog", sources)
        self.assertEqual(Decimal(sources["catalog"]["revenue"]), Decimal("6700000.00"))
        self.assertEqual(sources["catalog"]["source_label"], "Katalogdan sotuv")

    def test_catalog_sale_accepts_historical_sold_at(self):
        item = self._sell_catalog(sold_at="2026-07-10T12:00:00+05:00")
        item.refresh_from_db()
        self.assertEqual(item.sold_at.isoformat()[:10], "2026-07-10")
        history = CatalogHistory.objects.get(catalog_item=item, action="sold")
        self.assertEqual(timezone.localtime(history.created_at).date().isoformat(), "2026-07-10")
        response = self.client.get("/api/analytics/?from=2026-07-01&to=2026-07-31")
        row = next(day for day in response.data["daily_stats"] if day["date"] == "2026-07-10")
        self.assertEqual(Decimal(row["catalog_revenue"]), Decimal("6700000.00"))

    def test_stock_movement_accepts_historical_created_at(self):
        response = self.client.post(f"/api/stock-batches/{self.batch.id}/movement/", {"movement_type": "waste", "quantity_stems": 3, "reason": "test", "created_at": "2026-07-05T10:00:00+05:00"}, format="json")
        self.assertEqual(response.status_code, 200)
        movement = StockMovement.objects.get(id=response.data["id"])
        self.assertEqual(timezone.localtime(movement.created_at).date().isoformat(), "2026-07-05")

    def test_lead_accepts_historical_created_at(self):
        response = self.client.post("/api/leads/", {"customer_name": "Tarixiy", "customer_phone": "901110009", "request_uz": "Test", "created_at": "2026-07-03T09:00:00+05:00"}, format="json")
        self.assertEqual(response.status_code, 201)
        lead = Lead.objects.get(id=response.data["id"])
        self.assertEqual(timezone.localtime(lead.created_at).date().isoformat(), "2026-07-03")

    def test_packaging_movement_accepts_historical_created_at(self):
        packaging = Packaging.objects.create(packaging_type="wrap", name_uz="Test qog‘oz", cost_price=Decimal("1000"), sale_price=Decimal("2000"), quantity=50)
        response = self.client.post(f"/api/materials/{packaging.id}/movement/", {"movement_type": "in", "quantity": 10, "created_at": "2026-07-04T08:00:00+05:00"}, format="json")
        self.assertEqual(response.status_code, 200)
        movement = PackagingMovement.objects.get(id=response.data["id"])
        self.assertEqual(timezone.localtime(movement.created_at).date().isoformat(), "2026-07-04")

    def _florist(self, username="fl-stock"):
        user = User.objects.create_user(username, password="p", first_name="Stock")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        # standart katalogda haq hajm tarifidan olinadi, shuning uchun tarif kerak
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="M", default_stems=25, florist_fee=Decimal("50000"))
        return profile

    def test_issue_stock_to_florist_moves_balance(self):
        florist = self._florist()
        before = self.batch.remaining_stems
        response = self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 30, "reason": "Ish uchun"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["kind_label"], "Chiqarildi")
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 30)
        balance = FloristStockBalance.objects.get(florist=florist, batch=self.batch)
        self.assertEqual(balance.remaining_stems, 30)
        self.assertTrue(StockMovement.objects.filter(reference_type="florist_issue").exists())

    def test_bulk_issue_stock_to_florist_moves_multiple_batches(self):
        florist = self._florist("bulk-florist")
        second = StockBatch.objects.create(
            variant=self.batch.variant, batch_number="API-BULK-2", height_cm=60, stems_per_bunch=20,
            received_stems=100, remaining_stems=100, cost_per_stem=10000,
            sale_price_per_stem=20000, sale_price_per_bunch=400000,
        )
        response = self.client.post("/api/florist-stock-issues/bulk-issue/", {
            "florist": florist.id,
            "items": [
                {"batch": self.batch.id, "quantity_stems": 10},
                {"batch": second.id, "quantity_stems": 15},
            ],
            "reason": "Bulk test",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(len(response.data), 2)
        self.batch.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 90)
        self.assertEqual(second.remaining_stems, 85)
        self.assertEqual(FloristStockBalance.objects.get(florist=florist, batch=self.batch).remaining_stems, 10)
        self.assertEqual(FloristStockBalance.objects.get(florist=florist, batch=second).remaining_stems, 15)

    def test_catalog_restoration_wastes_old_flowers_and_issues_new_to_florist(self):
        florist = self._florist("restore-florist")
        new_batch = StockBatch.objects.create(
            variant=self.batch.variant, batch_number="API-RESTORE-2", height_cm=60, stems_per_bunch=20,
            received_stems=100, remaining_stems=100, cost_per_stem=12000,
            sale_price_per_stem=22000, sale_price_per_bunch=440000,
        )
        item = CatalogItem.objects.create(name_uz="Restavratsa buket", arrangement_type="bouquet", price=Decimal("600000"), quantity_total=1, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=20, quantity_bunches=Decimal("1.00"))
        response = self.client.post(f"/api/catalog/{item.id}/restore-flowers/", {
            "florist": florist.id,
            "old_batch": self.batch.id,
            "new_batch": new_batch.id,
            "quantity_stems": 15,
            "reason": "Restavratsa",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(CatalogComposition.objects.get(catalog_item=item, stock_batch=self.batch).quantity_stems, 5)
        self.assertEqual(CatalogComposition.objects.get(catalog_item=item, stock_batch=new_batch).quantity_stems, 15)
        new_batch.refresh_from_db()
        self.assertEqual(new_batch.remaining_stems, 85)
        self.assertEqual(FloristStockBalance.objects.get(florist=florist, batch=new_batch).remaining_stems, 0)
        self.assertTrue(FloristStockIssue.objects.filter(florist=florist, batch=new_batch, quantity_stems=15).exists())
        self.assertTrue(StockMovement.objects.filter(batch=self.batch, movement_type="waste", quantity_stems=-15, reference_type="catalog_restoration", reference_id=item.id).exists())

    def test_issue_rejects_more_than_stock(self):
        florist = self._florist()
        response = self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 99999}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("qolgan", response.data["detail"])

    def test_florist_can_return_unused_stock(self):
        florist = self._florist()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        response = self.client.post("/api/florist-stock-issues/return/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 10, "kind": "return"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before + 10)
        self.assertEqual(FloristStockBalance.objects.get(florist=florist, batch=self.batch).remaining_stems, 20)

    def test_florist_waste_does_not_return_to_stock(self):
        florist = self._florist()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        response = self.client.post("/api/florist-stock-issues/return/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 5, "kind": "waste", "reason": "so‘ldi"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before)
        self.assertEqual(FloristStockBalance.objects.get(florist=florist, batch=self.batch).remaining_stems, 25)

    def test_return_rejects_more_than_florist_has(self):
        florist = self._florist()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 10}, format="json")
        response = self.client.post("/api/florist-stock-issues/return/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 50}, format="json")
        self.assertEqual(response.status_code, 400)

    def test_catalog_with_florist_consumes_florist_stock_not_warehouse(self):
        florist = self._florist()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        self.batch.refresh_from_db()
        stock_before = self.batch.remaining_stems
        response = self.client.post("/api/catalog/", {"name_uz": "Florist buketi", "arrangement_type": "bouquet", "volume": "M", "price": "300000", "quantity_total": 1, "status": "available", "florist": florist.id, "composition": [{"stock_batch": self.batch.id, "quantity_stems": 20}]}, format="json")
        self.assertEqual(response.status_code, 201)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, stock_before)
        self.assertEqual(FloristStockBalance.objects.get(florist=florist, batch=self.batch).remaining_stems, 10)

    def test_catalog_with_florist_rejects_when_florist_lacks_stock(self):
        florist = self._florist()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 5}, format="json")
        response = self.client.post("/api/catalog/", {"name_uz": "Yetmaydi", "arrangement_type": "bouquet", "volume": "M", "price": "300000", "quantity_total": 1, "status": "available", "florist": florist.id, "composition": [{"stock_batch": self.batch.id, "quantity_stems": 20}]}, format="json")
        self.assertEqual(response.status_code, 400)

    def test_catalog_without_florist_still_uses_warehouse(self):
        before = self.batch.remaining_stems
        response = self.client.post("/api/catalog/", {"name_uz": "Skladdan", "arrangement_type": "bouquet", "price": "300000", "quantity_total": 1, "status": "available", "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}]}, format="json")
        self.assertEqual(response.status_code, 201)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 10)

    def test_florist_stock_balance_list_filters_by_florist(self):
        florist = self._florist()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": florist.id, "batch": self.batch.id, "quantity_stems": 12}, format="json")
        response = self.client.get(f"/api/florist-stock-balances/?florist={florist.id}")
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["remaining_stems"], 12)
        self.assertEqual(response.data["results"][0]["batch_detail"]["batch_number"], self.batch.batch_number)

    def test_florist_day_off_crud(self):
        florist = self._florist()
        response = self.client.post("/api/florist-days-off/", {"florist": florist.id, "work_date": "2026-08-02", "kind": "weekend", "note": "Yakshanba"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["kind_label"], "Dam kuni")
        listed = self.client.get(f"/api/florist-days-off/?florist={florist.id}")
        self.assertEqual(listed.data["count"], 1)
        duplicate = self.client.post("/api/florist-days-off/", {"florist": florist.id, "work_date": "2026-08-02", "kind": "sick"}, format="json")
        self.assertEqual(duplicate.status_code, 400)

    def _parkent(self):
        return Branch.objects.get(name="Parkent filiali")

    def _main_catalog(self, quantity=5, price="300000"):
        item = CatalogItem.objects.create(name_uz="Filial buketi", arrangement_type="bouquet", price=Decimal(price), quantity_total=quantity, quantity_stock_deducted=quantity, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=4)
        sync_catalog_financials(item)
        return item

    def test_transfer_moves_part_of_catalog_to_branch(self):
        item = self._main_catalog(quantity=5, price="300000")
        parkent = self._parkent()
        response = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": parkent.id, "quantity": 2, "price": "450000", "note": "Parkentga"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["quantity"], 2)
        self.assertEqual(Decimal(response.data["source_price"]), Decimal("300000.00"))
        self.assertEqual(Decimal(response.data["target_price"]), Decimal("450000.00"))
        item.refresh_from_db()
        self.assertEqual(item.quantity_total, 3)
        target = CatalogItem.objects.get(id=response.data["target_item"])
        self.assertEqual(target.branch_id, parkent.id)
        self.assertEqual(target.quantity_total, 2)
        self.assertEqual(target.price, Decimal("450000.00"))
        self.assertEqual(target.source_price, Decimal("300000.00"))
        self.assertEqual(target.source_item_id, item.id)

    def test_catalog_can_be_created_directly_for_branch_with_stems(self):
        # asosiy filialdan turib Parkent uchun katalog qo'shiladi, gul soni bilan
        parkent = self._parkent()
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        response = self.client.post("/api/catalog/", {
            "name_uz": "Parkent uchun buket", "arrangement_type": "bouquet", "volume": "M",
            "branch": parkent.id, "price": "500000", "quantity_total": 2, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 30}],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(response.json()["branch"], parkent.id)
        self.assertEqual(response.json()["branch_name"], "Parkent filiali")
        item = CatalogItem.objects.get(id=response.json()["id"])
        self.assertEqual(item.composition.get().quantity_stems, 30)
        # gul asosiy filial skladidan yechiladi
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 60)
        # kelib chiqish narxi o'rniga bir donaga to'g'ri keladigan tannarx yoziladi
        self.assertEqual(item.source_price, (item.calculated_cost_price / 2).quantize(Decimal("0.01")))

    def test_branch_user_sees_directly_created_catalog(self):
        parkent = self._parkent()
        created = self.client.post("/api/catalog/", {
            "name_uz": "To‘g‘ridan buket", "arrangement_type": "bouquet", "volume": "M",
            "branch": parkent.id, "price": "500000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}],
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        user = User.objects.create_user("parkent-direct", password="p")
        UserProfile.objects.create(user=user, role="operator", branch=parkent)
        PagePermission.objects.create(user=user, page="catalog", can_view=True, can_control=True)
        client = APIClient()
        client.force_authenticate(user)
        listed = client.get("/api/catalog/")
        self.assertEqual(listed.data["count"], 1)
        self.assertEqual(listed.data["results"][0]["name_uz"], "To‘g‘ridan buket")
        # asosiy filial ro'yxatida ko'rinmaydi
        main = self.client.get("/api/catalog/")
        self.assertNotIn("To‘g‘ridan buket", [row["name_uz"] for row in main.data["results"]])

    def test_branch_report_counts_directly_created_catalog(self):
        parkent = self._parkent()
        created = self.client.post("/api/catalog/", {
            "name_uz": "Hisobot buketi", "arrangement_type": "bouquet", "volume": "M",
            "branch": parkent.id, "price": "500000", "quantity_total": 2, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}],
        }, format="json")
        item = CatalogItem.objects.get(id=created.json()["id"])
        mark_catalog_sold(item, self.user, quantity=1)
        response = self.client.get("/api/branch-report/")
        row = next(r for r in response.data["branches"] if r["branch_name"] == "Parkent filiali")
        self.assertEqual(row["received_quantity"], 0)
        self.assertEqual(row["direct_quantity"], 2)
        self.assertEqual(row["incoming_quantity"], 2)
        self.assertEqual(row["sold_quantity"], 1)
        # ustama endi haqiqiy foyda: sotuv narxi - bir donalik tannarx
        self.assertEqual(Decimal(row["source_value"]), item.source_price)
        self.assertEqual(Decimal(row["markup_total"]), Decimal("500000.00") - item.source_price)

    def test_transfer_does_not_touch_warehouse_stock(self):
        item = self._main_catalog()
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 2}, format="json")
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before)

    def test_transfer_rejects_more_than_available(self):
        item = self._main_catalog(quantity=2)
        response = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 5}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("atigi 2 dona", response.data["detail"])

    def test_branch_catalogs_do_not_mix(self):
        item = self._main_catalog()
        parkent = self._parkent()
        self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": parkent.id, "quantity": 2}, format="json")
        # asosiy filial admini faqat asosiy katalogni ko'radi
        main_list = self.client.get("/api/catalog/")
        names = {row["id"] for row in main_list.data["results"]}
        self.assertIn(item.id, names)
        self.assertEqual(len(names), 1)
        # filial foydalanuvchisi faqat o'z filialini ko'radi
        user = User.objects.create_user("parkent-user", password="p")
        UserProfile.objects.create(user=user, role="operator", branch=parkent)
        PagePermission.objects.create(user=user, page="catalog", can_view=True, can_control=True)
        client = APIClient()
        client.force_authenticate(user)
        branch_list = client.get("/api/catalog/")
        self.assertEqual(branch_list.data["count"], 1)
        self.assertEqual(branch_list.data["results"][0]["branch_name"], "Parkent filiali")

    def _parkent_client(self, username="parkent-eye"):
        parkent = self._parkent()
        user = User.objects.create_user(username, password="p")
        UserProfile.objects.create(user=user, role="operator", branch=parkent)
        for page in ["catalog", "dashboard"]:
            PagePermission.objects.create(user=user, page=page, can_view=True, can_control=True)
        client = APIClient()
        client.force_authenticate(user)
        return client

    def test_branch_cannot_see_main_branch_price_and_cost(self):
        item = self._main_catalog(quantity=5, price="300000")
        transfer = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 2, "price": "450000"}, format="json")
        target_id = transfer.data["target_item"]
        client = self._parkent_client("parkent-price-eye")
        seen = client.get(f"/api/catalog/{target_id}/").data
        # o'z narxini ko'radi
        self.assertEqual(Decimal(seen["price"]), Decimal("450000.00"))
        # asosiy filial narxi va tannarxi ko'rinmaydi
        for key in ["source_price", "calculated_cost_price", "calculated_component_price",
                    "florist_fee", "florist_salary_amount", "profit", "discount_amount", "discount_percent"]:
            self.assertNotIn(key, seen, f"{key} filialga ko‘rinib turibdi")
        # gul narxlari ham ko'rinmaydi
        batch = seen["composition"][0]["batch_detail"]
        for key in ["cost_per_stem", "sale_price_per_stem", "cost_per_bunch", "sale_price_per_bunch", "supplier", "remaining_stems"]:
            self.assertNotIn(key, batch, f"batch_detail.{key} filialga ko‘rinib turibdi")

    def test_main_branch_still_sees_everything(self):
        item = self._main_catalog(quantity=5, price="300000")
        seen = self.client.get(f"/api/catalog/{item.id}/").data
        for key in ["calculated_cost_price", "florist_fee", "profit"]:
            self.assertIn(key, seen)
        self.assertIn("cost_per_stem", seen["composition"][0]["batch_detail"])

    def test_branch_transfer_record_hides_source_price(self):
        item = self._main_catalog(quantity=5, price="300000")
        self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 2, "price": "450000"}, format="json")
        client = self._parkent_client("parkent-transfer-eye")
        rows = client.get("/api/catalog-transfers/").data["results"]
        self.assertEqual(len(rows), 1)
        self.assertNotIn("source_price", rows[0])
        self.assertNotIn("source_item", rows[0])
        self.assertEqual(Decimal(rows[0]["target_price"]), Decimal("450000.00"))
        self.assertEqual(rows[0]["quantity"], 2)
        # asosiy filial admini ikkalasini ham ko'radi
        admin_rows = self.client.get("/api/catalog-transfers/").data["results"]
        self.assertEqual(Decimal(admin_rows[0]["source_price"]), Decimal("300000.00"))

    def test_branch_sees_only_its_own_transfers(self):
        other = Branch.objects.create(name="Chirchiq filiali", is_main=False, is_active=True)
        item = self._main_catalog(quantity=6, price="300000")
        self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 2}, format="json")
        self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": other.id, "quantity": 2}, format="json")
        client = self._parkent_client("parkent-own-transfers")
        rows = client.get("/api/catalog-transfers/").data["results"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["branch_name"], "Parkent filiali")
        self.assertEqual(self.client.get("/api/catalog-transfers/").data["count"], 2)

    def test_branch_can_still_change_its_own_price(self):
        item = self._main_catalog(quantity=5, price="300000")
        transfer = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 2, "price": "450000"}, format="json")
        target_id = transfer.data["target_item"]
        client = self._parkent_client("parkent-price-edit")
        changed = client.patch(f"/api/catalog/{target_id}/", {"price": "500000"}, format="json")
        self.assertEqual(changed.status_code, 200, changed.data)
        self.assertEqual(Decimal(changed.data["price"]), Decimal("500000.00"))
        self.assertNotIn("source_price", changed.data)

    def test_branch_user_cannot_create_catalog(self):
        parkent = self._parkent()
        user = User.objects.create_user("parkent-creator", password="p")
        UserProfile.objects.create(user=user, role="operator", branch=parkent)
        PagePermission.objects.create(user=user, page="catalog", can_view=True, can_control=True)
        client = APIClient()
        client.force_authenticate(user)
        response = client.post("/api/catalog/", {"name_uz": "Yangi", "arrangement_type": "bouquet", "price": "100000", "quantity_total": 1}, format="json")
        self.assertEqual(response.status_code, 400)

    def test_branch_can_change_price_then_sell_with_discount(self):
        item = self._main_catalog(quantity=3, price="300000")
        parkent = self._parkent()
        transfer = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": parkent.id, "quantity": 2, "price": "450000"}, format="json")
        target_id = transfer.data["target_item"]
        user = User.objects.create_user("parkent-seller", password="p")
        UserProfile.objects.create(user=user, role="operator", branch=parkent)
        PagePermission.objects.create(user=user, page="catalog", can_view=True, can_control=True)
        client = APIClient()
        client.force_authenticate(user)
        # narxni yana o'zgartiradi
        changed = client.patch(f"/api/catalog/{target_id}/", {"price": "500000"}, format="json")
        self.assertEqual(changed.status_code, 200)
        # keyin chegirma bilan sotadi
        sold = client.post(f"/api/catalog/{target_id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", "quantity": 1, "sale_price": "420000", "discount_reason": "Doimiy mijoz"}, format="json")
        self.assertEqual(sold.status_code, 200)
        history = CatalogHistory.objects.get(catalog_item_id=target_id, action="sold")
        self.assertEqual(history.listed_unit_price, Decimal("500000.00"))
        self.assertEqual(history.sold_unit_price, Decimal("420000.00"))
        self.assertEqual(history.discount_amount, Decimal("80000.00"))
        self.assertEqual(history.discount_reason, "Doimiy mijoz")

    def test_branch_report_counts_transfers_sales_and_discounts(self):
        item = self._main_catalog(quantity=5, price="300000")
        parkent = self._parkent()
        transfer = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": parkent.id, "quantity": 3, "price": "450000"}, format="json")
        target = CatalogItem.objects.get(id=transfer.data["target_item"])
        mark_catalog_sold(target, self.user, quantity=1)
        mark_catalog_sold(target, self.user, quantity=1, sale_price=Decimal("400000"), discount_reason="Aksiya")
        response = self.client.get("/api/branch-report/")
        self.assertEqual(response.status_code, 200)
        row = next(r for r in response.data["branches"] if r["branch_name"] == "Parkent filiali")
        self.assertEqual(row["received_quantity"], 3)
        self.assertEqual(row["sold_quantity"], 2)
        self.assertEqual(Decimal(row["sold_revenue"]), Decimal("850000.00"))
        self.assertEqual(Decimal(row["source_value"]), Decimal("600000.00"))
        self.assertEqual(Decimal(row["markup_total"]), Decimal("250000.00"))
        self.assertEqual(row["discounted_sales_count"], 1)
        self.assertEqual(row["discounted_quantity"], 1)
        self.assertEqual(Decimal(row["discount_total"]), Decimal("50000.00"))

    def _sold_in_both_branches(self):
        """Asosiy filialda 1 ta, Parkentda 1 ta sotuv qoldiradi."""
        item = self._main_catalog(quantity=4, price="300000")
        parkent = self._parkent()
        transfer = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": parkent.id, "quantity": 2, "price": "500000"}, format="json")
        target = CatalogItem.objects.get(id=transfer.data["target_item"])
        mark_catalog_sold(item, self.user, quantity=1, payment_type="cash")
        mark_catalog_sold(target, self.user, quantity=1, payment_type="card")
        return item, target, parkent

    def _branch_row(self, response, name):
        return next(row for row in response.data["by_branch"] if row["branch_name"] == name)

    def test_accounting_total_includes_branch_sales(self):
        # umumiy yig'indi ikkala filialni qamraydi
        self._sold_in_both_branches()
        response = self.client.get("/api/accounting/")
        self.assertEqual(Decimal(response.data["summary"]["total_sales"]), Decimal("800000.00"))
        self.assertEqual(response.data["summary"]["sales_count"], 2)
        self.assertEqual(response.data["summary"]["total_quantity"], 2)

    def test_accounting_splits_total_by_branch(self):
        # umumiyga qo'shilsa ham qaysi filialdan qanchaligi aniq ko'rinadi
        self._sold_in_both_branches()
        response = self.client.get("/api/accounting/")
        main = self._branch_row(response, "Toshkent (asosiy filial)")
        parkent = self._branch_row(response, "Parkent filiali")
        self.assertEqual(Decimal(main["total_sales"]), Decimal("300000.00"))
        self.assertEqual(Decimal(parkent["total_sales"]), Decimal("500000.00"))
        self.assertEqual(Decimal(main["share_percent"]), Decimal("37.50"))
        self.assertEqual(Decimal(parkent["share_percent"]), Decimal("62.50"))
        self.assertTrue(main["is_main"])
        self.assertFalse(parkent["is_main"])
        # filial qatorlari yig'indisi umumiyga teng
        self.assertEqual(
            sum(Decimal(row["total_sales"]) for row in response.data["by_branch"]),
            Decimal(response.data["summary"]["total_sales"]),
        )

    def test_accounting_splits_cash_and_card_per_branch(self):
        # naqd va karta har filialda alohida chiqadi
        self._sold_in_both_branches()
        response = self.client.get("/api/accounting/")
        main = self._branch_row(response, "Toshkent (asosiy filial)")
        parkent = self._branch_row(response, "Parkent filiali")
        self.assertEqual(Decimal(main["cash_total"]), Decimal("300000.00"))
        self.assertEqual(main["cash_count"], 1)
        self.assertEqual(Decimal(main["card_total"]), Decimal("0"))
        self.assertEqual(Decimal(parkent["card_total"]), Decimal("500000.00"))
        self.assertEqual(parkent["card_count"], 1)
        self.assertEqual(Decimal(parkent["cash_total"]), Decimal("0"))
        self.assertEqual(Decimal(response.data["summary"]["cash_total"]), Decimal("300000.00"))
        self.assertEqual(Decimal(response.data["summary"]["card_total"]), Decimal("500000.00"))

    def test_accounting_counts_sold_flower_stems_per_branch(self):
        # filialga o'tkazilganda tarkib nusxalanadi, shuning uchun gul donasi u yerda ham sanaladi
        self._sold_in_both_branches()
        response = self.client.get("/api/accounting/")
        main = self._branch_row(response, "Toshkent (asosiy filial)")
        parkent = self._branch_row(response, "Parkent filiali")
        self.assertEqual(main["flower_stems"], 4)
        self.assertEqual(parkent["flower_stems"], 4)
        self.assertEqual(response.data["summary"]["flower_stems"], 8)

    def test_accounting_branch_filter_narrows_report(self):
        _, _, parkent = self._sold_in_both_branches()
        only_main = self.client.get("/api/accounting/?branch=main")
        self.assertEqual(Decimal(only_main.data["summary"]["total_sales"]), Decimal("300000.00"))
        self.assertEqual(only_main.data["branch_filter"]["mode"], "main")
        only_branch = self.client.get(f"/api/accounting/?branch={parkent.id}")
        self.assertEqual(Decimal(only_branch.data["summary"]["total_sales"]), Decimal("500000.00"))
        self.assertEqual(only_branch.data["branch_filter"]["branch_name"], "Parkent filiali")
        self.assertEqual(len(only_branch.data["by_branch"]), 1)

    def test_accounting_history_rows_carry_branch(self):
        self._sold_in_both_branches()
        response = self.client.get("/api/accounting/")
        branches = sorted(row["branch_name"] for row in response.data["history"])
        self.assertEqual(branches, ["Parkent filiali", "Toshkent (asosiy filial)"])

    def test_branch_accounting_and_dashboard_are_separate(self):
        # filial foydalanuvchisi faqat o'z filialini ko'radi va buni kengaytira olmaydi
        _, _, parkent = self._sold_in_both_branches()
        user = User.objects.create_user("parkent-acc", password="p")
        UserProfile.objects.create(user=user, role="operator", branch=parkent)
        for page in ["catalog", "dashboard"]:
            PagePermission.objects.create(user=user, page=page, can_view=True, can_control=True)
        client = APIClient()
        client.force_authenticate(user)
        branch_acc = client.get("/api/accounting/?branch=all")
        self.assertEqual(Decimal(branch_acc.data["summary"]["total_sales"]), Decimal("500000.00"))
        self.assertEqual(len(branch_acc.data["by_branch"]), 1)
        branch_dash = client.get("/api/dashboard/")
        self.assertEqual(Decimal(branch_dash.data["catalog_sales_revenue_today"]), Decimal("500000.00"))

    def test_accounting_waste_stays_with_main_branch(self):
        # chiqit faqat asosiy skladda bo'ladi, filial hisobiga tushmaydi
        self._sold_in_both_branches()
        StockMovement.objects.create(batch=self.batch, movement_type="waste", quantity_stems=-5, quantity_bunches=Decimal("0.5"), reason="Qurib qoldi")
        response = self.client.get("/api/accounting/")
        main = self._branch_row(response, "Toshkent (asosiy filial)")
        parkent = self._branch_row(response, "Parkent filiali")
        self.assertEqual(main["waste_stems"], 5)
        self.assertEqual(parkent["waste_stems"], 0)
        self.assertEqual(response.data["summary"]["waste_stems"], 5)
        branch_only = self.client.get(f"/api/accounting/?branch={self._parkent().id}")
        self.assertEqual(branch_only.data["summary"]["waste_stems"], 0)

    def test_volume_rate_requires_florist(self):
        response = self.client.post("/api/florist-volume-rates/", {"arrangement_type": "bouquet", "volume": "M", "default_stems": 25, "florist_fee": "60000"}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("florist", response.data)

    def test_volume_rate_is_per_florist(self):
        user_a = User.objects.create_user("fl-a", password="p")
        user_b = User.objects.create_user("fl-b", password="p")
        a = FloristProfile.objects.create(user=user_a, staff_type="florist")
        b = FloristProfile.objects.create(user=user_b, staff_type="florist")
        for profile, fee, stems in [(a, "60000", 25), (b, "90000", 35)]:
            response = self.client.post("/api/florist-volume-rates/", {"florist": profile.id, "arrangement_type": "bouquet", "volume": "M", "default_stems": stems, "florist_fee": fee}, format="json")
            self.assertEqual(response.status_code, 201)
        item_a = CatalogItem.objects.create(name_uz="A buket", arrangement_type="bouquet", volume="M", florist=a, price=Decimal("500000"), quantity_total=1, status="available")
        item_b = CatalogItem.objects.create(name_uz="B buket", arrangement_type="bouquet", volume="M", florist=b, price=Decimal("500000"), quantity_total=1, status="available")
        from .inventory_services import apply_volume_rate
        self.assertEqual(apply_volume_rate(item_a).florist_salary_amount, Decimal("60000"))
        self.assertEqual(apply_volume_rate(item_b).florist_salary_amount, Decimal("90000"))

    def test_catalog_detail_returns_unit_profit(self):
        item = CatalogItem.objects.create(name_uz="Foyda buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=2, status="available", florist_fee=Decimal("20000"))
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=5)
        sync_catalog_financials(item)
        response = self.client.get(f"/api/catalog/{item.id}/")
        self.assertEqual(response.status_code, 200)
        profit = response.data["profit"]
        # 2 dona uchun tannarx: 2*5*10000 + 2*20000 = 140000, bittasiga 70000
        self.assertEqual(Decimal(profit["unit_cost"]), Decimal("70000.00"))
        self.assertEqual(Decimal(profit["unit_price"]), Decimal("300000.00"))
        self.assertEqual(Decimal(profit["unit_profit"]), Decimal("230000.00"))
        self.assertEqual(Decimal(profit["total_potential_profit"]), Decimal("460000.00"))
        self.assertEqual(Decimal(profit["realized_profit"]), Decimal("0.00"))

    def test_florist_own_dashboard_hides_sales_and_profit(self):
        profile = self._florist_with_history()
        UserProfile.objects.update_or_create(user=profile.user, defaults={"role": "florist"})
        PagePermission.objects.create(user=profile.user, page="florists", can_view=True, can_control=False)
        client = APIClient()
        client.force_authenticate(profile.user)
        response = client.get("/api/florists/me/dashboard/")
        self.assertEqual(response.status_code, 200)
        summary = response.data["summary"]
        # Florist faqat nechta yasagani va qancha ish haqi olganini ko'radi
        self.assertIn("salary_total", summary)
        self.assertIn("catalog_count", summary)
        self.assertNotIn("sale_revenue", summary)
        self.assertNotIn("sold_quantity", summary)
        self.assertNotIn("unsold_quantity", summary)
        for row in response.data["salary_entries"]:
            self.assertNotIn("sale_revenue", row)
            self.assertNotIn("listed_price", row)
            self.assertNotIn("is_sold", row)
        for row in response.data["by_volume"]:
            self.assertNotIn("sale_revenue", row)

    def test_admin_florist_stats_still_shows_sales(self):
        profile = self._florist_with_history()
        response = self.client.get(f"/api/florists/{profile.id}/stats/")
        self.assertIn("sale_revenue", response.data["summary"])
        self.assertIn("sold_quantity", response.data["summary"])

    def test_supplier_payment_crud_and_rollups(self):
        supplier = Supplier.objects.create(name="Gul Import")
        variant = self.batch.variant
        StockBatch.objects.create(variant=variant, supplier=supplier, batch_number="SUP-1", height_cm=50, stems_per_bunch=25, received_stems=100, remaining_stems=100, cost_per_stem=Decimal("7000"), sale_price_per_stem=15000, sale_price_per_bunch=375000)
        StockBatch.objects.create(variant=variant, supplier=supplier, batch_number="SUP-2", height_cm=50, stems_per_bunch=25, received_stems=50, remaining_stems=50, cost_per_stem=Decimal("8000"), sale_price_per_stem=16000, sale_price_per_bunch=400000)
        created = self.client.post("/api/supplier-payments/", {"supplier": supplier.id, "amount": "500000.00", "paid_at": "2026-07-29", "method": "cash", "note": "Iyul oyi uchun"}, format="json")
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["method_label"], "Naqd")
        self.assertEqual(created.data["supplier_detail"]["name"], "Gul Import")
        self.client.post("/api/supplier-payments/", {"supplier": supplier.id, "amount": "200000.00", "paid_at": "2026-07-30", "method": "transfer"}, format="json")
        listed = self.client.get(f"/api/supplier-payments/?supplier={supplier.id}")
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed.data["count"], 2)
        detail = self.client.get(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(detail.status_code, 200)
        # 100*7000 + 50*8000 = 1 100 000
        self.assertEqual(Decimal(detail.data["purchase_total"]), Decimal("1100000.00"))
        self.assertEqual(Decimal(detail.data["paid_total"]), Decimal("700000.00"))
        # qarz ko'rsatkichi yo'q — postavshikdan har safar to'liq to'lab olinadi
        self.assertNotIn("outstanding", detail.data)
        self.assertEqual(str(detail.data["last_payment_at"]), "2026-07-30")

    def test_supplier_without_payments_reports_zero_paid(self):
        supplier = Supplier.objects.create(name="Yangi Postavshik")
        response = self.client.get(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(Decimal(response.data["purchase_total"]), Decimal("0"))
        self.assertEqual(Decimal(response.data["paid_total"]), Decimal("0"))
        self.assertNotIn("outstanding", response.data)

    def test_supplier_shows_total_purchased_without_debt(self):
        # postavshikdan har safar to'liq to'lab olinadi: faqat umumiy sotib olingan turadi
        supplier = Supplier.objects.create(name="Umumiy xarid")
        StockBatch.objects.create(
            variant=self.batch.variant, supplier=supplier, batch_number="BUY-1", height_cm=50,
            stems_per_bunch=25, received_stems=100, remaining_stems=100,
            cost_per_stem=7000, sale_price_per_stem=12000, sale_price_per_bunch=300000,
        )
        StockBatch.objects.create(
            variant=self.batch.variant, supplier=supplier, batch_number="BUY-2", height_cm=50,
            stems_per_bunch=25, received_stems=50, remaining_stems=50,
            cost_per_stem=8000, sale_price_per_stem=14000, sale_price_per_bunch=350000,
        )
        detail = self.client.get(f"/api/suppliers/{supplier.id}/")
        # 100*7000 + 50*8000 = 1 100 000
        self.assertEqual(Decimal(detail.data["purchase_total"]), Decimal("1100000.00"))
        self.assertEqual(detail.data["total_received_stems"], 150)
        self.assertEqual(detail.data["batches_count"], 2)
        self.assertNotIn("outstanding", detail.data)

    def test_free_batch_not_counted_in_total_purchased(self):
        supplier = Supplier.objects.create(name="Tekin qo‘shgan")
        StockBatch.objects.create(
            variant=self.batch.variant, supplier=supplier, batch_number="BUY-3", height_cm=50,
            stems_per_bunch=25, received_stems=100, remaining_stems=100,
            cost_per_stem=5000, sale_price_per_stem=10000, sale_price_per_bunch=250000,
        )
        StockBatch.objects.create(
            variant=self.batch.variant, supplier=supplier, batch_number="BUY-4", height_cm=50,
            stems_per_bunch=25, received_stems=40, remaining_stems=40, is_free=True,
            cost_per_stem=0, sale_price_per_stem=10000, sale_price_per_bunch=250000,
        )
        detail = self.client.get(f"/api/suppliers/{supplier.id}/")
        # faqat sotib olingani: 100 * 5 000
        self.assertEqual(Decimal(detail.data["purchase_total"]), Decimal("500000.00"))
        self.assertEqual(detail.data["total_received_stems"], 140)

    def test_suppliers_can_be_ordered_by_total_purchased(self):
        response = self.client.get("/api/suppliers/?ordering=-purchase_total")
        self.assertEqual(response.status_code, 200)
        bad = self.client.get("/api/suppliers/?ordering=outstanding")
        self.assertEqual(bad.status_code, 200)

    def test_supplier_payment_rejects_non_positive_amount(self):
        supplier = Supplier.objects.create(name="Test Postavshik")
        response = self.client.post("/api/supplier-payments/", {"supplier": supplier.id, "amount": "0", "paid_at": "2026-07-29", "method": "cash"}, format="json")
        self.assertEqual(response.status_code, 400)

    def test_accounting_summary_splits_cost_and_reports_waste(self):
        packaging = Packaging.objects.create(packaging_type="wrap", name_uz="Qog‘oz", cost_price=Decimal("5000"), sale_price=Decimal("9000"), quantity=100)
        item = CatalogItem.objects.create(name_uz="Hisob buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=1, status="available", florist_fee=Decimal("20000"))
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=10)
        CatalogMaterialUsage.objects.create(catalog_item=item, packaging=packaging, quantity=2)
        sync_catalog_financials(item)
        mark_catalog_sold(item, self.user)
        StockMovement.objects.create(batch=self.batch, movement_type="waste", quantity_stems=-4, reason="so‘lgan")
        response = self.client.get("/api/accounting/")
        self.assertEqual(response.status_code, 200)
        summary = response.data["summary"]
        self.assertEqual(Decimal(summary["flower_cost_total"]), Decimal("100000.00"))
        self.assertEqual(Decimal(summary["material_cost_total"]), Decimal("10000.00"))
        self.assertEqual(Decimal(summary["florist_fee_cost_total"]), Decimal("20000.00"))
        self.assertEqual(Decimal(summary["cost_total"]), Decimal("130000.00"))
        self.assertEqual(Decimal(summary["waste_cost_total"]), Decimal("40000.00"))
        self.assertEqual(summary["waste_stems"], 4)
        row = response.data["history"][0]
        self.assertEqual(Decimal(row["flower_cost"]) + Decimal(row["material_cost"]) + Decimal(row["florist_fee_cost"]), Decimal(row["cost_total"]))

    def test_stock_movement_exposes_cost_value(self):
        StockMovement.objects.create(batch=self.batch, movement_type="waste", quantity_stems=-3, reason="chiqit")
        response = self.client.get("/api/stock-movements/")
        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(Decimal(row["cost_value"]), Decimal("30000.00"))
        self.assertEqual(Decimal(row["sale_value"]), Decimal("60000.00"))

    def test_dashboard_requires_authentication(self):
        response = APIClient().get("/api/dashboard/")
        self.assertEqual(response.status_code, 401)

    def test_dashboard_includes_daily_chart_stats_for_default_month(self):
        customer = Customer.objects.create(name="Chart User", phone="+998901234567", instagram_user_id="chart-user")
        conversation = Conversation.objects.create(customer=customer)
        lead = Lead.objects.create(customer=customer, conversation=conversation, request_uz="Chart lead", arrangement_type="catalog")
        today = timezone.localdate()
        created_at = timezone.now()
        Conversation.objects.filter(id=conversation.id).update(created_at=created_at)
        Lead.objects.filter(id=lead.id).update(created_at=created_at)
        response = self.client.get("/api/dashboard/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data["daily_stats"]), 30)
        self.assertEqual(data["daily_stats"][-1]["date"], today.isoformat())
        self.assertGreaterEqual(data["daily_stats"][-1]["leads"], 1)
        self.assertGreaterEqual(data["daily_stats"][-1]["conversations"], 1)

    @override_settings(INSTAGRAM_VERIFY_TOKEN="verify")
    def test_instagram_webhook_verification(self):
        response = APIClient().get("/api/instagram/webhook/?hub.verify_token=verify&hub.challenge=123")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), 123)

    def test_customers_list_hides_incomplete_placeholders(self):
        Customer.objects.create(instagram_user_id="placeholder")
        Customer.objects.create(instagram_user_id="complete", name="Ahmad", phone="+998901234567")
        response = self.client.get("/api/customers/")
        self.assertEqual(response.status_code, 200)
        ids = [row["instagram_user_id"] for row in response.json()["results"]]
        self.assertIn("complete", ids)
        self.assertNotIn("placeholder", ids)
        response = self.client.get("/api/customers/?include_incomplete=true")
        ids = [row["instagram_user_id"] for row in response.json()["results"]]
        self.assertIn("placeholder", ids)

    def test_customer_delete_archives_when_leads_exist(self):
        customer = Customer.objects.create(instagram_user_id="delete-me", name="Ahmad", phone="+998901234567")
        Lead.objects.create(customer=customer, request_uz="Test buyurtma")
        response = self.client.delete(f"/api/customers/{customer.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["archived"])
        customer.refresh_from_db()
        self.assertEqual(customer.name, "")
        self.assertEqual(customer.phone, "")
        self.assertEqual(customer.instagram_user_id, f"deleted:{customer.id}")
        response = self.client.get("/api/customers/")
        ids = [row["id"] for row in response.json()["results"]]
        self.assertNotIn(customer.id, ids)

    def test_social_post_response_includes_lead_ids(self):
        post = SocialPost.objects.create(post_type="story", title_uz="Story buket", title_ru="Story bouquet", is_active=True)
        customer = Customer.objects.create(name="Madina", phone="+998901234567", instagram_user_id="ig-lead")
        lead = Lead.objects.create(customer=customer, social_post=post, status="won", request_uz="Storydagi buket", arrangement_type="catalog", estimated_price=400000)
        item = CatalogItem.objects.create(social_post=post, name_uz="Qizil buket", arrangement_type="bouquet", price=400000, quantity_total=4, status="available")
        lead.catalog_usage.create(catalog_item=item, quantity=1)
        response = self.client.get(f"/api/social-posts/{post.id}/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["lead_count"], 1)
        self.assertEqual(data["leads"][0]["id"], lead.id)
        self.assertEqual(data["leads"][0]["customer_name"], "Madina")
        self.assertEqual(data["leads"][0]["catalog_items"][0]["id"], item.id)

    def test_social_post_create_accepts_catalog_composition(self):
        response = self.client.post("/api/social-posts/", {
                        "post_type": "post",
            "media_id": "api-post-composition",
            "title_uz": "Atirgul post",
            "title_ru": "Розовый пост",
            "price": "400000.00",
            "flower_count": 3,
            "is_active": True,
            "catalog_items": [{
                "name_uz": "Qizil atirgul buket",
                "arrangement_type": "bouquet",
                "price": "400000.00",
                "quantity_total": 4,
                "status": "available",
                "composition": [{
                    "stock_batch": self.batch.id,
                    "quantity_stems": 3,
                    "quantity_bunches": "0.15"
                }]
            }]
        }, format="json")
        self.assertEqual(response.status_code, 201)
        post = SocialPost.objects.get(media_id="api-post-composition")
        item = post.catalog_items.get()
        composition = item.composition.get()
        self.assertEqual(item.quantity_total, 4)
        self.assertEqual(composition.stock_batch_id, self.batch.id)
        self.assertEqual(composition.quantity_stems, 3)
        self.assertEqual(response.json()["catalog_items"][0]["composition"][0]["stock_batch"], self.batch.id)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 88)

    def test_social_post_custom_catalog_merges_multiple_flower_payloads_into_one_item(self):
        flower = Flower.objects.create(name_uz="Pion API", slug="pion-api")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Sarah", color_uz="Pushti")
        second_batch = StockBatch.objects.create(variant=variant, batch_number="API-2", height_cm=55, stems_per_bunch=10, received_stems=80, remaining_stems=80, cost_per_stem=50000, sale_price_per_stem=80000, sale_price_per_bunch=800000)
        response = self.client.post("/api/social-posts/", {
                        "post_type": "post",
            "media_id": "api-post-custom-merge",
            "title_uz": "Custom",
            "title_ru": "Custom",
            "price": "0.00",
            "is_active": True,
            "catalog_items": [
                {
                    "name_uz": "Atirgul custom",
                    "arrangement_type": "bouquet",
                    "catalog_kind": "custom",
                    "price": "120000.00",
                    "florist_salary_amount": "40000.00",
                    "discount_reason": "Custom jamlangan skidka",
                    "quantity_total": 1,
                    "composition": [{"stock_batch": self.batch.id, "quantity_stems": 4, "quantity_bunches": "0.20"}],
                },
                {
                    "name_uz": "Pion custom",
                    "arrangement_type": "bouquet",
                    "catalog_kind": "custom",
                    "price": "240000.00",
                    "florist_salary_amount": "60000.00",
                    "discount_reason": "Custom jamlangan skidka",
                    "quantity_total": 1,
                    "composition": [{"stock_batch": second_batch.id, "quantity_stems": 3, "quantity_bunches": "0.30"}],
                },
            ],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        post = SocialPost.objects.get(media_id="api-post-custom-merge")
        self.assertEqual(post.catalog_items.count(), 1)
        item = post.catalog_items.get()
        self.assertEqual(item.catalog_kind, "custom")
        self.assertEqual(item.status, "draft")
        self.assertEqual(item.quantity_sold, 0)
        self.assertIsNone(item.sold_at)
        self.assertEqual(item.price, Decimal("360000.00"))
        self.assertEqual(item.florist_salary_amount, Decimal("100000.00"))
        self.assertEqual(item.composition.count(), 2)
        self.assertTrue(item.composition.filter(stock_batch=self.batch, quantity_stems=4).exists())
        self.assertTrue(item.composition.filter(stock_batch=second_batch, quantity_stems=3).exists())

    def test_catalog_create_deducts_flowers_and_materials_then_delete_restores_unsold(self):
        material = Packaging.objects.create(packaging_type="wrap", name_uz="Koreya qogoz", quantity=10, sale_price=50000)
        payload = {
            "name_uz": "Materialli buket",
            "arrangement_type": "bouquet",
            "price": "450000.00",
            "quantity_total": 3,
            "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 4}],
            "materials": [{"packaging": material.id, "quantity": 2}],
        }
        response = self.client.post("/api/catalog/", payload, format="json")
        self.assertEqual(response.status_code, 201)
        item_id = response.json()["id"]
        self.batch.refresh_from_db()
        material.refresh_from_db()
        item = CatalogItem.objects.get(id=item_id)
        self.assertEqual(self.batch.remaining_stems, 88)
        self.assertEqual(material.quantity, 4)
        self.assertEqual(item.quantity_stock_deducted, 3)
        self.assertTrue(CatalogMaterialUsage.objects.filter(catalog_item=item, packaging=material, quantity=2).exists())
        response = self.client.patch(f"/api/catalog/{item_id}/", {"quantity_total": 2}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        material.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 92)
        self.assertEqual(material.quantity, 6)
        self.assertEqual(item.quantity_stock_deducted, 2)
        response = self.client.delete(f"/api/catalog/{item_id}/")
        self.assertEqual(response.status_code, 204)
        self.batch.refresh_from_db()
        material.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 100)
        self.assertEqual(material.quantity, 10)

    def test_catalog_create_merges_duplicate_composition_and_material_rows(self):
        material = Packaging.objects.create(packaging_type="wrap", name_uz="Dubl qogoz", quantity=20, sale_price=50000)
        response = self.client.post("/api/catalog/", {
            "name_uz": "Dubl rows buket",
            "arrangement_type": "bouquet",
            "price": "450000.00",
            "quantity_total": 2,
            "status": "available",
            "composition": [
                {"stock_batch": self.batch.id, "quantity_stems": 2, "quantity_bunches": "0.10"},
                {"stock_batch": self.batch.id, "quantity_stems": 3, "quantity_bunches": "0.15"},
            ],
            "materials": [
                {"packaging": material.id, "quantity": 1},
                {"packaging": material.id, "quantity": 2},
            ],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        item = CatalogItem.objects.get(id=response.json()["id"])
        self.assertEqual(item.composition.count(), 1)
        self.assertEqual(item.materials.count(), 1)
        self.assertEqual(item.composition.get().quantity_stems, 5)
        self.assertEqual(item.composition.get().quantity_bunches, Decimal("0.25"))
        self.assertEqual(item.materials.get().quantity, 3)
        self.batch.refresh_from_db()
        material.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 90)
        self.assertEqual(material.quantity, 14)

    def test_catalog_create_adds_decoration_salary_from_selected_florist(self):
        maker_user = User.objects.create_user("maker-decoration", password="password")
        decorator_user = User.objects.create_user("decorator-decoration", password="password")
        maker = FloristProfile.objects.create(user=maker_user, staff_type="florist")
        decorator = FloristProfile.objects.create(user=decorator_user, staff_type="florist", decoration_fee=Decimal("75000"))
        FloristVolumeRate.objects.create(florist=maker, arrangement_type="bouquet", volume="M", florist_fee=Decimal("50000"), is_active=True)
        from .inventory_services import issue_stock_to_florist
        issue_stock_to_florist(maker, self.batch, 10, user=self.user)
        response = self.client.post("/api/catalog/", {
            "name_uz": "Oformleniya buket",
            "arrangement_type": "bouquet",
            "volume": "M",
            "price": "500000.00",
            "quantity_total": 2,
            "status": "available",
            "florist": maker.id,
            "decoration_florist": decorator.id,
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 5}],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        item = CatalogItem.objects.get(id=response.json()["id"])
        self.assertEqual(item.decoration_salary_amount, Decimal("75000.00"))
        self.assertTrue(FloristSalaryEntry.objects.filter(florist=decorator, catalog_item=item, source="decoration", amount=Decimal("150000.00")).exists())
        self.assertTrue(FloristSalaryEntry.objects.filter(florist=maker, catalog_item=item, source="catalog", amount=Decimal("100000.00")).exists())

    def test_catalog_sell_can_deduct_extra_material_and_add_decoration_salary(self):
        material = Packaging.objects.create(packaging_type="wrap", name_uz="Sotuv qogoz", quantity=10, sale_price=50000)
        decorator_user = User.objects.create_user("sale-decorator", password="password")
        decorator = FloristProfile.objects.create(user=decorator_user, staff_type="florist", decoration_fee=Decimal("40000"))
        item = CatalogItem.objects.create(name_uz="Sotuv material buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=3, status="available")
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 2,
            "sale_price": "300000.00",
            "materials": [{"packaging": material.id, "quantity": 2}],
            "decoration_florist": decorator.id,
        }, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        item.refresh_from_db()
        material.refresh_from_db()
        self.assertEqual(item.quantity_sold, 2)
        self.assertEqual(material.quantity, 6)
        history = item.history.get(action="sold")
        self.assertEqual(history.snapshot["sale_materials"][0]["quantity"], 4)
        self.assertEqual(history.snapshot["sale_decoration"]["amount"], "80000.00")
        self.assertTrue(PackagingMovement.objects.filter(packaging=material, reference_type="catalog_sale", reference_id=history.id, quantity=-4).exists())
        self.assertTrue(FloristSalaryEntry.objects.filter(florist=decorator, catalog_item=item, source="sale_decoration", amount=Decimal("80000.00")).exists())

    def test_catalog_restore_sale_returns_item_to_available_and_removes_accounting_sale(self):
        item = CatalogItem.objects.create(name_uz="Qaytariladigan buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=1, status="available")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        response = self.client.post(f"/api/catalog/{item.id}/restore-sale/", {"reason": "Xato sotildi qilingan"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 0)
        self.assertEqual(item.status, "available")
        self.assertFalse(CatalogHistory.objects.filter(catalog_item=item, action="sold").exists())
        self.assertTrue(CatalogHistory.objects.filter(catalog_item=item, action="sale_restored").exists())
        report = self.client.get("/api/accounting/")
        self.assertEqual(Decimal(report.data["summary"]["total_sales"]), Decimal("0"))

    def test_catalog_restore_sale_partially_restores_material_and_decoration_salary(self):
        material = Packaging.objects.create(packaging_type="wrap", name_uz="Qaytariladigan qogoz", quantity=10, sale_price=50000)
        decorator_user = User.objects.create_user("restore-decorator", password="password")
        decorator = FloristProfile.objects.create(user=decorator_user, staff_type="florist", decoration_fee=Decimal("40000"))
        item = CatalogItem.objects.create(name_uz="Partial qaytarish buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=3, status="available")
        sell = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 2,
            "payment_type": "cash",
            "materials": [{"packaging": material.id, "quantity": 2}],
            "decoration_florist": decorator.id,
        }, format="json")
        self.assertEqual(sell.status_code, 200, sell.data)
        history = CatalogHistory.objects.get(catalog_item=item, action="sold")
        response = self.client.post(f"/api/catalog/{item.id}/restore-sale/", {
            "sale_history": history.id,
            "quantity": 1,
            "reason": "Bitta dona xato sotildi",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        item.refresh_from_db()
        material.refresh_from_db()
        history.refresh_from_db()
        self.assertEqual(item.quantity_sold, 1)
        self.assertEqual(history.quantity, 1)
        self.assertEqual(material.quantity, 8)
        self.assertTrue(PackagingMovement.objects.filter(packaging=material, reference_type="catalog_sale_restore", quantity=2).exists())
        self.assertEqual(FloristSalaryEntry.objects.get(florist=decorator, catalog_item=item, source="sale_decoration").amount, Decimal("40000.00"))

    def test_catalog_sell_api_accepts_discounted_price_with_reason(self):
        item = CatalogItem.objects.create(name_uz="API skidka buket", arrangement_type="bouquet", price=500000, quantity_total=2, status="available")
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", "quantity": 1, "sale_price": "450000.00"}, format="json")
        self.assertEqual(response.status_code, 400)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", "quantity": 1, "sale_price": "450000.00", "discount_reason": "VIP mijoz", "payment_type": "card"}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 1)
        history = item.history.get(action="sold")
        self.assertEqual(history.discount_amount, Decimal("50000.00"))
        self.assertEqual(history.discount_reason, "VIP mijoz")
        self.assertEqual(history.snapshot["payment_type"], "card")

    def test_conversation_response_includes_source(self):
        instagram_customer = Customer.objects.create(name="Instagram", phone="+998901234567", instagram_user_id="ig-source")
        telegram_customer = Customer.objects.create(name="Telegram", phone="+998901234568", instagram_user_id="telegram:123")
        instagram_conversation = Conversation.objects.create(customer=instagram_customer)
        telegram_conversation = Conversation.objects.create(customer=telegram_customer)
        response = self.client.get(f"/api/conversations/{instagram_conversation.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["source"], "instagram")
        self.assertEqual(response.json()["source_label"], "Instagram")
        response = self.client.get(f"/api/conversations/{telegram_conversation.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["source"], "telegram")
        self.assertEqual(response.json()["source_label"], "Telegram")

    def test_operator_send_uses_telegram_for_telegram_conversation(self):
        customer = Customer.objects.create(name="Telegram", phone="+998901234568", instagram_user_id="telegram:123")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="Salom", instagram_message_id="telegram:555:77")
        from unittest.mock import patch
        with patch("core.views.telegram_send", return_value={"ok": True}) as telegram_mock, patch("core.views.instagram_send", return_value={"ok": True}) as instagram_mock:
            response = self.client.post(f"/api/conversations/{conversation.id}/send/", {"text": "Javob"}, format="json")
        self.assertEqual(response.status_code, 200)
        telegram_mock.assert_called_once_with("555", "Javob")
        instagram_mock.assert_not_called()

    def test_operator_send_stores_instagram_message_id(self):
        customer = Customer.objects.create(name="Instagram", phone="+998901234567", instagram_user_id="ig-source")
        conversation = Conversation.objects.create(customer=customer)
        from unittest.mock import patch
        with patch("core.views.instagram_send", return_value={"message_id": "mid-crm-send-1"}):
            response = self.client.post(f"/api/conversations/{conversation.id}/send/", {"text": "Javob"}, format="json")
        self.assertEqual(response.status_code, 200)
        message = conversation.messages.get(sender="operator", text="Javob")
        self.assertEqual(message.instagram_message_id, "mid-crm-send-1")

    def test_operator_send_records_failed_delivery_when_instagram_rejects(self):
        customer = Customer.objects.create(name="Instagram", phone="+998901234567", instagram_user_id="ig-source")
        conversation = Conversation.objects.create(customer=customer)
        response_obj = requests.Response()
        response_obj.status_code = 403
        response_obj._content = b'{"error":{"message":"Forbidden"}}'
        error = requests.HTTPError("403 Client Error", response=response_obj)
        from unittest.mock import patch
        with patch("core.views.instagram_send", side_effect=error):
            response = self.client.post(f"/api/conversations/{conversation.id}/send/", {"text": "Javob"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["delivery_status"], "failed")
        self.assertEqual(response.json()["platform_status"], 403)
        message = conversation.messages.get(sender="operator", text="Javob")
        self.assertEqual(message.metadata["delivery_status"], "failed")

    def test_branches_endpoint_lists_seeded_branches(self):
        response = self.client.get("/api/branches/")
        self.assertEqual(response.status_code, 200)
        names = {row["name"] for row in response.data["results"]}
        self.assertIn("Asosiy filial", names)
        self.assertIn("Parkent filiali", names)

    def test_admin_permission_matrix_uses_saved_rows(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="users", can_view=True, can_control=True)
        PagePermission.objects.create(user=self.user, page="catalog", can_view=False, can_control=False)
        response = self.client.get(f"/api/users/{self.user.id}/")
        self.assertEqual(response.status_code, 200)
        permissions = {row["page"]: row for row in response.json()["permissions"]}
        self.assertEqual(permissions["catalog"]["can_view"], False)
        self.assertEqual(permissions["catalog"]["can_control"], False)
        developer = User.objects.create_user("developer", password="password")
        UserProfile.objects.create(user=developer, role="developer")
        developer_permissions = {row["page"]: row for row in permission_matrix(developer)}
        self.assertTrue(developer_permissions["catalog"]["can_view"])
        self.assertTrue(developer_permissions["catalog"]["can_control"])
        admin_permissions = {row["page"]: row for row in permission_matrix(self.user)}
        self.assertNotIn("ai_settings", admin_permissions)
        self.assertNotIn("integrations", admin_permissions)
        self.assertIn("audit", admin_permissions)

    def test_admin_cannot_grant_developer_only_permissions(self):
        UserProfile.objects.create(user=self.user, role="admin")
        operator = User.objects.create_user("limited", password="password")
        UserProfile.objects.create(user=operator, role="operator")
        response = self.client.post("/api/permissions/", {"user": operator.id, "page": "ai_settings", "can_view": True, "can_control": True}, format="json")
        self.assertEqual(response.status_code, 400)
        response = self.client.post("/api/users/", {"username": "bad-user", "password": "password", "role": "operator", "permissions": [{"page": "integrations", "can_view": True, "can_control": True}]}, format="json")
        self.assertEqual(response.status_code, 400)
        response = self.client.get("/api/audit/")
        self.assertEqual(response.status_code, 200)

    def test_branchless_mode_hides_branch_fields(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="users", can_view=True, can_control=True)
        response = self.client.post("/api/users/", {
            "username": "branchless-user",
            "password": "Password123!",
            "role": "operator",
            "permissions": [{"page": "conversations", "can_view": True, "can_control": True}]
        }, format="json")
        self.assertEqual(response.status_code, 201)
        created = User.objects.get(username="branchless-user")
        self.assertFalse(hasattr(created.profile, "branches"))
        self.assertNotIn("branches", response.json()["profile"])
        self.assertIsNone(created.profile.branch_id)
        response = self.client.post("/api/packaging/", {"packaging_type": "basket", "name_uz": "Branchsiz savat", "quantity": 1, "sale_price": "100000.00"}, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertNotIn("branch", response.json())
        response = self.client.get("/api/dashboard/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("net_profit", response.json())
        self.assertIn("batch_inventory_stats", response.json())
        self.assertIn("florist_production_stats", response.json())
        self.assertNotIn("branch_stock", response.json())

    def test_user_can_change_own_password(self):
        UserProfile.objects.create(user=self.user, role="admin")
        response = self.client.post("/api/me/change-password/", {
            "old_password": "wrong",
            "new_password": "NewPassword123!",
            "new_password_confirm": "NewPassword123!",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        response = self.client.post("/api/me/change-password/", {
            "old_password": "password",
            "new_password": "NewPassword123!",
            "new_password_confirm": "NewPassword123!",
        }, format="json")
        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password("NewPassword123!"))

    def test_florist_notifications_are_targeted(self):
        UserProfile.objects.create(user=self.user, role="admin")
        florist_user = User.objects.create_user("target-florist", password="password", first_name="Ali")
        UserProfile.objects.create(user=florist_user, role="florist")
        FloristProfile.objects.create(user=florist_user, staff_type="florist")
        PagePermission.objects.create(user=florist_user, page="notifications", can_view=True, can_control=False)
        Notification.objects.create(notification_type="lead", title_uz="Global", title_ru="Global", body_uz="Global", body_ru="Global")
        target = Notification.objects.create(target_user=florist_user, notification_type="florist_salary", title_uz="Target", title_ru="Target", body_uz="Target", body_ru="Target")
        assigned = Notification.objects.create(target_user=florist_user, notification_type="florist_catalog", title_uz="Yangi ish biriktirildi", title_ru="Yangi ish biriktirildi", body_uz="Yangi ish biriktirildi", body_ru="Yangi ish biriktirildi")
        Notification.objects.create(target_user=florist_user, notification_type="florist_catalog", title_uz="Katalog sotildi", title_ru="Katalog sotildi", body_uz="Katalog sotildi", body_ru="Katalog sotildi")
        self.client.force_authenticate(florist_user)
        response = self.client.get("/api/notifications/")
        self.assertEqual(response.status_code, 200)
        ids = [row["id"] for row in response.json()["results"]]
        self.assertEqual(ids, [assigned.id, target.id])
        self.client.force_authenticate(self.user)
        response = self.client.get("/api/notifications/")
        self.assertEqual(response.status_code, 200)
        titles = [row["title_uz"] for row in response.json()["results"]]
        self.assertIn("Global", titles)
        self.assertNotIn("Target", titles)

    def test_florist_can_only_see_own_profile_and_salary(self):
        florist_user = User.objects.create_user("own-florist", password="password", first_name="Ali")
        other_user = User.objects.create_user("other-florist", password="password", first_name="Vali")
        UserProfile.objects.create(user=florist_user, role="florist")
        UserProfile.objects.create(user=other_user, role="florist")
        own_profile = FloristProfile.objects.create(user=florist_user, staff_type="florist")
        other_profile = FloristProfile.objects.create(user=other_user, staff_type="florist")
        FloristSalaryEntry.objects.create(florist=own_profile, amount=Decimal("100000"), work_date=timezone.localdate(), source="manual")
        FloristSalaryEntry.objects.create(florist=other_profile, amount=Decimal("200000"), work_date=timezone.localdate(), source="manual")
        PagePermission.objects.create(user=florist_user, page="florists", can_view=True, can_control=False)
        self.client.force_authenticate(florist_user)
        response = self.client.get("/api/florists/")
        self.assertEqual(response.status_code, 200)
        profile_ids = [row["id"] for row in response.json()["results"]]
        self.assertEqual(profile_ids, [own_profile.id])
        response = self.client.get("/api/florist-salary/")
        self.assertEqual(response.status_code, 200)
        salary_florist_ids = [row["florist"] for row in response.json()["results"]]
        self.assertEqual(salary_florist_ids, [own_profile.id])

    def test_florist_role_cannot_access_global_business_pages_even_with_permissions(self):
        florist_user = User.objects.create_user("restricted-florist", password="password", first_name="Ali")
        UserProfile.objects.create(user=florist_user, role="florist")
        FloristProfile.objects.create(user=florist_user, staff_type="florist")
        for page in ["dashboard", "catalog", "expenses", "audit"]:
            PagePermission.objects.create(user=florist_user, page=page, can_view=True, can_control=True)
        for page in ["florists", "attendance", "notifications"]:
            PagePermission.objects.create(user=florist_user, page=page, can_view=True, can_control=False)
        self.client.force_authenticate(florist_user)
        for path in ["/api/dashboard/", "/api/analytics/", "/api/catalog/", "/api/expenses/", "/api/audit/"]:
            response = self.client.get(path)
            self.assertEqual(response.status_code, 403, path)
        response = self.client.get("/api/florists/me/dashboard/")
        self.assertEqual(response.status_code, 200)
        pages = [row["page"] for row in permission_matrix(florist_user)]
        self.assertEqual(set(pages), {"florists", "attendance", "notifications"})
        self.assertEqual(len(pages), 3)

    def test_notification_mark_read_marks_single_notification(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="notifications", can_view=True, can_control=True)
        first = Notification.objects.create(notification_type="lead", title_uz="First", title_ru="First", body_uz="", body_ru="")
        second = Notification.objects.create(notification_type="lead", title_uz="Second", title_ru="Second", body_uz="", body_ru="")
        self.client.force_authenticate(self.user)
        response = self.client.post(f"/api/notifications/{first.id}/mark-read/")
        self.assertEqual(response.status_code, 200)
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertTrue(first.is_read)
        self.assertFalse(second.is_read)

    def test_lead_status_api_does_not_expose_name_ru(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="crm", can_view=True, can_control=True)
        LeadStatus.objects.create(key="test-status", name_uz="Test status", color="#111111", order=1)
        self.client.force_authenticate(self.user)
        response = self.client.get("/api/lead-statuses/")
        self.assertEqual(response.status_code, 200)
        row = response.json()["results"][0]
        self.assertIn("name_uz", row)
        self.assertNotIn("name_ru", row)

    def test_admin_gets_notification_when_florist_checks_in(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="notifications", can_view=True, can_control=False)
        florist_user = User.objects.create_user("checkin-florist", password="password", first_name="Ali")
        UserProfile.objects.create(user=florist_user, role="florist")
        FloristProfile.objects.create(user=florist_user, staff_type="florist")
        PagePermission.objects.create(user=florist_user, page="attendance", can_view=True, can_control=True)
        self.client.force_authenticate(florist_user)
        response = self.client.post("/api/florist-attendance/check-in/", {"checked_at": "2026-07-27T09:00:00+05:00"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.client.force_authenticate(self.user)
        response = self.client.get("/api/notifications/?notification_type=attendance")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(any("ishga keldi" in row["title_uz"] for row in response.json()["results"]))

    def test_florist_check_in_accepts_precise_mobile_location(self):
        florist_user = User.objects.create_user("precise-location", password="password", first_name="Ali")
        UserProfile.objects.create(user=florist_user, role="florist")
        FloristProfile.objects.create(user=florist_user, staff_type="florist", shop_latitude=Decimal("41.2954351234"), shop_longitude=Decimal("69.2503551234"))
        PagePermission.objects.create(user=florist_user, page="attendance", can_view=True, can_control=True)
        self.client.force_authenticate(florist_user)
        response = self.client.post("/api/florist-attendance/check-in/", {
            "checked_at": "2026-07-27T09:00:00+05:00",
            "latitude": "41.29543512345678",
            "longitude": "69.25035512345678",
        }, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["check_in_latitude"], "41.2954351235")
        self.assertEqual(response.json()["check_in_longitude"], "69.2503551235")

    def test_admin_does_not_see_developer_notifications(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="notifications", can_view=True, can_control=False)
        developer = User.objects.create_user("hidden-developer", password="password", first_name="Developer")
        UserProfile.objects.create(user=developer, role="developer")
        developer_profile = FloristProfile.objects.create(user=developer, staff_type="florist")
        attendance = FloristAttendance.objects.create(florist=developer_profile, work_date=timezone.localdate(), check_in_at=timezone.now())
        Notification.objects.create(notification_type="attendance", title_uz="Developer ishga keldi", title_ru="Developer ishga keldi", body_uz="Developer ishga keldi", body_ru="Developer ishga keldi", reference_type="attendance", reference_id=attendance.id)
        Notification.objects.create(target_user=developer, notification_type="attendance", title_uz="Developer target", title_ru="Developer target", body_uz="Developer target", body_ru="Developer target")
        self.client.force_authenticate(self.user)
        response = self.client.get("/api/notifications/")
        self.assertEqual(response.status_code, 200)
        titles = [row["title_uz"] for row in response.json()["results"]]
        self.assertFalse(any("Developer" in title for title in titles))

    def test_developer_check_in_does_not_create_visible_notifications(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="notifications", can_view=True, can_control=False)
        developer = User.objects.create_user("checkin-developer", password="password", first_name="Developer")
        UserProfile.objects.create(user=developer, role="developer")
        FloristProfile.objects.create(user=developer, staff_type="florist")
        PagePermission.objects.create(user=developer, page="attendance", can_view=True, can_control=True)
        self.client.force_authenticate(developer)
        response = self.client.post("/api/florist-attendance/check-in/", {"checked_at": "2026-07-27T09:00:00+05:00"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Notification.objects.filter(title_uz__icontains="Developer").exists())
        self.client.force_authenticate(self.user)
        response = self.client.get("/api/notifications/?notification_type=attendance")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["results"], [])

    def test_permissions_pagination_does_not_conflict_with_permission_page_filter(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="users", can_view=True, can_control=True)
        users = [self.user]
        for index in range(3):
            user = User.objects.create_user(f"operator-{index}", password="password")
            UserProfile.objects.create(user=user, role="operator")
            users.append(user)
        for user in users:
            for page, _ in PagePermission.PAGE_CHOICES:
                PagePermission.objects.get_or_create(user=user, page=page, defaults={"can_view": True, "can_control": True})
        response = self.client.get("/api/permissions/?page=2")
        self.assertEqual(response.status_code, 200)
        self.assertIn("results", response.json())
        response = self.client.get("/api/permissions/?permission_page=users")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(all(row["page"] == "users" for row in response.json()["results"]))

    def test_packaging_movement_updates_quantity(self):
        response = self.client.post("/api/packaging/", {"packaging_type": "basket", "name_uz": "API savat", "quantity": 4, "sale_price": "90000.00"}, format="json")
        self.assertEqual(response.status_code, 201)
        api_packaging = Packaging.objects.get(id=response.json()["id"])
        self.assertTrue(PackagingMovement.objects.filter(packaging=api_packaging, movement_type="in", quantity=4).exists())
        response = self.client.patch(f"/api/packaging/{api_packaging.id}/", {"quantity": 6}, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(PackagingMovement.objects.filter(packaging=api_packaging, movement_type="adjustment", quantity=2).exists())
        packaging = Packaging.objects.create(packaging_type="basket", name_uz="Test savat", quantity=10, sale_price=100000)
        response = self.client.post(f"/api/packaging/{packaging.id}/movement/", {"movement_type": "out", "quantity": 3, "reason": "test"}, format="json")
        self.assertEqual(response.status_code, 200)
        packaging.refresh_from_db()
        self.assertEqual(packaging.quantity, 7)
        self.assertTrue(PackagingMovement.objects.filter(packaging=packaging, quantity=-3).exists())
        response = self.client.post(f"/api/packaging/{packaging.id}/movement/", {"movement_type": "adjustment", "quantity": -2, "reason": "count"}, format="json")
        self.assertEqual(response.status_code, 200)
        packaging.refresh_from_db()
        self.assertEqual(packaging.quantity, 5)
        response = self.client.post(f"/api/packaging/{packaging.id}/movement/", {"movement_type": "out", "quantity": 99}, format="json")
        self.assertEqual(response.status_code, 400)

    def test_stock_batch_create_calculates_received_stems_from_bunches(self):
        flower = Flower.objects.create(name_uz="Gortenziya API", slug="gortenziya-api")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Golland", color_uz="Moviy")
        response = self.client.post("/api/stock-batches/", {
            "variant": variant.id,
            "batch_number": "BUNCH-1",
            "height_cm": 50,
            "stems_per_bunch": 5,
            "received_bunches": "3.00",
            "cost_per_stem": "50000.00",
            "sale_price_per_stem": "80000.00",
            "sale_price_per_bunch": "400000.00",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(response.json()["received_stems"], 15)
        self.assertEqual(response.json()["remaining_stems"], 15)
        self.assertEqual(response.json()["remaining_bunches"], "3.00")

    def _florist_with_leftover(self, issued=100, per_item=25, items=3, quantity_total=1):
        """Skladdan gul olib, standart bo'yicha katalog yasagan florist."""
        self._leftover_seq = getattr(self, "_leftover_seq", 0) + 1
        user = User.objects.create_user(f"fl-leftover-{self._leftover_seq}", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="M", default_stems=per_item, florist_fee=Decimal("50000"))
        # bir nechta florist ketma-ket sinalganda skladda gul tugab qolmasin
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + issued)
        self.batch.refresh_from_db()
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": issued}, format="json")
        made = []
        for index in range(items):
            response = self.client.post("/api/catalog/", {
                "name_uz": f"Buket {index + 1}", "arrangement_type": "bouquet", "volume": "M",
                "florist": profile.id, "price": "500000", "quantity_total": quantity_total, "status": "available",
                "composition": [{"stock_batch": self.batch.id, "quantity_stems": per_item}],
            }, format="json")
            self.assertEqual(response.status_code, 201, response.json())
            made.append(CatalogItem.objects.get(id=response.json()["id"]))
        return profile, made

    def _balance(self, profile):
        row = FloristStockBalance.objects.filter(florist=profile, batch=self.batch).first()
        return row.remaining_stems if row else 0

    def test_leftover_splits_evenly_into_catalog_items(self):
        # 100 dona olindi, standart 25 dan 3 ta buket = 75. Qolgani 25, uchga bo'linsa 8+8+9
        profile, made = self._florist_with_leftover(issued=100, per_item=25, items=3)
        self.assertEqual(self._balance(profile), 25)
        response = self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(self._balance(profile), 0)
        stems = sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True))
        self.assertEqual(stems, [33, 33, 34])
        self.assertEqual(sum(stems), 100)

    def test_leftover_gives_extra_to_one_item_when_not_divisible(self):
        # 90 olindi, 25 dan 3 ta = 75, qolgani 15 -> 5+5+5 teng bo'linadi
        profile, made = self._florist_with_leftover(issued=90, per_item=25, items=3)
        self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id}, format="json")
        self.assertEqual(sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True)), [30, 30, 30])
        # 92 olinganda qolgani 17 -> 5+6+6, kimdir bittaga ko'p oladi
        profile2, made2 = self._florist_with_leftover(issued=92, per_item=25, items=3, quantity_total=1)
        self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile2.id}, format="json")
        stems = sorted(CatalogComposition.objects.filter(catalog_item__in=made2).values_list("quantity_stems", flat=True))
        self.assertEqual(stems, [30, 31, 31])
        self.assertEqual(sum(stems), 92)
        self.assertEqual(self._balance(profile2), 0)

    def test_leftover_counts_units_not_items(self):
        # 2 donadan 2 ta katalog = 4 dona buket. Bo'lish katalog emas, dona hisobida borishi kerak
        profile, made = self._florist_with_leftover(issued=92, per_item=23, items=2, quantity_total=2)
        self.assertEqual(self._balance(profile), 0)
        issued = self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 8}, format="json")
        self.assertEqual(issued.status_code, 201, issued.json())
        self.assertEqual(self._balance(profile), 8)
        self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id}, format="json")
        # 4 dona buket, 8 gul -> har donaga +2, ya'ni tarkib 23 -> 25
        self.assertEqual(sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True)), [25, 25])
        self.assertEqual(self._balance(profile), 0)

    def test_leftover_raises_costs_including_sold_items(self):
        profile, made = self._florist_with_leftover(issued=100, per_item=25, items=3)
        mark_catalog_sold(made[0], self.user)
        before = CatalogItem.objects.get(pk=made[0].pk).calculated_cost_price
        self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id}, format="json")
        after = CatalogItem.objects.get(pk=made[0].pk).calculated_cost_price
        self.assertGreater(after, before)
        # hisob-kitobdagi tannarx ham o'sadi
        report = self.client.get("/api/accounting/")
        self.assertGreater(Decimal(report.data["summary"]["cost_total"]), Decimal("0"))

    def test_leftover_blocked_when_no_catalog_uses_the_flower(self):
        user = User.objects.create_user("fl-no-catalog", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 10}, format="json")
        response = self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("katalog topilmadi", response.data["detail"])
        self.assertEqual(self._balance(profile), 10)

    def test_adjust_preview_changes_nothing(self):
        profile, made = self._florist_with_leftover(issued=100, per_item=25, items=3)
        response = self.client.get(f"/api/florist-stock-balances/adjust-preview/?florist={profile.id}")
        self.assertEqual(response.status_code, 200)
        row = response.data["batches"][0]
        self.assertEqual(row["florist_stems_now"], 25)
        self.assertEqual(sorted(item["change_per_item"] for item in row["items"]), [8, 8, 9])
        self.assertEqual(self._balance(profile), 25)
        self.assertEqual(sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True)), [25, 25, 25])

    def test_reverse_returns_stems_from_catalog_to_florist(self):
        # standart 25 edi, florist 23 tadan ishlatgan: 3 buketdan 6 dona ortdi
        profile, made = self._florist_with_leftover(issued=75, per_item=25, items=3)
        self.assertEqual(self._balance(profile), 0)
        response = self.client.post("/api/florist-stock-balances/adjust/", {
            "florist": profile.id, "batch": self.batch.id, "direction": "to_florist", "quantity_stems": 6,
        }, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True)), [23, 23, 23])
        self.assertEqual(self._balance(profile), 6)

    def test_reverse_can_empty_the_row_completely(self):
        # chiqim yopishni butunlay orqaga qaytarish: tarkib 0 ga tushadi
        profile, made = self._florist_with_leftover(issued=75, per_item=25, items=3)
        response = self.client.post("/api/florist-stock-balances/adjust/", {
            "florist": profile.id, "batch": self.batch.id, "direction": "to_florist", "quantity_stems": 75,
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True)), [0, 0, 0])
        self.assertEqual(self._balance(profile), 75)

    def test_reverse_lowers_catalog_cost(self):
        profile, made = self._florist_with_leftover(issued=75, per_item=25, items=3)
        before = CatalogItem.objects.get(pk=made[0].pk).calculated_cost_price
        self.client.post("/api/florist-stock-balances/adjust/", {
            "florist": profile.id, "batch": self.batch.id, "direction": "to_florist", "quantity_stems": 6,
        }, format="json")
        after = CatalogItem.objects.get(pk=made[0].pk).calculated_cost_price
        self.assertLess(after, before)

    def test_reverse_requires_batch_and_quantity(self):
        profile, _ = self._florist_with_leftover(issued=75, per_item=25, items=3)
        no_batch = self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id, "direction": "to_florist", "quantity_stems": 3}, format="json")
        self.assertEqual(no_batch.status_code, 400)
        self.assertIn("batch", no_batch.data)
        no_quantity = self.client.post("/api/florist-stock-balances/adjust/", {"florist": profile.id, "batch": self.batch.id, "direction": "to_florist"}, format="json")
        self.assertEqual(no_quantity.status_code, 400)
        self.assertIn("quantity_stems", no_quantity.data)

    def test_reverse_rejects_more_than_catalog_holds(self):
        profile, made = self._florist_with_leftover(issued=75, per_item=25, items=3)
        response = self.client.post("/api/florist-stock-balances/adjust/", {
            "florist": profile.id, "batch": self.batch.id, "direction": "to_florist", "quantity_stems": 500,
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("yetmayapti", response.data["detail"])
        # hech narsa o'zgarmadi
        self.assertEqual(sorted(CatalogComposition.objects.filter(catalog_item__in=made).values_list("quantity_stems", flat=True)), [25, 25, 25])
        self.assertEqual(self._balance(profile), 0)

    def test_florist_catalog_checks_florist_balance_not_warehouse(self):
        # gul floristning qo'liga chiqarilgach skladda qolmaydi, lekin katalog qo'shilishi kerak
        user = User.objects.create_user("fl-empty-warehouse", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="M", default_stems=25, florist_fee=Decimal("50000"))
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 0)
        response = self.client.post("/api/catalog/", {
            "name_uz": "Sklad bo‘sh buket", "arrangement_type": "bouquet", "volume": "M", "florist": profile.id,
            "price": "500000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 25}],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 75)

    def test_florist_catalog_rejected_when_florist_has_too_few(self):
        user = User.objects.create_user("fl-short", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="M", default_stems=25, florist_fee=Decimal("50000"))
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 10}, format="json")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Ko‘p gulli buket", "arrangement_type": "bouquet", "volume": "M", "florist": profile.id,
            "price": "500000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 25}],
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("floristdagi gul yetarli emas", response.data["detail"])

    def _florist_with_rates(self, name):
        user = User.objects.create_user(name, password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        for volume, stems in [("S", 15), ("M", 25), ("L", 40)]:
            FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume=volume, default_stems=stems, florist_fee=Decimal("50000"))
        return profile

    def _make_sized_catalog(self, profile, volume, count=1, quantity_total=1, batch=None):
        """Florist katalogi: gul tanlanadi, soni yozilmaydi."""
        batch = batch or self.batch
        made = []
        for index in range(count):
            response = self.client.post("/api/catalog/", {
                "name_uz": f"{volume} buket {index + 1}", "arrangement_type": "bouquet", "volume": volume,
                "florist": profile.id, "price": "500000", "quantity_total": quantity_total, "status": "available",
                "composition": [{"stock_batch": batch.id}],
            }, format="json")
            self.assertEqual(response.status_code, 201, response.json())
            made.append(CatalogItem.objects.get(id=response.json()["id"]))
        return made

    def test_catalog_takes_flower_without_stem_count(self):
        # florist gulni tanlaydi, sonini yozmaydi — son 0 bo'lib turadi
        profile = self._florist_with_rates("fl-size-only")
        made = self._make_sized_catalog(profile, "M")
        row = made[0].composition.get()
        self.assertEqual(row.stock_batch_id, self.batch.id)
        self.assertEqual(row.quantity_stems, 0)
        self.assertEqual(made[0].volume, "M")

    def test_florist_catalog_requires_flower(self):
        profile = self._florist_with_rates("fl-no-flower")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Gulsiz", "arrangement_type": "bouquet", "volume": "M",
            "florist": profile.id, "price": "500000", "quantity_total": 1,
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("composition", response.data)

    def test_operator_catalog_still_requires_stem_count(self):
        response = self.client.post("/api/catalog/", {
            "name_uz": "Operator buketi", "arrangement_type": "bouquet",
            "price": "500000", "quantity_total": 1,
            "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("composition", response.data)

    def test_closing_one_flower_does_not_touch_other_flower_catalogs(self):
        # florist ikki xil gul olgan: qizil tugadi deyilsa faqat qizil buketlarga tushadi
        profile = self._florist_with_rates("fl-two-flowers")
        second = StockBatch.objects.create(
            variant=self.batch.variant, batch_number="API-2", height_cm=60, stems_per_bunch=20,
            received_stems=300, remaining_stems=300, cost_per_stem=10000,
            sale_price_per_stem=20000, sale_price_per_bunch=400000,
        )
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 200)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 200}, format="json")
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": second.id, "quantity_stems": 300}, format="json")
        first_items = self._make_sized_catalog(profile, "M", count=2, batch=self.batch)
        second_items = self._make_sized_catalog(profile, "M", count=3, batch=second)
        # birinchi gulni yopamiz
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(sorted(item.composition.get().quantity_stems for item in first_items), [100, 100])
        # ikkinchi guldan yasalganlarga tegilmadi
        self.assertEqual(sorted(item.composition.get().quantity_stems for item in second_items), [0, 0, 0])
        # endi ikkinchisini ham yopamiz
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": second.id}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(sorted(item.composition.get().quantity_stems for item in second_items), [100, 100, 100])
        self.assertEqual(sorted(item.composition.get().quantity_stems for item in first_items), [100, 100])
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=second).remaining_stems, 0)

    def test_florist_catalog_requires_volume(self):
        profile = self._florist_with_rates("fl-no-volume")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Hajmsiz", "arrangement_type": "bouquet",
            "florist": profile.id, "price": "500000", "quantity_total": 1,
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("volume", response.data)

    def test_closing_issue_shares_stems_by_volume_standard(self):
        # 600 dona chiqarildi, 3 ta S + 2 ta M + 1 ta L yasaldi
        profile = self._florist_with_rates("fl-close-1")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 600)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 600}, format="json")
        small = self._make_sized_catalog(profile, "S", count=3)
        medium = self._make_sized_catalog(profile, "M", count=2)
        large = self._make_sized_catalog(profile, "L", count=1)
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(response.data["shared_stems"], 600)
        stems = {item.id: item.composition.first().quantity_stems for item in small + medium + large}
        # ulush: S=15/135, M=25/135, L=40/135
        self.assertEqual(sorted(stems[item.id] for item in small), [66, 67, 67])
        self.assertEqual(sorted(stems[item.id] for item in medium), [111, 111])
        self.assertEqual(stems[large[0].id], 178)
        self.assertEqual(sum(stems.values()), 600)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)

    def test_closing_issue_falls_back_to_florist_fee_weight(self):
        # hajm tarifida dona soni kiritilmagan bo'lsa og'irlik florist haqidan olinadi
        user = User.objects.create_user("fl-fee-weight", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        for volume, fee in [("small", "10000"), ("medium", "15000")]:
            FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume=volume, default_stems=0, florist_fee=Decimal(fee))
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 250)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 250}, format="json")
        small = self._make_sized_catalog(profile, "small", count=1, quantity_total=7)
        medium = self._make_sized_catalog(profile, "medium", count=1, quantity_total=1)
        preview = self.client.get(f"/api/florist-stock-balances/close-issue-preview/?florist={profile.id}&batch={self.batch.id}")
        self.assertEqual(preview.data["weight_source"], "florist_fee")
        self.assertEqual(preview.data["missing_rates"], [])
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(response.data["shared_stems"], 250)
        self.assertEqual(response.data["weight_source"], "florist_fee")
        # 7 ta small (10 000) + 1 ta medium (15 000) -> og'irlik jami 85 000
        small_stems = small[0].composition.get().quantity_stems
        medium_stems = medium[0].composition.get().quantity_stems
        self.assertEqual(small_stems * 7 + medium_stems, 250)
        self.assertGreater(medium_stems, small_stems)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)

    def test_closing_issue_prefers_default_stems_over_fee(self):
        profile = self._florist_with_rates("fl-stems-weight")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        self._make_sized_catalog(profile, "M", count=2)
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.data["weight_source"], "default_stems")

    def test_closing_issue_blocks_when_rates_partially_filled(self):
        # bir hajmda dona soni bor, boshqasida yo'q — aralashtirilsa taqsimot buziladi
        user = User.objects.create_user("fl-partial-rate", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="small", default_stems=15, florist_fee=Decimal("10000"))
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="medium", default_stems=0, florist_fee=Decimal("15000"))
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        self._make_sized_catalog(profile, "small")
        self._make_sized_catalog(profile, "medium")
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("medium", response.data["detail"])
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 100)

    def test_closing_issue_returns_leftover_to_warehouse_first(self):
        profile = self._florist_with_rates("fl-close-2")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        self.batch.refresh_from_db()
        warehouse_before = self.batch.remaining_stems
        made = self._make_sized_catalog(profile, "M", count=2)
        response = self.client.post("/api/florist-stock-balances/close-issue/", {
            "florist": profile.id, "batch": self.batch.id, "return_stems": 20,
        }, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(response.data["returned_stems"], 20)
        self.assertEqual(response.data["shared_stems"], 80)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, warehouse_before + 20)
        self.assertEqual(sorted(item.composition.first().quantity_stems for item in made), [40, 40])
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)

    def test_closing_issue_counts_units_inside_one_catalog(self):
        profile = self._florist_with_rates("fl-close-3")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 200)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 200}, format="json")
        made = self._make_sized_catalog(profile, "M", count=1, quantity_total=4)
        self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        # 4 dona buket, 200 gul -> har donaga 50
        self.assertEqual(made[0].composition.first().quantity_stems, 50)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)

    def test_closing_issue_sets_catalog_cost(self):
        profile = self._florist_with_rates("fl-close-4")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        made = self._make_sized_catalog(profile, "M", count=2)
        before = CatalogItem.objects.get(pk=made[0].pk).calculated_cost_price
        self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        after = CatalogItem.objects.get(pk=made[0].pk).calculated_cost_price
        self.assertGreater(after, before)

    def test_closing_issue_needs_volume_rate(self):
        # tarif katalog qo'shilayotgandayoq talab qilinadi, keyin ham tekshiriladi
        user = User.objects.create_user("fl-close-norate", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="M", default_stems=25, florist_fee=Decimal("50000"))
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        blocked = self.client.post("/api/catalog/", {
            "name_uz": "XL buket", "arrangement_type": "bouquet", "volume": "XL",
            "florist": profile.id, "price": "500000", "quantity_total": 1,
            "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("hajm tarifi belgilanmagan", str(blocked.data))
        # tarif o'chirilgan bo'lsa yopishda ham xato beriladi
        self._make_sized_catalog(profile, "M")
        FloristVolumeRate.objects.filter(florist=profile, volume="M").update(is_active=False)
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("hajm tarifi to‘liq emas", response.data["detail"])
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 100)

    def test_closing_issue_without_catalog_is_rejected(self):
        profile = self._florist_with_rates("fl-close-5")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 50)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 50}, format="json")
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("yasalgan katalog topilmadi", response.data["detail"])

    def test_closing_issue_can_return_everything(self):
        profile = self._florist_with_rates("fl-close-6")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 30)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        self.batch.refresh_from_db()
        warehouse_before = self.batch.remaining_stems
        response = self.client.post("/api/florist-stock-balances/close-issue/", {
            "florist": profile.id, "batch": self.batch.id, "return_stems": 30,
        }, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(response.data["shared_stems"], 0)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, warehouse_before + 30)

    def test_close_issue_preview_changes_nothing(self):
        profile = self._florist_with_rates("fl-close-7")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        made = self._make_sized_catalog(profile, "M", count=2)
        response = self.client.get(f"/api/florist-stock-balances/close-issue-preview/?florist={profile.id}&batch={self.batch.id}&return_stems=20")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["share_stems"], 80)
        self.assertEqual(sorted(row["stems_per_item"] for row in response.data["items"]), [40, 40])
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 100)
        # gul tanlangan, lekin soni hali yozilmagan
        self.assertEqual(sorted(item.composition.get().quantity_stems for item in made), [0, 0])

    def test_close_issue_absorbs_remainder_by_default(self):
        profile = self._florist_with_rates("fl-close-rem")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 74)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 74}, format="json")
        item = self._make_sized_catalog(profile, "M", count=1, quantity_total=7, batch=self.batch)[0]
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(response.data["shared_stems"], 74)
        self.assertEqual(response.data["unplaced_stems"], 0)
        self.assertEqual(response.data["absorbed_remainder"], 4)
        self.assertEqual(response.data["rounded_extra_stems"], 3)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)
        item.refresh_from_db()
        self.assertEqual(item.composition.get().quantity_stems, 11)

    def test_close_all_absorbs_small_remainders(self):
        profile = self._florist_with_rates("fl-close-all")
        second = StockBatch.objects.create(
            variant=self.batch.variant, batch_number="API-REM-2", height_cm=60, stems_per_bunch=20,
            received_stems=100, remaining_stems=100, cost_per_stem=10000,
            sale_price_per_stem=20000, sale_price_per_bunch=400000,
        )
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 75)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 75}, format="json")
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": second.id, "quantity_stems": 74}, format="json")
        first_item = self._make_sized_catalog(profile, "M", count=1, quantity_total=7, batch=self.batch)[0]
        second_item = self._make_sized_catalog(profile, "M", count=1, quantity_total=7, batch=second)[0]
        first_row = first_item.composition.get()
        first_row.quantity_stems = 10
        first_row.save(update_fields=["quantity_stems", "updated_at"])
        second_row = second_item.composition.get()
        second_row.quantity_stems = 10
        second_row.save(update_fields=["quantity_stems", "updated_at"])
        FloristStockBalance.objects.filter(florist=profile, batch=self.batch).update(remaining_stems=5)
        FloristStockBalance.objects.filter(florist=profile, batch=second).update(remaining_stems=4)
        response = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "close_all": True}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.assertEqual(response.data["closed_batches"], 2)
        self.assertEqual(response.data["unplaced_stems"], 0)
        self.assertEqual(response.data["absorbed_remainder"], 9)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=second).remaining_stems, 0)
        first_item.refresh_from_db()
        second_item.refresh_from_db()
        self.assertEqual(first_item.composition.get().quantity_stems, 11)
        self.assertEqual(second_item.composition.get().quantity_stems, 11)

    def test_closing_issue_rejects_return_bigger_than_held(self):
        profile = self._florist_with_rates("fl-close-8")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 40)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 40}, format="json")
        response = self.client.post("/api/florist-stock-balances/close-issue/", {
            "florist": profile.id, "batch": self.batch.id, "return_stems": 500,
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 40)

    def test_catalog_quantity_can_be_edited_before_closing_issue(self):
        # soni xato yozilgan bo'lsa chiqim yopilishidan oldin tuzatiladi
        profile = self._florist_with_rates("fl-edit-1")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 300)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 300}, format="json")
        made = self._make_sized_catalog(profile, "M", count=1, quantity_total=2)
        item = made[0]
        # faqat son
        one = self.client.patch(f"/api/catalog/{item.id}/", {"quantity_total": 3}, format="json")
        self.assertEqual(one.status_code, 200, one.json())
        # son va tarkib birga
        two = self.client.patch(f"/api/catalog/{item.id}/", {
            "quantity_total": 6, "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        self.assertEqual(two.status_code, 200, two.json())
        item.refresh_from_db()
        self.assertEqual(item.quantity_total, 6)
        self.assertEqual(item.composition.get().quantity_stems, 0)
        # keyin chiqim yopiladi
        closed = self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        self.assertEqual(closed.status_code, 200, closed.json())
        item.refresh_from_db()
        self.assertEqual(item.composition.get().quantity_stems, 50)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)

    def test_catalog_quantity_edit_keeps_warehouse_in_balance(self):
        # florist tanlanmagan katalogda ham son va tarkibni birga tahrirlash ishlaydi
        before = self.batch.remaining_stems
        created = self.client.post("/api/catalog/", {
            "name_uz": "Sklad buketi", "arrangement_type": "bouquet", "price": "300000",
            "quantity_total": 2, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}],
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 20)
        updated = self.client.patch(f"/api/catalog/{created.json()['id']}/", {
            "quantity_total": 3, "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}],
        }, format="json")
        self.assertEqual(updated.status_code, 200, updated.json())
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 30)

    def test_standard_catalog_ignores_manual_florist_fee(self):
        # standart katalogda haq faqat hajm tarifidan olinadi
        profile = self._florist_with_rates("fl-fee-1")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Tarifli buket", "arrangement_type": "bouquet", "volume": "M",
            "florist": profile.id, "price": "500000", "quantity_total": 1,
            "florist_salary_amount": "999000",
            "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        # 999 000 emas, M hajm tarifidagi 50 000
        self.assertEqual(Decimal(response.json()["florist_salary_amount"]), Decimal("50000.00"))

    def test_standard_catalog_needs_volume_rate(self):
        profile = self._florist_with_rates("fl-fee-2")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Tarifsiz buket", "arrangement_type": "bouquet", "volume": "XXL",
            "florist": profile.id, "price": "500000", "quantity_total": 1,
            "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("hajm tarifi belgilanmagan", str(response.data))

    def test_custom_catalog_still_takes_manual_florist_fee(self):
        profile = self._florist_with_rates("fl-fee-3")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Custom buket", "arrangement_type": "bouquet", "volume": "M",
            "catalog_kind": "custom", "florist": profile.id, "price": "500000", "quantity_total": 1,
            "florist_salary_amount": "125000",
            "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(Decimal(response.json()["florist_salary_amount"]), Decimal("125000.00"))

    def test_florist_issue_can_be_edited(self):
        user = User.objects.create_user("fl-issue-edit", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        created = self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        issue_id = created.json()["id"]
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 30)
        # 30 emas 50 ekan
        response = self.client.patch(f"/api/florist-stock-issues/{issue_id}/edit/", {"quantity_stems": 50, "reason": "Tuzatildi"}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 50)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 50)
        # endi kamaytiramiz
        response = self.client.patch(f"/api/florist-stock-issues/{issue_id}/edit/", {"quantity_stems": 20}, format="json")
        self.assertEqual(response.status_code, 200, response.json())
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before - 20)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 20)

    def test_florist_issue_edit_rejects_more_than_warehouse(self):
        user = User.objects.create_user("fl-issue-edit-2", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        created = self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        response = self.client.patch(f"/api/florist-stock-issues/{created.json()['id']}/edit/", {"quantity_stems": 99999}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 30)

    def test_florist_issue_can_be_cancelled(self):
        user = User.objects.create_user("fl-issue-cancel", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        created = self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 40}, format="json")
        response = self.client.delete(f"/api/florist-stock-issues/{created.json()['id']}/cancel/")
        self.assertEqual(response.status_code, 204)
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before)
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 0)
        self.assertFalse(FloristStockIssue.objects.filter(pk=created.json()["id"]).exists())

    def test_florist_issue_cancel_rejected_when_already_used(self):
        profile = self._florist_with_rates("fl-issue-cancel-2")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        created = self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        self._make_sized_catalog(profile, "M", count=2)
        self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": self.batch.id}, format="json")
        response = self.client.delete(f"/api/florist-stock-issues/{created.json()['id']}/cancel/")
        self.assertEqual(response.status_code, 400)
        self.assertIn("bekor qilib bo‘lmaydi", response.data["detail"])

    def test_material_delivery_is_created_then_materials_received(self):
        supplier = Supplier.objects.create(name="Qadoq Servis")
        created = self.client.post("/api/material-deliveries/", {
            "number": "M-1", "received_at": "2026-08-01", "supplier": supplier.id, "note": "Qog‘oz va savat",
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        delivery_id = created.json()["id"]
        wrap = Packaging.objects.create(packaging_type="wrap", name_uz="Buket qog‘ozi", cost_price=Decimal("5000"), sale_price=Decimal("12000"), quantity=10)
        basket = Packaging.objects.create(packaging_type="basket", name_uz="O‘rta savat", cost_price=Decimal("12000"), sale_price=Decimal("30000"), quantity=0)
        one = self.client.post(f"/api/material-deliveries/{delivery_id}/receive/", {"packaging": wrap.id, "quantity": 100, "cost_price": "6000"}, format="json")
        two = self.client.post(f"/api/material-deliveries/{delivery_id}/receive/", {"packaging": basket.id, "quantity": 30, "cost_price": "15000"}, format="json")
        self.assertEqual(one.status_code, 201, one.json())
        self.assertEqual(two.status_code, 201, two.json())
        wrap.refresh_from_db()
        basket.refresh_from_db()
        # material bitta qator bo'lib qoladi: soni oshadi, tannarxi yangilanadi
        self.assertEqual(wrap.quantity, 110)
        self.assertEqual(wrap.cost_price, Decimal("6000.00"))
        self.assertEqual(basket.quantity, 30)
        self.assertEqual(basket.cost_price, Decimal("15000.00"))
        detail = self.client.get(f"/api/material-deliveries/{delivery_id}/")
        self.assertEqual(detail.data["item_count"], 2)
        self.assertEqual(detail.data["total_quantity"], 130)
        self.assertEqual(Decimal(detail.data["total_cost"]), Decimal("1050000.00"))
        items = self.client.get(f"/api/material-deliveries/{delivery_id}/items/")
        self.assertEqual(len(items.data), 2)

    def _debt_catalog(self, name="Qarz buketi", price="300000", quantity=2, stems=25):
        # ketma-ket bir nechta katalog sinalganda skladda gul tugab qolmasin
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + stems * quantity)
        created = self.client.post("/api/catalog/", {
            "name_uz": name, "arrangement_type": "bouquet", "price": price,
            "quantity_total": quantity, "status": "available", "image_url": "https://example.com/b.jpg",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": stems}],
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        return CatalogItem.objects.get(id=created.json()["id"])

    def test_debt_sale_creates_debt_with_new_customer(self):
        item = self._debt_catalog()
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt",
            "customer_name": "Aziz Karimov", "customer_phone": "+998901234567",
            "debt_note": "Hafta oxirida to‘laydi",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        debt = Debt.objects.get(catalog_item=item)
        self.assertEqual(debt.customer.name, "Aziz Karimov")
        self.assertEqual(debt.amount, Decimal("300000.00"))
        self.assertEqual(debt.quantity, 1)
        self.assertEqual(debt.note, "Hafta oxirida to‘laydi")
        self.assertFalse(debt.is_paid)

    def test_debt_sale_attaches_to_existing_customer(self):
        customer = Customer.objects.create(name="Malika", phone="+998939876543")
        item = self._debt_catalog(name="Bor mijoz buketi")
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt", "customer": customer.id,
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(Customer.objects.filter(name="Malika").count(), 1)
        self.assertEqual(Debt.objects.get(catalog_item=item).customer_id, customer.id)

    def test_debt_sale_requires_customer_or_contact(self):
        item = self._debt_catalog(name="Mijozsiz")
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", "quantity": 1, "payment_type": "debt"}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("customer", response.data)
        self.assertEqual(Debt.objects.count(), 0)

    def test_debt_uses_discounted_price(self):
        item = self._debt_catalog(name="Chegirmali qarz", price="300000")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "sale_price": "250000", "discount_reason": "Doimiy mijoz",
            "payment_type": "debt", "customer_name": "Sardor", "customer_phone": "+998901112233",
        }, format="json")
        self.assertEqual(Debt.objects.get(catalog_item=item).amount, Decimal("250000.00"))

    def test_unpaid_debt_stays_out_of_accounting(self):
        item = self._debt_catalog(name="Hisobsiz qarz")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt",
            "customer_name": "Qarzdor", "customer_phone": "+998900000001",
        }, format="json")
        report = self.client.get("/api/accounting/")
        self.assertEqual(Decimal(report.data["summary"]["total_sales"]), Decimal("0"))
        self.assertEqual(report.data["summary"]["sales_count"], 0)

    def test_paid_debt_enters_accounting_on_payment_day(self):
        item = self._debt_catalog(name="To‘langan qarz")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt",
            "customer_name": "Bekzod", "customer_phone": "+998900000002",
        }, format="json")
        debt = Debt.objects.get(catalog_item=item)
        paid = self.client.post(f"/api/debts/{debt.id}/pay/", {"method": "card"}, format="json")
        self.assertEqual(paid.status_code, 200, paid.data)
        report = self.client.get("/api/accounting/")
        self.assertEqual(Decimal(report.data["summary"]["total_sales"]), Decimal("300000.00"))
        # to'lov usuli karta bo'lgani uchun karta ustuniga tushadi
        self.assertEqual(Decimal(report.data["summary"]["card_total"]), Decimal("300000.00"))
        self.assertEqual(Decimal(report.data["summary"]["cash_total"]), Decimal("0"))

    def test_debt_cannot_be_paid_twice(self):
        item = self._debt_catalog(name="Ikki marta")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt",
            "customer_name": "Jasur", "customer_phone": "+998900000003",
        }, format="json")
        debt = Debt.objects.get(catalog_item=item)
        self.client.post(f"/api/debts/{debt.id}/pay/", {"method": "cash"}, format="json")
        again = self.client.post(f"/api/debts/{debt.id}/pay/", {"method": "cash"}, format="json")
        self.assertEqual(again.status_code, 400)
        self.assertIn("allaqachon to‘langan", again.data["detail"])

    def test_debtors_page_groups_by_customer(self):
        first = self._debt_catalog(name="Qarz A", price="300000")
        second = self._debt_catalog(name="Qarz B", price="150000")
        for item in [first, second]:
            self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
                "quantity": 1, "payment_type": "debt",
                "customer_name": "Aziz Karimov", "customer_phone": "+998901234567",
            }, format="json")
        third = self._debt_catalog(name="Qarz C", price="200000")
        self.client.post(f"/api/catalog/{third.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt",
            "customer_name": "Malika", "customer_phone": "+998939999999",
        }, format="json")
        response = self.client.get("/api/debts/by-customer/")
        self.assertEqual(response.status_code, 200)
        rows = response.data["customers"]
        self.assertEqual(len(rows), 2)
        aziz = next(row for row in rows if row["name"] == "Aziz Karimov")
        self.assertEqual(aziz["debt_count"], 2)
        self.assertEqual(aziz["unpaid_total"], Decimal("450000.00"))
        self.assertEqual(len(aziz["items"]), 2)
        # rasm, gul soni va summa qatorda bo'ladi
        row = aziz["items"][0]
        self.assertEqual(row["catalog_detail"]["image_url"], "https://example.com/b.jpg")
        self.assertEqual(row["catalog_detail"]["stems_per_item"], 25)
        self.assertEqual(response.data["totals"]["unpaid_total"], Decimal("650000.00"))
        self.assertEqual(response.data["totals"]["customer_count"], 2)

    def test_debtors_page_hides_paid_by_default(self):
        item = self._debt_catalog(name="Yopilgan qarz")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"sale_image_url": "https://example.com/sale.jpg", 
            "quantity": 1, "payment_type": "debt",
            "customer_name": "To‘lagan", "customer_phone": "+998900000004",
        }, format="json")
        debt = Debt.objects.get(catalog_item=item)
        self.client.post(f"/api/debts/{debt.id}/pay/", {"method": "cash"}, format="json")
        hidden = self.client.get("/api/debts/by-customer/")
        self.assertEqual(hidden.data["customers"], [])
        shown = self.client.get("/api/debts/by-customer/?include_paid=true")
        self.assertEqual(len(shown.data["customers"]), 1)
        self.assertEqual(shown.data["customers"][0]["paid_total"], Decimal("300000.00"))

    def test_suppliers_can_be_filtered_by_type(self):
        Supplier.objects.create(name="Gul postavshigi", supplier_type="flower")
        Supplier.objects.create(name="Qadoq postavshigi", supplier_type="material")
        Supplier.objects.create(name="Ikkalasi", supplier_type="both")
        materials = self.client.get("/api/suppliers/?supplier_type=material")
        self.assertEqual([row["name"] for row in materials.data["results"]], ["Qadoq postavshigi"])
        flowers = self.client.get("/api/suppliers/?supplier_type=flower")
        self.assertIn("Gul postavshigi", [row["name"] for row in flowers.data["results"]])
        self.assertNotIn("Qadoq postavshigi", [row["name"] for row in flowers.data["results"]])

    def test_material_can_be_added_straight_into_delivery(self):
        supplier = Supplier.objects.create(name="Qadoq yuk", supplier_type="material")
        delivery = self.client.post("/api/material-deliveries/", {
            "number": "M-10", "received_at": "2026-08-01", "supplier": supplier.id,
        }, format="json").json()
        response = self.client.post("/api/materials/", {
            "packaging_type": "wrap", "name_uz": "Flizilin", "unit": "bunch", "units_per_bunch": 20,
            "delivery": delivery["id"], "bunches": 10, "cost_per_bunch": "45000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        material = Packaging.objects.get(id=response.json()["id"])
        self.assertEqual(material.quantity, 200)
        self.assertEqual(material.cost_price, Decimal("2250.00"))
        # kirim yozuvi bir marta yaratiladi va yukka bog'lanadi
        movements = PackagingMovement.objects.filter(packaging=material)
        self.assertEqual(movements.count(), 1)
        self.assertEqual(movements.first().delivery_id, delivery["id"])
        detail = self.client.get(f"/api/material-deliveries/{delivery['id']}/")
        self.assertEqual(detail.data["total_quantity"], 200)
        self.assertEqual(Decimal(detail.data["total_cost"]), Decimal("450000.00"))

    def test_material_without_delivery_still_works(self):
        response = self.client.post("/api/materials/", {
            "packaging_type": "other", "name_uz": "Lenta", "unit": "piece",
            "quantity": 30, "cost_price": "4000", "sale_price": "0",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        material = Packaging.objects.get(id=response.json()["id"])
        self.assertEqual(material.quantity, 30)
        self.assertEqual(PackagingMovement.objects.filter(packaging=material).count(), 1)

    def test_receiving_by_bunches_computes_pieces_and_cost(self):
        delivery = self.client.post("/api/material-deliveries/", {"number": "M-11", "received_at": "2026-08-01"}, format="json").json()
        sponge = Packaging.objects.create(packaging_type="other", name_uz="Gupka", unit="bunch", units_per_bunch=20, cost_price=Decimal("0"), sale_price=Decimal("0"), quantity=0)
        response = self.client.post(f"/api/material-deliveries/{delivery['id']}/receive/", {
            "packaging": sponge.id, "bunches": 5, "cost_per_bunch": "60000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        sponge.refresh_from_db()
        self.assertEqual(sponge.quantity, 100)
        self.assertEqual(sponge.cost_price, Decimal("3000.00"))

    def test_receive_needs_quantity_or_bunches(self):
        delivery = self.client.post("/api/material-deliveries/", {"number": "M-12", "received_at": "2026-08-01"}, format="json").json()
        lak = Packaging.objects.create(packaging_type="other", name_uz="Lak", unit="piece", cost_price=Decimal("0"), sale_price=Decimal("0"), quantity=0)
        response = self.client.post(f"/api/material-deliveries/{delivery['id']}/receive/", {"packaging": lak.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("quantity", response.data)

    def test_basket_size_is_validated(self):
        good = self.client.post("/api/materials/", {
            "packaging_type": "basket", "name_uz": "Yog‘ochli savat", "basket_material": "wooden",
            "size": "M", "cost_price": "0", "sale_price": "0", "quantity": 0,
        }, format="json")
        self.assertEqual(good.status_code, 201, good.json())
        self.assertEqual(good.json()["size"], "m")
        self.assertEqual(good.json()["basket_material_label"], "Yog‘ochli")
        bad = self.client.post("/api/materials/", {
            "packaging_type": "basket", "name_uz": "Katta savat", "basket_material": "woven",
            "size": "ultra", "cost_price": "0", "sale_price": "0", "quantity": 0,
        }, format="json")
        self.assertEqual(bad.status_code, 400)
        self.assertIn("size", bad.data)

    def test_materials_can_be_filtered_by_basket_material_and_size(self):
        for material, size in [("wooden", "s"), ("wooden", "l"), ("woven", "s")]:
            Packaging.objects.create(packaging_type="basket", name_uz=f"{material} {size}", basket_material=material, size=size, cost_price=Decimal("0"), sale_price=Decimal("0"), quantity=0)
        wooden = self.client.get("/api/materials/?basket_material=wooden")
        self.assertEqual(wooden.data["count"], 2)
        small = self.client.get("/api/materials/?basket_material=wooden&size=s")
        self.assertEqual(small.data["count"], 1)

    def test_material_keeps_cost_when_not_given(self):
        delivery = self.client.post("/api/material-deliveries/", {"number": "M-2", "received_at": "2026-08-01"}, format="json").json()
        wrap = Packaging.objects.create(packaging_type="wrap", name_uz="Plyonka", cost_price=Decimal("4000"), sale_price=Decimal("9000"), quantity=5)
        response = self.client.post(f"/api/material-deliveries/{delivery['id']}/receive/", {"packaging": wrap.id, "quantity": 20}, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        wrap.refresh_from_db()
        self.assertEqual(wrap.quantity, 25)
        self.assertEqual(wrap.cost_price, Decimal("4000.00"))

    def test_material_shows_last_delivery_and_supplier(self):
        supplier = Supplier.objects.create(name="Gupka Savdo")
        delivery = self.client.post("/api/material-deliveries/", {
            "number": "M-3", "received_at": "2026-08-01", "supplier": supplier.id,
        }, format="json").json()
        sponge = Packaging.objects.create(packaging_type="other", name_uz="Gupka", cost_price=Decimal("3000"), sale_price=Decimal("8000"), quantity=0)
        self.client.post(f"/api/material-deliveries/{delivery['id']}/receive/", {"packaging": sponge.id, "quantity": 40, "cost_price": "3500"}, format="json")
        listed = self.client.get(f"/api/materials/{sponge.id}/")
        last = listed.data["last_delivery"]
        self.assertEqual(last["number"], "M-3")
        self.assertEqual(last["supplier"], "Gupka Savdo")
        self.assertEqual(last["quantity"], 40)
        self.assertEqual(Decimal(last["unit_cost"]), Decimal("3500.00"))

    def test_material_receive_rejects_zero(self):
        delivery = self.client.post("/api/material-deliveries/", {"number": "M-4", "received_at": "2026-08-01"}, format="json").json()
        sponge = Packaging.objects.create(packaging_type="other", name_uz="Lenta", cost_price=Decimal("1000"), sale_price=Decimal("3000"), quantity=0)
        response = self.client.post(f"/api/material-deliveries/{delivery['id']}/receive/", {"packaging": sponge.id, "quantity": 0}, format="json")
        self.assertEqual(response.status_code, 400)
        sponge.refresh_from_db()
        self.assertEqual(sponge.quantity, 0)

    def test_material_movement_carries_delivery(self):
        supplier = Supplier.objects.create(name="Karton Plus")
        delivery = self.client.post("/api/material-deliveries/", {
            "number": "M-5", "received_at": "2026-08-01", "supplier": supplier.id,
        }, format="json").json()
        box = Packaging.objects.create(packaging_type="box", name_uz="Quti", cost_price=Decimal("8000"), sale_price=Decimal("20000"), quantity=0)
        self.client.post(f"/api/material-deliveries/{delivery['id']}/receive/", {"packaging": box.id, "quantity": 15, "cost_price": "9000"}, format="json")
        movements = self.client.get(f"/api/material-movements/?packaging={box.id}")
        row = movements.data["results"][0]
        self.assertEqual(row["delivery"], delivery["id"])
        self.assertEqual(row["movement_type"], "in")
        self.assertEqual(row["quantity"], 15)
        self.assertEqual(Decimal(row["unit_cost"]), Decimal("9000.00"))

    def test_delivery_is_created_then_flowers_added_into_it(self):
        supplier = Supplier.objects.create(name="Golland Flowers")
        created = self.client.post("/api/stock-deliveries/", {
            "number": "7", "received_at": "2026-08-01", "supplier": supplier.id, "note": "Chorshanba yuki",
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        delivery_id = created.json()["id"]
        flower = Flower.objects.create(name_uz="Lola partiya", slug="lola-partiya")
        first = FlowerVariant.objects.create(flower=flower, name_uz="Qizil nav", color_uz="Qizil")
        second = FlowerVariant.objects.create(flower=flower, name_uz="Sariq nav", color_uz="Sariq")
        for variant in [first, second]:
            response = self.client.post("/api/stock-batches/", {
                "delivery": delivery_id, "variant": variant.id, "height_cm": 50,
                "stems_per_bunch": 25, "received_stems": 100,
                "cost_per_bunch": "250000", "sale_price_per_bunch": "500000",
            }, format="json")
            self.assertEqual(response.status_code, 201, response.json())
            # partiya raqami, sanasi va postavshigi partiyadan olinadi
            self.assertEqual(response.json()["batch_number"], "7")
            self.assertEqual(response.json()["received_at"], "2026-08-01")
            self.assertEqual(response.json()["supplier"], supplier.id)
        listed = self.client.get(f"/api/stock-deliveries/{delivery_id}/")
        self.assertEqual(listed.data["batch_count"], 2)
        self.assertEqual(listed.data["total_stems"], 200)
        inside = self.client.get(f"/api/stock-deliveries/{delivery_id}/batches/")
        self.assertEqual(len(inside.data), 2)

    def test_bunch_price_fills_stem_price_automatically(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "AUTO-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 50,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(Decimal(response.json()["cost_per_stem"]), Decimal("1000.00"))
        self.assertEqual(Decimal(response.json()["sale_price_per_stem"]), Decimal("2000.00"))

    def test_stem_price_is_rounded_to_hundred(self):
        # 24 950 / 25 = 998 -> 1 000,  26 500 / 25 = 1 060 -> 1 100
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "AUTO-2", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 50,
            "cost_per_bunch": "24950", "sale_price_per_bunch": "26500",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(Decimal(response.json()["cost_per_stem"]), Decimal("1000.00"))
        self.assertEqual(Decimal(response.json()["sale_price_per_stem"]), Decimal("1100.00"))
        # pochka narxi kiritilgani o'zgarmay saqlanadi
        self.assertEqual(Decimal(response.json()["cost_per_bunch"]), Decimal("24950.00"))

    def test_exact_price_is_kept_next_to_rounded(self):
        # 24 950 / 25 = 998 -> yaxlitlangani 1 000, aniq hisob 998 saqlanadi
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "EXACT-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_bunch": "24950", "sale_price_per_bunch": "26500",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        data = response.json()
        self.assertEqual(Decimal(data["cost_per_stem"]), Decimal("1000.00"))
        self.assertEqual(Decimal(data["cost_per_stem_exact"]), Decimal("998.0000"))
        self.assertEqual(Decimal(data["sale_price_per_stem"]), Decimal("1100.00"))
        self.assertEqual(Decimal(data["sale_price_per_stem_exact"]), Decimal("1060.0000"))
        rounding = data["rounding"]["cost"]
        self.assertEqual(Decimal(rounding["per_stem_diff"]), Decimal("2.0000"))
        self.assertEqual(Decimal(rounding["total_exact"]), Decimal("99800.00"))
        self.assertEqual(Decimal(rounding["total_rounded"]), Decimal("100000.00"))
        self.assertEqual(Decimal(rounding["total_diff"]), Decimal("200.00"))
        self.assertTrue(rounding["is_rounded"])

    def test_exact_price_equals_rounded_when_it_divides(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "EXACT-2", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 25,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(Decimal(response.json()["cost_per_stem_exact"]), Decimal("1000.0000"))
        self.assertFalse(response.json()["rounding"]["cost"]["is_rounded"])

    def test_exact_price_kept_when_stem_price_typed_by_hand(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "EXACT-3", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 24, "received_stems": 24,
            "cost_per_stem": "1041", "sale_price_per_stem": "2000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        # qo'lda kiritilgan narx yaxlitlanmaydi va aniq hisob ham o'shaning o'zi
        self.assertEqual(Decimal(response.json()["cost_per_stem"]), Decimal("1041.00"))
        self.assertEqual(Decimal(response.json()["cost_per_stem_exact"]), Decimal("1041.0000"))
        self.assertFalse(response.json()["rounding"]["cost"]["is_rounded"])

    def test_delivery_detail_shows_both_totals(self):
        delivery = self.client.post("/api/stock-deliveries/", {"number": "EX-1", "received_at": "2026-08-01"}, format="json").json()
        self.client.post("/api/stock-batches/", {
            "delivery": delivery["id"], "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_bunch": "24950", "sale_price_per_bunch": "50000",
        }, format="json")
        detail = self.client.get(f"/api/stock-deliveries/{delivery['id']}/")
        self.assertEqual(Decimal(detail.data["total_cost"]), Decimal("99800.00"))
        self.assertEqual(Decimal(detail.data["total_cost_exact"]), Decimal("99800.00"))
        self.assertEqual(Decimal(detail.data["rounding_diff"]), Decimal("200.00"))

    def test_supplier_rollup_uses_exact_delivery_cost(self):
        supplier = Supplier.objects.create(name="Exact supplier")
        StockBatch.objects.create(
            variant=self.batch.variant, supplier=supplier, batch_number="EXACT-SUP", height_cm=50,
            stems_per_bunch=15, received_stems=285, remaining_stems=285,
            cost_per_bunch=50000, cost_per_stem=3300, cost_per_stem_exact=Decimal("3333.3333"),
            sale_price_per_stem=10000, sale_price_per_bunch=150000,
        )
        detail = self.client.get(f"/api/suppliers/{supplier.id}/")
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(Decimal(detail.data["purchase_total"]), Decimal("950000.00"))

    def _batch_with_usage(self, received=100, used=30):
        created = self.client.post("/api/stock-batches/", {
            "batch_number": "EDIT-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": received,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json").json()
        if used:
            self.client.post(f"/api/stock-batches/{created['id']}/movement/", {
                "movement_type": "out", "quantity_stems": used, "reason": "test",
            }, format="json")
        return StockBatch.objects.get(id=created["id"])

    def test_batch_received_stems_can_be_edited_upwards(self):
        # xato kiritilgan son tuzatilganda ishlatilgan gul unutilmasligi kerak
        batch = self._batch_with_usage(received=100, used=30)
        self.assertEqual(batch.remaining_stems, 70)
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"received_stems": 120}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.received_stems, 120)
        self.assertEqual(batch.remaining_stems, 90)

    def test_batch_received_stems_can_be_edited_downwards(self):
        batch = self._batch_with_usage(received=100, used=30)
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"received_stems": 80}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.received_stems, 80)
        self.assertEqual(batch.remaining_stems, 50)

    def test_batch_received_stems_cannot_drop_below_used(self):
        batch = self._batch_with_usage(received=100, used=30)
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"received_stems": 10}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("received_stems", response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.received_stems, 100)
        self.assertEqual(batch.remaining_stems, 70)

    def test_batch_edit_updates_incoming_movement(self):
        batch = self._batch_with_usage(received=100, used=30)
        self.client.patch(f"/api/stock-batches/{batch.id}/", {"received_stems": 120}, format="json")
        incoming = StockMovement.objects.filter(batch=batch, movement_type="in").order_by("created_at", "id").first()
        self.assertEqual(incoming.quantity_stems, 120)
        # chiqim yozuvi tegilmaydi
        outgoing = StockMovement.objects.filter(batch=batch, movement_type="out").first()
        self.assertEqual(outgoing.quantity_stems, -30)

    def test_batch_remaining_can_still_be_set_directly(self):
        batch = self._batch_with_usage(received=100, used=30)
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"received_stems": 120, "remaining_stems": 65}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.received_stems, 120)
        self.assertEqual(batch.remaining_stems, 65)

    def test_batch_edit_without_quantity_change_keeps_remaining(self):
        batch = self._batch_with_usage(received=100, used=30)
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"sale_price_per_bunch": "60000"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.remaining_stems, 70)
        self.assertEqual(batch.sale_price_per_bunch, Decimal("60000.00"))

    def test_unused_batch_variant_can_be_fixed(self):
        # xato nav tanlangan bo'lsa, hali ishlatilmagan partiyada tuzatiladi
        other = FlowerVariant.objects.create(flower=self.batch.variant.flower, name_uz="Boshqa nav", color_uz="Oq")
        created = self.client.post("/api/stock-batches/", {
            "batch_number": "VAR-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 50,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json").json()
        response = self.client.patch(f"/api/stock-batches/{created['id']}/", {"variant": other.id}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(StockBatch.objects.get(id=created["id"]).variant_id, other.id)

    def test_used_batch_variant_cannot_be_changed(self):
        other = FlowerVariant.objects.create(flower=self.batch.variant.flower, name_uz="Almashtirilmas", color_uz="Sariq")
        batch = self._batch_with_usage(received=100, used=30)
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"variant": other.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("change-flower", str(response.data["flower"]))
        batch.refresh_from_db()
        self.assertEqual(batch.variant_id, self.batch.variant_id)

    def test_batch_used_in_catalog_cannot_change_variant(self):
        other = FlowerVariant.objects.create(flower=self.batch.variant.flower, name_uz="Katalogli", color_uz="Pushti")
        created = self.client.post("/api/stock-batches/", {
            "batch_number": "VAR-2", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json").json()
        self.client.post("/api/catalog/", {
            "name_uz": "Navli buket", "arrangement_type": "bouquet", "price": "300000",
            "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": created["id"], "quantity_stems": 20}],
        }, format="json")
        response = self.client.patch(f"/api/stock-batches/{created['id']}/", {"variant": other.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("flower", response.data)

    def _used_batch_with_sale(self, name="Nav buketi"):
        other = FlowerVariant.objects.create(flower=self.batch.variant.flower, name_uz="Yangi nav", color_uz="Oq")
        created = self.client.post("/api/stock-batches/", {
            "batch_number": "CHG-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json").json()
        item = self.client.post("/api/catalog/", {
            "name_uz": name, "arrangement_type": "bouquet", "price": "300000",
            "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": created["id"], "quantity_stems": 20}],
        }, format="json").json()
        self.client.post(f"/api/catalog/{item['id']}/sell/", {"sale_image_url": "https://example.com/sale.jpg", "quantity": 1, "payment_type": "cash"}, format="json")
        return StockBatch.objects.get(id=created["id"]), CatalogItem.objects.get(id=item["id"]), other

    def test_florist_issue_can_be_backdated(self):
        # o'tib ketgan kun uchun chiqim sanasini belgilash
        user = User.objects.create_user("fl-backdate", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        when = "2026-07-28T10:00:00+05:00"
        response = self.client.post("/api/florist-stock-issues/issue/", {
            "florist": profile.id, "batch": self.batch.id, "quantity_stems": 20, "created_at": when,
        }, format="json")
        self.assertEqual(response.status_code, 201, response.data)
        issue = FloristStockIssue.objects.get(id=response.data["id"])
        self.assertEqual(issue.created_at.date().isoformat(), "2026-07-28")
        # sklad harakati ham o'sha kunga tushadi
        movement = StockMovement.objects.get(reference_type="florist_issue", reference_id=issue.id)
        self.assertEqual(movement.created_at.date().isoformat(), "2026-07-28")
        # qoldiqlar odatdagidek o'zgaradi
        self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=self.batch).remaining_stems, 20)

    def test_florist_issue_without_date_uses_now(self):
        user = User.objects.create_user("fl-nodate", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        response = self.client.post("/api/florist-stock-issues/issue/", {
            "florist": profile.id, "batch": self.batch.id, "quantity_stems": 10,
        }, format="json")
        issue = FloristStockIssue.objects.get(id=response.data["id"])
        self.assertEqual(issue.created_at.date(), timezone.localdate())

    def test_florist_return_can_be_backdated(self):
        user = User.objects.create_user("fl-ret-backdate", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 30}, format="json")
        response = self.client.post("/api/florist-stock-issues/return/", {
            "florist": profile.id, "batch": self.batch.id, "quantity_stems": 10,
            "created_at": "2026-07-29T09:00:00+05:00",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(FloristStockIssue.objects.get(id=response.data["id"]).created_at.date().isoformat(), "2026-07-29")

    def test_bulk_issue_can_be_backdated(self):
        user = User.objects.create_user("fl-bulk-backdate", password="p")
        profile = FloristProfile.objects.create(user=user, staff_type="florist")
        second = StockBatch.objects.create(
            variant=self.batch.variant, batch_number="BD-2", height_cm=50, stems_per_bunch=25,
            received_stems=100, remaining_stems=100, cost_per_stem=5000,
            sale_price_per_stem=10000, sale_price_per_bunch=250000,
        )
        response = self.client.post("/api/florist-stock-issues/bulk-issue/", {
            "florist": profile.id,
            "items": [{"batch": self.batch.id, "quantity_stems": 10}, {"batch": second.id, "quantity_stems": 20}],
            "created_at": "2026-07-27T08:00:00+05:00",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.data)
        for row in response.data:
            self.assertEqual(FloristStockIssue.objects.get(id=row["id"]).created_at.date().isoformat(), "2026-07-27")

    def test_catalog_can_be_backdated(self):
        profile = self._florist_with_rates("fl-cat-backdate")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        response = self.client.post("/api/catalog/", {
            "name_uz": "O‘tgan kungi buket", "arrangement_type": "bouquet", "volume": "M",
            "florist": profile.id, "price": "500000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id}],
            "created_at": "2026-07-30T12:00:00+05:00",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        item = CatalogItem.objects.get(id=response.json()["id"])
        self.assertEqual(item.created_at.date().isoformat(), "2026-07-30")
        # tarix yozuvi ham o'sha kunga tushadi
        self.assertEqual(CatalogHistory.objects.filter(catalog_item=item).first().created_at.date().isoformat(), "2026-07-30")

    def test_backdated_catalog_moves_florist_salary_date(self):
        profile = self._florist_with_rates("fl-salary-backdate")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        response = self.client.post("/api/catalog/", {
            "name_uz": "Ish haqi sanasi", "arrangement_type": "bouquet", "volume": "M",
            "florist": profile.id, "price": "500000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id}],
            "created_at": "2026-07-25T12:00:00+05:00",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        entry = FloristSalaryEntry.objects.get(catalog_item_id=response.json()["id"])
        self.assertEqual(entry.work_date.isoformat(), "2026-07-25")

    def test_catalog_without_date_uses_now(self):
        created = self.client.post("/api/catalog/", {
            "name_uz": "Bugungi buket", "arrangement_type": "bouquet", "price": "300000",
            "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}],
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        self.assertEqual(CatalogItem.objects.get(id=created.json()["id"]).created_at.date(), timezone.localdate())

    def test_catalog_date_can_be_fixed_later(self):
        created = self.client.post("/api/catalog/", {
            "name_uz": "Sanasi tuzatiladi", "arrangement_type": "bouquet", "price": "300000",
            "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 10}],
        }, format="json").json()
        response = self.client.patch(f"/api/catalog/{created['id']}/", {"created_at": "2026-07-26T15:00:00+05:00"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(CatalogItem.objects.get(id=created["id"]).created_at.date().isoformat(), "2026-07-26")

    def test_sale_works_without_image(self):
        # rasm hozircha ixtiyoriy — busiz ham sotiladi, guruhga xabar ketmaydi
        item = self._debt_catalog(name="Rasmsiz sotuv", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 1)
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        self.assertIsNone((history.snapshot or {}).get("sale_image_url"))

    def test_sale_image_is_stored_on_history(self):
        item = self._debt_catalog(name="Rasmli sotuv", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "card", "sale_image_url": "https://example.com/buket.jpg",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        self.assertEqual(history.snapshot.get("sale_image_url"), "https://example.com/buket.jpg")

    def test_sale_group_caption_shows_payment_and_amount(self):
        from .inventory_services import sale_group_caption

        item = self._debt_catalog(name="Guruh xabari", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "cash", "sale_image_url": "https://example.com/x.jpg",
        }, format="json")
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        caption = sale_group_caption(item, history, "cash", "https://example.com/x.jpg")
        self.assertIn("Guruh xabari", caption)
        self.assertIn("Naqd", caption)
        self.assertIn("300 000", caption)

    def test_sale_group_caption_shows_card_and_discount(self):
        from .inventory_services import sale_group_caption

        item = self._debt_catalog(name="Chegirmali", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "sale_price": "250000", "discount_reason": "Doimiy mijoz",
            "payment_type": "card", "sale_image_url": "https://example.com/x.jpg",
        }, format="json")
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        caption = sale_group_caption(item, history, "card", "https://example.com/x.jpg")
        self.assertIn("Karta", caption)
        self.assertIn("250 000", caption)
        self.assertIn("Chegirma", caption)
        self.assertIn("Doimiy mijoz", caption)

    def test_sale_works_when_group_not_configured(self):
        # bot sozlanmagan bo'lsa ham sotuv bajariladi
        item = self._debt_catalog(name="Sozlanmagan guruh", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "cash", "sale_image_url": "https://example.com/x.jpg",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 1)

    def _three_issued_batches(self, profile, stems=100):
        batches = []
        for index in range(3):
            batch = StockBatch.objects.create(
                variant=self.batch.variant, batch_number=f"MULTI-{index}", height_cm=50, stems_per_bunch=25,
                received_stems=stems, remaining_stems=stems, cost_per_stem=1000,
                sale_price_per_stem=2000, sale_price_per_bunch=50000,
            )
            self.client.post("/api/florist-stock-issues/issue/", {
                "florist": profile.id, "batch": batch.id, "quantity_stems": stems,
            }, format="json")
            batches.append(batch)
        return batches

    def _florist_catalog(self, profile, batches, name="Ko‘p gulli buket"):
        response = self.client.post("/api/catalog/", {
            "name_uz": name, "arrangement_type": "bouquet", "volume": "M",
            "florist": profile.id, "price": "500000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": b.id} for b in batches],
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        return CatalogItem.objects.get(id=response.json()["id"])

    def test_third_flower_can_be_added_and_closed(self):
        # 2 ta gul bilan katalog qo'shilib, keyin 3-chisi qo'shiladi
        profile = self._florist_with_rates("fl-multi-1")
        batches = self._three_issued_batches(profile)
        item = self._florist_catalog(profile, batches[:2])
        response = self.client.patch(f"/api/catalog/{item.id}/", {
            "composition": [{"stock_batch": b.id} for b in batches],
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(item.composition.count(), 3)
        closed = self.client.post("/api/florist-stock-balances/close-issue/", {
            "florist": profile.id, "batch": batches[2].id,
        }, format="json")
        self.assertEqual(closed.status_code, 200, closed.data)

    def test_editing_catalog_keeps_already_shared_stems(self):
        # yopilgan gul soni katalog tahrirlanganda yo'qolmasligi kerak
        profile = self._florist_with_rates("fl-multi-2")
        batches = self._three_issued_batches(profile)
        item = self._florist_catalog(profile, batches[:2])
        for batch in batches[:2]:
            self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": batch.id}, format="json")
        before = {row.stock_batch_id: row.quantity_stems for row in item.composition.all()}
        self.assertEqual(sorted(before.values()), [100, 100])

        response = self.client.patch(f"/api/catalog/{item.id}/", {
            "composition": [{"stock_batch": b.id} for b in batches],
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        after = {row.stock_batch_id: row.quantity_stems for row in item.composition.all()}
        # eski ikkitasi saqlanadi, yangisi 0 bo'lib qo'shiladi
        self.assertEqual(after[batches[0].id], 100)
        self.assertEqual(after[batches[1].id], 100)
        self.assertEqual(after[batches[2].id], 0)

    def test_editing_catalog_keeps_florist_balance_correct(self):
        profile = self._florist_with_rates("fl-multi-3")
        batches = self._three_issued_batches(profile)
        item = self._florist_catalog(profile, batches[:2])
        for batch in batches[:2]:
            self.client.post("/api/florist-stock-balances/close-issue/", {"florist": profile.id, "batch": batch.id}, format="json")
        for batch in batches[:2]:
            self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=batch).remaining_stems, 0)
        self.client.patch(f"/api/catalog/{item.id}/", {
            "composition": [{"stock_batch": b.id} for b in batches],
        }, format="json")
        # gul katalogda qolgani uchun floristga qaytib kelmasligi kerak
        for batch in batches[:2]:
            self.assertEqual(FloristStockBalance.objects.get(florist=profile, batch=batch).remaining_stems, 0)

    def test_catalog_can_be_wasted_without_deleting(self):
        # sotilmay qolgan buket chiqitga chiqadi, katalog o'chmaydi
        item = self._debt_catalog(name="Chiqit savat", price="800000", quantity=5, stems=39)
        for _ in range(4):
            self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 4)

        response = self.client.post(f"/api/catalog/{item.id}/waste/", {"quantity": 1, "reason": "Gul so‘lidi"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        item.refresh_from_db()
        self.assertEqual(item.quantity_wasted, 1)
        self.assertEqual(item.quantity_total, 5)
        self.assertEqual(response.data["quantity_remaining"], 0)
        # katalog o'chmadi
        self.assertTrue(CatalogItem.objects.filter(pk=item.pk).exists())
        # tarixda chiqit yozuvi bor
        row = CatalogHistory.objects.filter(catalog_item=item, action="wasted").first()
        self.assertIsNotNone(row)
        self.assertEqual(row.quantity, 1)
        self.assertEqual(row.snapshot.get("waste_reason"), "Gul so‘lidi")

    def test_waste_cannot_exceed_remaining(self):
        item = self._debt_catalog(name="Ko‘p chiqit", price="300000", quantity=2)
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        response = self.client.post(f"/api/catalog/{item.id}/waste/", {"quantity": 5}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("atigi 1 dona", response.data["detail"])
        item.refresh_from_db()
        self.assertEqual(item.quantity_wasted, 0)

    def test_waste_reduces_profit_by_unit_cost(self):
        item = self._debt_catalog(name="Yo‘qotish", price="300000", quantity=2)
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        before = self.client.get("/api/accounting/").data["summary"]
        unit_cost = (CatalogItem.objects.get(pk=item.pk).calculated_cost_price / 2).quantize(Decimal("0.01"))
        self.client.post(f"/api/catalog/{item.id}/waste/", {"quantity": 1, "reason": "Sinib qoldi"}, format="json")
        after = self.client.get("/api/accounting/").data["summary"]
        self.assertEqual(Decimal(after["catalog_waste_total"]) - Decimal(before["catalog_waste_total"]), unit_cost)
        self.assertEqual(after["catalog_waste_quantity"] - before["catalog_waste_quantity"], 1)
        # savdo o'zgarmaydi, foyda esa tannarx qadar kamayadi
        self.assertEqual(Decimal(after["total_sales"]), Decimal(before["total_sales"]))
        self.assertEqual(Decimal(before["net_profit"]) - Decimal(after["net_profit"]), unit_cost)

    def test_waste_does_not_touch_stock(self):
        # gul katalog yasalganda yechilgan, chiqitda yana yechilmaydi
        item = self._debt_catalog(name="Skladga tegmaydi", price="300000", quantity=2)
        self.batch.refresh_from_db()
        before = self.batch.remaining_stems
        self.client.post(f"/api/catalog/{item.id}/waste/", {"quantity": 1}, format="json")
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, before)

    def test_wasted_catalog_is_not_sellable_anymore(self):
        item = self._debt_catalog(name="Sotib bo‘lmaydi", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/waste/", {"quantity": 1}, format="json")
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.assertEqual(response.status_code, 400)

    def test_delivery_fee_is_subtracted_from_sale(self):
        # mijoz 300 000 to'laydi, shundan 20 000 kuryerga ketadi
        item = self._debt_catalog(name="Dastafkali", price="300000", quantity=1)
        before = self.client.get("/api/accounting/").data["summary"]
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "cash", "delivery_amount": "20000",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        after = self.client.get("/api/accounting/").data["summary"]
        # savdoga dastafkasiz 280 000 kiradi
        self.assertEqual(Decimal(after["total_sales"]) - Decimal(before["total_sales"]), Decimal("280000.00"))
        self.assertEqual(Decimal(after["delivery_total"]) - Decimal(before["delivery_total"]), Decimal("20000.00"))
        # kassaga mijoz bergan 300 000 tushadi
        self.assertEqual(Decimal(after["cash_total"]) - Decimal(before["cash_total"]), Decimal("300000.00"))
        self.assertEqual(Decimal(after["received_total"]),
                         Decimal(after["total_sales"]) + Decimal(after["delivery_total"]))
        self.assertEqual(after["delivery_count"] - before["delivery_count"], 1)
        row = self.client.get("/api/catalog/sales/").data["results"][0]
        self.assertEqual(row["sale_total"], Decimal("280000.00"))
        self.assertEqual(row["received_total"], Decimal("300000.00"))
        self.assertEqual(row["delivery_amount"], Decimal("20000.00"))

    def test_delivery_fee_lowers_profit_by_its_amount(self):
        # dastafka kuryerga ketgani uchun foyda o'sha summaga kam bo'ladi
        item = self._debt_catalog(name="Foydaga tegadi", price="300000", quantity=1)
        before = self.client.get("/api/accounting/").data["summary"]
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "card", "delivery_amount": "25000",
        }, format="json")
        after = self.client.get("/api/accounting/").data["summary"]
        gain = Decimal(after["net_profit"]) - Decimal(before["net_profit"])
        sale_gain = Decimal(after["total_sales"]) - Decimal(before["total_sales"])
        cost_gain = Decimal(after["cost_total"]) - Decimal(before["cost_total"])
        self.assertEqual(sale_gain, Decimal("275000.00"))
        self.assertEqual(gain, sale_gain - cost_gain)

    def test_delivery_fee_with_mixed_payment(self):
        item = self._debt_catalog(name="Aralash dastafka", price="300000", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed",
            "cash_amount": "100000", "card_amount": "200000", "delivery_amount": "20000",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        row = self.client.get("/api/catalog/sales/").data["results"][0]
        self.assertEqual(row["delivery_amount"], Decimal("20000.00"))
        self.assertEqual(row["received_total"], Decimal("300000.00"))
        self.assertEqual(row["sale_total"], Decimal("280000.00"))
        self.assertEqual(row["payment_breakdown"]["cash"], Decimal("100000.00"))
        self.assertEqual(row["payment_breakdown"]["card"], Decimal("200000.00"))

    def test_mixed_payment_must_equal_received_total(self):
        item = self._debt_catalog(name="Yig‘indi to‘g‘ri kelmadi", price="300000", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed",
            "cash_amount": "150000", "card_amount": "130000", "delivery_amount": "20000",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("dastafka", response.data["detail"])
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 0)

    def test_delivery_cannot_swallow_whole_sale(self):
        item = self._debt_catalog(name="Dastafka katta", price="300000", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "cash", "delivery_amount": "300000",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("dastafka", response.data["delivery_amount"].lower())
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 0)

    def test_sale_without_delivery_reports_zero(self):
        item = self._debt_catalog(name="Dastafkasiz", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        row = self.client.get("/api/catalog/sales/").data["results"][0]
        self.assertEqual(row["delivery_amount"], Decimal("0"))
        self.assertEqual(row["received_total"], row["sale_total"])

    def test_mixed_payment_sale_splits_money(self):
        # yarmi naqd, yarmi karta
        item = self._debt_catalog(name="Aralash to‘lov", price="300000", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed",
            "cash_amount": "100000", "card_amount": "200000",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        report = self.client.get("/api/accounting/")
        summary = report.data["summary"]
        self.assertEqual(Decimal(summary["total_sales"]), Decimal("300000.00"))
        self.assertEqual(Decimal(summary["cash_total"]), Decimal("100000.00"))
        self.assertEqual(Decimal(summary["card_total"]), Decimal("200000.00"))
        # sotuv soni ikki marta sanalmaydi
        self.assertEqual(summary["sales_count"], 1)
        self.assertEqual(summary["cash_count"] + summary["card_count"], 1)
        self.assertEqual(summary["mixed_count"], 1)

    def test_mixed_payment_must_match_sale_total(self):
        item = self._debt_catalog(name="Notog‘ri aralash", price="300000", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed",
            "cash_amount": "100000", "card_amount": "100000",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("teng emas", response.data["detail"])
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 0)

    def test_mixed_payment_needs_both_amounts(self):
        item = self._debt_catalog(name="Yarim aralash", price="300000", quantity=1)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed", "cash_amount": "300000",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("cash_amount", response.data)

    def test_mixed_payment_works_with_discount_and_quantity(self):
        item = self._debt_catalog(name="Aralash chegirma", price="300000", quantity=2)
        response = self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 2, "sale_price": "250000", "discount_reason": "Aksiya",
            "payment_type": "mixed", "cash_amount": "200000", "card_amount": "300000",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        summary = self.client.get("/api/accounting/").data["summary"]
        self.assertEqual(Decimal(summary["total_sales"]), Decimal("500000.00"))
        self.assertEqual(Decimal(summary["cash_total"]), Decimal("200000.00"))
        self.assertEqual(Decimal(summary["card_total"]), Decimal("300000.00"))
        self.assertEqual(summary["total_quantity"], 2)

    def test_mixed_payment_shows_in_sales_history(self):
        item = self._debt_catalog(name="Aralash tarix", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed", "cash_amount": "120000", "card_amount": "180000",
        }, format="json")
        row = self.client.get("/api/catalog/sales/").data["results"][0]
        self.assertEqual(row["payment_type"], "mixed")
        self.assertEqual(row["payment_label"], "Aralash")
        self.assertEqual(row["payment_breakdown"]["cash"], Decimal("120000.00"))
        self.assertEqual(row["payment_breakdown"]["card"], Decimal("180000.00"))
        totals = self.client.get("/api/catalog/sales/").data["totals"]
        self.assertEqual(totals["cash_total"], Decimal("120000.00"))
        self.assertEqual(totals["card_total"], Decimal("180000.00"))
        self.assertEqual(totals["mixed_count"], 1)

    def test_plain_payments_still_have_no_breakdown(self):
        item = self._debt_catalog(name="Oddiy naqd", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        row = self.client.get("/api/catalog/sales/").data["results"][0]
        self.assertEqual(row["payment_type"], "cash")
        self.assertIsNone(row["payment_breakdown"])

    def test_catalog_sales_history_lists_every_sale(self):
        # bitta dona sotilgan bo'lsa ham tarixda chiqishi kerak
        first = self._debt_catalog(name="Tarix A", price="300000", quantity=3)
        second = self._debt_catalog(name="Tarix B", price="150000", quantity=1)
        self.client.post(f"/api/catalog/{first.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.client.post(f"/api/catalog/{first.id}/sell/", {"quantity": 2, "payment_type": "card"}, format="json")
        self.client.post(f"/api/catalog/{second.id}/sell/", {"quantity": 1, "payment_type": "card"}, format="json")

        response = self.client.get("/api/catalog/sales/")
        self.assertEqual(response.status_code, 200)
        rows = response.data["results"]
        self.assertEqual(len(rows), 3)
        # yangisidan boshlanadi
        self.assertEqual(rows[0]["catalog_name"], "Tarix B")
        totals = response.data["totals"]
        self.assertEqual(totals["sales_count"], 3)
        self.assertEqual(totals["quantity"], 4)
        self.assertEqual(totals["revenue"], Decimal("1050000.00"))
        self.assertEqual(totals["cash_total"], Decimal("300000.00"))
        self.assertEqual(totals["card_total"], Decimal("750000.00"))

    def test_catalog_sales_row_has_page_fields(self):
        item = self._debt_catalog(name="Qatordagi maydonlar", price="300000", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "card", "sale_image_url": "https://example.com/s.jpg",
        }, format="json")
        row = self.client.get("/api/catalog/sales/").data["results"][0]
        self.assertEqual(row["catalog_item"], item.id)
        self.assertEqual(row["catalog_name"], "Qatordagi maydonlar")
        self.assertEqual(row["quantity"], 1)
        self.assertEqual(row["sale_total"], Decimal("300000.00"))
        self.assertEqual(row["payment_label"], "Karta")
        self.assertEqual(row["sale_image_url"], "https://example.com/s.jpg")
        self.assertEqual(row["branch_name"], "Toshkent (asosiy filial)")
        self.assertTrue(row["created_at"].endswith("+05:00"))

    def test_single_catalog_sales_history(self):
        first = self._debt_catalog(name="Faqat shu", price="300000", quantity=2)
        other = self._debt_catalog(name="Boshqasi", price="150000", quantity=1)
        self.client.post(f"/api/catalog/{first.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.client.post(f"/api/catalog/{first.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.client.post(f"/api/catalog/{other.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        response = self.client.get(f"/api/catalog/{first.id}/sales/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["results"]), 2)
        self.assertEqual(response.data["totals"]["quantity"], 2)
        self.assertEqual(response.data["totals"]["revenue"], Decimal("600000.00"))

    def test_catalog_sales_can_be_filtered(self):
        item = self._debt_catalog(name="Filtrli", price="300000", quantity=2)
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "card"}, format="json")
        cash = self.client.get("/api/catalog/sales/?payment_type=cash")
        self.assertEqual(cash.data["totals"]["sales_count"], 1)
        self.assertEqual(cash.data["totals"]["cash_total"], Decimal("300000.00"))
        named = self.client.get("/api/catalog/sales/?search=Filtrli")
        self.assertEqual(named.data["totals"]["sales_count"], 2)
        empty = self.client.get("/api/catalog/sales/?date_from=2020-01-01&date_to=2020-01-02")
        self.assertEqual(empty.data["totals"]["sales_count"], 0)

    def test_branch_user_sees_only_own_catalog_sales(self):
        item = self._main_catalog(quantity=4, price="300000")
        transfer = self.client.post(f"/api/catalog/{item.id}/transfer/", {"branch": self._parkent().id, "quantity": 2, "price": "500000"}, format="json")
        target = CatalogItem.objects.get(id=transfer.data["target_item"])
        mark_catalog_sold(item, self.user, quantity=1)
        mark_catalog_sold(target, self.user, quantity=1)
        client = self._parkent_client("parkent-sales-eye")
        rows = client.get("/api/catalog/sales/").data["results"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["branch_name"], "Parkent filiali")
        self.assertEqual(len(self.client.get("/api/catalog/sales/").data["results"]), 1)

    def test_api_returns_local_time_everywhere(self):
        # sotuv vaqti hamma endpointda bir xil, mahalliy vaqtda (+05:00) chiqishi kerak
        item = self._debt_catalog(name="Vaqt buketi", quantity=1)
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "cash", "sale_image_url": "https://example.com/x.jpg",
        }, format="json")
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        local = timezone.localtime(history.created_at)

        catalog = self.client.get(f"/api/catalog/{item.id}/")
        self.assertTrue(catalog.data["created_at"].endswith("+05:00"), catalog.data["created_at"])

        report = self.client.get("/api/accounting/")
        row = next(r for r in report.data["history"] if r["catalog_id"] == item.id)
        self.assertIn("+05:00", str(row["sold_at"]), str(row["sold_at"]))
        self.assertNotIn("Z", str(row["sold_at"]))
        self.assertIn(local.strftime("%H:%M"), str(row["sold_at"]))

    def test_json_safe_keeps_local_time(self):
        from .views import json_safe

        now = timezone.now()
        value = json_safe({"at": now})["at"]
        self.assertTrue(value.endswith("+05:00"), value)
        self.assertIn(timezone.localtime(now).strftime("%H:%M"), value)

    def test_export_labels_are_in_uzbek(self):
        from .views import arrangement_text, catalog_kind_text, volume_text

        self.assertEqual(volume_text("small"), "Kichik")
        self.assertEqual(volume_text("medium"), "O‘rta")
        self.assertEqual(volume_text("large"), "Katta")
        self.assertEqual(volume_text("M"), "M")
        self.assertEqual(volume_text(""), "Belgilanmagan")
        self.assertEqual(catalog_kind_text("standard"), "Standart")
        self.assertEqual(catalog_kind_text("custom"), "Maxsus")
        self.assertEqual(arrangement_text("bouquet"), "Buket")
        self.assertEqual(arrangement_text("basket"), "Savat")

    def test_florist_stats_carry_uzbek_volume_label(self):
        profile = self._florist_with_rates("fl-uz-label")
        StockBatch.objects.filter(pk=self.batch.pk).update(remaining_stems=F("remaining_stems") + 100)
        self.client.post("/api/florist-stock-issues/issue/", {"florist": profile.id, "batch": self.batch.id, "quantity_stems": 100}, format="json")
        FloristVolumeRate.objects.create(florist=profile, arrangement_type="bouquet", volume="small", default_stems=15, florist_fee=Decimal("10000"))
        self.client.post("/api/catalog/", {
            "name_uz": "Kichik buket", "arrangement_type": "bouquet", "volume": "small",
            "florist": profile.id, "price": "300000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id}],
        }, format="json")
        response = self.client.get(f"/api/florists/{profile.id}/stats/")
        row = next(r for r in response.data["salary_entries"] if r["catalog_name"] == "Kichik buket")
        self.assertEqual(row["volume"], "small")
        self.assertEqual(row["volume_label"], "Kichik")
        self.assertEqual(row["catalog_kind_label"], "Standart")
        self.assertEqual(row["arrangement_label"], "Buket")

    def test_florist_excel_export_opens(self):
        profile = self._florist_with_rates("fl-excel-uz")
        response = self.client.get(f"/api/exports/florists/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheet", response["Content-Type"])

    def test_lists_show_newest_first(self):
        # oxirgi qo'shilgani ro'yxatda birinchi turishi kerak
        first = self.client.post("/api/stock-deliveries/", {"number": "ORD-1", "received_at": "2026-08-01"}, format="json").json()
        second = self.client.post("/api/stock-deliveries/", {"number": "ORD-2", "received_at": "2026-08-01"}, format="json").json()
        rows = self.client.get("/api/stock-deliveries/").data["results"]
        self.assertEqual(rows[0]["id"], second["id"])
        self.assertEqual(rows[1]["id"], first["id"])

        payload = {
            "delivery": second["id"], "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 25,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }
        batch_one = self.client.post("/api/stock-batches/", payload, format="json").json()
        # bo'yi boshqa — shuning uchun birinchi qatorga qo'shilmaydi, yangi qator ochiladi
        batch_two = self.client.post("/api/stock-batches/", {**payload, "height_cm": 60}, format="json").json()
        batches = self.client.get("/api/stock-batches/").data["results"]
        self.assertEqual(batches[0]["id"], batch_two["id"])
        self.assertEqual(batches[1]["id"], batch_one["id"])

        catalog_payload = {
            "name_uz": "Tartib buketi", "arrangement_type": "bouquet", "price": "300000",
            "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 5}],
        }
        cat_one = self.client.post("/api/catalog/", catalog_payload, format="json").json()
        cat_two = self.client.post("/api/catalog/", catalog_payload, format="json").json()
        catalogs = self.client.get("/api/catalog/").data["results"]
        self.assertEqual(catalogs[0]["id"], cat_two["id"])
        self.assertEqual(catalogs[1]["id"], cat_one["id"])

    def test_ordering_can_still_be_reversed(self):
        first = self.client.post("/api/stock-deliveries/", {"number": "ORD-A", "received_at": "2026-08-01"}, format="json").json()
        second = self.client.post("/api/stock-deliveries/", {"number": "ORD-B", "received_at": "2026-08-01"}, format="json").json()
        rows = self.client.get("/api/stock-deliveries/?ordering=created_at").data["results"]
        self.assertEqual(rows[0]["id"], first["id"])
        self.assertEqual(rows[-1]["id"], second["id"])

    def test_material_deliveries_show_newest_first(self):
        first = self.client.post("/api/material-deliveries/", {"number": "MORD-1", "received_at": "2026-08-01"}, format="json").json()
        second = self.client.post("/api/material-deliveries/", {"number": "MORD-2", "received_at": "2026-08-01"}, format="json").json()
        rows = self.client.get("/api/material-deliveries/").data["results"]
        self.assertEqual(rows[0]["id"], second["id"])
        self.assertEqual(rows[1]["id"], first["id"])

    def test_batch_usage_summary_is_reported(self):
        batch, item, _ = self._used_batch_with_sale()
        response = self.client.get(f"/api/stock-batches/{batch.id}/usage/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["is_used"])
        self.assertEqual(response.data["catalog_items"], 1)
        self.assertEqual(response.data["sold_catalog_items"], 1)
        self.assertEqual(response.data["used_stems"], 20)

    def test_used_batch_variant_can_be_changed_with_reason(self):
        batch, item, other = self._used_batch_with_sale()
        old_cost = CatalogItem.objects.get(pk=item.pk).calculated_cost_price
        response = self.client.post(f"/api/stock-batches/{batch.id}/change-variant/", {
            "variant": other.id, "reason": "Kirimda xato yozilgan",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.variant_id, other.id)
        # katalog tarkibi ham yangi navni ko'rsatadi
        self.assertEqual(item.composition.get().stock_batch.variant_id, other.id)
        # pul o'zgarmaydi
        self.assertEqual(CatalogItem.objects.get(pk=item.pk).calculated_cost_price, old_cost)

    def test_variant_change_updates_frozen_sale_snapshot(self):
        batch, item, other = self._used_batch_with_sale()
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        self.assertIn("Freedom", json.dumps(history.snapshot, ensure_ascii=False))
        self.client.post(f"/api/stock-batches/{batch.id}/change-variant/", {
            "variant": other.id, "reason": "Xato nav",
        }, format="json")
        history.refresh_from_db()
        self.assertIn("Yangi nav", json.dumps(history.snapshot, ensure_ascii=False))
        self.assertNotIn("Freedom", json.dumps(history.snapshot, ensure_ascii=False))

    def test_variant_change_requires_reason(self):
        batch, _, other = self._used_batch_with_sale()
        response = self.client.post(f"/api/stock-batches/{batch.id}/change-variant/", {"variant": other.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("reason", response.data)
        batch.refresh_from_db()
        self.assertEqual(batch.variant_id, self.batch.variant_id)

    def test_variant_change_rejects_same_variant(self):
        batch, _, _ = self._used_batch_with_sale()
        response = self.client.post(f"/api/stock-batches/{batch.id}/change-variant/", {
            "variant": batch.variant_id, "reason": "Bir xil",
        }, format="json")
        self.assertEqual(response.status_code, 400)

    def test_variant_change_is_written_to_audit(self):
        batch, _, other = self._used_batch_with_sale()
        self.client.post(f"/api/stock-batches/{batch.id}/change-variant/", {
            "variant": other.id, "reason": "Kirimda xato yozilgan",
        }, format="json")
        row = AuditLog.objects.filter(action="stock_batch_variant_changed").first()
        self.assertIsNotNone(row)
        self.assertEqual(row.after["reason"], "Kirimda xato yozilgan")
        self.assertEqual(row.after["usage"]["catalog_items"], 1)

    def test_plain_patch_still_blocks_variant_and_points_to_action(self):
        batch, _, other = self._used_batch_with_sale()
        response = self.client.patch(f"/api/stock-batches/{batch.id}/", {"variant": other.id}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("change-flower", str(response.data["flower"]))

    def test_changing_stems_per_bunch_recalculates_stem_prices(self):
        created = self.client.post("/api/stock-batches/", {
            "batch_number": "BUNCH-CH", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json").json()
        self.assertEqual(Decimal(created["cost_per_stem"]), Decimal("1000.00"))
        self.assertEqual(Decimal(created["sale_price_per_stem"]), Decimal("2000.00"))
        response = self.client.patch(f"/api/stock-batches/{created['id']}/", {"stems_per_bunch": 50}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        batch = StockBatch.objects.get(id=created["id"])
        # 25 000 / 50 = 500,  50 000 / 50 = 1 000
        self.assertEqual(batch.cost_per_stem, Decimal("500.00"))
        self.assertEqual(batch.sale_price_per_stem, Decimal("1000.00"))
        # pochka narxlari o'zgarmaydi
        self.assertEqual(batch.cost_per_bunch, Decimal("25000.00"))
        self.assertEqual(batch.sale_price_per_bunch, Decimal("50000.00"))

    def test_typed_stem_price_wins_over_recalculation(self):
        created = self.client.post("/api/stock-batches/", {
            "batch_number": "BUNCH-CH2", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_bunch": "25000", "sale_price_per_bunch": "50000",
        }, format="json").json()
        response = self.client.patch(f"/api/stock-batches/{created['id']}/", {
            "stems_per_bunch": 50, "cost_per_stem": "700",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(StockBatch.objects.get(id=created["id"]).cost_per_stem, Decimal("700.00"))

    def test_free_batch_needs_no_cost_price(self):
        # postavshik tekinga qo'shib bergan gul: faqat sotuv narxi yoziladi
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "FREE-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 100, "is_free": True,
            "sale_price_per_bunch": "50000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        data = response.json()
        self.assertTrue(data["is_free"])
        self.assertEqual(Decimal(data["cost_per_stem"]), Decimal("0.00"))
        self.assertEqual(Decimal(data["cost_per_bunch"]), Decimal("0.00"))
        self.assertEqual(Decimal(data["cost_per_stem_exact"]), Decimal("0.0000"))
        # sotuv narxi odatdagidek hisoblanadi
        self.assertEqual(Decimal(data["sale_price_per_stem"]), Decimal("2000.00"))

    def test_free_batch_ignores_typed_cost(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "FREE-2", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 50, "is_free": True,
            "cost_per_bunch": "99000", "sale_price_per_bunch": "50000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(Decimal(response.json()["cost_per_stem"]), Decimal("0.00"))
        self.assertEqual(Decimal(response.json()["cost_per_bunch"]), Decimal("0.00"))

    def test_free_batch_needs_sale_price(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "FREE-3", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 50, "is_free": True,
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("sale_price_per_bunch", response.data)

    def test_free_batch_adds_nothing_to_supplier_debt(self):
        supplier = Supplier.objects.create(name="Tekin beruvchi")
        self.client.post("/api/stock-batches/", {
            "batch_number": "FREE-4", "variant": self.batch.variant_id, "supplier": supplier.id,
            "height_cm": 50, "stems_per_bunch": 25, "received_stems": 100, "is_free": True,
            "sale_price_per_bunch": "50000",
        }, format="json")
        row = next(r for r in self.client.get("/api/suppliers/").data["results"] if r["name"] == "Tekin beruvchi")
        self.assertEqual(Decimal(row["purchase_total"]), Decimal("0.00"))

    def test_free_batch_makes_catalog_flower_cost_zero(self):
        free = StockBatch.objects.create(
            variant=self.batch.variant, batch_number="FREE-5", height_cm=50, stems_per_bunch=25,
            received_stems=100, remaining_stems=100, is_free=True, cost_per_stem=0,
            sale_price_per_stem=2000, sale_price_per_bunch=50000,
        )
        created = self.client.post("/api/catalog/", {
            "name_uz": "Tekin guldan buket", "arrangement_type": "bouquet",
            "price": "300000", "quantity_total": 1, "status": "available",
            "composition": [{"stock_batch": free.id, "quantity_stems": 20}],
        }, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        item = CatalogItem.objects.get(id=created.json()["id"])
        from .inventory_services import catalog_cost_breakdown
        self.assertEqual(catalog_cost_breakdown(item)["flower_cost"], Decimal("0"))

    def test_free_batches_can_be_filtered(self):
        StockBatch.objects.create(
            variant=self.batch.variant, batch_number="FREE-6", height_cm=50, stems_per_bunch=25,
            received_stems=50, remaining_stems=50, is_free=True, cost_per_stem=0,
            sale_price_per_stem=2000, sale_price_per_bunch=50000,
        )
        free = self.client.get("/api/stock-batches/?is_free=true")
        self.assertEqual([row["batch_number"] for row in free.data["results"]], ["FREE-6"])
        paid = self.client.get("/api/stock-batches/?is_free=false")
        self.assertIn("API-1", [row["batch_number"] for row in paid.data["results"]])

    def test_batch_without_any_price_is_rejected(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "NOPRICE-1", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 25,
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("cost_per_bunch", response.data)

    def test_batch_without_sale_price_is_rejected(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "NOPRICE-2", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 25, "received_stems": 25, "cost_per_bunch": "25000",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("sale_price_per_bunch", response.data)

    def test_stem_price_fills_bunch_price_backwards(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "AUTO-3", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 10, "received_stems": 50,
            "cost_per_stem": "3000", "sale_price_per_stem": "5000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertEqual(Decimal(response.json()["cost_per_bunch"]), Decimal("30000.00"))
        self.assertEqual(Decimal(response.json()["sale_price_per_bunch"]), Decimal("50000.00"))

    def test_batch_without_delivery_gets_one_automatically(self):
        response = self.client.post("/api/stock-batches/", {
            "batch_number": "AUTO-4", "variant": self.batch.variant_id, "height_cm": 50,
            "stems_per_bunch": 10, "received_stems": 20,
            "cost_per_stem": "1000", "sale_price_per_stem": "2000", "sale_price_per_bunch": "20000",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.json())
        self.assertIsNotNone(response.json()["delivery"])
        self.assertEqual(response.json()["delivery_detail"]["number"], "AUTO-4")

    def test_same_delivery_reused_for_same_number_and_date(self):
        payload = {
            "batch_number": "AUTO-5", "variant": self.batch.variant_id, "height_cm": 50,
            "received_at": "2026-08-01", "stems_per_bunch": 10, "received_stems": 20,
            "cost_per_stem": "1000", "sale_price_per_stem": "2000", "sale_price_per_bunch": "20000",
        }
        one = self.client.post("/api/stock-batches/", payload, format="json")
        two = self.client.post("/api/stock-batches/", payload, format="json")
        self.assertEqual(one.status_code, 201)
        self.assertEqual(two.status_code, 201)
        self.assertEqual(one.json()["delivery"], two.json()["delivery"])
        self.assertEqual(StockDelivery.objects.filter(number="AUTO-5").count(), 1)

    def test_stock_batch_number_can_repeat(self):
        # bir xil raqamli partiyalar turli gul va turli kunlarda kelaveradi
        flower = Flower.objects.create(name_uz="Xrizantema API", slug="xrizantema-api")
        first = FlowerVariant.objects.create(flower=flower, name_uz="Oq", color_uz="Oq")
        second = FlowerVariant.objects.create(flower=flower, name_uz="Sariq", color_uz="Sariq")
        payload = {
            "batch_number": "1",
            "height_cm": 50,
            "stems_per_bunch": 5,
            "received_stems": 20,
            "cost_per_stem": "10000.00",
            "sale_price_per_stem": "20000.00",
            "sale_price_per_bunch": "100000.00",
        }
        one = self.client.post("/api/stock-batches/", {**payload, "variant": first.id}, format="json")
        self.assertEqual(one.status_code, 201, one.json())
        two = self.client.post("/api/stock-batches/", {**payload, "variant": second.id}, format="json")
        self.assertEqual(two.status_code, 201, two.json())
        self.assertEqual(StockBatch.objects.filter(batch_number="1").count(), 2)
        # ayni gulni ayni raqam bilan yana kiritsa yangi qator ochilmaydi, soni qo'shiladi
        three = self.client.post("/api/stock-batches/", {**payload, "variant": first.id}, format="json")
        self.assertEqual(three.status_code, 201, three.json())
        self.assertEqual(StockBatch.objects.filter(batch_number="1").count(), 2)
        self.assertEqual(three.json()["id"], one.json()["id"])
        self.assertEqual(three.json()["received_stems"], 40)

    @override_settings(OPENAI_API_KEY="")
    def test_mini_app_custom_quote_ai_returns_final_price_note(self):
        BusinessSettings.objects.update_or_create(pk=1, defaults={"default_florist_fee": Decimal("50000")})
        result = mini_app_custom_quote_ai("10 ta atirgul buketga", "bouquet")
        self.assertEqual(result["ai_note"], mini_app_quote_note(result["estimated_price"]))
        self.assertEqual(result["ai_note"], "Taxminiy narx 50 000 so'm. Operatorlarimiz aloqaga chiqib, sizga batafsil ma'lumot berishadi.")

    def test_mini_app_lead_history_returns_customer_orders(self):
        CatalogItem.objects.create(name_uz="Mini katalog", arrangement_type="bouquet", price=250000, status="available")
        init_data = 'user={"id":777,"first_name":"Ali"}'
        payload = {"init_data": init_data, "arrangement_type": "basket", "request_text": "7 ta gortenziya savatga", "name": "Ali", "phone": "901234567", "note": "Bugun kerak"}
        from unittest.mock import patch
        quote = {"lines": [{"type": "custom_text", "request_text": "7 ta gortenziya savatga"}], "packaging": None, "florist_fee": "50000", "estimated_price": "750000", "price_is_estimate": True, "ai_note": "Taxminiy narx"}
        with patch("core.views.mini_app_custom_quote_ai", return_value=quote):
            response = APIClient().post("/api/mini-app/leads/", payload, format="json")
        self.assertEqual(response.status_code, 201)
        lead = Lead.objects.get(source="mini_app")
        self.assertEqual(lead.customer.instagram_user_id, "miniapp:777")
        self.assertEqual(lead.customer.phone, "+998901234567")
        self.assertEqual(lead.details["lines"][0]["request_text"], "7 ta gortenziya savatga")
        self.assertTrue(lead.details["price_is_estimate"])
        response = APIClient().get("/api/mini-app/me/", {"init_data": init_data})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["customer"]["name"], "Ali")
        self.assertEqual(len(data["orders"]), 1)
        self.assertEqual(data["orders"][0]["details"]["lines"][0]["type"], "custom_text")
        response = APIClient().get("/api/mini-app/leads/", {"init_data": init_data})
        self.assertEqual(response.status_code, 200)
        leads_data = response.json()
        self.assertEqual(leads_data["customer"]["name"], "Ali")
        self.assertEqual(len(leads_data["orders"]), 1)
        self.assertEqual(leads_data["orders"][0]["status"], "new")
        response = APIClient().get("/api/mini-app/catalog/", {"init_data": init_data})
        self.assertEqual(response.status_code, 200)
        catalog_data = response.json()
        self.assertEqual(len(catalog_data["orders"]), 1)
        self.assertNotIn("stock", catalog_data)
        self.assertNotIn("packaging", catalog_data)
        self.assertEqual(catalog_data["catalog"][0]["name_uz"], "Mini katalog")

    def test_catalog_create_rejects_short_stock_for_total_quantity(self):
        payload = {
                        "name_uz": "Kop buket",
            "arrangement_type": "bouquet",
            "price": "100000.00",
            "status": "available",
            "quantity_total": 40,
            "composition": [{"stock_batch": self.batch.id, "quantity_stems": 3}],
        }
        response = self.client.post("/api/catalog/", payload, format="json")
        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertIsInstance(data["detail"], str)
        self.assertIn("Katalogni saqlash uchun sklad qoldig'i yetarli emas.", data["detail"])
        self.assertIn("Gul: Atirgul API Freedom Qizil", data["detail"])
        self.assertIn("Kerak: 120 dona", data["detail"])
        self.assertIn("Bor: 100 dona", data["detail"])
        self.assertIn("Yetmayapti: 20 dona", data["detail"])

    def test_flower_delete_archives_when_variants_exist(self):
        flower = Flower.objects.create(name_uz="Liliya", slug="lily-delete")
        FlowerVariant.objects.create(flower=flower, name_uz="Oriental", color_uz="Oq")
        response = self.client.delete(f"/api/flowers/{flower.id}/")
        self.assertEqual(response.status_code, 204)
        flower.refresh_from_db()
        self.assertFalse(flower.is_active)
        self.assertFalse(flower.variants.first().is_active)
        self.assertTrue(AuditLog.objects.filter(action="flower_archived", entity_id=str(flower.id)).exists())

    def test_flower_variant_delete_archives_when_stock_batches_exist(self):
        variant_id = self.batch.variant_id
        response = self.client.delete(f"/api/flower-variants/{variant_id}/")
        self.assertEqual(response.status_code, 204)
        variant = FlowerVariant.objects.get(id=variant_id)
        self.assertFalse(variant.is_active)
        self.assertTrue(AuditLog.objects.filter(action="flowervariant_archived", entity_id=str(variant_id)).exists())

    def test_stock_batch_accepts_height_range_without_height_cm(self):
        payload = {
            "variant": self.batch.variant_id,
            "batch_number": "API-RANGE-1",
            "height_from_cm": 50,
            "height_to_cm": 60,
            "stems_per_bunch": 20,
            "received_stems": 40,
            "remaining_stems": 40,
            "cost_per_stem": "10000.00",
            "sale_price_per_stem": "20000.00",
            "sale_price_per_bunch": "400000.00",
            "minimum_sale_stems": 5,
        }
        response = self.client.post("/api/stock-batches/", payload, format="json")
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["height_cm"], 50)
        self.assertEqual(data["height_from_cm"], 50)
        self.assertEqual(data["height_to_cm"], 60)
        self.assertEqual(data["height_label"], "50-60 sm")

    def test_ai_stock_rows_include_variant_description(self):
        variant = self.batch.variant
        variant.description_uz = "Premium import gortenziya, boshi yirikroq va rangi to‘qroq."
        variant.description_ru = "Премиальная импортная гортензия."
        variant.save(update_fields=["description_uz", "description_ru", "updated_at"])
        rows = ai_stock_rows("premium import")
        self.assertEqual(rows[0]["description_uz"], "Premium import gortenziya, boshi yirikroq va rangi to‘qroq.")
        self.assertEqual(rows[0]["description_ru"], "Премиальная импортная гортензия.")

    def test_ai_stock_rows_include_height_range_label(self):
        self.batch.height_from_cm = 50
        self.batch.height_to_cm = 60
        self.batch.save(update_fields=["height_from_cm", "height_to_cm", "updated_at"])
        rows = ai_stock_rows("Freedom")
        self.assertEqual(rows[0]["height_label"], "50-60 sm")
        self.assertEqual(rows[0]["height_from_cm"], 50)
        self.assertEqual(rows[0]["height_to_cm"], 60)

    def test_ai_stock_rows_include_image_url(self):
        self.batch.image_url = "https://example.com/freedom.jpg"
        self.batch.save(update_fields=["image_url", "updated_at"])
        rows = ai_stock_rows("Freedom")
        self.assertTrue(rows[0]["has_image"])
        self.assertEqual(rows[0]["image_url"], "https://example.com/freedom.jpg")

    def test_manual_lead_create_customer_does_not_deduct_stock_when_won(self):
        packaging = Packaging.objects.create(packaging_type="basket", name_uz="Lead savat", quantity=2, sale_price=50000)
        payload = {
                        "status": "new",
            "request_uz": "Manual lead",
            "arrangement_type": "basket",
            "estimated_price": "250000.00",
            "florist_fee": "50000.00",
            "customer_name": "Vali",
            "customer_phone": "901112233",
            "stock_usage_input": [{"stock_batch": self.batch.id, "quantity_stems": 4}],
            "packaging_usage_input": [{"packaging": packaging.id, "quantity": 1}],
        }
        response = self.client.post("/api/leads/", payload, format="json")
        self.assertEqual(response.status_code, 201)
        lead_id = response.json()["id"]
        self.batch.refresh_from_db()
        packaging.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 100)
        self.assertEqual(packaging.quantity, 2)
        response = self.client.patch(f"/api/leads/{lead_id}/", {"status": "won"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        packaging.refresh_from_db()
        lead = Lead.objects.get(id=lead_id)
        self.assertEqual(self.batch.remaining_stems, 100)
        self.assertEqual(packaging.quantity, 2)
        self.assertIsNone(lead.stock_deducted_at)
        self.assertFalse(StockMovement.objects.filter(reference_type="lead", reference_id=lead_id).exists())
        self.assertFalse(PackagingMovement.objects.filter(reference_type="lead", reference_id=lead_id).exists())
        self.assertEqual(Customer.objects.get(phone="+998901112233").leads.count(), 1)
        response = self.client.patch(f"/api/leads/{lead_id}/", {"status": "lost"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        packaging.refresh_from_db()
        lead = Lead.objects.get(id=lead_id)
        self.assertEqual(self.batch.remaining_stems, 100)
        self.assertEqual(packaging.quantity, 2)
        self.assertIsNone(lead.stock_deducted_at)
        self.assertFalse(StockMovement.objects.filter(reference_type="lead", reference_id=lead_id).exists())
        self.assertFalse(PackagingMovement.objects.filter(reference_type="lead", reference_id=lead_id).exists())
        response = self.client.patch(f"/api/leads/{lead_id}/", {"status": "won"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        packaging.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 100)
        self.assertEqual(packaging.quantity, 2)

    def test_catalog_lead_does_not_sell_catalog_when_won_is_reverted(self):
        item = CatalogItem.objects.create(name_uz="Catalog buket", arrangement_type="bouquet", price=300000, quantity_total=3, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=5)
        deduct_catalog_stock(item, self.user)
        payload = {
                        "status": "new",
            "request_uz": "Catalog lead",
            "arrangement_type": "catalog",
            "customer_name": "Sardor",
            "customer_phone": "901234000",
            "catalog_usage_input": [{"catalog_item": item.id, "quantity": 2}],
        }
        response = self.client.post("/api/leads/", payload, format="json")
        self.assertEqual(response.status_code, 201)
        lead_id = response.json()["id"]
        response = self.client.post(f"/api/leads/{lead_id}/move/", {"status": "won"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 85)
        self.assertEqual(item.quantity_sold, 0)
        self.assertEqual(item.quantity_stock_deducted, 3)
        response = self.client.post(f"/api/leads/{lead_id}/move/", {"status": "new"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 85)
        self.assertEqual(item.quantity_sold, 0)
        self.assertEqual(item.quantity_stock_deducted, 3)
        self.assertEqual(item.status, "available")

    def test_catalog_lead_delete_does_not_restore_when_won_was_not_deducted(self):
        item = CatalogItem.objects.create(name_uz="Delete buket", arrangement_type="bouquet", price=300000, quantity_total=3, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=5)
        deduct_catalog_stock(item, self.user)
        payload = {
                        "status": "new",
            "request_uz": "Delete lead",
            "arrangement_type": "catalog",
            "customer_name": "Sardor",
            "customer_phone": "901234001",
            "catalog_usage_input": [{"catalog_item": item.id, "quantity": 2}],
        }
        response = self.client.post("/api/leads/", payload, format="json")
        self.assertEqual(response.status_code, 201)
        lead_id = response.json()["id"]
        response = self.client.patch(f"/api/leads/{lead_id}/", {"status": "won"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.batch.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 85)
        self.assertEqual(item.quantity_sold, 0)
        response = self.client.delete(f"/api/leads/{lead_id}/")
        self.assertEqual(response.status_code, 204)
        self.batch.refresh_from_db()
        item.refresh_from_db()
        self.assertEqual(self.batch.remaining_stems, 85)
        self.assertEqual(item.quantity_sold, 0)
        self.assertEqual(item.quantity_stock_deducted, 3)
        self.assertFalse(Lead.objects.filter(id=lead_id).exists())
        audit = AuditLog.objects.filter(action="lead_deleted", entity_id=str(lead_id)).first()
        self.assertIsNotNone(audit)
        self.assertEqual(audit.before["status"], "won")
        self.assertEqual(audit.after["restored_before_delete"]["stock_deducted_at"], None)

    def test_audit_hides_developer_logs_from_non_developer(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="audit", can_view=True, can_control=False)
        developer = User.objects.create_user("developer-audit", password="password")
        UserProfile.objects.create(user=developer, role="developer")
        AuditLog.objects.create(user=developer, action="secret", entity_type="AISettings", entity_id="1")
        AuditLog.objects.create(user=self.user, action="visible", entity_type="Lead", entity_id="1")
        response = self.client.get("/api/audit/")
        self.assertEqual(response.status_code, 200)
        actions = [row["action"] for row in response.json()["results"]]
        self.assertIn("visible", actions)
        self.assertNotIn("secret", actions)
        self.assertEqual(response.json()["results"][0]["action_label"], "Visible")

    def test_audit_can_filter_by_user_without_exposing_developer_logs(self):
        UserProfile.objects.create(user=self.user, role="admin")
        PagePermission.objects.create(user=self.user, page="audit", can_view=True, can_control=False)
        operator = User.objects.create_user("operator-audit", password="password", first_name="Operator")
        UserProfile.objects.create(user=operator, role="operator")
        developer = User.objects.create_user("developer-filter-audit", password="password")
        UserProfile.objects.create(user=developer, role="developer")
        AuditLog.objects.create(user=self.user, action="admin_action", entity_type="Lead", entity_id="1")
        AuditLog.objects.create(user=operator, action="operator_action", entity_type="Lead", entity_id="2")
        AuditLog.objects.create(user=developer, action="developer_action", entity_type="AISettings", entity_id="3")
        response = self.client.get(f"/api/audit/?user={operator.id}")
        self.assertEqual(response.status_code, 200)
        actions = [row["action"] for row in response.json()["results"]]
        self.assertEqual(actions, ["operator_action"])
        self.assertEqual(response.json()["results"][0]["actor_name"], "Operator")
        response = self.client.get(f"/api/audit/?user_id={self.user.id}")
        self.assertEqual(response.status_code, 200)
        actions = [row["action"] for row in response.json()["results"]]
        self.assertEqual(actions, ["admin_action"])
        response = self.client.get(f"/api/audit/?user={developer.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["results"], [])

    def test_analytics_and_dashboard_include_top_selling_flowers(self):
        item = CatalogItem.objects.create(name_uz="Analytics buket", arrangement_type="bouquet", price=300000, quantity_total=5, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=4)
        customer = Customer.objects.create(name="Analytics", phone="+998901234501", instagram_user_id="analytics")
        lead = Lead.objects.create(customer=customer, status="won", request_uz="Analytics lead", arrangement_type="catalog", estimated_price=600000, source="instagram")
        LeadCatalogUsage.objects.create(lead=lead, catalog_item=item, quantity=2)
        mark_catalog_sold(item, self.user, quantity=1, sale_price=250000, discount_reason="Analytics skidka")
        response = self.client.get("/api/analytics/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data["daily_stats"]), 30)
        # orders endi lead va katalog sotuvini jamlaydi: 1 won lead + 1 katalog sotuvi
        self.assertEqual(data["summary"]["orders"], 2)
        self.assertEqual(data["summary"]["lead_orders"], 1)
        self.assertEqual(data["summary"]["catalog_sales_orders"], 1)
        self.assertEqual(Decimal(str(data["summary"]["lead_revenue"])), Decimal("600000.00"))
        self.assertEqual(Decimal(str(data["summary"]["catalog_sales_revenue"])), Decimal("250000.00"))
        self.assertEqual(Decimal(str(data["summary"]["revenue"])), Decimal("850000.00"))
        self.assertEqual(data["top_selling_flowers"][0]["name_uz"], "Atirgul API")
        self.assertEqual(data["top_selling_flowers"][0]["stems"], 8)
        self.assertEqual(data["top_catalog_items"][0]["catalog_item__name_uz"], "Analytics buket")
        # katalogdan 1 ta + lead orqali 2 ta
        self.assertEqual(data["top_catalog_items"][0]["quantity"], 3)
        self.assertEqual(data["recent_top_catalog_items"][0]["catalog_item__name_uz"], "Analytics buket")
        self.assertEqual(data["recent_top_catalog_items"][0]["orders"], 2)
        self.assertEqual(data["summary"]["discounted_catalog_sales_count"], 1)
        self.assertEqual(data["summary"]["discounted_catalog_quantity"], 1)
        self.assertEqual(Decimal(str(data["summary"]["discounted_catalog_amount"])), Decimal("50000.00"))
        response = self.client.get("/api/accounting/")
        self.assertEqual(response.status_code, 200)
        accounting = response.json()
        self.assertEqual(accounting["summary"]["discounted_sales_count"], 1)
        self.assertEqual(accounting["summary"]["total_quantity"], 1)
        self.assertEqual(accounting["history"][0]["catalog_name"], "Analytics buket")
        response = self.client.get("/api/dashboard/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["top_selling_flowers"][0]["stems"], 8)
        self.assertEqual(Decimal(str(response.json()["discounted_catalog_amount"])), Decimal("50000.00"))

    def test_lead_move_keeps_kanban_position_between_two_leads(self):
        customer = Customer.objects.create(name="Kanban", phone="+998901234567", instagram_user_id="kanban")
        first = Lead.objects.create(customer=customer, status="new", request_uz="First", sort_order=Decimal("1000"))
        moving = Lead.objects.create(customer=customer, status="new", request_uz="Moving", sort_order=Decimal("2000"))
        last = Lead.objects.create(customer=customer, status="new", request_uz="Last", sort_order=Decimal("3000"))
        response = self.client.post(f"/api/leads/{moving.id}/move/", {"status": "new", "before": first.id, "after": last.id}, format="json")
        self.assertEqual(response.status_code, 200)
        moving.refresh_from_db()
        self.assertEqual(moving.sort_order, Decimal("2000.000000"))
        ids = list(Lead.objects.filter(status="new").order_by("sort_order").values_list("id", flat=True))
        self.assertEqual(ids, [first.id, moving.id, last.id])

    def test_leads_can_be_paginated_by_status_for_kanban_column(self):
        customer = Customer.objects.create(name="Kanban page", phone="+998901234569", instagram_user_id="kanban-page")
        first = Lead.objects.create(customer=customer, status="new", request_uz="First", sort_order=Decimal("1000"))
        second = Lead.objects.create(customer=customer, status="new", request_uz="Second", sort_order=Decimal("2000"))
        Lead.objects.create(customer=customer, status="qualified", request_uz="Other column", sort_order=Decimal("1000"))
        response = self.client.get("/api/leads/", {"status": "new", "page": 1, "page_size": 1})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 2)
        self.assertEqual(len(data["results"]), 1)
        self.assertEqual(data["results"][0]["id"], first.id)
        self.assertTrue(data["next"])
        response = self.client.get("/api/leads/", {"status": "new", "page": 2, "page_size": 1})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["results"][0]["id"], second.id)
        response = self.client.get("/api/leads/")
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.json()["count"], 3)

    def test_lead_reorder_column_accepts_full_column_ids(self):
        customer = Customer.objects.create(name="Kanban full", phone="+998901234568", instagram_user_id="kanban-full")
        first = Lead.objects.create(customer=customer, status="new", request_uz="First", sort_order=Decimal("1000"))
        second = Lead.objects.create(customer=customer, status="new", request_uz="Second", sort_order=Decimal("2000"))
        third = Lead.objects.create(customer=customer, status="new", request_uz="Third", sort_order=Decimal("3000"))
        response = self.client.post("/api/leads/reorder-column/", {"status": "new", "lead_ids": [third.id, first.id, second.id]}, format="json")
        self.assertEqual(response.status_code, 200)
        ids = list(Lead.objects.filter(status="new").order_by("sort_order").values_list("id", flat=True))
        self.assertEqual(ids, [third.id, first.id, second.id])
        third.refresh_from_db()
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(third.sort_order, Decimal("1000.000000"))
        self.assertEqual(first.sort_order, Decimal("2000.000000"))
        self.assertEqual(second.sort_order, Decimal("3000.000000"))
        response = self.client.post("/api/leads/reorder-column/", {"status": "new", "lead_ids": [third.id, first.id]}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn(second.id, response.json()["missing_ids"])


class ExpenseApiTests(TestCase):
    """Rasxodlar sahifasi: qo'lda kiritish, sana, yig'indi va hisob-kitob."""

    def setUp(self):
        self.user = User.objects.create_user("expense-admin", password="password", is_superuser=True, is_staff=True)
        UserProfile.objects.create(user=self.user, role="admin")
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _create(self, **overrides):
        payload = {"amount": "150000", "destination": "Ijara — avgust"}
        payload.update(overrides)
        return self.client.post("/api/expenses/", payload, format="json")

    def test_expense_without_date_takes_current_time(self):
        response = self._create()
        self.assertEqual(response.status_code, 201, response.json())
        expense = Expense.objects.get(pk=response.json()["id"])
        self.assertLess((timezone.now() - expense.spent_at).total_seconds(), 60)
        self.assertEqual(expense.created_by, self.user)
        self.assertEqual(expense.payment_method, "cash")
        self.assertEqual(response.json()["payment_method_label"], "Naqd")

    def test_expense_keeps_chosen_date(self):
        chosen = timezone.now() - timedelta(days=9)
        response = self._create(spent_at=chosen.isoformat())
        self.assertEqual(response.status_code, 201, response.json())
        expense = Expense.objects.get(pk=response.json()["id"])
        self.assertEqual(expense.spent_at.date(), chosen.date())

    def test_expense_requires_destination_and_positive_amount(self):
        self.assertEqual(self._create(destination="   ").status_code, 400)
        self.assertEqual(self._create(amount="0").status_code, 400)
        self.assertEqual(self._create(amount="-5000").status_code, 400)

    def test_expense_summary_groups_by_method_and_day(self):
        today = timezone.now()
        self._create(amount="150000", destination="Ijara")
        self._create(amount="50000", destination="Kuryer", payment_method="card")
        self._create(amount="30000", destination="Benzin", spent_at=(today - timedelta(days=40)).isoformat())
        response = self.client.get("/api/expenses/summary/", {"date_from": (today - timedelta(days=3)).date().isoformat()})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["totals"]["expense_count"], 2)
        self.assertEqual(Decimal(data["totals"]["total"]), Decimal("200000"))
        self.assertNotIn("by_category", data)
        methods = {row["payment_method"]: Decimal(row["total"]) for row in data["by_payment_method"]}
        self.assertEqual(methods, {"cash": Decimal("150000"), "card": Decimal("50000")})
        self.assertEqual(len(data["by_day"]), 1)

    def test_expense_list_filters_by_date_and_search(self):
        self._create(destination="Svet puli", spent_at=(timezone.now() - timedelta(days=30)).isoformat())
        self._create(destination="Reklama", payment_method="card")
        response = self.client.get("/api/expenses/", {"date_from": timezone.now().date().isoformat()})
        self.assertEqual([row["destination"] for row in response.json()["results"]], ["Reklama"])
        response = self.client.get("/api/expenses/", {"search": "svet"})
        self.assertEqual([row["destination"] for row in response.json()["results"]], ["Svet puli"])
        response = self.client.get("/api/expenses/", {"payment_method": "card"})
        self.assertEqual([row["destination"] for row in response.json()["results"]], ["Reklama"])

    def test_expense_options_endpoint_lists_payment_methods(self):
        response = self.client.get("/api/expenses/options/")
        self.assertEqual(response.status_code, 200)
        values = [row["value"] for row in response.json()["payment_methods"]]
        self.assertEqual(values, ["cash", "card", "transfer"])
        self.assertNotIn("categories", response.json())

    def test_expense_can_be_edited_and_deleted(self):
        created = self._create().json()
        response = self.client.patch(f"/api/expenses/{created['id']}/", {"amount": "200000", "note": "Tuzatildi"}, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Decimal(response.json()["amount"]), Decimal("200000"))
        self.assertEqual(self.client.delete(f"/api/expenses/{created['id']}/").status_code, 204)
        self.assertFalse(Expense.objects.filter(pk=created["id"]).exists())

    def test_expense_page_is_closed_without_permission(self):
        other = User.objects.create_user("expense-operator", password="password")
        UserProfile.objects.create(user=other, role="operator")
        client = APIClient()
        client.force_authenticate(other)
        self.assertEqual(client.get("/api/expenses/").status_code, 403)
        self.assertEqual(client.post("/api/expenses/", {"amount": "1000", "destination": "X"}, format="json").status_code, 403)
        PagePermission.objects.create(user=other, page="expenses", can_view=True, can_control=False)
        self.assertEqual(client.get("/api/expenses/").status_code, 200)
        self.assertEqual(client.post("/api/expenses/", {"amount": "1000", "destination": "X"}, format="json").status_code, 403)

    def test_accounting_reports_expenses_separately(self):
        flower = Flower.objects.create(name_uz="Atirgul rasxod", slug="rose-expense")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Freedom", color_uz="Qizil")
        batch = StockBatch.objects.create(variant=variant, batch_number="EXP-1", height_cm=60, stems_per_bunch=20, received_stems=100, remaining_stems=100, cost_per_stem=1000, sale_price_per_stem=5000, sale_price_per_bunch=100000)
        item = CatalogItem.objects.create(name_uz="Rasxod buket", arrangement_type="bouquet", catalog_kind="standard", price=Decimal("500000"), quantity_total=1, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=batch, quantity_stems=10)
        mark_catalog_sold(item, self.user, payment_type="cash")
        self._create(amount="120000", destination="Ijara")
        self._create(amount="80000", destination="Kuryer")
        response = self.client.get("/api/accounting/")
        self.assertEqual(response.status_code, 200)
        summary = response.json()["summary"]
        self.assertEqual(summary["expense_count"], 2)
        self.assertEqual(Decimal(summary["expense_total"]), Decimal("200000"))
        self.assertEqual(
            Decimal(summary["net_profit_after_expenses"]),
            Decimal(summary["net_profit"]) - Decimal("200000"),
        )
        rows = response.json()["expenses"]
        self.assertEqual([row["destination"] for row in rows], ["Kuryer", "Ijara"])
        self.assertEqual(sum(Decimal(row["amount"]) for row in rows), Decimal("200000"))

    def test_accounting_expenses_follow_date_filter(self):
        self._create(amount="70000", destination="Eski rasxod", spent_at=(timezone.now() - timedelta(days=20)).isoformat())
        self._create(amount="40000", destination="Bugungi rasxod")
        today = timezone.now().date().isoformat()
        response = self.client.get("/api/accounting/", {"date_from": today, "date_to": today})
        self.assertEqual(Decimal(response.json()["summary"]["expense_total"]), Decimal("40000"))


class SaleGroupMessageTests(TestCase):
    """Sotilganda guruhga ketadigan xabar."""

    def setUp(self):
        self.user = User.objects.create_user("sale-admin", password="password", first_name="Diyor", last_name="A", is_superuser=True, is_staff=True)
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        flower = Flower.objects.create(name_uz="Atirgul guruh", slug="rose-group")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Freedom", color_uz="Qizil")
        self.batch = StockBatch.objects.create(variant=variant, batch_number="GRP-1", height_cm=60, stems_per_bunch=20, received_stems=100, remaining_stems=100, cost_per_stem=1000, sale_price_per_stem=5000, sale_price_per_bunch=100000)

    def _configure_group(self):
        IntegrationSettings.objects.update_or_create(pk=1, defaults={"sale_bot_token": "test-token", "sale_group_chat_id": "-100500"})

    def _item(self, **overrides):
        data = {"name_uz": "Qizil buket", "arrangement_type": "bouquet", "catalog_kind": "standard",
                "price": Decimal("300000"), "quantity_total": 2, "status": "available",
                "image_url": "https://example.com/katalog.jpg"}
        data.update(overrides)
        item = CatalogItem.objects.create(**data)
        CatalogComposition.objects.create(catalog_item=item, stock_batch=self.batch, quantity_stems=5)
        return item

    def test_caption_shows_amount_payment_and_delivery(self):
        from .inventory_services import sale_group_caption
        item = self._item()
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "cash", "delivery_amount": "50000",
        }, format="json")
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        caption = sale_group_caption(item, history, "cash")
        self.assertIn("Qizil buket", caption)
        self.assertIn("Savdo", caption)
        self.assertIn("250", caption)          # 300 000 - 50 000 dastafka
        self.assertIn("Dastafka", caption)
        self.assertIn("Jami olingan", caption)
        self.assertIn("Naqd", caption)
        self.assertIn("Sotdi: Diyor A", caption)
        self.assertIn("\U0001f338", caption)   # emoji bor

    def test_caption_splits_mixed_payment(self):
        from .inventory_services import sale_group_caption
        item = self._item()
        self.client.post(f"/api/catalog/{item.id}/sell/", {
            "quantity": 1, "payment_type": "mixed", "cash_amount": "100000", "card_amount": "200000",
        }, format="json")
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        caption = sale_group_caption(item, history, "mixed")
        self.assertIn("Aralash", caption)
        self.assertIn("100", caption)
        self.assertIn("200", caption)

    def test_terminal_payment_is_accepted_and_shown(self):
        from unittest.mock import patch
        from .inventory_services import sale_group_caption
        self._configure_group()
        item = self._item(image_url="")
        with patch("core.platform_services.telegram_send_message_with") as sender:
            sender.return_value = {"ok": True}
            response = self.client.post(f"/api/catalog/{item.id}/sell/", {
                "quantity": 1,
                "payment_type": "terminal",
            }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        self.assertEqual((history.snapshot or {}).get("payment_type"), "terminal")
        self.assertIn("Terminal", sale_group_caption(item, history, "terminal"))
        self.assertIn("Terminal", sender.call_args[0][2])

    def test_branch_sale_goes_to_branch_group_only(self):
        from unittest.mock import patch
        # asosiy filial sozlangan, Parkentniki boshqa bot
        self._configure_group()
        branch = Branch.objects.create(name="Parkent", sale_bot_token="parkent-token", sale_group_chat_id="-900900")
        item = self._item(branch=branch)
        with patch("core.platform_services.telegram_send_message_with") as sender:
            sender.return_value = {"ok": True}
            mark_catalog_sold(item, self.user, 1, payment_type="cash")
            from .inventory_services import notify_sale_to_group
            history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
            item.refresh_from_db()
            notify_sale_to_group(item, history, "cash")
        self.assertEqual(sender.call_args[0][0], "parkent-token")
        self.assertEqual(sender.call_args[0][1], "-900900")

    @override_settings(SALE_TELEGRAM_BOT_TOKEN="env-token", SALE_TELEGRAM_GROUP_CHAT_ID="-5409867283")
    def test_sale_group_env_overrides_branch_and_db_settings(self):
        from unittest.mock import patch
        self._configure_group()
        branch = Branch.objects.create(name="Parkent", sale_bot_token="parkent-token", sale_group_chat_id="-900900")
        item = self._item(branch=branch)
        with patch("core.platform_services.telegram_send_message_with") as sender:
            sender.return_value = {"ok": True}
            mark_catalog_sold(item, self.user, 1, payment_type="cash")
            from .inventory_services import notify_sale_to_group
            history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
            notify_sale_to_group(item, history, "cash")
        self.assertEqual(sender.call_args[0][0], "env-token")
        self.assertEqual(sender.call_args[0][1], "-5409867283")

    def test_branch_without_group_sends_nothing(self):
        from unittest.mock import patch
        self._configure_group()
        branch = Branch.objects.create(name="Chilonzor")
        item = self._item(branch=branch)
        with patch("core.platform_services.telegram_send_photo_with") as photo_sender, \
                patch("core.platform_services.telegram_send_message_with") as text_sender:
            mark_catalog_sold(item, self.user, 1, payment_type="cash")
            from .inventory_services import notify_sale_to_group
            history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
            item.refresh_from_db()
            notify_sale_to_group(item, history, "cash")
        self.assertEqual(photo_sender.call_count, 0)
        self.assertEqual(text_sender.call_count, 0)

    def test_main_branch_sale_uses_main_settings(self):
        from unittest.mock import patch
        self._configure_group()
        Branch.objects.create(name="Parkent", sale_bot_token="parkent-token", sale_group_chat_id="-900900")
        item = self._item()
        with patch("core.platform_services.telegram_send_message_with") as sender:
            sender.return_value = {"ok": True}
            self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.assertEqual(sender.call_args[0][0], "test-token")
        self.assertEqual(sender.call_args[0][1], "-100500")

    def test_branch_token_is_not_exposed_in_api(self):
        branch = Branch.objects.create(name="Parkent", sale_bot_token="maxfiy-token", sale_group_chat_id="-900900")
        row = self.client.get(f"/api/branches/{branch.id}/").json()
        self.assertNotIn("sale_bot_token", row)
        self.assertNotIn("sale_group_chat_id", row)
        self.assertTrue(row["sale_group_configured"])

    def test_caption_shows_branch_name(self):
        from .inventory_services import sale_group_caption
        branch = Branch.objects.create(name="Parkent")
        item = self._item(branch=branch)
        mark_catalog_sold(item, self.user, 1, payment_type="card")
        history = CatalogHistory.objects.filter(catalog_item=item, action="sold").first()
        item.refresh_from_db()
        self.assertIn("Parkent", sale_group_caption(item, history, "card"))

    def test_text_message_is_sent_when_no_sale_photo(self):
        from unittest.mock import patch
        self._configure_group()
        item = self._item()
        with patch("core.platform_services.telegram_send_photo_with") as photo_sender, \
                patch("core.platform_services.telegram_send_message_with") as text_sender:
            text_sender.return_value = {"ok": True}
            response = self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(photo_sender.call_count, 0)
        self.assertEqual(text_sender.call_count, 1)
        self.assertEqual(text_sender.call_args[0][0], "test-token")
        self.assertEqual(text_sender.call_args[0][1], "-100500")
        self.assertIn("Qizil buket", text_sender.call_args[0][2])

    def test_uploaded_sale_photo_is_sent_to_group_as_one_photo_message(self):
        from unittest.mock import patch
        self._configure_group()
        item = self._item()
        uploaded = SimpleUploadedFile("sold.jpg", b"sale-photo-bytes", content_type="image/jpeg")
        with patch("core.platform_services.telegram_send_photo_with") as sender:
            sender.return_value = {"ok": True}
            response = self.client.post(f"/api/catalog/{item.id}/sell/", {
                "quantity": 1,
                "payment_type": "mixed",
                "cash_amount": "100000",
                "card_amount": "200000",
                "sale_image": uploaded,
            }, format="multipart")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(sender.call_count, 1)
        args = sender.call_args[0]
        self.assertEqual(args[0], "test-token")
        self.assertEqual(args[1], "-100500")
        self.assertEqual(args[2], b"sale-photo-bytes")
        self.assertIn("Qizil buket", args[3])
        self.assertIn("Aralash", args[3])
        self.assertIn("100 000", args[3])
        self.assertIn("200 000", args[3])

    def test_nothing_is_sent_without_token(self):
        from unittest.mock import patch
        item = self._item()
        with patch("core.platform_services.telegram_send_photo_with") as photo_sender, \
                patch("core.platform_services.telegram_send_message_with") as text_sender:
            self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.assertEqual(photo_sender.call_count, 0)
        self.assertEqual(text_sender.call_count, 0)

    def test_sale_still_works_when_telegram_fails(self):
        from unittest.mock import patch
        self._configure_group()
        item = self._item()
        with patch("core.platform_services.telegram_send_message_with", side_effect=RuntimeError("tarmoq")):
            response = self.client.post(f"/api/catalog/{item.id}/sell/", {"quantity": 1, "payment_type": "cash"}, format="json")
        self.assertEqual(response.status_code, 200)
        item.refresh_from_db()
        self.assertEqual(item.quantity_sold, 1)

    def test_sale_group_chat_id_retries_with_and_without_100_prefix(self):
        from unittest.mock import patch
        from .platform_services import telegram_send_photo_with

        class BadResponse:
            def raise_for_status(self):
                raise requests.HTTPError("bad chat")

        class GoodResponse:
            def raise_for_status(self):
                return None

            def json(self):
                return {"ok": True}

        with patch("core.platform_services.requests.post", side_effect=[BadResponse(), GoodResponse()]) as sender:
            result = telegram_send_photo_with("token", "-1005409867283", "https://example.com/x.jpg", "caption")
        self.assertEqual(result["ok"], True)
        self.assertEqual(sender.call_args_list[0].kwargs["json"]["chat_id"], "-1005409867283")
        self.assertEqual(sender.call_args_list[1].kwargs["json"]["chat_id"], "-5409867283")


class BranchExpenseTests(TestCase):
    """Filial rasxodi asosiy filialnikidan ajratiladi."""

    def setUp(self):
        self.branch = Branch.objects.create(name="Parkent")
        self.admin = User.objects.create_user("exp-bosh", password="password", is_superuser=True, is_staff=True)
        UserProfile.objects.create(user=self.admin, role="admin")
        self.branch_user = User.objects.create_user("exp-parkent", password="password")
        UserProfile.objects.create(user=self.branch_user, role="admin", branch=self.branch)
        PagePermission.objects.create(user=self.branch_user, page="expenses", can_view=True, can_control=True)
        self.client = APIClient()
        self.client.force_authenticate(self.admin)
        self.branch_client = APIClient()
        self.branch_client.force_authenticate(self.branch_user)

    def test_branch_user_expense_is_tagged_and_isolated(self):
        self.client.post("/api/expenses/", {"amount": "100000", "destination": "Bosh ofis ijarasi"}, format="json")
        created = self.branch_client.post("/api/expenses/", {"amount": "70000", "destination": "Parkent svet"}, format="json")
        self.assertEqual(created.status_code, 201, created.json())
        self.assertEqual(created.json()["branch"], self.branch.id)
        self.assertEqual(created.json()["branch_name"], "Parkent")
        rows = self.branch_client.get("/api/expenses/").json()["results"]
        self.assertEqual([row["destination"] for row in rows], ["Parkent svet"])
        self.assertEqual(self.branch_client.get("/api/expenses/summary/").json()["totals"]["expense_count"], 1)
        all_rows = self.client.get("/api/expenses/").json()["results"]
        self.assertEqual(len(all_rows), 2)

    def test_accounting_splits_branch_expense(self):
        self.client.post("/api/expenses/", {"amount": "100000", "destination": "Bosh ofis"}, format="json")
        self.branch_client.post("/api/expenses/", {"amount": "70000", "destination": "Parkent svet"}, format="json")
        data = self.client.get("/api/accounting/", {"branch": "all"}).json()
        self.assertEqual(Decimal(data["summary"]["expense_total"]), Decimal("170000"))
        buckets = {row["branch_name"]: row for row in data["by_branch"]}
        self.assertEqual(Decimal(buckets["Parkent"]["expense_total"]), Decimal("70000"))
        only_branch = self.client.get("/api/accounting/", {"branch": self.branch.id}).json()
        self.assertEqual(Decimal(only_branch["summary"]["expense_total"]), Decimal("70000"))


class StockIntakeMergeTests(TestCase):
    """Partiya kirimi: nav so'ralmaydi, bo'yi va tannarxi bir xil qatorlar qo'shiladi."""

    def setUp(self):
        self.user = User.objects.create_user("sklad", password="password", is_superuser=True, is_staff=True)
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.rose = Flower.objects.create(name_uz="Atirgul", slug="atirgul-merge")
        self.chrysanthemum = Flower.objects.create(name_uz="Xrizantema", slug="xrizantema-merge")
        self.supplier = Supplier.objects.create(name="Gollandiya")
        self.delivery = StockDelivery.objects.create(number="PT-114", received_at="2026-08-05", supplier=self.supplier)

    def _post(self, **overrides):
        payload = {
            "delivery": self.delivery.id, "flower": self.rose.id, "height_cm": 40,
            "stems_per_bunch": 25, "received_stems": 100,
            "cost_per_stem": "8000", "sale_price_per_stem": "15000",
        }
        payload.update(overrides)
        return self.client.post("/api/stock-batches/", payload, format="json")

    def test_intake_takes_flower_without_variant(self):
        response = self._post()
        self.assertEqual(response.status_code, 201, response.data)
        batch = StockBatch.objects.get(id=response.json()["id"])
        self.assertEqual(batch.variant.flower_id, self.rose.id)
        self.assertTrue(batch.variant.is_general)
        self.assertEqual(batch.variant.name_uz, "")
        self.assertEqual(batch.variant.color_uz, "")
        # navsiz qatorning nomi faqat gul nomi bo'ladi, ortiqcha ajratgich qolmaydi
        self.assertEqual(str(batch.variant), "Atirgul")
        self.assertEqual(batch.flower_name, "Atirgul")
        self.assertEqual(batch.title, "Atirgul 40 sm")

    def test_intake_without_flower_is_rejected(self):
        response = self.client.post("/api/stock-batches/", {
            "delivery": self.delivery.id, "height_cm": 40, "stems_per_bunch": 25,
            "received_stems": 100, "cost_per_stem": "8000", "sale_price_per_stem": "15000",
        }, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("flower", response.data)

    def test_same_flower_height_and_cost_are_added_together(self):
        first = self._post(received_stems=100).json()
        second = self._post(received_stems=130).json()
        third = self._post(received_stems=150).json()
        self.assertEqual(second["id"], first["id"])
        self.assertEqual(third["id"], first["id"])
        self.assertEqual(StockBatch.objects.filter(delivery=self.delivery).count(), 1)
        batch = StockBatch.objects.get(id=first["id"])
        self.assertEqual(batch.received_stems, 380)
        self.assertEqual(batch.remaining_stems, 380)
        self.assertFalse(first["merged"])
        self.assertTrue(second["merged"])
        self.assertEqual(second["merged_stems"], 130)
        self.assertEqual(third["merged_stems"], 150)

    def test_only_one_general_variant_is_created_per_flower(self):
        self._post(received_stems=100)
        self._post(received_stems=50, height_cm=30)
        self._post(received_stems=50, cost_per_stem="9000")
        self.assertEqual(FlowerVariant.objects.filter(flower=self.rose, is_general=True).count(), 1)

    def test_different_height_stays_a_separate_row(self):
        first = self._post(height_cm=40).json()
        second = self._post(height_cm=30).json()
        self.assertNotEqual(first["id"], second["id"])
        self.assertEqual(StockBatch.objects.filter(delivery=self.delivery).count(), 2)

    def test_different_cost_stays_a_separate_row(self):
        first = self._post(cost_per_stem="8000").json()
        second = self._post(cost_per_stem="9000").json()
        self.assertNotEqual(first["id"], second["id"])

    def test_different_flower_stays_a_separate_row(self):
        first = self._post(flower=self.rose.id).json()
        second = self._post(flower=self.chrysanthemum.id).json()
        self.assertNotEqual(first["id"], second["id"])

    def test_free_flowers_do_not_merge_with_paid_ones(self):
        paid = self._post(cost_per_stem="0").json()
        free = self._post(is_free=True, cost_per_stem="0").json()
        self.assertNotEqual(paid["id"], free["id"])

    def test_other_delivery_stays_a_separate_row(self):
        other = StockDelivery.objects.create(number="PT-115", received_at="2026-08-06", supplier=self.supplier)
        first = self._post().json()
        second = self._post(delivery=other.id).json()
        self.assertNotEqual(first["id"], second["id"])

    def test_last_sale_price_wins_for_the_whole_row(self):
        first = self._post(received_stems=100, sale_price_per_stem="15000").json()
        second = self._post(received_stems=100, sale_price_per_stem="18000").json()
        self.assertEqual(second["id"], first["id"])
        batch = StockBatch.objects.get(id=first["id"])
        self.assertEqual(batch.sale_price_per_stem, Decimal("18000.00"))
        # pochka narxi ham yangi dona narxidan qayta hisoblanadi
        self.assertEqual(batch.sale_price_per_bunch, Decimal("450000.00"))
        self.assertEqual(batch.cost_per_bunch, Decimal("200000.00"))

    def test_used_stems_are_kept_when_more_arrives(self):
        created = self._post(received_stems=100).json()
        batch = StockBatch.objects.get(id=created["id"])
        batch.remaining_stems = 60
        batch.save(update_fields=["remaining_stems"])
        self._post(received_stems=50)
        batch.refresh_from_db()
        self.assertEqual(batch.received_stems, 150)
        self.assertEqual(batch.remaining_stems, 110)

    def test_each_intake_writes_its_own_stock_movement(self):
        created = self._post(received_stems=100).json()
        self._post(received_stems=130)
        movements = StockMovement.objects.filter(batch_id=created["id"], movement_type="in").order_by("id")
        self.assertEqual([row.quantity_stems for row in movements], [100, 130])
        self.assertEqual(StockBatch.objects.get(id=created["id"]).received_stems, sum(row.quantity_stems for row in movements))

    def test_merge_is_written_to_audit(self):
        self._post(received_stems=100)
        self._post(received_stems=130)
        row = AuditLog.objects.filter(action="stock_batch_merged").first()
        self.assertIsNotNone(row)
        self.assertEqual(row.after["added_stems"], 130)
        self.assertEqual(row.after["received_stems"], 230)
        self.assertIn("Atirgul 40 sm", row.summary)

    def test_editing_received_stems_keeps_movement_total_in_step(self):
        created = self._post(received_stems=100).json()
        self._post(received_stems=130)
        response = self.client.patch(f"/api/stock-batches/{created['id']}/", {"received_stems": 200}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        movements = StockMovement.objects.filter(batch_id=created["id"], movement_type="in").order_by("id")
        self.assertEqual(sum(row.quantity_stems for row in movements), 200)
        self.assertEqual([row.quantity_stems for row in movements], [100, 100])

    def test_general_variants_are_hidden_from_the_variant_list(self):
        self._post()
        FlowerVariant.objects.create(flower=self.rose, name_uz="Freedom", color_uz="Qizil")
        rows = self.client.get("/api/flower-variants/").json()["results"]
        self.assertEqual([row["name_uz"] for row in rows], ["Freedom"])
        everything = self.client.get("/api/flower-variants/", {"is_general": "true"}).json()["results"]
        self.assertEqual(len(everything), 1)
        self.assertTrue(everything[0]["is_general"])

    def test_stock_list_can_be_filtered_by_flower(self):
        self._post(flower=self.rose.id)
        self._post(flower=self.chrysanthemum.id)
        rows = self.client.get("/api/stock-batches/", {"flower": self.rose.id}).json()["results"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["flower_detail"]["name_uz"], "Atirgul")
        self.assertEqual(rows[0]["title"], "Atirgul 40 sm")

    def test_change_flower_moves_a_used_row_to_another_flower(self):
        created = self._post().json()
        response = self.client.post(f"/api/stock-batches/{created['id']}/change-flower/", {
            "flower": self.chrysanthemum.id, "reason": "Kirimda xato yozilgan",
        }, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(StockBatch.objects.get(id=created["id"]).variant.flower_id, self.chrysanthemum.id)
        self.assertEqual(response.json()["variant_change"]["new_variant"], "Xrizantema")

    def test_change_flower_needs_a_flower(self):
        created = self._post().json()
        response = self.client.post(f"/api/stock-batches/{created['id']}/change-flower/", {"reason": "Sabab"}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("flower", response.data)

    def test_ai_sees_the_flower_name_without_a_variant(self):
        created = self._post().json()
        batch = StockBatch.objects.get(id=created["id"])
        row = stock_batch_ai_row(batch)
        self.assertEqual(row["display_name_uz"], "Atirgul")
        self.assertEqual(row["flower_uz"], "Atirgul")
        self.assertEqual(row["variant_uz"], "")
        stock = ai_stock_rows()
        self.assertEqual([item["display_name_uz"] for item in stock], ["Atirgul"])
        variants = ai_flower_variant_rows()
        self.assertEqual([item["display_name_uz"] for item in variants], ["Atirgul"])

    def test_catalog_composition_summary_has_no_double_spaces(self):
        created = self._post().json()
        batch = StockBatch.objects.get(id=created["id"])
        item = CatalogItem.objects.create(name_uz="Buket", arrangement_type="bouquet", price=Decimal("300000"), quantity_total=1, status="available")
        CatalogComposition.objects.create(catalog_item=item, stock_batch=batch, quantity_stems=20)
        self.assertEqual([row["name_uz"] for row in catalog_composition_summary(item)], ["Atirgul"])

    def test_ai_sees_every_height_and_price_of_one_flower(self):
        self._post(height_cm=40, cost_per_stem="8000", sale_price_per_stem="15000")
        self._post(height_cm=60, cost_per_stem="12000", sale_price_per_stem="25000")
        self._post(height_cm=40, cost_per_stem="9000", sale_price_per_stem="17000")
        rows = ai_stock_rows()
        self.assertEqual(
            [(row["height_label"], row["price_per_stem"]) for row in rows],
            [("40 sm", "15000.00"), ("60 sm", "25000.00"), ("40 sm", "17000.00")],
        )
        variants = ai_flower_variant_rows()
        self.assertEqual(len(variants[0]["active_stock"]), 3)

    def test_ai_does_not_repeat_the_same_offer_twice(self):
        other = StockDelivery.objects.create(number="PT-116", received_at="2026-08-07", supplier=self.supplier)
        self._post(height_cm=40, sale_price_per_stem="15000")
        self._post(delivery=other.id, height_cm=40, sale_price_per_stem="15000")
        self.assertEqual(StockBatch.objects.count(), 2)
        rows = ai_stock_rows()
        self.assertEqual([(row["height_label"], row["price_per_stem"]) for row in rows], [("40 sm", "15000.00")])


class FloristExtraDecorationTests(TestCase):
    """Florist detalidan qo'lda oformleniya haqi yozish."""

    def setUp(self):
        self.user = User.objects.create_user("boss", password="password", is_superuser=True, is_staff=True)
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        florist_user = User.objects.create_user("isroil", password="password", first_name="Isroil")
        self.florist = FloristProfile.objects.create(user=florist_user, staff_type="florist", decoration_fee=Decimal("5000"))

    def _add(self, **payload):
        return self.client.post(f"/api/florists/{self.florist.id}/decoration/", payload, format="json")

    def test_count_is_multiplied_by_the_profile_fee(self):
        response = self._add(count=3)
        self.assertEqual(response.status_code, 201, response.data)
        entry = FloristSalaryEntry.objects.get(id=response.json()["id"])
        self.assertEqual(entry.source, "extra_decoration")
        self.assertEqual(entry.quantity, 3)
        self.assertEqual(entry.unit_amount, Decimal("5000.00"))
        self.assertEqual(entry.amount, Decimal("15000.00"))
        self.assertEqual(entry.work_date, timezone.localdate())
        self.assertIsNone(entry.catalog_item_id)

    def test_unit_amount_can_be_given_instead_of_the_profile_fee(self):
        response = self._add(count=2, unit_amount="7000")
        entry = FloristSalaryEntry.objects.get(id=response.json()["id"])
        self.assertEqual(entry.unit_amount, Decimal("7000.00"))
        self.assertEqual(entry.amount, Decimal("14000.00"))
        # profildagi narx o'zgarmaydi
        self.florist.refresh_from_db()
        self.assertEqual(self.florist.decoration_fee, Decimal("5000"))

    def test_same_day_and_same_price_add_up_in_one_row(self):
        first = self._add(count=3)
        second = self._add(count=2)
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.json()["id"], first.json()["id"])
        entry = FloristSalaryEntry.objects.get(id=first.json()["id"])
        self.assertEqual(entry.quantity, 5)
        self.assertEqual(entry.amount, Decimal("25000.00"))
        self.assertEqual(FloristSalaryEntry.objects.filter(source="extra_decoration").count(), 1)

    def test_different_unit_price_opens_a_separate_row(self):
        first = self._add(count=3)
        second = self._add(count=2, unit_amount="7000")
        self.assertNotEqual(first.json()["id"], second.json()["id"])
        self.assertEqual(FloristSalaryEntry.objects.filter(source="extra_decoration").count(), 2)

    def test_another_day_opens_a_separate_row(self):
        self._add(count=3)
        response = self._add(count=2, work_date="2026-08-01")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(FloristSalaryEntry.objects.filter(source="extra_decoration").count(), 2)

    def test_zero_count_is_rejected(self):
        response = self._add(count=0)
        self.assertEqual(response.status_code, 400)
        self.assertIn("count", response.data)

    def test_florist_without_a_fee_must_be_given_a_price(self):
        self.florist.decoration_fee = Decimal("0")
        self.florist.save(update_fields=["decoration_fee"])
        response = self._add(count=3)
        self.assertEqual(response.status_code, 400)
        self.assertIn("narx", str(response.data["detail"]).lower())
        self.assertEqual(self._add(count=3, unit_amount="6000").status_code, 201)

    def test_profile_decoration_fee_can_be_changed(self):
        response = self.client.patch(f"/api/florists/{self.florist.id}/", {"decoration_fee": "8000"}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.florist.refresh_from_db()
        self.assertEqual(self.florist.decoration_fee, Decimal("8000.00"))
        entry = FloristSalaryEntry.objects.get(id=self._add(count=2).json()["id"])
        self.assertEqual(entry.amount, Decimal("16000.00"))

    def test_count_can_be_edited_and_the_amount_follows(self):
        created = self._add(count=3).json()
        response = self.client.patch(f"/api/florist-salary/{created['id']}/", {"quantity": 5}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        entry = FloristSalaryEntry.objects.get(id=created["id"])
        self.assertEqual(entry.quantity, 5)
        self.assertEqual(entry.amount, Decimal("25000.00"))

    def test_unit_price_can_be_edited_and_the_amount_follows(self):
        created = self._add(count=3).json()
        self.client.patch(f"/api/florist-salary/{created['id']}/", {"unit_amount": "6000"}, format="json")
        entry = FloristSalaryEntry.objects.get(id=created["id"])
        self.assertEqual(entry.amount, Decimal("18000.00"))

    def test_amount_written_by_hand_wins_over_the_multiplication(self):
        created = self._add(count=3).json()
        self.client.patch(f"/api/florist-salary/{created['id']}/", {"amount": "20000"}, format="json")
        entry = FloristSalaryEntry.objects.get(id=created["id"])
        self.assertEqual(entry.amount, Decimal("20000.00"))
        self.assertEqual(entry.quantity, 3)

    def test_extra_decoration_lands_in_the_decoration_column(self):
        self._add(count=3)
        stats = self.client.get(f"/api/florists/{self.florist.id}/stats/").json()
        self.assertEqual(Decimal(stats["summary"]["decoration_salary_total"]), Decimal("15000"))
        self.assertEqual(Decimal(stats["summary"]["manual_salary_total"]), Decimal("0"))
        self.assertEqual(Decimal(stats["summary"]["salary_total"]), Decimal("15000"))
        sources = {row["source"]: row for row in stats["by_source"]}
        self.assertEqual(sources["extra_decoration"]["source_label"], "Qo‘shimcha oformleniya")

    def test_florist_sees_it_and_gets_a_notification(self):
        created = self._add(count=3).json()
        self.assertTrue(Notification.objects.filter(target_user=self.florist.user, reference_type="florist_salary", reference_id=created["id"]).exists())
        self.assertTrue(AuditLog.objects.filter(action="florist_decoration_added").exists())

    def test_florist_cannot_write_it_for_themselves(self):
        UserProfile.objects.update_or_create(user=self.florist.user, defaults={"role": "florist"})
        self.client.force_authenticate(self.florist.user)
        self.assertEqual(self._add(count=3).status_code, 403)
        self.assertFalse(FloristSalaryEntry.objects.filter(source="extra_decoration").exists())


class PaginationTotalsTests(TestCase):
    """Katalog, sklad va floristlar ro'yxatlarining sahifalanishi va umumiy sonlari.

    Asosiy talab: sahifada nechta yozuv ko'rinishidan qat'i nazar `totals`
    butun filtr bo'yicha bo'lishi kerak.
    """

    def setUp(self):
        self.user = User.objects.create_user("pager", password="password", is_superuser=True, is_staff=True)
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        flower = Flower.objects.create(name_uz="Atirgul P", slug="rose-pager")
        self.variant = FlowerVariant.objects.create(flower=flower, name_uz="Freedom P", color_uz="Qizil")
        self.supplier = Supplier.objects.create(name="Hoji P", phone="+998901110000")
        self.batches = [
            StockBatch.objects.create(
                variant=self.variant, supplier=self.supplier, batch_number=f"P-{index}",
                height_cm=50 + index, stems_per_bunch=25, received_stems=100, remaining_stems=60,
                cost_per_stem=Decimal("5000"), sale_price_per_stem=Decimal("9000"), sale_price_per_bunch=Decimal("225000"),
            )
            for index in range(3)
        ]
        florist_user = User.objects.create_user("pager-florist", password="password", first_name="Bekzod")
        self.florist = FloristProfile.objects.create(user=florist_user, staff_type="florist", phone="+998901110001")
        for index in range(35):
            CatalogItem.objects.create(
                name_uz=f"Buket {index}", arrangement_type="bouquet", price=Decimal("100000"),
                quantity_total=2, quantity_sold=1, status="available", florist=self.florist,
                calculated_cost_price=Decimal("40000"),
            )

    def test_catalog_list_returns_page_meta_and_whole_list_totals(self):
        response = self.client.get("/api/catalog/")
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["count"], 35)
        self.assertEqual(body["page"], 1)
        self.assertEqual(body["page_size"], 30)
        self.assertEqual(body["total_pages"], 2)
        self.assertTrue(body["has_next"])
        self.assertFalse(body["has_previous"])
        self.assertEqual(len(body["results"]), 30)
        totals = body["totals"]
        # sahifada 30 ta, jami sonlar esa 35 tasi bo'yicha
        self.assertEqual(totals["items"], 35)
        self.assertEqual(totals["quantity_total"], 70)
        self.assertEqual(totals["quantity_sold"], 35)
        self.assertEqual(totals["quantity_remaining"], 35)
        self.assertEqual(Decimal(totals["remaining_value"]), Decimal("3500000"))
        self.assertEqual(Decimal(totals["sold_value"]), Decimal("3500000"))
        self.assertEqual(Decimal(totals["cost_total"]), Decimal("1400000"))
        self.assertEqual(totals["by_status"], {"available": 35})
        self.assertEqual(totals["status_counts"]["all"], 35)
        self.assertEqual(totals["status_counts"]["available"], 35)
        self.assertEqual(totals["status_counts"]["sold"], 0)
        self.assertEqual(totals["available_count"], 35)
        self.assertEqual(totals["sold_count"], 0)
        self.assertEqual(totals["archived_count"], 0)

    def test_second_page_totals_stay_the_same(self):
        first = self.client.get("/api/catalog/").json()
        second = self.client.get("/api/catalog/?page=2").json()
        self.assertEqual(second["page"], 2)
        self.assertEqual(len(second["results"]), 5)
        self.assertFalse(second["has_next"])
        self.assertTrue(second["has_previous"])
        self.assertEqual(first["totals"], second["totals"])

    def test_totals_follow_the_filter(self):
        CatalogItem.objects.filter(name_uz="Buket 0").update(status="sold")
        body = self.client.get("/api/catalog/?status=sold").json()
        self.assertEqual(body["count"], 1)
        self.assertEqual(body["totals"]["items"], 1)
        self.assertEqual(body["totals"]["quantity_total"], 2)
        self.assertEqual(body["totals"]["by_status"], {"sold": 1})
        self.assertEqual(body["totals"]["status_counts"]["all"], 35)
        self.assertEqual(body["totals"]["status_counts"]["available"], 34)
        self.assertEqual(body["totals"]["status_counts"]["sold"], 1)
        self.assertEqual(body["totals"]["available_count"], 34)
        self.assertEqual(body["totals"]["sold_count"], 1)

    def test_page_size_all_returns_everything_in_one_page(self):
        body = self.client.get("/api/catalog/?page_size=all").json()
        self.assertEqual(body["count"], 35)
        self.assertEqual(body["total_pages"], 1)
        self.assertFalse(body["has_next"])
        self.assertEqual(len(body["results"]), 35)
        self.assertEqual(body["totals"]["items"], 35)

    def test_page_size_all_works_on_an_empty_list(self):
        CatalogItem.objects.all().delete()
        body = self.client.get("/api/catalog/?page_size=all").json()
        self.assertEqual(body["count"], 0)
        self.assertEqual(body["results"], [])
        self.assertEqual(body["totals"]["items"], 0)
        self.assertEqual(Decimal(body["totals"]["remaining_value"]), Decimal("0"))

    def test_stock_batch_totals_count_stems_and_money(self):
        totals = self.client.get("/api/stock-batches/").json()["totals"]
        self.assertEqual(totals["batches"], 3)
        self.assertEqual(totals["received_stems"], 300)
        self.assertEqual(totals["remaining_stems"], 180)
        self.assertEqual(totals["used_stems"], 120)
        self.assertEqual(totals["flowers"], 1)
        self.assertEqual(totals["suppliers"], 1)
        self.assertEqual(Decimal(totals["remaining_cost"]), Decimal("900000"))
        self.assertEqual(Decimal(totals["remaining_sale_value"]), Decimal("1620000"))
        self.assertEqual(Decimal(totals["received_cost"]), Decimal("1500000"))

    def test_florist_totals_gather_salary_catalog_and_stock(self):
        FloristSalaryEntry.objects.create(florist=self.florist, amount=Decimal("60000"), source="catalog", work_date="2026-08-01")
        FloristSalaryEntry.objects.create(florist=self.florist, amount=Decimal("40000"), source="manual", work_date="2026-08-02")
        FloristStockBalance.objects.create(florist=self.florist, batch=self.batches[0], remaining_stems=25)
        totals = self.client.get("/api/florists/").json()["totals"]
        self.assertEqual(totals["florists"], 1)
        self.assertEqual(totals["active"], 1)
        self.assertEqual(totals["by_staff_type"], {"florist": 1})
        self.assertEqual(Decimal(totals["salary_total"]), Decimal("100000"))
        self.assertEqual(totals["catalog_quantity"], 70)
        self.assertEqual(totals["catalog_remaining"], 35)
        self.assertEqual(totals["stock_stems"], 25)

    def test_florist_stock_issue_totals_separate_issue_return_and_waste(self):
        FloristStockIssue.objects.create(florist=self.florist, batch=self.batches[0], kind="issue", quantity_stems=50)
        FloristStockIssue.objects.create(florist=self.florist, batch=self.batches[1], kind="issue", quantity_stems=30)
        FloristStockIssue.objects.create(florist=self.florist, batch=self.batches[0], kind="return", quantity_stems=10)
        FloristStockIssue.objects.create(florist=self.florist, batch=self.batches[0], kind="waste", quantity_stems=5)
        totals = self.client.get("/api/florist-stock-issues/").json()["totals"]
        self.assertEqual(totals["rows"], 4)
        self.assertEqual(totals["issued_stems"], 80)
        self.assertEqual(totals["returned_stems"], 10)
        self.assertEqual(totals["wasted_stems"], 5)
        self.assertEqual(totals["net_stems"], 65)
        self.assertEqual(totals["florists"], 1)
        self.assertEqual(totals["batches"], 2)

    def test_florist_stock_balance_totals_include_cost(self):
        FloristStockBalance.objects.create(florist=self.florist, batch=self.batches[0], remaining_stems=20)
        FloristStockBalance.objects.create(florist=self.florist, batch=self.batches[1], remaining_stems=30)
        totals = self.client.get("/api/florist-stock-balances/").json()["totals"]
        self.assertEqual(totals["rows"], 2)
        self.assertEqual(totals["remaining_stems"], 50)
        self.assertEqual(Decimal(totals["cost_total"]), Decimal("250000"))

    def test_salary_totals_split_by_source(self):
        FloristSalaryEntry.objects.create(florist=self.florist, amount=Decimal("60000"), quantity=1, source="catalog", work_date="2026-08-01")
        FloristSalaryEntry.objects.create(florist=self.florist, amount=Decimal("40000"), quantity=2, source="manual", work_date="2026-08-02")
        totals = self.client.get("/api/florist-salary/").json()["totals"]
        self.assertEqual(totals["entries"], 2)
        self.assertEqual(Decimal(totals["amount_total"]), Decimal("100000"))
        self.assertEqual(totals["quantity_total"], 3)
        self.assertEqual(totals["by_source"]["catalog"], {"count": 1, "amount": "60000.00"})
        self.assertEqual(totals["by_source"]["manual"], {"count": 1, "amount": "40000.00"})

    def test_materials_and_movement_journals_carry_totals(self):
        # bazada standart materiallar seed qilingan bo'lishi mumkin, shuning
        # uchun jami sonlar faqat shu testning materiali bo'yicha tekshiriladi
        Packaging.objects.all().delete()
        PackagingMovement.objects.all().delete()
        StockMovement.objects.all().delete()
        packaging = Packaging.objects.create(packaging_type="box", name_uz="Quti P", quantity=10, cost_price=Decimal("15000"), sale_price=Decimal("30000"))
        PackagingMovement.objects.create(packaging=packaging, movement_type="in", quantity=10, unit_cost=Decimal("15000"))
        StockMovement.objects.create(batch=self.batches[0], movement_type="in", quantity_stems=100)
        # chiqim bazada manfiy yoziladi, javobda esa musbat ko'rinishi kerak
        StockMovement.objects.create(batch=self.batches[0], movement_type="out", quantity_stems=-40)
        materials = self.client.get("/api/materials/").json()["totals"]
        self.assertEqual(materials["items"], 1)
        self.assertEqual(materials["quantity_total"], 10)
        self.assertEqual(Decimal(materials["cost_value"]), Decimal("150000"))
        self.assertEqual(Decimal(materials["sale_value"]), Decimal("300000"))
        stock_journal = self.client.get("/api/stock-movements/").json()["totals"]
        self.assertEqual(stock_journal["rows"], 2)
        self.assertEqual(stock_journal["in_stems"], 100)
        self.assertEqual(stock_journal["out_stems"], 40)
        self.assertEqual(stock_journal["net_stems"], 60)
        material_journal = self.client.get("/api/material-movements/").json()["totals"]
        self.assertEqual(material_journal["in_quantity"], 10)
        self.assertEqual(Decimal(material_journal["cost_total"]), Decimal("150000"))


class OperatorHandoffTests(TestCase):
    """Sklad AI dan olib tashlangach so'rovlar operatorga qanday topshirilishi."""

    def setUp(self):
        self.customer = Customer.objects.create(instagram_user_id="telegram:7001", name="Ahmad", phone="+998901112233")
        self.conversation = Conversation.objects.create(customer=self.customer)

    def create_lead(self, **overrides):
        arguments = {
            "customer_name": None,
            "phone": None,
            "request_text": "Mijoz Jumila pushti atirguldan katta buket yasatmoqchi",
            "arrangement_type": "bouquet",
            "estimated_price": None,
            "florist_fee": None,
            "fulfillment": None,
            "delivery_address": None,
            "desired_date": None,
            "desired_time": None,
            "catalog_items": [],
            "note": "Tug'ilgan kunga sovg'a",
            "topic": "custom_order",
            "flowers_text": "jumila pushti",
            "size_text": "51 dona",
            "photo_urls": [],
        }
        arguments.update(overrides)
        return execute_ai_tool("client_lead_create", arguments, self.conversation)

    def test_custom_order_lead_keeps_flowers_size_and_note(self):
        result = self.create_lead()
        self.assertTrue(result["ok"])
        details = Lead.objects.get(id=result["lead_id"]).details
        self.assertEqual(details["topic"], "custom_order")
        self.assertEqual(details["flowers_text"], "jumila pushti")
        self.assertEqual(details["size_text"], "51 dona")
        self.assertEqual(details["note"], "Tug'ilgan kunga sovg'a")

    def test_custom_order_lead_carries_no_price(self):
        """Yasatma buyurtmada narxni operator qo'yadi, AI emas."""
        lead = Lead.objects.get(id=self.create_lead()["lead_id"])
        self.assertIsNone(lead.estimated_price)
        self.assertEqual(lead.florist_fee, Decimal("0"))

    def test_photo_request_lead_stores_the_link_without_downloading_it(self):
        url = "https://api.telegram.org/file/bot123/photos/file_9.jpg"
        result = self.create_lead(
            topic="photo_request",
            request_text="Mijoz rasm yubordi va shu buketdan bormi deb so'radi",
            flowers_text=None,
            size_text=None,
            photo_urls=[url, url, "not-a-link"],
        )
        details = Lead.objects.get(id=result["lead_id"]).details
        self.assertEqual(details["photo_urls"], [url])
        self.assertEqual(details["flowers_text"], "")

    def test_photo_urls_are_capped(self):
        urls = [f"https://cdn.example.com/{index}.jpg" for index in range(9)]
        result = self.create_lead(topic="photo_request", photo_urls=urls)
        self.assertEqual(Lead.objects.get(id=result["lead_id"]).details["photo_urls"], urls[:5])

    def test_lead_creation_writes_the_conversation_summary(self):
        self.create_lead()
        self.conversation.refresh_from_db()
        summary = self.conversation.ai_summary
        self.assertIn("Yasatma buyurtma", summary)
        self.assertIn("buket", summary)
        self.assertIn("jumila pushti", summary)
        self.assertIn("51 dona", summary)

    def test_question_lead_summary_names_the_topic(self):
        self.create_lead(topic="question", request_text="Mijoz to'y bezagi haqida so'radi", flowers_text=None, size_text=None, arrangement_type=None)
        self.conversation.refresh_from_db()
        self.assertTrue(self.conversation.ai_summary.startswith("Savol"))
        self.assertIn("to'y bezagi", self.conversation.ai_summary)

    def test_lead_edit_keeps_earlier_details(self):
        lead_id = self.create_lead()["lead_id"]
        execute_ai_tool("client_lead_edit", {
            "lead_id": lead_id,
            "customer_name": None,
            "phone": None,
            "request_text": None,
            "status": None,
            "arrangement_type": None,
            "estimated_price": None,
            "florist_fee": None,
            "fulfillment": "pickup",
            "delivery_address": None,
            "desired_date": None,
            "desired_time": None,
            "catalog_items": None,
            "note": None,
            "topic": None,
            "flowers_text": None,
            "size_text": "80 dona",
            "photo_urls": [],
        }, self.conversation)
        lead = Lead.objects.get(id=lead_id)
        self.assertEqual(lead.fulfillment, "pickup")
        self.assertEqual(lead.details["size_text"], "80 dona")
        self.assertEqual(lead.details["note"], "Tug'ilgan kunga sovg'a")

    def test_lead_still_needs_a_name_and_a_phone(self):
        stranger = Customer.objects.create(instagram_user_id="telegram:7002")
        conversation = Conversation.objects.create(customer=stranger)
        result = execute_ai_tool("client_lead_create", {
            "customer_name": None, "phone": None, "request_text": "Savol bor", "arrangement_type": None,
            "estimated_price": None, "fulfillment": None, "delivery_address": None,
            "desired_date": None, "desired_time": None, "catalog_items": [], "note": None,
            "topic": "question", "flowers_text": None, "size_text": None, "photo_urls": [],
        }, conversation)
        self.assertEqual(result["detail"], "customer_name_required")

    def test_customer_photo_links_reach_the_ai_context(self):
        self.conversation.messages.create(sender="customer", text="Salom")
        self.conversation.messages.create(
            sender="customer",
            text="Shundan bormi\nMijoz yuborgan rasm: https://cdn.example.com/a.jpg",
            metadata={"attachments": [{"kind": "photo", "url": "https://cdn.example.com/a.jpg"}]},
        )
        self.conversation.messages.create(sender="ai", text="Operatorlarimiz aniq javob berishadi", metadata={"attachments": [{"kind": "photo", "url": "https://cdn.example.com/ours.jpg"}]})
        rows = customer_attachment_rows(list(self.conversation.messages.order_by("created_at", "id")))
        self.assertEqual(rows, [{"kind": "photo", "url": "https://cdn.example.com/a.jpg"}])

    def test_media_match_requires_customer_media(self):
        customer = Customer.objects.create(instagram_user_id="ig-media-empty")
        conversation = Conversation.objects.create(customer=customer)
        result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertFalse(result["ok"])
        self.assertFalse(result["allow_send"])
        self.assertEqual(result["detail"], "no_customer_media")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_media_match_uses_ai_catalog_instagram_link_without_vision(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Pion Story", arrangement_type="bouquet", price=800000, quantity=1, image_url="https://cdn.example.com/pion.jpg", instagram_link="https://www.instagram.com/reel/ABC123/")
        customer = Customer.objects.create(instagram_user_id="ig-link-match")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(
            sender="customer",
            text="shu nechpul\nReel link: https://www.instagram.com/reel/ABC123/?igsh=test",
            metadata={"attachments": [{"kind": "reel", "url": "https://www.instagram.com/reel/ABC123/?igsh=test"}]},
        )
        with patch("core.vision_services.OpenAI") as openai_class:
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["ok"])
        self.assertTrue(result["allow_send"])
        self.assertEqual(result["matches"][0]["catalog_id"], item.id)
        openai_class.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key")
    def test_media_match_does_not_send_plain_reel_permalink_to_vision(self):
        from unittest.mock import patch
        AICatalogItem.objects.create(name="Pion Story", arrangement_type="bouquet", price=800000, quantity=1, image_url="https://cdn.example.com/pion.jpg")
        customer = Customer.objects.create(instagram_user_id="ig-reel-no-image")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(
            sender="customer",
            text="shu nechpul\nReel link: https://www.instagram.com/reel/XYZ999/",
            metadata={"attachments": [{"kind": "reel", "url": "https://www.instagram.com/reel/XYZ999/"}]},
        )
        with patch("core.vision_services.OpenAI") as openai_class:
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertFalse(result["ok"])
        self.assertFalse(result["allow_send"])
        self.assertEqual(result["detail"], "media_url_not_image")
        openai_class.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key", OPENAI_VISION_MODEL="vision-test")
    def test_media_match_analyses_one_image_then_verifies_each_candidate(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(
            name="Katalina Gulidan Savat Kompazitsia",
            arrangement_type="basket",
            price=800000,
            quantity=1,
            image_url="https://cdn.example.com/katalina.jpg",
            note="pionavidniy katalina",
            **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], color_pattern="solid", container="basket"),
        )
        conversation = media_conversation("ig-two-stage")
        with patch("core.vision_services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.side_effect = [
                SimpleNamespace(output_text=json.dumps(vision_fingerprint(flower_form="peony_rose", dominant_colors=["yellow"], color_pattern="solid", container="basket", region_requested=True, region_description="tepadan ikkinchisi")), status="completed"),
                SimpleNamespace(output_text=json.dumps(verdict_payload()), status="completed"),
            ]
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "tepadan 2chisi nechpul"}, conversation)
        self.assertTrue(result["ok"])
        self.assertTrue(result["allow_send"])
        self.assertFalse(result["allow_group"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [item.id])
        self.assertEqual(result["matches"][0]["price_text"], "800 000 so'm")
        self.assertTrue(result["region_requested"])
        self.assertEqual(client.responses.create.call_count, 2)
        source_call = client.responses.create.call_args_list[0].kwargs
        self.assertEqual(source_call["model"], "vision-test")
        source_content = source_call["input"][0]["content"]
        # Birinchi bosqichda faqat bitta rasm — mijozniki. Katalog rasmlari bormaydi.
        self.assertEqual([part["type"] for part in source_content], ["input_text", "input_image"])
        self.assertEqual(source_content[1]["image_url"], "https://cdn.example.com/customer.jpg")
        self.assertEqual(source_content[1]["detail"], "high")
        self.assertIn("tepadan 2chisi", source_content[0]["text"])
        # Tekshiruv bosqichida har nomzod alohida: aynan ikkita rasm.
        verify_content = client.responses.create.call_args_list[1].kwargs["input"][0]["content"]
        self.assertEqual([part["image_url"] for part in verify_content if part["type"] == "input_image"], ["https://cdn.example.com/customer.jpg", "https://cdn.example.com/katalina.jpg"])
        stored = conversation.messages.filter(sender="system").order_by("-id").first().metadata
        self.assertIn("ai_catalog_media_match", stored)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_media_match_refuses_when_the_flowers_only_look_similar(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(
            name="Katalina Gulidan Savat Kompazitsia",
            arrangement_type="basket",
            price=800000,
            quantity=1,
            image_url="https://cdn.example.com/katalina.jpg",
            **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket"),
        )
        conversation = media_conversation("ig-similar-only")
        with patch_vision(vision_fingerprint(flower_form="peony_rose", dominant_colors=["cream", "pink"], container="basket"), {item.id: verdict_payload(verdict="similar_only", color_match=False, differences="katalogda sariq")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        # Aynan o'shasi emas: rasm yuborilmaydi va "topdim" deyilmaydi. Mijoz quruq
        # qaytmasin — butun katalog ko'rsatiladi va mijoz Telegram akkauntga yo'naltiriladi.
        self.assertFalse(result["allow_send"])
        self.assertEqual(result["matches"], [])
        self.assertEqual(result["detail"], "similar_only")
        self.assertTrue(result["show_whole_catalog"])
        self.assertEqual(result["group_matches"], [])
        self.assertEqual([row["catalog_id"] for row in result["near_matches"]], [item.id])
        self.assertIn("business.operator_telegram", result["instruction_uz"])
        self.assertIn("Telefon raqami SO'RAMA", result["instruction_uz"])
        self.assertIn("topdim", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_media_match_refuses_same_product_when_a_check_failed(self):
        """Model 'same_product' desa ham rang mos kelmasa backend rad etadi."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(
            name="London Gulidan Kompazitsia Savat",
            arrangement_type="basket",
            price=1000000,
            quantity=1,
            image_url="https://cdn.example.com/london.jpg",
            **catalog_fingerprint_fields("https://cdn.example.com/london.jpg", flower_form="peony_rose", dominant_colors=["cream", "pink"], container="basket"),
        )
        conversation = media_conversation("ig-model-overconfident")
        with patch_vision(vision_fingerprint(flower_form="peony_rose", dominant_colors=["cream", "pink"], container="basket"), {item.id: verdict_payload(color_match=False, differences="rangi boshqa")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertFalse(result["allow_send"])
        # Model o'z javobida ziddiyatga tushdi: mahsulotni "same_product" dedi, ammo
        # rangi mos emas dedi. Rasm yuborilmaydi, lekin "bunday gul yo'q" ham deyilmaydi —
        # ball baland, bu o'sha mahsulot bo'lishi mumkin.
        self.assertEqual(result["detail"], "close_matches")
        self.assertEqual([row["catalog_id"] for row in result["group_matches"]], [item.id])
        self.assertIn("eng mos", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_look_alike_catalog_items_are_offered_as_an_album(self):
        """Bir xil guldan turli o'lchamdagi mahsulotlarni rasmdan ajratib bo'lmaydi."""
        small = AICatalogItem.objects.create(name="Buket Jumila", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/small.jpg", **catalog_fingerprint_fields("https://cdn.example.com/small.jpg", flower_form="spray_rose", dominant_colors=["cream", "pink"], container="unwrapped_bouquet"))
        big = AICatalogItem.objects.create(name="Buket Jumila 100 Tali", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/big.jpg", **catalog_fingerprint_fields("https://cdn.example.com/big.jpg", flower_form="spray_rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet"))
        conversation = media_conversation("ig-look-alike")
        source = vision_fingerprint(flower_form="spray_rose", dominant_colors=["cream", "pink"], container="unwrapped_bouquet")
        # Rasmda ajratib bo'lmagani uchun model ikkalasini ham aynan shu mahsulot deydi.
        with patch_vision(source, {small.id: verdict_payload(), big.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["ok"])
        self.assertFalse(result["allow_send"])
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["detail"], "several_look_the_same")
        self.assertEqual(result["matches"], [])
        self.assertEqual(sorted(row["catalog_id"] for row in result["group_matches"]), sorted([small.id, big.id]))
        self.assertIn("send_catalog_album", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_media_tool_is_not_forced_when_the_model_already_called_it(self):
        """Ikki marta chaqirish har javobga o'ttiz soniya va bir so'rov qo'shardi."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket"))
        conversation = media_conversation("ig-no-double-call")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["yellow"], container="basket")
        calls = []

        class Response:
            def __init__(self, output, text=""):
                self.output = output
                self.id = "resp-%s" % len(calls)
                self.output_text = text

        class Call:
            type = "function_call"
            name = "match_ai_catalog_by_media"
            call_id = "call-1"
            arguments = '{"source_url": null, "user_text": "shu nechpul"}'

        def fake_create(**kwargs):
            calls.append(kwargs)
            if len(calls) == 1:
                return Response([Call()])
            return Response([], json.dumps({"reply": "Savat Katalina, 800 000 so'm", "handoff": False, "lead_ready": False, "phone": None, "customer_name": None, "detected_language": "uz", "estimated_price": None, "arrangement_type": None, "lead_request": None, "catalog_items": [], "stock_items": []}))

        with patch_vision(source, {item.id: verdict_payload()}):
            with patch("core.services.OpenAI") as client_cls:
                client_cls.return_value.responses.create.side_effect = fake_create
                with patch("core.services.send_image_to_customer", return_value=(True, "mocked", {})):
                    result = ai_reply(conversation)
        matched = [row for row in result["tool_results"] if row["name"] == "match_ai_catalog_by_media"]
        self.assertEqual(len(matched), 1)
        self.assertNotIn("forced_by_backend", matched[0]["arguments"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_story_answers_from_what_the_shop_wrote_on_it(self):
        """Storyda nomi va narxi turibdi — rasmni tahlil qilish faqat xato qo'shadi."""
        AICatalogItem.objects.create(name="Alfalob 100 Tali", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/alfalob.jpg", **catalog_fingerprint_fields("https://cdn.example.com/alfalob.jpg", flower_form="peony_rose", dominant_colors=["hot_pink"], container="unwrapped_bouquet"))
        story_url = "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=18118106525486311&signature=abc"
        SocialPost.objects.create(
            post_type="story", media_id="story-share-3969235178066479407",
            webhook_story_id="18118106525486311",
            permalink="https://www.instagram.com/stories/euroflowers.premium/3969235178066479407",
            title_uz="Alfalob 200 tali", description_uz="200 tali alfalob atirguldan yasalgan buket",
            price=1600000, flower_count=200, image_url="https://cdn.example.com/story.jpg", is_active=True,
        )
        conversation = media_conversation("ig-own-story", image_url=story_url, kind="story")
        result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "Nechpul"}, conversation)
        self.assertEqual(result["detail"], "own_story_matched")
        self.assertTrue(result["own_post"])
        self.assertFalse(result["allow_send"])
        self.assertEqual(result["story"]["title"], "Alfalob 200 tali")
        self.assertEqual(result["story"]["price_text"], "1 600 000 so'm")
        self.assertIn("O'xshagan", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_story_photo_can_be_sent_back_but_only_that_story(self):
        from unittest.mock import patch
        post = SocialPost.objects.create(post_type="story", media_id="story-1", webhook_story_id="111", title_uz="Alfalob 200 tali", price=1600000, image_url="https://cdn.example.com/story.jpg", is_active=True)
        other = SocialPost.objects.create(post_type="story", media_id="story-2", webhook_story_id="222", title_uz="Boshqa story", price=500000, image_url="https://cdn.example.com/other.jpg", is_active=True)
        conversation = media_conversation("ig-story-image")
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": True, "detail": "own_story_matched", "allow_send": False, "story": {"social_post_id": post.id}}}]
        with patch("core.services.send_image_to_customer", return_value=(True, "mocked", {"mocked": True})) as send_mock:
            allowed = execute_ai_tool("send_post_image", {"social_post_id": post.id}, conversation, tool_results=tool_results)
        # Keyingi navbatda tool natijasi yo'q, lekin story hali suhbatning mavzusi.
        conversation.messages.create(sender="system", text="", metadata={"ai_catalog_media_match": {"detail": "own_story_matched", "story": {"social_post_id": post.id}}})
        with patch("core.services.send_image_to_customer", return_value=(True, "mocked", {"mocked": True})):
            later = execute_ai_tool("send_post_image", {"social_post_id": post.id}, conversation, tool_results=[])
        self.assertTrue(later["ok"])
        self.assertTrue(allowed["ok"])
        send_mock.assert_called_once()
        with patch("core.services.send_image_to_customer") as send_mock:
            refused = execute_ai_tool("send_post_image", {"social_post_id": other.id}, conversation, tool_results=tool_results)
        self.assertFalse(refused["ok"])
        self.assertEqual(refused["detail"], "story_not_matched")
        send_mock.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_story_without_a_price_still_goes_through_the_picture(self):
        """Operator storyga narx yozmagan bo'lsa, eski yo'l — rasmni tahlil qilish."""
        item = AICatalogItem.objects.create(name="Alfalob 100 Tali", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/alfalob.jpg", **catalog_fingerprint_fields("https://cdn.example.com/alfalob.jpg", flower_form="peony_rose", dominant_colors=["hot_pink"], container="unwrapped_bouquet"))
        story_url = "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=999888777&signature=abc"
        SocialPost.objects.create(post_type="story", media_id="story-noprice", webhook_story_id="999888777", title_uz="", price=None, is_active=True)
        conversation = media_conversation("ig-story-no-price", image_url=story_url, kind="story")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["hot_pink"], container="unwrapped_bouquet")
        with patch_vision(source, {item.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "nechpul"}, conversation)
        self.assertEqual(result["detail"], "matched")
        self.assertTrue(result["allow_send"])
        # Bizning storyimiz bo'lgani uchun "o'xshagan" demaydi.
        self.assertTrue(result["own_post"])
        self.assertIn("aynan o'sha mahsulotning o'zi", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_every_catalog_item_on_a_shared_reel_is_offered_as_an_album(self):
        """Bitta reelga yetti xil mahsulot qo'yilgan — qaysi birini so'raganini link aytmaydi."""
        reel = "https://www.instagram.com/reel/DXHQrOliE8f/?igsi=MWtoazNkZHMxODd5aQ=="
        first = AICatalogItem.objects.create(name="Qizil Atir Gul", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/red.jpg", instagram_link=reel)
        second = AICatalogItem.objects.create(name="Oq Jumila", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/white.jpg", instagram_link=reel)
        conversation = media_conversation("ig-reel-album", image_url=reel, kind="reel")
        result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shundan bormi"}, conversation)
        self.assertFalse(result["allow_send"])
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["detail"], "instagram_link_group")
        self.assertEqual(sorted(row["catalog_id"] for row in result["group_matches"]), sorted([first.id, second.id]))

    @override_settings(OPENAI_API_KEY="test-key")
    def test_one_catalog_item_on_a_shared_reel_is_sent_on_its_own(self):
        reel = "https://www.instagram.com/reel/DIjgRABNbSf/"
        item = AICatalogItem.objects.create(name="Buket Alfalob", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/alfalob.jpg", instagram_link=reel)
        conversation = media_conversation("ig-reel-single", image_url=reel, kind="reel")
        result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shundan bormi"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual(result["detail"], "instagram_link_matched")
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [item.id])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_screenshot_falls_back_to_the_reel_the_customer_shared_earlier(self):
        """Screenshot'ning o'z havolasi yo'q, lekin reel hali suhbatda turadi."""
        reel = "https://www.instagram.com/reel/DXHQrOliE8f/"
        first = AICatalogItem.objects.create(name="Qizil Atir Gul", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/red.jpg", instagram_link=reel, **catalog_fingerprint_fields("https://cdn.example.com/red.jpg", flower_form="rose", dominant_colors=["red"], container="unwrapped_bouquet"))
        second = AICatalogItem.objects.create(name="Oq Jumila", arrangement_type="bouquet", price=900000, quantity=1, image_url="https://cdn.example.com/white.jpg", instagram_link=reel, **catalog_fingerprint_fields("https://cdn.example.com/white.jpg", flower_form="rose", dominant_colors=["white"], container="unwrapped_bouquet"))
        conversation = media_conversation("ig-reel-screenshot", image_url=reel, kind="reel")
        conversation.messages.create(sender="customer", text="", metadata={"attachments": [{"kind": "photo", "url": "https://cdn.example.com/screenshot.jpg"}]})
        conversation.messages.create(sender="customer", text="shundagisi bormi")
        source = vision_fingerprint(flower_form="tulip", dominant_colors=["yellow"], container="vase")
        with patch_vision(source, {}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shundagisi bormi"}, conversation)
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["detail"], "instagram_link_fallback")
        self.assertEqual(sorted(row["catalog_id"] for row in result["group_matches"]), sorted([first.id, second.id]))
        self.assertIn("reeldan", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_closest_catalog_items_are_shown_when_the_exact_flower_is_missing(self):
        """Aynan o'shasi yo'q bo'lsa ham mijoz quruq ketmasin."""
        close = AICatalogItem.objects.create(name="Savat London", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/london.jpg", **catalog_fingerprint_fields("https://cdn.example.com/london.jpg", flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket"))
        far = AICatalogItem.objects.create(name="Qizil Quti", arrangement_type="", price=400000, quantity=1, image_url="https://cdn.example.com/red.jpg", **catalog_fingerprint_fields("https://cdn.example.com/red.jpg", flower_form="rose", dominant_colors=["red"], container="hat_box"))
        conversation = media_conversation("ig-similar-album")
        # Rangi ham, hajmi ham boshqa: bu haqiqatan "o'xshash", "o'sha" emas.
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["purple", "lavender"], container="basket", size="small", count_bucket="25_to_50")
        with patch_vision(source, {close.id: verdict_payload(verdict="similar_only"), far.id: verdict_payload(verdict="similar_only")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertFalse(result["allow_send"])
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["detail"], "similar_only")
        # Uzoqdan o'xshagan ikkitasini ko'rsatish o'rniga butun katalog yuboriladi.
        self.assertTrue(result["show_whole_catalog"])
        self.assertEqual(result["group_matches"], [])
        self.assertEqual([row["catalog_id"] for row in result["near_matches"]], [close.id])
        # Rasm bo'yicha so'rov buyurtma emas: raqam so'ralmaydi, lead ochilmaydi.
        self.assertIn("business.operator_telegram", result["instruction_uz"])
        self.assertIn("lead yaratma", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_same_arrangement_in_another_colour_counts_as_similar(self):
        """Binafsha savat katalogda yo'q — savatlarimizni ko'rsatish kerak."""
        basket = AICatalogItem.objects.create(name="Savat London", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/london.jpg", **catalog_fingerprint_fields("https://cdn.example.com/london.jpg", flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket"))
        conversation = media_conversation("ig-violet-basket")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["purple", "lavender"], container="basket")
        with patch_vision(source, {basket.id: verdict_payload(verdict="different", color_match=False, differences="rangi butunlay boshqa")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shunaqasi bormi"}, conversation)
        self.assertFalse(result["allow_send"])
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["detail"], "similar_only")
        self.assertTrue(result["show_whole_catalog"])
        self.assertEqual([row["catalog_id"] for row in result["near_matches"]], [basket.id])

    def test_a_customer_returning_the_next_day_brings_their_history(self):
        """Ertaga qaytgan mijozni birinchi marta ko'rgandek kutib olmaslik uchun."""
        from datetime import timedelta as _timedelta
        customer = Customer.objects.create(instagram_user_id="ig-returning", name="Ahmad", phone="+998901112233")
        conversation = Conversation.objects.create(customer=customer)
        old = conversation.messages.create(sender="ai", text="Rahmat, kuningiz xayrli o'tsin.")
        conversation.messages.filter(id=old.id).update(created_at=timezone.now() - _timedelta(days=2))
        conversation.messages.create(sender="customer", text="salom")
        history = list(conversation.messages.order_by("created_at", "id"))
        row = services.previous_visit_context(conversation, history)
        self.assertEqual(row["days_since_previous_message"], 2)
        self.assertEqual(row["previous_message_date"], (timezone.localdate() - _timedelta(days=2)).isoformat())

    def test_yesterdays_replies_do_not_count_as_this_sessions_greeting(self):
        """24 soatdan keyin qaytgan mijoz uchun bu yangi suhbat — salomlashiladi."""
        from datetime import timedelta as _timedelta
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-fresh-session", name="Ahmad", phone="+998901112233")
        conversation = Conversation.objects.create(customer=customer)
        old_customer = conversation.messages.create(sender="customer", text="rahmat")
        old_ai = conversation.messages.create(sender="ai", text="Rahmat, kuningiz xayrli o'tsin.")
        conversation.messages.filter(id__in=[old_customer.id, old_ai.id]).update(created_at=timezone.now() - _timedelta(days=2))
        conversation.messages.create(sender="customer", text="salom")
        captured = {}

        class Response:
            output = []
            id = "resp-1"
            output_text = json.dumps({"reply": "Assalomu alaykum, Ahmad!", "handoff": False, "lead_ready": False, "phone": None, "customer_name": None, "detected_language": "uz", "estimated_price": None, "arrangement_type": None, "lead_request": None, "catalog_items": [], "stock_items": []})

        def fake_create(**kwargs):
            captured.update(kwargs)
            return Response()

        with override_settings(OPENAI_API_KEY="test-key"):
            with patch("core.services.OpenAI") as client_cls:
                client_cls.return_value.responses.create.side_effect = fake_create
                ai_reply(conversation)
        context = json.loads(captured["input"][0]["content"].split("REAL_CONTEXT_JSON:\n", 1)[1])
        self.assertTrue(context["conversation"]["fresh_session"])
        self.assertFalse(context["conversation"]["has_ai_reply_in_session"])
        self.assertEqual(context["customer"]["days_since_previous_message"], 2)

    def test_a_first_message_has_no_previous_visit(self):
        customer = Customer.objects.create(instagram_user_id="ig-brand-new")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="salom")
        history = list(conversation.messages.order_by("created_at", "id"))
        row = services.previous_visit_context(conversation, history)
        self.assertIsNone(row["days_since_previous_message"])

    def test_our_own_album_echo_does_not_pause_the_ai(self):
        """Instagram yuborgan albomni bizga qaytaradi — u operator javobi emas."""
        from .webhook_services import instagram_sent_message_exists
        customer = Customer.objects.create(instagram_user_id="ig-echo")
        conversation = Conversation.objects.create(customer=customer)
        services.SENT_INSTAGRAM_MESSAGE_IDS.clear()
        services.remember_sent_instagram_message({"message_id": "mid-album-1", "recipient_id": "x"})
        self.assertTrue(instagram_sent_message_exists(conversation, "mid-album-1"))
        self.assertFalse(instagram_sent_message_exists(conversation, "mid-operator-typed"))

    def test_a_story_photo_echo_is_recognised_from_the_saved_result(self):
        from .webhook_services import instagram_sent_message_exists
        customer = Customer.objects.create(instagram_user_id="ig-echo-story")
        conversation = Conversation.objects.create(customer=customer)
        services.SENT_INSTAGRAM_MESSAGE_IDS.clear()
        conversation.messages.create(sender="system", text="", metadata={"post_image_result": {"social_post_id": 1, "sent": {"message_id": "mid-story-9"}}})
        self.assertTrue(instagram_sent_message_exists(conversation, "mid-story-9"))

    def test_an_ad_banner_is_not_treated_as_a_photo_the_customer_sent(self):
        """Reklama banneri mijozning har bir xabariga o'zi qo'shiladi."""
        customer = Customer.objects.create(instagram_user_id="ig-ad-banner")
        conversation = Conversation.objects.create(customer=customer)
        ad_one = "https://www.facebook.com/ads/image/?d=AQJNuJz-first"
        ad_two = "https://www.facebook.com/ads/image/?d=AQJNuJz-second"
        conversation.messages.create(sender="customer", text="salom", metadata={"attachments": [{"kind": "photo", "url": ad_one}]})
        conversation.messages.create(sender="ai", text="Assalomu alaykum")
        conversation.messages.create(sender="customer", text="Adres qayerda", metadata={"attachments": [{"kind": "photo", "url": ad_two}]})
        rows = customer_attachment_rows(conversation.messages.order_by("created_at", "id"))
        self.assertEqual([row["url"] for row in rows], [ad_one])
        self.assertEqual(rows[0]["kind"], "ad")

    def test_a_real_photo_after_an_ad_banner_still_counts(self):
        customer = Customer.objects.create(instagram_user_id="ig-ad-then-photo")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="salom", metadata={"attachments": [{"kind": "photo", "url": "https://www.facebook.com/ads/image/?d=AQJ-banner"}]})
        conversation.messages.create(sender="ai", text="Assalomu alaykum")
        conversation.messages.create(sender="customer", text="shu nechpul", metadata={"attachments": [{"kind": "photo", "url": "https://cdn.example.com/real.jpg"}]})
        rows = customer_attachment_rows(conversation.messages.order_by("created_at", "id"))
        # Reklama banneri "ad" bo'lib qoladi, mijoz yuborgan rasm esa oddiy media.
        self.assertEqual([row["kind"] for row in rows], ["ad", "photo"])
        self.assertEqual(rows[-1]["url"], "https://cdn.example.com/real.jpg")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_same_catalog_photo_is_not_sent_twice(self):
        """"Yana qanaqalari bor" degan savolga o'sha rasmni qayta yuborish javob emas."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Qizil Atir Gul", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/red.jpg")
        conversation = media_conversation("ig-no-resend")
        with patch("core.services.send_image_to_customer", return_value=(True, "mocked", {"mocked": True})) as send_mock:
            first = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation)
        self.assertTrue(first["ok"])
        send_mock.assert_called_once()
        with patch("core.services.send_image_to_customer") as send_mock:
            with patch("core.services.send_catalog_album_chunk", return_value=(True, "mocked", None)) as album_mock:
                second = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation)
        # Rad etib qo'yish yetarli emas — o'rniga mijoz so'ragan katalog yuboriladi.
        self.assertTrue(second["ok"])
        self.assertEqual(second["detail"], "catalog_sent_instead")
        send_mock.assert_not_called()
        album_mock.assert_called()
        self.assertTrue(services.whole_catalog_already_sent(conversation))
        with patch("core.services.send_image_to_customer") as send_mock:
            with patch("core.services.send_catalog_album_chunk") as album_mock:
                third = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation)
        self.assertFalse(third["ok"])
        self.assertEqual(third["detail"], "catalog_image_already_sent")
        self.assertIn("operator_telegram", third["instruction_uz"])
        album_mock.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_safeguard_does_not_resend_a_photo_from_an_earlier_turn(self):
        """G'olib rasmi oldin yuborilgan bo'lsa, xavfsizlik chorasi ham qayta yubormaydi."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Qizil Atir Gul", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/red.jpg")
        conversation = media_conversation("ig-safeguard-once")
        conversation.messages.create(sender="system", text="", metadata={"image_tool_result": {"catalog_id": item.id, "delivered": True, "detail": "sent"}})
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": True, "allow_send": True, "matches": [{"catalog_id": item.id}], "group_matches": [], "near_matches": []}}]
        with patch("core.services.send_image_to_customer") as send_mock:
            apply_media_match_safeguard(conversation, {"reply": "..."}, tool_results)
        send_mock.assert_not_called()
        self.assertEqual([row["name"] for row in tool_results], ["match_ai_catalog_by_media"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_whole_catalog_is_not_sent_twice(self):
        """Mijoz katalogni allaqachon ko'rgan — qayta yuborish yangi hech narsa bermaydi."""
        from unittest.mock import patch
        for index in range(7):
            AICatalogItem.objects.create(name=f"Buket {index}", arrangement_type="bouquet", price=200000, quantity=1, image_url=f"https://cdn.example.com/{index}.jpg")
        conversation = media_conversation("ig-catalog-twice")
        with patch("core.services.send_catalog_album_chunk", return_value=(True, "mocked", None)):
            first = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertTrue(first["ok"])
        with patch("core.services.send_catalog_album_chunk") as album_mock:
            second = execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
        self.assertFalse(second["ok"])
        self.assertEqual(second["detail"], "catalog_already_sent")
        self.assertIn("operator_telegram", second["instruction_uz"])
        album_mock.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_named_group_album_is_still_allowed_after_the_whole_catalog(self):
        """Butun katalogdan keyin ham aniq ikkita mahsulotni ko'rsatish mumkin."""
        from unittest.mock import patch
        items = [AICatalogItem.objects.create(name=f"Buket {i}", arrangement_type="bouquet", price=200000, quantity=1, image_url=f"https://cdn.example.com/{i}.jpg") for i in range(7)]
        conversation = media_conversation("ig-group-after-catalog")
        with patch("core.services.send_catalog_album_chunk", return_value=(True, "mocked", None)):
            execute_ai_tool("send_catalog_album", {"catalog_ids": []}, conversation)
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": [items[0].id, items[1].id]}, conversation)
        self.assertTrue(result["ok"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_photo_already_seen_in_an_album_is_not_sent_again(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Savat London", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/london.jpg")
        conversation = media_conversation("ig-album-then-image")
        conversation.messages.create(sender="system", text="", metadata={"catalog_album_result": {"ok": True, "whole_catalog": True, "items": [{"catalog_id": item.id, "delivered": True}]}})
        with patch("core.services.send_image_to_customer") as send_mock:
            result = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation)
        # Butun katalog ham, bu rasm ham ko'rilgan — endi operatorga uzatiladi.
        self.assertFalse(result["ok"])
        self.assertEqual(result["detail"], "catalog_image_already_sent")
        self.assertIn("operator_telegram", result["instruction_uz"])
        send_mock.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_high_scoring_reject_is_offered_as_a_closest_match(self):
        """Ball baland bo'lsa "bizda bunday gul yo'q" deyish yolg'on bo'lardi."""
        item = AICatalogItem.objects.create(name="Savat London", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/london.jpg", **catalog_fingerprint_fields("https://cdn.example.com/london.jpg", flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket"))
        conversation = media_conversation("ig-close-match")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket")
        with patch_vision(source, {item.id: verdict_payload(verdict="similar_only")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertEqual(result["detail"], "close_matches")
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["no_match_reason"], "")
        self.assertGreaterEqual(result["group_matches"][0]["score"], services.CLOSE_MATCH_SCORE)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_nothing_close_enough_still_goes_to_the_operator(self):
        item = AICatalogItem.objects.create(name="Qizil Quti", arrangement_type="", price=400000, quantity=1, image_url="https://cdn.example.com/red.jpg", **catalog_fingerprint_fields("https://cdn.example.com/red.jpg", flower_form="rose", dominant_colors=["red"], container="hat_box"))
        conversation = media_conversation("ig-nothing-close")
        source = vision_fingerprint(flower_form="tulip", dominant_colors=["yellow"], container="vase")
        with patch_vision(source, {item.id: verdict_payload(verdict="different")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertFalse(result["allow_send"])
        self.assertFalse(result["allow_group"])
        self.assertIn(result["detail"], {"not_confident", "no_similar_catalog_item"})
        self.assertIn("operator_telegram", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_merely_similar_item_is_not_put_in_the_group(self):
        """Model "o'xshash" degani "boshqa mahsulot" degani — uni ko'rsatish chalg'itadi."""
        winner = AICatalogItem.objects.create(name="Savat London", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/london.jpg", **catalog_fingerprint_fields("https://cdn.example.com/london.jpg", flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket"))
        other = AICatalogItem.objects.create(name="Savat Bables", arrangement_type="basket", price=1500000, quantity=1, image_url="https://cdn.example.com/bables.jpg", **catalog_fingerprint_fields("https://cdn.example.com/bables.jpg", flower_form="peony_rose", dominant_colors=["pink", "hot_pink"], container="basket"))
        conversation = media_conversation("ig-similar-not-group")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket")
        with patch_vision(source, {winner.id: verdict_payload(), other.id: verdict_payload(verdict="similar_only")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [winner.id])
        self.assertEqual([row["catalog_id"] for row in result["near_matches"]], [other.id])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_clearly_weaker_candidate_is_not_put_in_the_group(self):
        """G'olibdan ancha orqada qolgan mahsulotni "qaysi biri" deb so'rash keraksiz."""
        winner = AICatalogItem.objects.create(name="Savat London", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/london.jpg", **catalog_fingerprint_fields("https://cdn.example.com/london.jpg", flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket", color_pattern="two_tone"))
        weaker = AICatalogItem.objects.create(name="Savat Jumila", arrangement_type="basket", price=900000, quantity=1, image_url="https://cdn.example.com/jumila.jpg", **catalog_fingerprint_fields("https://cdn.example.com/jumila.jpg", flower_form="peony_rose", dominant_colors=["white", "cream"], container="basket", color_pattern="solid"))
        conversation = media_conversation("ig-weaker-not-group")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket", color_pattern="two_tone")
        with patch_vision(source, {winner.id: verdict_payload(), weaker.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [winner.id])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_middling_score_is_enough_for_a_single_product_photo(self):
        item = AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket", color_pattern="solid", size="medium", count_bucket="over_100"))
        conversation = media_conversation("ig-single-middling")
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["yellow", "cream"], container="basket", color_pattern="two_tone", size="large", count_bucket="50_to_100")
        with patch_vision(source, {item.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertLess(result["matches"][0]["score"], vision_services.CROWDED_PHOTO_MIN_SCORE)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_same_middling_score_is_not_enough_in_a_crowded_photo(self):
        """Kadrda beshta buket turganda model gul turini chalkashtiradi — chegara balandroq."""
        item = AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket", color_pattern="solid", size="medium", count_bucket="over_100"))
        conversation = media_conversation("ig-crowded-middling")
        source = vision_fingerprint(
            flower_form="peony_rose", dominant_colors=["yellow", "cream"], container="basket",
            color_pattern="two_tone", size="large", count_bucket="50_to_100",
            region_requested=True, multiple_products_visible=True,
            visible_products=[{"position": 1, "where": "top", "short_description": "a"}, {"position": 2, "where": "bottom", "short_description": "b"}],
            chosen_position=2,
        )
        with patch_vision(source, {item.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "chizganim qancha"}, conversation)
        self.assertFalse(result["allow_send"])
        self.assertEqual(result["detail"], "ask_for_crop")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_pointed_at_flower_in_a_crowded_photo_gets_a_crop_request(self):
        """Ko'p gulli rasmdan bittasi ko'rsatilib topilmasa, kesib yuborishni so'raymiz."""
        AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket"))
        conversation = media_conversation("ig-needs-crop")
        source = vision_fingerprint(
            flower_form="rose",
            dominant_colors=["red"],
            container="wrapped_bouquet",
            region_requested=True,
            multiple_products_visible=True,
            visible_products=[
                {"position": 1, "where": "top", "short_description": "red roses"},
                {"position": 2, "where": "bottom", "short_description": "yellow basket"},
            ],
            chosen_position=1,
        )
        with patch_vision(source, {}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "chizib belgilaganim qancha"}, conversation)
        self.assertEqual(result["detail"], "ask_for_crop")
        self.assertTrue(result["ask_for_crop"])
        self.assertFalse(result["allow_send"])
        self.assertFalse(result["allow_group"])
        self.assertIn("kesib", result["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_crop_is_asked_for_only_once_in_a_conversation(self):
        """Mijoz kesa olmasa ikkinchi marta qiynamaymiz — operatorga uzatiladi."""
        conversation = media_conversation("ig-crop-once")
        source = vision_fingerprint(
            flower_form="rose",
            dominant_colors=["red"],
            container="wrapped_bouquet",
            region_requested=True,
            multiple_products_visible=True,
            chosen_position=1,
        )
        AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket"))
        with patch_vision(source, {}):
            first = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "chizganim qancha"}, conversation)
            second = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "chizganim qancha"}, conversation)
        self.assertEqual(first["detail"], "ask_for_crop")
        self.assertEqual(second["detail"], "not_confident")
        self.assertIn("operator_telegram", second["instruction_uz"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_single_product_photo_never_gets_a_crop_request(self):
        """Bitta gul turgan rasmni kesib berish hech narsani o'zgartirmaydi."""
        AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket"))
        conversation = media_conversation("ig-single-no-crop")
        source = vision_fingerprint(flower_form="rose", dominant_colors=["red"], container="wrapped_bouquet", region_requested=True, multiple_products_visible=False)
        with patch_vision(source, {}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertEqual(result["detail"], "not_confident")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_catalog_image_is_still_blocked_while_a_crop_is_being_asked_for(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Savat Katalina", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        conversation = media_conversation("ig-crop-gate")
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": True, "allow_send": False, "allow_group": False, "ask_for_crop": True, "matches": [], "group_matches": [], "near_matches": [{"catalog_id": item.id}]}}]
        with patch("core.services.send_image_to_customer") as send_mock:
            result = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation, tool_results=tool_results)
        self.assertFalse(result["ok"])
        self.assertEqual(result["detail"], "media_match_needs_a_crop")
        send_mock.assert_not_called()

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_spray_rose_bouquet_is_not_offered_for_a_classic_rose_photo(self):
        """Shoxli gul bir novdada ko'p kichik gul, klassik atir gul bitta yirik bosh."""
        classic = AICatalogItem.objects.create(name="Oq Jumila 100 Tali", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/classic.jpg", **catalog_fingerprint_fields("https://cdn.example.com/classic.jpg", flower_form="rose", dominant_colors=["cream", "pink"], container="unwrapped_bouquet", size="large"))
        spray = AICatalogItem.objects.create(name="Buket Shoxli Bambastic", arrangement_type="bouquet", price=900000, quantity=1, image_url="https://cdn.example.com/spray.jpg", **catalog_fingerprint_fields("https://cdn.example.com/spray.jpg", flower_form="spray_rose", dominant_colors=["cream", "pink"], container="unwrapped_bouquet", size="large"))
        conversation = media_conversation("ig-spray-vs-classic")
        source = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="unwrapped_bouquet", size="large")
        with patch_vision(source, {classic.id: verdict_payload(), spray.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shundan bormi"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [classic.id])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_peony_rose_still_matches_when_the_model_calls_it_a_spray_rose(self):
        """Pionavidniy gulni model bir safar shoxli, bir safar pionavidniy deb ataydi."""
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow"], container="basket"))
        conversation = media_conversation("ig-peony-called-spray")
        source = vision_fingerprint(flower_form="spray_rose", dominant_colors=["yellow"], container="basket")
        with patch_vision(source, {item.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "sariq savat qancha"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [item.id])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_25_stem_bouquet_is_not_offered_for_a_100_stem_photo(self):
        """Production xatosi: 199 000 lik buket 1 000 000 lik rasmga qo'shib yuborilgan edi."""
        big = AICatalogItem.objects.create(name="Oq Jumila 100 Tali", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/big.jpg", **catalog_fingerprint_fields("https://cdn.example.com/big.jpg", flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="extra_large", count_bucket="over_100"))
        small = AICatalogItem.objects.create(name="Buket Jumila", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/small.jpg", **catalog_fingerprint_fields("https://cdn.example.com/small.jpg", flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="small", count_bucket="25_to_50"))
        conversation = media_conversation("ig-stem-count")
        source = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="extra_large", count_bucket="over_100")
        with patch_vision(source, {big.id: verdict_payload(), small.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shundan bormi"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [big.id])

    def test_a_much_smaller_product_cannot_reach_a_confident_score(self):
        source = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="extra_large", count_bucket="over_100")
        same = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="extra_large", count_bucket="over_100")
        smaller = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="small", count_bucket="25_to_50")
        self.assertGreater(vision_services.fingerprint_score(source, same), 85)
        self.assertLessEqual(vision_services.fingerprint_score(source, smaller), vision_services.DIFFERENT_SIZE_CEILING)

    def test_a_crowded_photo_is_analysed_again_with_a_deeper_budget(self):
        """Kadrda beshta buket bo'lsa "low" uchdan bir hollarda adashadi."""
        from unittest.mock import patch
        crowded = vision_services.clean_fingerprint(vision_fingerprint(
            region_requested=True, multiple_products_visible=True,
            visible_products=[{"position": 1, "where": "top", "short_description": "a"}, {"position": 2, "where": "bottom", "short_description": "b"}],
            chosen_position=2,
        ))
        with patch.object(vision_services, "vision_json", return_value=crowded) as call:
            with override_settings(OPENAI_VISION_REASONING="low", OPENAI_VISION_CROWDED_REASONING="medium"):
                vision_services.analyze_image("https://cdn.example.com/grid.jpg", context_text="tepadan 2chisi", with_region=True, api_key="test-key")
        self.assertEqual(call.call_count, 2)
        self.assertEqual(call.call_args_list[0].kwargs.get("reasoning", ""), "")
        self.assertEqual(call.call_args_list[1].kwargs["reasoning"], "medium")

    def test_a_single_product_photo_is_analysed_once(self):
        from unittest.mock import patch
        plain = vision_services.clean_fingerprint(vision_fingerprint(region_requested=False, multiple_products_visible=False))
        with patch.object(vision_services, "vision_json", return_value=plain) as call:
            with override_settings(OPENAI_VISION_REASONING="low", OPENAI_VISION_CROWDED_REASONING="medium"):
                vision_services.analyze_image("https://cdn.example.com/one.jpg", context_text="shu nechpul", with_region=True, api_key="test-key")
        self.assertEqual(call.call_count, 1)

    def test_a_low_basket_full_of_flowers_still_matches_its_own_photo(self):
        """Savat katalogda "medium" (bo'yi past), rasmda esa model uni "extra_large" deydi."""
        catalog = vision_fingerprint(flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket", size="medium", count_bucket="over_100")
        photo = vision_fingerprint(flower_form="peony_rose", dominant_colors=["pink", "peach"], container="basket", size="extra_large", count_bucket="over_100")
        self.assertTrue(vision_services.sizes_can_match(photo, catalog))
        self.assertGreater(vision_services.fingerprint_score(photo, catalog), vision_services.DIFFERENT_SIZE_CEILING)

    def test_one_size_step_apart_is_still_the_same_product(self):
        """Rasmdan hajmni aniq o'lchab bo'lmaydi, bir pog'ona farq jazolanmaydi."""
        source = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="large", count_bucket="50_to_100")
        target = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "pink"], container="wrapped_bouquet", size="extra_large", count_bucket="over_100")
        self.assertGreater(vision_services.fingerprint_score(source, target), vision_services.DIFFERENT_SIZE_CEILING)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_photo_inventory_reaches_the_tool_result(self):
        """Operator "nega shuni tanladi" deb so'raganda javob shu maydonlarda turadi."""
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg", **catalog_fingerprint_fields("https://cdn.example.com/katalina.jpg", flower_form="peony_rose", dominant_colors=["yellow", "cream"], container="basket"))
        conversation = media_conversation("ig-photo-inventory")
        source = vision_fingerprint(
            flower_form="peony_rose",
            dominant_colors=["yellow", "cream"],
            container="basket",
            region_requested=True,
            region_description="Second from the top",
            multiple_products_visible=True,
            visible_products=[
                {"position": 1, "where": "top", "short_description": "red roses in a hat box"},
                {"position": 2, "where": "second from top", "short_description": "yellow basket"},
            ],
            chosen_position=2,
        )
        with patch_vision(source, {item.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "tepadan 2chisi qancha"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertTrue(result["region_requested"])
        self.assertEqual(result["chosen_position"], 2)
        self.assertEqual([row["where"] for row in result["visible_products"]], ["top", "second from top"])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_basket_photo_never_matches_a_hand_held_bouquet(self):
        """Model savatni buketga "same_product" desa ham idishi boshqa — o'tmaydi."""
        bouquet = AICatalogItem.objects.create(name="Buket Bambastic", arrangement_type="bouquet", price=900000, quantity=1, image_url="https://cdn.example.com/bouquet.jpg", **catalog_fingerprint_fields("https://cdn.example.com/bouquet.jpg", flower_form="spray_rose", dominant_colors=["cream", "pink"], container="unwrapped_bouquet"))
        conversation = media_conversation("ig-basket-vs-bouquet")
        source = vision_fingerprint(flower_form="spray_rose", dominant_colors=["cream", "pink"], container="basket")
        with patch_vision(source, {bouquet.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertFalse(result["allow_send"])
        self.assertFalse(result["allow_group"])
        self.assertEqual(result["detail"], "not_confident")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_wrapped_bouquet_still_matches_a_hat_box(self):
        """Quticha bilan o'ralgan buketni model ikki xil ataydi, ular qo'shni bo'lib qoladi."""
        box = AICatalogItem.objects.create(name="Qizil Atir Gul", arrangement_type="", price=400000, quantity=1, image_url="https://cdn.example.com/box.jpg", **catalog_fingerprint_fields("https://cdn.example.com/box.jpg", flower_form="rose", dominant_colors=["red"], container="hat_box"))
        conversation = media_conversation("ig-box-vs-wrap")
        source = vision_fingerprint(flower_form="rose", dominant_colors=["red"], container="wrapped_bouquet")
        with patch_vision(source, {box.id: verdict_payload()}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [box.id])

    @override_settings(OPENAI_API_KEY="test-key")
    def test_a_look_alike_with_the_same_price_is_not_worth_asking_about(self):
        """Narxi bir xil bo'lsa mijoz qaysi birini tanlasa ham javob o'zgarmaydi."""
        winner = AICatalogItem.objects.create(name="Savat Jumila", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/jumila-a.jpg", **catalog_fingerprint_fields("https://cdn.example.com/jumila-a.jpg", flower_form="rose", dominant_colors=["cream", "peach"], container="basket"))
        twin = AICatalogItem.objects.create(name="Savat Jumila 100 Tali", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/jumila-b.jpg", **catalog_fingerprint_fields("https://cdn.example.com/jumila-b.jpg", flower_form="rose", dominant_colors=["cream", "peach"], container="basket"))
        conversation = media_conversation("ig-same-price-twin")
        source = vision_fingerprint(flower_form="rose", dominant_colors=["cream", "peach"], container="basket")
        with patch_vision(source, {winner.id: verdict_payload(), twin.id: verdict_payload(verdict="similar_only")}):
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "shu nechpul"}, conversation)
        self.assertTrue(result["allow_send"])
        self.assertFalse(result["allow_group"])
        self.assertEqual([row["catalog_id"] for row in result["matches"]], [winner.id])

    def test_the_album_of_look_alike_items_is_allowed_but_a_single_image_is_not(self):
        from unittest.mock import patch
        first = AICatalogItem.objects.create(name="Buket Bir", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/one.jpg")
        second = AICatalogItem.objects.create(name="Buket Ikki", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/two.jpg")
        conversation = media_conversation("ig-group-gate")
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": True, "allow_send": False, "allow_group": True, "matches": [], "group_matches": [{"catalog_id": first.id}, {"catalog_id": second.id}], "near_matches": []}}]
        with patch("core.services.send_catalog_album", return_value={"ok": True, "items": []}) as album_mock:
            allowed = execute_ai_tool("send_catalog_album", {"catalog_ids": [first.id, second.id]}, conversation, tool_results=tool_results)
        self.assertTrue(allowed["ok"])
        album_mock.assert_called_once()
        with patch("core.services.send_image_to_customer") as send_mock:
            blocked = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": first.id}, conversation, tool_results=tool_results)
        self.assertFalse(blocked["ok"])
        self.assertEqual(blocked["detail"], "media_match_needs_a_group")
        send_mock.assert_not_called()

    def test_send_catalog_image_is_blocked_after_a_failed_media_match(self):
        """Production xatosi: mos kelmagan bo'lsa ham ikkita katalog rasmi yuborilgan edi."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        conversation = media_conversation("ig-blocked-image")
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": False, "allow_send": False, "matches": [], "near_matches": [{"catalog_id": item.id}]}}]
        with patch("core.services.send_image_to_customer") as send_mock:
            result = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation, tool_results=tool_results)
        self.assertFalse(result["ok"])
        self.assertEqual(result["detail"], "media_match_not_confident")
        send_mock.assert_not_called()

    def test_send_catalog_album_is_blocked_for_the_failed_candidates(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        conversation = media_conversation("ig-blocked-album")
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": False, "allow_send": False, "matches": [], "near_matches": [{"catalog_id": item.id}]}}]
        with patch("core.services.send_catalog_album") as album_mock:
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": [item.id]}, conversation, tool_results=tool_results)
        self.assertFalse(result["ok"])
        self.assertEqual(result["detail"], "media_match_not_confident")
        album_mock.assert_not_called()

    def test_whole_catalog_album_still_works_after_a_failed_media_match(self):
        """Mijoz "boshqa nima bor" desa katalogni ko'rsatish taqiqlanmaydi."""
        from unittest.mock import patch
        matched = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        other = AICatalogItem.objects.create(name="London Buket", arrangement_type="bouquet", price=400000, quantity=1, image_url="https://cdn.example.com/london.jpg")
        conversation = media_conversation("ig-album-allowed")
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": False, "allow_send": False, "matches": [], "near_matches": [{"catalog_id": matched.id}]}}]
        with patch("core.services.send_catalog_album", return_value={"ok": True, "items": []}) as album_mock:
            result = execute_ai_tool("send_catalog_album", {"catalog_ids": [other.id]}, conversation, tool_results=tool_results)
        self.assertTrue(result["ok"])
        album_mock.assert_called_once()

    def test_catalog_image_is_allowed_when_no_media_match_ran(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        conversation = media_conversation("ig-no-match-tool")
        with patch("core.services.send_image_to_customer", return_value=(True, "mocked", {})):
            result = execute_ai_tool("send_catalog_image", {"query": "", "catalog_id": item.id}, conversation, tool_results=[])
        self.assertTrue(result["ok"])

    def test_media_match_safeguard_sends_the_image_without_rewriting_the_reply(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        conversation = media_conversation("ig-media-safe")
        result = {"reply": "Katalina Savat, narxi 800 000 so'm. Qachonga kerak edi?"}
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": True, "allow_send": True, "matches": [{"catalog_id": item.id, "name": item.name, "price_text": "800 000 so'm"}]}}]
        with patch("core.services.send_image_to_customer", return_value=(True, "mocked", {"mocked": True})) as send_mock:
            fixed = apply_media_match_safeguard(conversation, result, tool_results)
        self.assertEqual(fixed["reply"], "Katalina Savat, narxi 800 000 so'm. Qachonga kerak edi?")
        self.assertEqual(send_mock.call_count, 1)
        self.assertEqual(tool_results[-1]["name"], "send_catalog_image")

    def test_media_match_safeguard_leaves_a_failed_match_alone(self):
        item = AICatalogItem.objects.create(name="Katalina Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/katalina.jpg")
        conversation = media_conversation("ig-media-low")
        result = {"reply": "Telefon raqamingizni yozib yuboraolasizmi?"}
        tool_results = [{"name": "match_ai_catalog_by_media", "arguments": {}, "output": {"ok": False, "allow_send": False, "matches": [], "near_matches": [{"catalog_id": item.id}]}}]
        fixed = apply_media_match_safeguard(conversation, result, tool_results)
        self.assertEqual(fixed["reply"], "Telefon raqamingizni yozib yuboraolasizmi?")
        self.assertEqual(len(tool_results), 1)


    def test_telegram_photo_is_labelled_as_a_customer_photo(self):
        from unittest.mock import patch
        payload = {
            "update_id": 7100,
            "message": {
                "message_id": 91,
                "chat": {"id": 7100},
                "from": {"id": 7100, "first_name": "Aziz"},
                "photo": [{"file_id": "photo-small", "file_size": 100}, {"file_id": "photo-big", "file_size": 900}],
                "caption": "Shundan bormi",
            },
        }
        with patch("core.webhook_services.telegram_file_url", return_value="https://api.telegram.org/file/bot1/photo.jpg"):
            jobs = resolve_telegram_update(payload)
        self.assertEqual(len(jobs), 1)
        message = Message.objects.get(id=jobs[0]["message_id"])
        self.assertEqual(message.metadata["attachments"][0]["kind"], "photo")
        self.assertIn("Mijoz yuborgan rasm: https://api.telegram.org/file/bot1/photo.jpg", message.text)


class NoStockPromptTests(TestCase):
    """Prompt keyingi migratsiyalarda sklad qoidalariga qaytib ketmasin."""

    def setUp(self):
        self.prompt = AISettings.objects.get(pk=1).system_prompt

    def test_prompt_no_longer_mentions_the_stock_tools(self):
        for name in ["get_stock", "calculate_custom_arrangement_price", "send_stock_image", "get_flower_variant_info", "batch_id", "price_per_stem"]:
            self.assertNotIn(name, self.prompt)

    def test_prompt_forbids_showing_the_stock_list(self):
        section = self.prompt.split("4. SKLAD SENDA YO'Q", 1)[1].split("5. KATALOG", 1)[0]
        self.assertIn("senga BERILMAYDI", section)
        self.assertIn("YASATMA TARTIBI", section)
        self.assertIn("flowers_text", section)
        self.assertIn("size_text", section)
        self.assertIn("Aniq narxni faqat operator aytadi", section)
        self.assertNotIn("Skladimizda hozir quyidagi gullar bor", self.prompt)

    def test_prompt_describes_the_customer_photo_flow(self):
        section = self.prompt.split("6A. MIJOZ RASM YUBORSA", 1)[1].split("7. NARX", 1)[0]
        self.assertIn("photo_request", section)
        self.assertIn("photo_urls", section)
        self.assertIn("customer_attachments", section)
        self.assertIn("KO'RMAYSAN", section)

    def test_prompt_routes_unknown_topics_to_the_shop_telegram(self):
        """Javob berolmagan savol lead emas — mijoz Telegram akkauntga yo'naltiriladi."""
        section = self.prompt.split("8B. JAVOB BEROLMAGAN SAVOL", 1)[1].split("9. DO'KON MA'LUMOTLARI", 1)[0]
        self.assertIn("TELEFON RAQAMI SO'RALMAYDI VA LEAD YARATILMAYDI", section)
        self.assertIn("business.operator_telegram", section)
        self.assertIn("client_lead_create CHAQIRILMAYDI", section)
        self.assertIn("Lead faqat buyurtma uchun ochiladi", section)

    def test_prompt_lists_every_lead_topic(self):
        section = self.prompt.split("8. BUYURTMA", 1)[1].split("8A. OPERATORGA ULASH", 1)[0]
        for topic in ["catalog_order", "custom_order", "photo_request", "question", "other"]:
            self.assertIn(topic, section)

    def test_prompt_does_not_promise_an_operator_for_off_topic_questions(self):
        """Ob-havo yoki bizda yo'q mahsulot uchun operator va'da qilinmasin."""
        section = self.prompt.split("8B. JAVOB BEROLMAGAN SAVOL", 1)[1].split("9. DO'KON MA'LUMOTLARI", 1)[0]
        self.assertIn("B. DO'KONGA UMUMAN TEGISHLI EMAS", section)
        self.assertIn("operatorga TOPSHIRMA, lead ham YARATMA", section)
        self.assertIn("Operatorlarimiz havoga oid aniq ma'lumot berishadi", section)
        self.assertIn("C. BIZDA YO'Q MAHSULOT", section)

    def test_prompt_keeps_the_contact_block_to_one_message(self):
        section = self.prompt.split("8A. OPERATORGA ULASH", 1)[1].split("8B. JAVOB BEROLMAGAN", 1)[0]
        self.assertIn("BIR SUHBATDA BIR MARTA", section)
        self.assertIn("Ketma-ket ikki xabarda takrorlash qat'iy taqiqlanadi", section)
        self.assertIn("Ism va telefon allaqachon olingan bo'lsa bu blokni umuman yozma", section)
        self.assertIn("IKKI MARTA so'rama", section)

    def test_prompt_hides_the_word_stock_from_the_customer(self):
        self.assertIn('MIJOZGA "SKLAD" SO\'ZINI AYTMA', self.prompt)
        self.assertIn("Biz skladdagi to'liq ro'yxatni yubormaymiz", self.prompt)

    def test_prompt_states_the_paid_container_price(self):
        section = self.prompt.split("7A. IDISH RANGI", 1)[1].split("8. BUYURTMA", 1)[0]
        self.assertIn("Qizil bo'ladi, u 100 000 so'm", section)
        # Yakuniy "operatorlar bog'lanadi" jumlasi narx qatorini yutib yubormasin.
        self.assertIn("yakuniy jumla bu qatorni almashtirmaydi", section)

    def test_prompt_keeps_the_flower_field_to_flower_names_only(self):
        """flowers_text ga butun jumla emas, faqat gul nomi tushsin."""
        self.assertIn("Butun jumlani yoki so'rovni bu yerga ko'chirma", self.prompt)
        self.assertIn('flowers_text ga "Jumila pushti atirgul", size_text ga "51 dona, katta"', self.prompt)


class VisionFingerprintTests(TestCase):
    def test_score_prefers_the_same_flower_over_a_similar_basket(self):
        """Production xatosi: krem-pushti atirgul savati sariq savat bilan mos deb topilgan."""
        source = vision_fingerprint(flower_form="peony_rose", dominant_colors=["cream", "pink"], color_pattern="two_tone", container="basket", size="extra_large", count_bucket="over_100")
        right = vision_fingerprint(flower_form="peony_rose", dominant_colors=["cream", "pink"], color_pattern="two_tone", container="basket", size="extra_large", count_bucket="over_100")
        wrong = vision_fingerprint(flower_form="peony_rose", dominant_colors=["yellow"], color_pattern="solid", container="basket", size="extra_large", count_bucket="over_100")
        right_score = vision_services.fingerprint_score(source, right)
        wrong_score = vision_services.fingerprint_score(source, wrong)
        self.assertGreater(right_score, wrong_score)
        self.assertGreaterEqual(right_score, vision_services.min_match_score())
        self.assertLess(wrong_score, vision_services.min_match_score())

    def test_score_separates_a_basket_from_a_bouquet(self):
        source = vision_fingerprint(container="basket")
        bouquet = vision_fingerprint(container="wrapped_bouquet")
        self.assertLess(vision_services.fingerprint_score(source, bouquet), vision_services.fingerprint_score(source, vision_fingerprint(container="basket")))

    def test_a_wrapped_bouquet_and_a_hat_box_are_not_punished_against_each_other(self):
        """Model bitta rasmni goh hat_box, goh wrapped_bouquet deydi — bu jazolanmasin."""
        source = vision_fingerprint(container="wrapped_bouquet")
        same_family = vision_services.fingerprint_score(source, vision_fingerprint(container="unwrapped_bouquet"))
        exact = vision_services.fingerprint_score(source, vision_fingerprint(container="wrapped_bouquet"))
        basket = vision_services.fingerprint_score(source, vision_fingerprint(container="basket"))
        self.assertEqual(same_family, exact)
        self.assertLess(basket, exact)

    def test_the_operator_arrangement_type_beats_the_model_guess(self):
        source = vision_fingerprint(container="basket")
        guessed_bouquet = vision_fingerprint(container="wrapped_bouquet")
        with_db_truth = vision_services.fingerprint_score(source, guessed_bouquet, target_arrangement_type="basket")
        without = vision_services.fingerprint_score(source, guessed_bouquet)
        self.assertGreater(with_db_truth, without)

    def test_red_is_not_treated_as_a_shade_of_hot_pink(self):
        red = vision_fingerprint(dominant_colors=["red"], color_pattern="solid")
        hot_pink = vision_fingerprint(dominant_colors=["hot_pink"], color_pattern="solid")
        self.assertLess(vision_services.fingerprint_score(red, hot_pink), vision_services.min_match_score())

    def test_near_colours_score_higher_than_opposite_colours(self):
        source = vision_fingerprint(dominant_colors=["cream"])
        near = vision_services.fingerprint_score(source, vision_fingerprint(dominant_colors=["white"]))
        far = vision_services.fingerprint_score(source, vision_fingerprint(dominant_colors=["burgundy"]))
        self.assertGreater(near, far)

    def test_model_answers_outside_the_allowed_values_are_dropped(self):
        cleaned = vision_services.clean_fingerprint({
            "flower_form": "beautiful pink roses",
            "dominant_colors": ["magenta", "pink", "pink", "white"],
            "container": "Basket",
            "size": "huge",
            "count_bucket": "over_100",
            "summary": "x",
        })
        self.assertEqual(cleaned["flower_form"], "")
        self.assertEqual(cleaned["dominant_colors"], ["pink", "white"])
        self.assertEqual(cleaned["container"], "basket")
        self.assertEqual(cleaned["size"], "")
        self.assertEqual(cleaned["count_bucket"], "over_100")

    def test_fingerprint_is_rebuilt_only_when_the_image_changes(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/a.jpg")
        with patch("core.vision_services.build_catalog_fingerprint", return_value=vision_services.clean_fingerprint(vision_fingerprint())) as build:
            vision_services.ensure_catalog_fingerprint(item)
            vision_services.ensure_catalog_fingerprint(item)
        self.assertEqual(build.call_count, 1)
        item.refresh_from_db()
        self.assertEqual(item.fingerprint_source_url, "https://cdn.example.com/a.jpg")
        self.assertIsNotNone(item.fingerprint_updated_at)
        item.image_url = "https://cdn.example.com/b.jpg"
        item.save(update_fields=["image_url"])
        self.assertTrue(vision_services.fingerprint_is_stale(item))

    @override_settings(OPENAI_API_KEY="test-key")
    def test_truncated_vision_json_is_retried_with_more_room(self):
        """reasoning byudjeti JSON ni kesib qo'ysa so'rov kengroq joy bilan qaytariladi."""
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/a.jpg")
        with patch("core.vision_services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.side_effect = [
                SimpleNamespace(output_text='{"flower_form": "peo', status="incomplete"),
                SimpleNamespace(output_text=json.dumps(vision_fingerprint()), status="completed"),
            ]
            fingerprint = vision_services.build_catalog_fingerprint(item)
        self.assertEqual(fingerprint["flower_form"], "peony_rose")
        first, second = client.responses.create.call_args_list
        self.assertEqual(second.kwargs["max_output_tokens"], first.kwargs["max_output_tokens"] * 2)
        self.assertEqual(second.kwargs["reasoning"]["effort"], "low")

    @override_settings(OPENAI_API_KEY="test-key")
    def test_vision_json_gives_up_after_the_retry(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/a.jpg")
        with patch("core.vision_services.OpenAI") as openai_class:
            openai_class.return_value.responses.create.return_value = SimpleNamespace(output_text="{", status="incomplete")
            # ensure_catalog_fingerprint xatoni yutadi va eski fingerprintni qoldiradi.
            self.assertEqual(vision_services.ensure_catalog_fingerprint(item), {})
        item.refresh_from_db()
        self.assertEqual(item.visual_fingerprint, {})

    @override_settings(OPENAI_API_KEY="test-key")
    def test_catalog_fingerprint_gets_the_name_and_note_as_context(self):
        from unittest.mock import patch
        item = AICatalogItem.objects.create(name="Katalina Gulidan Savat", arrangement_type="basket", price=800000, quantity=1, image_url="https://cdn.example.com/k.jpg", note="pionavidniy katalina ranglari tiniq sariq")
        with patch("core.vision_services.OpenAI") as openai_class:
            openai_class.return_value.responses.create.return_value = SimpleNamespace(output_text=json.dumps(vision_fingerprint()))
            vision_services.build_catalog_fingerprint(item)
        text = openai_class.return_value.responses.create.call_args.kwargs["input"][0]["content"][0]["text"]
        self.assertIn("Katalina Gulidan Savat", text)
        self.assertIn("pionavidniy katalina", text)


class NaturalSalesFlowTests(TestCase):
    """Real operator suhbatlaridan chiqqan holatlar: budjet, lead, til."""

    def setUp(self):
        self.customer = Customer.objects.create(instagram_user_id="ig-natural", instagram_username="mijoz", name="Ahmad", phone="+998901112233")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.cheap = AICatalogItem.objects.create(name="Buket Alfalob Gulidan", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/cheap.jpg", note="25 tacha guli bor, boyi 40 sm")
        self.middle = AICatalogItem.objects.create(name="Jumila Gulidan Kompazitsia", arrangement_type="bouquet", price=400000, quantity=1, image_url="https://cdn.example.com/mid.jpg")
        self.pricey = AICatalogItem.objects.create(name="London Gulidan Savat", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/big.jpg")

    def test_budget_query_returns_only_what_fits_cheapest_first(self):
        """«200 mingdan 500 minggacha boganlarini tashab bera olasizmi» — real savol."""
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": None, "min_price": 200000, "max_price": 500000}, self.conversation)
        self.assertEqual([row["name_uz"] for row in result["catalog"]], ["Jumila Gulidan Kompazitsia"])
        self.assertTrue(result["budget"]["exact_match"])
        # Budjetga mos mahsulot bor — arzonrog'ini eslatib mijozni pastga tortmaymiz.
        self.assertNotIn("cheapest_price", result["budget"])

    def test_budget_below_everything_still_shows_the_cheapest_and_says_so(self):
        """«250 mingga bomi» — bunday narx yo'q, lekin quruq «yo'q» deyish savdoni yopadi."""
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": None, "min_price": None, "max_price": 150000}, self.conversation)
        self.assertFalse(result["budget"]["exact_match"])
        self.assertEqual(result["budget"]["matched"], 0)
        self.assertEqual(result["budget"]["cheapest_price"], "199000.00")
        self.assertEqual(result["catalog"][0]["name_uz"], "Buket Alfalob Gulidan")

    def test_a_sum_we_do_not_carry_shows_what_is_around_it(self):
        """«А можете показать еще букеты за 350?» — 350 000 lik yo'q, 400 va 450 bor."""
        AICatalogItem.objects.create(name="Qizil Atir Guldan Kompazitsia", arrangement_type="bouquet",
                                     price=450000, quantity=1, image_url="https://cdn.example.com/450.jpg")
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": None,
                                                 "min_price": 350000, "max_price": 350000}, self.conversation)
        prices = sorted({row["price"] for row in result["catalog"]})
        self.assertEqual(prices, ["400000.00", "450000.00"])
        self.assertFalse(result["budget"]["exact_match"])
        self.assertEqual(result["budget"]["near_window_min"], "250000")
        self.assertEqual(result["budget"]["near_window_max"], "450000")
        # 199 000 ni eslatish mijozni o'zi tanlagan narxdan pastga tortadi.
        self.assertNotIn("cheapest_price", result["budget"])
        self.assertIn("HAMMASINI", result["budget"]["instruction_uz"])

    def test_nothing_within_the_window_falls_back_to_the_whole_catalog(self):
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": None,
                                                 "min_price": 3000000, "max_price": 3000000}, self.conversation)
        self.assertEqual(len(result["catalog"]), 3)
        self.assertEqual(result["budget"]["cheapest_price"], "199000.00")
        self.assertNotIn("near_window_min", result["budget"])

    def test_an_explicit_range_that_fits_is_left_alone(self):
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": None,
                                                 "min_price": 300000, "max_price": 500000}, self.conversation)
        self.assertEqual([row["price"] for row in result["catalog"]], ["400000.00"])
        self.assertTrue(result["budget"]["exact_match"])
        self.assertNotIn("near_window_min", result["budget"])

    def test_catalog_without_a_budget_keeps_the_newest_first_ordering(self):
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": None, "min_price": None, "max_price": None}, self.conversation)
        self.assertNotIn("budget", result)
        self.assertEqual(result["catalog"][0]["name_uz"], "London Gulidan Savat")

    def _lead_arguments(self, **overrides):
        arguments = {
            "customer_name": "Ahmad", "phone": "+998901112233", "request_text": "Katalogdan London savatini tanladi",
            "arrangement_type": "catalog", "estimated_price": 1000000,
            "fulfillment": "delivery", "delivery_address": "Chilonzor 5", "desired_date": "2026-08-25", "desired_time": "15:00",
            "catalog_items": [{"catalog_name": "London Gulidan Savat", "quantity": 1}], "note": "",
            "topic": "catalog_order", "flowers_text": None, "size_text": None, "photo_urls": [],
        }
        arguments.update(overrides)
        return arguments

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_new_lead_reaches_the_operators_telegram_group(self):
        """Lead bazada yotib qolmasin — operator uni Telegramda ko'radi."""
        from unittest.mock import patch
        with patch("core.services.telegram_send_rich_message_with", return_value={"ok": True}) as send:
            result = execute_ai_tool("client_lead_create", self._lead_arguments(), self.conversation)
        self.assertTrue(result["ok"])
        self.assertTrue(result["operators_notified"])
        send.assert_called_once()
        html = send.call_args.args[2]["html"]
        self.assertIn("Yangi lead", html)
        self.assertIn("London Gulidan Savat", html)
        self.assertIn("1 000 000", html)
        self.assertIn("Chilonzor 5", html)
        self.assertIn("+998901112233", html)
        self.assertTrue(send.call_args.args[2]["media"])

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_operator_message_uses_the_selected_ai_catalog_id(self):
        from unittest.mock import patch
        wrong = AICatalogItem.objects.create(name="Oq Jumila Atir Gulidan Yasalgan Kompazitsia 100 Tali", arrangement_type="bouquet", price=1000000, quantity=1, image_url="https://cdn.example.com/wrong.jpg")
        selected = AICatalogItem.objects.create(name="Buket Jumila Va Oq Atir Guldan Yasalgan Kompazitsia", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/right.jpg")
        arguments = self._lead_arguments(
            request_text="Buket Jumila Va Oq Atir Guldan Yasalgan Kompazitsia katalogdan 1 dona",
            estimated_price=199000,
            catalog_items=[{"catalog_id": selected.id, "catalog_name": selected.name, "quantity": 1}],
        )
        with patch("core.services.telegram_send_rich_message_with", return_value={"ok": True}) as send:
            result = execute_ai_tool("client_lead_create", arguments, self.conversation)
        lead = Lead.objects.get(id=result["lead_id"])
        self.assertEqual(lead.details["catalog_items"][0]["ai_catalog_item"], selected.id)
        self.assertNotEqual(lead.details["catalog_items"][0]["ai_catalog_item"], wrong.id)
        payload = send.call_args.args[2]
        self.assertIn(selected.name, payload["html"])
        self.assertNotIn(wrong.name, payload["html"])
        self.assertEqual(payload["media"][0]["media"]["media"], selected.image_url)

    @override_settings(OPENAI_API_KEY="test-key")
    def test_media_match_uses_instagram_ad_ids_before_vision(self):
        from unittest.mock import patch
        first = AICatalogItem.objects.create(name="Ad Buket Bir", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/1.jpg", instagram_ad_id="ad-1", instagram_ad_post_id="post-1")
        second = AICatalogItem.objects.create(name="Ad Buket Ikki", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/2.jpg", instagram_ad_id="ad-1", instagram_ad_post_id="post-1")
        conversation = Conversation.objects.create(customer=Customer.objects.create(instagram_user_id="ig-ad-match"))
        conversation.messages.create(
            sender="customer",
            text="Buyurtma bermoqchi edim\nMijoz yuborgan rasm: https://www.facebook.com/ads/image/?d=abc",
            metadata={
                "attachments": [{"kind": "photo", "url": "https://www.facebook.com/ads/image/?d=abc"}],
                "instagram_referral": {"ad_id": "ad-1", "ads_context_data": {"post_id": "post-1"}},
                "instagram_ad_id": "ad-1",
                "instagram_ad_post_id": "post-1",
            },
        )
        with patch("core.vision_services.OpenAI") as openai_class:
            result = execute_ai_tool("match_ai_catalog_by_media", {"source_url": None, "user_text": "buyurtma"}, conversation)
        self.assertTrue(result["ok"])
        self.assertTrue(result["allow_group"])
        self.assertEqual(result["detail"], "instagram_ad_group")
        self.assertEqual([row["catalog_id"] for row in result["group_matches"]], [first.id, second.id])
        openai_class.assert_not_called()

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_the_photo_the_customer_sent_goes_to_the_operators_with_the_lead(self):
        """Rasm bo'yicha so'rovda operator mijoz yuborgan rasmni ko'rishi shart."""
        from unittest.mock import patch
        photo = "https://cdn.example.com/customer-photo.jpg"
        self.conversation.messages.create(sender="customer", text="shu nechpul", metadata={"attachments": [{"kind": "photo", "url": photo}]})
        arguments = self._lead_arguments(topic="photo_request", catalog_items=[], photo_urls=[photo], request_text="Mijoz rasm yubordi, aynan shu gul narxini so'rayapti")
        with patch("core.services.telegram_send_rich_message_with", return_value={"ok": True}) as send:
            execute_ai_tool("client_lead_create", arguments, self.conversation)
        payload = send.call_args.args[2]
        # Rasm slideshow bo'lib ketadi; havolasi matn bo'lib yozilmaydi.
        self.assertEqual([row["media"]["media"] for row in payload["media"]], [photo])
        self.assertIn("tg://photo", payload["html"])
        self.assertNotIn(photo, payload["html"])

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_a_failing_telegram_send_never_loses_the_lead(self):
        from unittest.mock import patch
        with patch("core.services.telegram_send_rich_message_with", side_effect=RuntimeError("telegram down")), patch("core.services.telegram_send_with", side_effect=RuntimeError("still down")):
            result = execute_ai_tool("client_lead_create", self._lead_arguments(), self.conversation)
        self.assertTrue(result["ok"])
        self.assertFalse(result["operators_notified"])
        self.assertEqual(Lead.objects.filter(customer=self.customer).count(), 1)

    def test_uzbek_cyrillic_is_not_mistaken_for_russian(self):
        """«Доставка», «адрес», «заказ» — o'zbek mijozlar ham shunday yozadi."""
        from .services import detect_text_script
        for text in ["Доставка Канака булади", "Адрес каерда", "Заказ бермокчиман", "Цена канча", "Ассаламу алекум яхшимисиз"]:
            self.assertEqual(detect_text_script(text), "uz_cyril", text)

    def test_real_russian_is_still_detected(self):
        from .services import detect_text_script
        for text in ["Здравствуйте, какие цветы есть", "Сколько стоит букет", "Можно заказать букет на день рождения", "Добрый день, мне нужен букет для мамы"]:
            self.assertEqual(detect_text_script(text), "ru", text)

    def test_an_unmatched_photo_is_answered_with_the_catalog_not_a_dead_end(self):
        """Topilmasa ham mijoz quruq ketmasin: butun katalog va Telegram akkaunt."""
        from .services import MEDIA_MATCH_NOT_FOUND_INSTRUCTION
        self.assertIn("send_catalog_album", MEDIA_MATCH_NOT_FOUND_INSTRUCTION)
        self.assertIn("BO'SH massiv", MEDIA_MATCH_NOT_FOUND_INSTRUCTION)
        self.assertIn("business.operator_telegram", MEDIA_MATCH_NOT_FOUND_INSTRUCTION)
        # Rasm bo'yicha so'rov buyurtma emas — telefon ham, lead ham yo'q.
        self.assertIn("lead yaratma", MEDIA_MATCH_NOT_FOUND_INSTRUCTION)


class NaturalSalesPromptTests(TestCase):
    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0134_ai_prompt_natural_sales")

    def test_the_voice_block_wins_over_the_rule_sections(self):
        prompt = self.migration.TOP_BLOCK
        self.assertIn("HAMMA QOIDADAN USTUN", prompt)
        self.assertIn("BLOKLARNI USTMA-UST QO'YMA", prompt)
        self.assertIn("FAQAT SO'RALGANIGA JAVOB BER", prompt)

    def test_the_budget_block_explains_both_outcomes(self):
        prompt = self.migration.TOP_BLOCK
        self.assertIn("min_price", prompt)
        self.assertIn("max_price", prompt)
        self.assertIn("exact_match false", prompt)
        self.assertIn("cheapest_price", prompt)

    def test_the_note_is_retold_not_pasted(self):
        prompt = self.migration.TOP_BLOCK
        self.assertIn("note_uz", prompt)
        self.assertIn("o'z so'zing bilan", prompt)

    def test_the_contact_block_no_longer_fires_on_a_plain_contact_ask(self):
        self.assertIn("FAQAT shu ikki holatda", self.migration.CONTACT_BLOCK_NEW)
        self.assertNotIn("BIRINCHI marta ism va telefon so'raganingda", self.migration.CONTACT_BLOCK_NEW)

    def test_applying_the_migration_is_idempotent(self):
        settings_row = AISettings.objects.get_or_create(pk=1)[0]
        settings_row.system_prompt = "boshlanish\n" + self.migration.CONTACT_BLOCK_OLD + "\noxiri"
        settings_row.save()
        from django.apps import apps as installed_apps
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        settings_row.refresh_from_db()
        self.assertEqual(settings_row.system_prompt.count("00. QANDAY GAPIRASAN"), 1)
        self.assertIn(self.migration.CONTACT_BLOCK_NEW, settings_row.system_prompt)
        self.assertNotIn("sen BIRINCHI marta ism va telefon so'raganingda", settings_row.system_prompt)


class CatalogNeverLooksEmptyTests(TestCase):
    """Qidiruv so'zi topilmagani «bunday mahsulot yo'q» degani emas."""

    def setUp(self):
        self.conversation = Conversation.objects.create(customer=Customer.objects.create(instagram_user_id="ig-empty"))
        AICatalogItem.objects.create(name="Buket Bambastic", arrangement_type="bouquet", price=900000, quantity=1, image_url="https://cdn.example.com/b.jpg")
        AICatalogItem.objects.create(name="London Gulidan Savat", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/l.jpg")

    def test_a_russian_word_does_not_empty_the_catalog(self):
        """Katalog nomlari lotinda, mijoz esa «букет» deb so'raydi."""
        result = execute_ai_tool("get_catalog", {"query": "букет", "arrangement_type": None, "min_price": None, "max_price": None}, self.conversation)
        self.assertEqual(len(result["catalog"]), 2)
        self.assertFalse(result["query_matched"])
        self.assertIn("yo'q deb AYTMA", result["instruction_uz"])

    def test_a_matching_word_keeps_the_narrow_result(self):
        result = execute_ai_tool("get_catalog", {"query": "savat", "arrangement_type": None, "min_price": None, "max_price": None}, self.conversation)
        self.assertEqual([row["name_uz"] for row in result["catalog"]], ["London Gulidan Savat"])
        self.assertNotIn("query_matched", result)

    def test_an_arrangement_type_with_nothing_in_it_stays_empty(self):
        """Qutimiz yo'q bo'lsa «yo'q» deyish rost, buni yashirmaymiz."""
        result = execute_ai_tool("get_catalog", {"query": "", "arrangement_type": "box", "min_price": None, "max_price": None}, self.conversation)
        self.assertEqual(result["catalog"], [])
        self.assertNotIn("query_matched", result)


class BudgetAndContactTimingPromptTests(TestCase):
    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0135_ai_prompt_budget_and_contact_timing")
        self.earlier = importlib.import_module("core.migrations.0134_ai_prompt_natural_sales")

    def test_a_budget_that_exists_never_mentions_the_cheapest_item(self):
        """«1 millionlik savatingiz bormi» ga «eng arzoni 199 000» deb javob berilmasin."""
        self.assertIn("cheapest_price ni MUTLAQO tilga olma", self.migration.BUDGET_NEW)
        self.assertIn("FAQAT SHU HOLATDA", self.migration.BUDGET_NEW)

    def test_a_product_type_in_the_budget_question_is_passed_through(self):
        self.assertIn("arrangement_type basket", self.migration.BUDGET_NEW)

    def test_contact_is_asked_only_after_the_customer_picks_something(self):
        block = self.migration.CONTACT_TIMING
        self.assertIn("Mijoz hali savol berayotgan bo'lsa SO'RAMA", block)
        self.assertIn("chegirma so'radi", block)

    def test_it_rewrites_the_block_the_previous_migration_installed(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = ""
        row.save()
        self.earlier.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertIn(self.migration.BUDGET_OLD, row.system_prompt)
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertNotIn(self.migration.BUDGET_OLD, row.system_prompt)
        self.assertIn(self.migration.BUDGET_NEW, row.system_prompt)
        self.assertEqual(row.system_prompt.count("ISM VA TELEFONNI QACHON SO'RAYSAN"), 1)
        self.assertLess(row.system_prompt.index("ISM VA TELEFONNI QACHON SO'RAYSAN"), row.system_prompt.index("00A. BUDJET AYTILSA"))


class BudgetResultHidesTheCheapestWhenItFitsTests(TestCase):
    """Model ko'rmagan raqamni ayta olmaydi — promptdagi taqiq yetarli bo'lmadi."""

    def setUp(self):
        self.conversation = Conversation.objects.create(customer=Customer.objects.create(instagram_user_id="ig-budget"))
        AICatalogItem.objects.create(name="Arzon Buket", arrangement_type="bouquet", price=199000, quantity=1, image_url="https://cdn.example.com/a.jpg")
        AICatalogItem.objects.create(name="London Savat", arrangement_type="basket", price=1000000, quantity=1, image_url="https://cdn.example.com/b.jpg")

    def _budget(self, **kwargs):
        arguments = {"query": "", "arrangement_type": None, "min_price": None, "max_price": None}
        arguments.update(kwargs)
        return execute_ai_tool("get_catalog", arguments, self.conversation)["budget"]

    def test_a_budget_that_fits_never_exposes_the_cheapest_price(self):
        """«1 millionlik savatingiz bormi» — bor, arzonini eslatishning hojati yo'q."""
        budget = self._budget(min_price=900000, max_price=1100000)
        self.assertTrue(budget["exact_match"])
        self.assertNotIn("cheapest_price", budget)
        self.assertNotIn("instruction_uz", budget)

    def test_a_budget_that_fits_nothing_exposes_it_with_an_instruction(self):
        budget = self._budget(max_price=150000)
        self.assertFalse(budget["exact_match"])
        self.assertEqual(budget["cheapest_price"], "199000.00")
        self.assertIn("199 000", budget["instruction_uz"])


class BargainingPromptTests(TestCase):
    def test_the_bargaining_reply_does_not_ask_for_contact_details(self):
        migration = importlib.import_module("core.migrations.0136_ai_prompt_no_contact_ask_while_bargaining")
        self.assertIn("SAVDOLASHUV JAVOBIDA ISM VA TELEFON SO'RAMA", migration.INSERT)
        self.assertIn("ikki qatordan oshmasin", migration.INSERT)

    def test_it_lands_inside_the_bargaining_section(self):
        from django.apps import apps as installed_apps
        migration = importlib.import_module("core.migrations.0136_ai_prompt_no_contact_ask_while_bargaining")
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "C. CHEGIRMA SO'RASH\n...\n" + migration.ANCHOR + "\nqolgani"
        row.save()
        for _ in range(2):
            migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count("SAVDOLASHUV JAVOBIDA"), 1)
        self.assertLess(row.system_prompt.index("SAVDOLASHUV JAVOBIDA"), row.system_prompt.index(migration.ANCHOR))


class AlbumEchoIsRecognisedAcrossProcessesTests(TestCase):
    """Albomni bir celery jarayoni yuboradi, echo'ni boshqasi qabul qiladi."""

    def setUp(self):
        self.customer = Customer.objects.create(instagram_user_id="ig-echo")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.item = AICatalogItem.objects.create(name="Buket", arrangement_type="bouquet", price=500000, quantity=1, image_url="https://cdn.example.com/e.jpg")

    def test_the_album_message_ids_are_written_to_the_database(self):
        from unittest.mock import patch
        with patch("core.services.instagram_send_carousel", return_value={"message_id": "mid-album-1"}):
            services.send_catalog_album(self.conversation, [self.item], whole_catalog=True)
        stored = self.conversation.messages.filter(sender="system").last().metadata["catalog_album_result"]
        self.assertEqual(stored["sent_message_ids"], ["mid-album-1"])

    def test_our_own_album_echo_is_not_filed_as_an_operator_reply(self):
        from unittest.mock import patch
        from .webhook_services import instagram_sent_message_exists
        with patch("core.services.instagram_send_carousel", return_value={"message_id": "mid-album-2"}):
            services.send_catalog_album(self.conversation, [self.item], whole_catalog=True)
        # Boshqa jarayon: xotiradagi ro'yxat bo'sh, faqat baza qoladi.
        services.SENT_INSTAGRAM_MESSAGE_IDS.clear()
        self.assertTrue(instagram_sent_message_exists(self.conversation, "mid-album-2"))

    def test_a_real_operator_message_is_still_recognised_as_inbound(self):
        from .webhook_services import instagram_sent_message_exists
        services.SENT_INSTAGRAM_MESSAGE_IDS.clear()
        self.assertFalse(instagram_sent_message_exists(self.conversation, "mid-from-a-human"))


class SocialPostRaceTests(TestCase):
    """Instagram bitta reelni bir necha marta yuboradi, celery ularni parallel ishlaydi."""

    def test_a_concurrent_duplicate_does_not_break_the_webhook(self):
        from .webhook_services import social_post_upsert
        defaults = {"post_type": "reel", "permalink": "https://www.instagram.com/reel/AAA/", "title_uz": "Reel", "title_ru": "Reel", "is_active": True}
        first = social_post_upsert("media-race-1", defaults)
        second = social_post_upsert("media-race-1", dict(defaults, title_uz="Yangilangan"))
        self.assertEqual(first.id, second.id)
        self.assertEqual(SocialPost.objects.filter(media_id="media-race-1").count(), 1)
        self.assertEqual(second.title_uz, "Yangilangan")

    def test_an_inactive_post_with_the_same_media_id_is_reused_not_recreated(self):
        from .webhook_services import social_post_upsert
        SocialPost.objects.create(media_id="media-race-2", post_type="post", title_uz="Eski", title_ru="Eski", is_active=False)
        post = social_post_upsert("media-race-2", {"post_type": "reel", "permalink": "https://www.instagram.com/reel/BBB/", "title_uz": "Yangi", "title_ru": "Yangi", "is_active": True})
        self.assertEqual(SocialPost.objects.filter(media_id="media-race-2").count(), 1)
        self.assertTrue(post.is_active)

    def test_a_reel_share_survives_a_duplicate_delivery(self):
        """Avval shu yerda IntegrityError chiqib, mijozning reeli umuman qabul qilinmasdi."""
        from .webhook_services import social_post_from_ai_catalog_item
        item = AICatalogItem.objects.create(name="London Savat", arrangement_type="basket", price=1000000, quantity=1, instagram_link="https://www.instagram.com/reel/CCC/")
        event = SimpleNamespace(media_id="18424627252193879", story_id="", story_url="https://www.instagram.com/reel/CCC/", event_type="media_send")
        first = social_post_from_ai_catalog_item(item, event, "https://www.instagram.com/reel/CCC/")
        second = social_post_from_ai_catalog_item(item, event, "https://www.instagram.com/reel/CCC/")
        self.assertEqual(first.id, second.id)
        self.assertEqual(SocialPost.objects.filter(media_id="18424627252193879").count(), 1)


class SayFlowerNotProductTests(TestCase):
    def test_the_prompt_bans_the_internal_word(self):
        migration = importlib.import_module("core.migrations.0137_ai_prompt_say_gul_not_mahsulot")
        self.assertIn('MIJOZGA "MAHSULOT" DEMA', migration.INSERT)
        self.assertIn("Qaysi gulni nazarda tutyapsiz", migration.INSERT)
        self.assertIn("товар", migration.INSERT)

    def test_it_is_inserted_once_at_the_top(self):
        from django.apps import apps as installed_apps
        migration = importlib.import_module("core.migrations.0137_ai_prompt_say_gul_not_mahsulot")
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = migration.ANCHOR + "\n════════════════════════════════════\nqolgan matn"
        row.save()
        for _ in range(2):
            migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count('MIJOZGA "MAHSULOT" DEMA'), 1)


class NoBackendTriggersTests(TestCase):
    """AI faqat mijoz xabari, suhbat, system prompt va function natijalari bilan ishlaydi."""

    def test_the_backend_never_calls_a_tool_the_model_did_not_ask_for(self):
        source = Path(__file__).with_name("services.py").read_text(encoding="utf-8")
        for marker in ["forced_media_match", "forced_by_backend", "SYSTEM: Mijoz media yuborgan", "unanswered_customer_media"]:
            self.assertNotIn(marker, source, f"backend trigger qaytib kelgan: {marker}")

    def test_the_backend_never_rewrites_the_reply_text(self):
        source = Path(__file__).with_name("services.py").read_text(encoding="utf-8")
        rewrites = [line.strip() for line in source.splitlines() if 'result["reply"]' in line and "=" in line.split('result["reply"]')[1][:3]]
        # Yagona ruxsat etilgani — kirill suhbatda javobni kirillga qaytarish.
        self.assertEqual(rewrites, ['result["reply"] = latin_to_cyrillic(result["reply"])'])

    def test_the_tool_description_carries_the_obligation_instead(self):
        media_tool = next(tool for tool in ai_tool_definitions() if tool["name"] == "match_ai_catalog_by_media")
        self.assertIn("MAJBURIY", media_tool["description"])
        self.assertIn("Shubhalansang chaqir", media_tool["description"])


class WebhookSurvivesPostLinkingTests(TestCase):
    """Post bog'lash — qulaylik. Mijozning xabari esa buyurtma, u yo'qolmasligi kerak."""

    def test_a_failing_link_step_does_not_lose_the_media(self):
        from unittest.mock import patch
        from . import webhook_services
        event = SimpleNamespace(media_id="m-1", story_id="", story_url="https://www.instagram.com/reel/ZZZ/", event_type="media_send")
        with patch.object(webhook_services, "social_post_by_media_or_url", side_effect=RuntimeError("db down")), \
             patch.object(webhook_services, "link_story_post_from_event", side_effect=RuntimeError("api down")), \
             patch.object(webhook_services, "link_media_post_from_event", side_effect=RuntimeError("unique violation")):
            self.assertIsNone(webhook_services.resolve_social_post_safely("m-1", event))

    def test_a_later_step_still_wins_when_an_earlier_one_fails(self):
        from unittest.mock import patch
        from . import webhook_services
        post = SocialPost.objects.create(media_id="m-2", post_type="reel", title_uz="Reel", title_ru="Reel")
        event = SimpleNamespace(media_id="m-2", story_id="", story_url="", event_type="media_send")
        with patch.object(webhook_services, "social_post_by_media_or_url", side_effect=RuntimeError("boom")), \
             patch.object(webhook_services, "link_story_post_from_event", return_value=None), \
             patch.object(webhook_services, "link_media_post_from_event", return_value=post):
            self.assertEqual(webhook_services.resolve_social_post_safely("m-2", event), post)

    def test_adopting_a_media_id_someone_else_took_does_not_raise(self):
        from .webhook_services import adopt_media_id
        taken = SocialPost.objects.create(media_id="shared-id", post_type="reel", title_uz="Birinchi", title_ru="Birinchi")
        other = SocialPost.objects.create(media_id="own-id", post_type="reel", title_uz="Ikkinchi", title_ru="Ikkinchi")
        adopt_media_id(other, "shared-id")
        other.refresh_from_db()
        self.assertEqual(other.media_id, "own-id")
        self.assertEqual(SocialPost.objects.filter(media_id="shared-id").count(), 1)
        self.assertEqual(taken.id, SocialPost.objects.get(media_id="shared-id").id)

    def test_a_free_media_id_is_adopted(self):
        from .webhook_services import adopt_media_id
        post = SocialPost.objects.create(media_id="old-id", post_type="reel", title_uz="Reel", title_ru="Reel")
        adopt_media_id(post, "new-id")
        post.refresh_from_db()
        self.assertEqual(post.media_id, "new-id")


class AlbumEchoIsRecordedBeforeItReturnsTests(TestCase):
    """Echo albom yozuvidan oldin keladi, shuning uchun id darhol yozilishi kerak."""

    def test_each_chunk_is_recorded_as_soon_as_it_is_sent(self):
        from unittest.mock import patch
        from .webhook_services import instagram_sent_message_exists
        customer = Customer.objects.create(instagram_user_id="ig-fast-echo")
        conversation = Conversation.objects.create(customer=customer)
        item = AICatalogItem.objects.create(name="Buket", arrangement_type="bouquet", price=500000, quantity=1, image_url="https://cdn.example.com/x.jpg")
        seen = {}

        def carousel(*args, **kwargs):
            # Echo aynan shu payt keladi: albom yozuvi hali saqlanmagan.
            services.SENT_INSTAGRAM_MESSAGE_IDS.clear()
            seen["during_send"] = instagram_sent_message_exists(conversation, "mid-fast")
            return {"message_id": "mid-fast"}

        with patch("core.services.instagram_send_carousel", side_effect=lambda *a, **k: carousel()):
            services.send_catalog_album(conversation, [item], whole_catalog=True)
        services.SENT_INSTAGRAM_MESSAGE_IDS.clear()
        self.assertTrue(instagram_sent_message_exists(conversation, "mid-fast"))


class CustomOrderPromptTests(TestCase):
    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0138_ai_prompt_custom_order_and_bargaining")

    def test_the_phrasings_the_ai_missed_are_listed(self):
        insert = self.migration.INSERT
        for phrase in ["yasab berolislami", "man hohlaganimdek qb", "nechpul qberasla", "arzonroq qberaslami"]:
            self.assertIn(phrase, insert)

    def test_bargaining_is_separated_from_a_plain_price_question(self):
        self.assertIn("oddiy narx savoli deb tushunma", self.migration.INSERT)

    def test_the_custom_order_section_records_the_flowers_and_the_intent(self):
        block = self.migration.CUSTOM_ORDER
        self.assertIn("custom_order", block)
        self.assertIn("flowers_text", block)
        self.assertIn("o'zi yasattirmoqchi", block)
        self.assertIn("Narx AYTMA", block)

    def test_applying_it_twice_changes_nothing(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = self.migration.ANCHOR + "\n════════════════════════════════════\nqolgani"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count("MIJOZ NIMA DEMOQCHILIGINI"), 1)
        self.assertEqual(row.system_prompt.count("00C. YASATMA BUYURTMA"), 1)


class NoOperatorHandoffToolTests(TestCase):
    """Operatorga uzatish olib tashlandi. Mijoz Telegram akkauntga yo'naltiriladi."""

    def test_the_handoff_tool_is_gone(self):
        names = [tool["name"] for tool in ai_tool_definitions()]
        self.assertNotIn("handoff_media_to_operator", names)
        self.assertEqual(sorted(names), sorted([
            "client_leads_get", "client_lead_create", "client_lead_edit",
            "client_payment_update", "call_operator", "delivery_location_link",
            "match_ai_catalog_by_media", "get_catalog", "send_catalog_image",
            "send_post_image", "send_catalog_album",
        ]))

    def test_calling_it_anyway_is_refused(self):
        conversation = Conversation.objects.create(customer=Customer.objects.create(instagram_user_id="ig-no-handoff"))
        result = execute_ai_tool("handoff_media_to_operator", {"summary": "x", "phone": None, "customer_refused_phone": True}, conversation)
        self.assertEqual(result, {"ok": False, "detail": "unknown_tool"})

    def test_the_unmatched_photo_instruction_points_at_telegram(self):
        from .services import MEDIA_MATCH_NOT_FOUND_INSTRUCTION, MEDIA_MATCH_SIMILAR_INSTRUCTION
        for text in [MEDIA_MATCH_NOT_FOUND_INSTRUCTION, MEDIA_MATCH_SIMILAR_INSTRUCTION]:
            self.assertIn("operator_telegram", text)
            self.assertIn("SO'RAMA", text)
            self.assertNotIn("handoff_media_to_operator", text)

    def test_the_telegram_account_reaches_the_model_from_the_database(self):
        from unittest.mock import patch
        settings_row = BusinessSettings.objects.get_or_create(pk=1)[0]
        settings_row.operator_telegram = "@euroflowerspremium"
        settings_row.save()
        customer = Customer.objects.create(instagram_user_id="ig-ctx")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="salom")
        with patch("core.services.OpenAI") as openai_class, patch("core.services.openai_api_key", return_value="k"):
            openai_class.return_value.responses.create.return_value = SimpleNamespace(id="r", output=[], output_text=json.dumps({"reply": "ok", "detected_language": "uz", "customer_name": None, "phone": None, "handoff": False, "lead_ready": False, "lead_request": "", "estimated_price": None, "arrangement_type": None}))
            ai_reply(conversation)
        context = json.loads(openai_class.return_value.responses.create.call_args.kwargs["input"][0]["content"].split("REAL_CONTEXT_JSON:\n", 1)[1])
        self.assertEqual(context["business"]["operator_telegram"], "@euroflowerspremium")


class LeadOnlyForOrdersPromptTests(TestCase):
    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0140_ai_prompt_telegram_instead_of_handoff")

    def test_contact_details_are_only_for_an_order(self):
        self.assertIn("ISM VA TELEFON FAQAT BUYURTMA UCHUN", self.migration.TOP_RULE)
        self.assertIn("Lead — buyurtma, savol emas", self.migration.TOP_RULE)

    def test_the_question_flow_no_longer_takes_a_number_or_opens_a_lead(self):
        block = self.migration.NEW_BLOCK
        self.assertIn("TELEFON RAQAMI SO'RALMAYDI VA LEAD YARATILMAYDI", block)
        self.assertIn("business.operator_telegram", block)
        self.assertNotIn("client_lead_create chaqir, topic ga question", block)

    def test_it_replaces_the_old_block_once(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = self.migration.ANCHOR + "\n" + "═" * 36 + "\nbosh\n" + self.migration.OLD_BLOCK + "\noxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertNotIn(self.migration.OLD_BLOCK, row.system_prompt)
        self.assertEqual(row.system_prompt.count("TELEFON RAQAMI SO'RALMAYDI"), 1)
        self.assertEqual(row.system_prompt.count("ISM VA TELEFON FAQAT BUYURTMA UCHUN"), 1)


class FlowersInAVaseCanStillMatchTheCatalogTests(TestCase):
    """Bitta kompazitsiya goh vazada, goh qo'lda suratga olinadi.

    Katalogda "vaza" degan tur yo'q — hamma narsa "bouquet" yoki "basket". Mijoz
    vazada turgan rasmni yuborganida idish oilasi mos kelmay qolsa, katalogdagi
    birorta gul ham tekshiruvdan o'tolmaydi: mijoz o'z do'konimizning storysiga
    javob yozganida ham "bunday gulimiz yo'q" degan yolg'on javob chiqadi.
    """

    def test_a_vase_photo_can_match_a_bouquet_in_the_catalog(self):
        from . import vision_services
        source_family = vision_services.container_family({"container": "vase"})
        catalog_family = vision_services.container_family({"container": "unwrapped_bouquet"}, "bouquet")
        self.assertEqual(source_family, "vase")
        self.assertEqual(catalog_family, "bouquet")
        self.assertTrue(vision_services.families_can_match(source_family, catalog_family))
        self.assertTrue(vision_services.families_can_match(catalog_family, source_family))

    def test_a_basket_is_still_a_different_product(self):
        from . import vision_services
        self.assertFalse(vision_services.families_can_match("vase", "basket"))
        self.assertFalse(vision_services.families_can_match("basket", "vase"))

    def test_every_catalog_arrangement_type_is_reachable_from_a_vase_photo_or_a_basket_photo(self):
        from . import vision_services
        for arrangement in vision_services.ARRANGEMENT_FAMILIES:
            family = vision_services.ARRANGEMENT_FAMILIES[arrangement]
            reachable = any(
                vision_services.families_can_match(source, family)
                for source in ("vase", "basket", "bouquet", "box")
            )
            self.assertTrue(reachable, f"{arrangement} hech qaysi mijoz rasmiga mos kelolmaydi")


class StoryLookupCoversEveryConnectedAccountTests(TestCase):
    """Tizimga bir nechta Instagram akkaunt ulanadi va storyni qidirish hammasini ko'rishi kerak.

    Faqat asosiy akkauntdan qidirilganda ikkinchi akkauntning storysi "yo'q" bo'lib
    chiqadi: operator to'g'ri link qo'ygan bo'lsa ham post bog'lanmaydi va mijozning
    storyga yozgan javobi har safar rasm tahliliga tushib ketadi.
    """

    def setUp(self):
        from unittest.mock import patch
        from . import platform_services
        self.platform_services = platform_services
        self.calls = []
        self.stories = {
            "acc-birinchi": [{"id": "story-1", "permalink": "https://www.instagram.com/stories/birinchi/111", "media_url": "https://cdn/1.jpg"}],
            "acc-ikkinchi": [{"id": "story-2", "permalink": "https://www.instagram.com/stories/ikkinchi/222", "media_url": "https://cdn/2.jpg"}],
        }
        pairs = patch.object(platform_services, "instagram_account_token_pairs", return_value={"acc-birinchi": "tok-1", "acc-ikkinchi": "tok-2"})
        pairs.start()
        self.addCleanup(pairs.stop)

        class Response:
            def __init__(self, rows):
                self.rows = rows

            def raise_for_status(self):
                return None

            def json(self):
                return {"data": self.rows}

        def fake_get(url, params=None, timeout=None):
            account = url.rstrip("/").split("/")[-2]
            self.calls.append(account)
            return Response(self.stories.get(account, []))

        requests_patch = patch.object(platform_services.requests, "get", fake_get)
        requests_patch.start()
        self.addCleanup(requests_patch.stop)

    def test_it_asks_every_account_when_the_account_is_unknown(self):
        rows = self.platform_services.instagram_active_stories()
        self.assertEqual(sorted(self.calls), ["acc-birinchi", "acc-ikkinchi"])
        self.assertEqual([row["id"] for row in rows], ["story-1", "story-2"])

    def test_a_story_of_the_second_account_is_found(self):
        story = self.platform_services.find_active_story_by_media_url(
            "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=story-2&signature=x"
        )
        self.assertIsNotNone(story)
        self.assertEqual(story["permalink"], "https://www.instagram.com/stories/ikkinchi/222")

    def test_the_operators_link_resolves_on_the_second_account_too(self):
        story = self.platform_services.find_active_story_by_permalink("https://www.instagram.com/stories/ikkinchi/222")
        self.assertEqual(story["id"], "story-2")

    def test_a_known_account_is_asked_alone(self):
        rows = self.platform_services.instagram_active_stories("acc-ikkinchi")
        self.assertEqual(self.calls, ["acc-ikkinchi"])
        self.assertEqual([row["id"] for row in rows], ["story-2"])

    def test_one_broken_account_does_not_hide_the_others(self):
        def explode(url, params=None, timeout=None):
            raise RuntimeError("token expired")

        original = self.platform_services.requests.get

        def maybe_explode(url, params=None, timeout=None):
            if "acc-birinchi" in url:
                return explode(url)
            return original(url, params=params, timeout=timeout)

        from unittest.mock import patch
        with patch.object(self.platform_services.requests, "get", maybe_explode):
            rows = self.platform_services.instagram_active_stories()
        self.assertEqual([row["id"] for row in rows], ["story-2"])


class BargainingIsRecognisedByTheVerbTests(TestCase):
    """Mijoz savdolashuvni "arzonroq" so'zisiz, "berasiz" fe'li bilan so'raydi.

    Real suhbatda "Shuni nechpul qberas" ga oddiy narx, "Bolishi nechpul" ga esa
    yetkazib berish savoli qaytdi. Ro'yxatga yangi shakl qo'shish yetmaydi —
    modelga fe'l bo'yicha ajratishni o'rgatish kerak.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0141_ai_prompt_bargaining_verbs")

    def test_the_phrasings_from_the_real_chat_are_listed(self):
        block = self.migration.NEW_BARGAIN
        for phrase in ["nechpul qberas", "bolishi nechpul", "bo'lishi qancha", "qanchaga qo'yasiz"]:
            self.assertIn(phrase, block, f"savdolashuv shakli yo'q: {phrase}")

    def test_the_verb_tells_a_price_question_from_a_haggle(self):
        block = self.migration.NEW_BARGAIN
        self.assertIn("RO'YXATNI YOD OLMA, FE'LGA QARA", block)
        self.assertIn('"turadi", "narxi qancha", "nechpul" → oddiy narx savoli', block)
        self.assertIn('"berasiz", "qberasla", "qo\'yasiz", "qilib berasiz", "bo\'lishi" → savdolashuv', block)

    def test_the_two_real_mistakes_are_written_out(self):
        block = self.migration.NEW_BARGAIN
        self.assertIn('mijoz "Shuni nechpul qberas" dedi, sen 1 000 000 so\'m deding', block)
        self.assertIn('mijoz "Bolishi nechpul" dedi', block)

    def test_a_question_the_customer_asked_is_not_a_block_to_drop(self):
        block = self.migration.NEW_BLOCKS
        self.assertIn("Mijozning o'zi so'ragan savol", block)
        self.assertIn("ketma-ket ikki xabarda", block)
        self.assertIn('mijoz "Nechpul qberasla" va "Manzil qayoda" deb yozdi', block)

    def test_it_rewrites_each_anchor_once(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.OLD_BARGAIN + "orasi\n" + self.migration.OLD_BLOCKS + "oxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        prompt = row.system_prompt
        self.assertEqual(prompt.count("RO'YXATNI YOD OLMA, FE'LGA QARA"), 1)
        self.assertEqual(prompt.count("Mijozning o'zi so'ragan savol"), 1)
        self.assertIn("bolishi nechpul", prompt)
        self.assertNotIn(self.migration.OLD_BARGAIN, prompt)

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.OLD_BARGAIN + "orasi\n" + self.migration.OLD_BLOCKS + "oxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class AHaggleKeepsItsAnswerNextToAnotherQuestionTests(TestCase):
    """Savdolashuv savoli boshqa savol bilan birga kelsa ham savdolashuv bo'lib qoladi.

    "Nechpul qberasla" va "Manzil qayoda" birga kelganda AI ikkalasiga javob berdi,
    lekin narxni katalog narxi qilib qo'ydi — savdolashuv payqalmadi.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0142_ai_prompt_haggle_keeps_its_answer")

    def test_each_question_keeps_its_own_answer(self):
        block = self.migration.INSERT
        self.assertIn("har bir savol O'Z javobini oladi", block)
        self.assertIn("oddiy narx savoliga aylanib qolmaydi", block)

    def test_the_real_mistake_is_written_out(self):
        self.assertIn("sen 1 000 000 so'm va\nmanzilni aytding", self.migration.INSERT)
        self.assertIn("To'g'ri: 800 000 so'm va manzil.", self.migration.INSERT)

    def test_it_appends_once_after_the_anchor(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "oxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count("har bir savol O'Z javobini oladi"), 1)
        self.assertLess(row.system_prompt.index(self.migration.ANCHOR), row.system_prompt.index(self.migration.INSERT))

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "oxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class CheckTheVerbBeforeWritingAPriceTests(TestCase):
    """Savdolashuvni tanish qoidasi narx yozishdan oldingi tekshiruvga aylandi.

    Yolg'iz kelgan "Nechpul qberasla" ni model ishonchli tanidi, lekin yonida
    "Manzil qayoda" turganda oltita yurgizishdan ikkitasida katalog narxini yozdi.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0143_ai_prompt_check_the_verb_before_a_price")

    def test_the_check_runs_before_every_number(self):
        block = self.migration.INSERT
        self.assertIn("NARX YOZISHDAN OLDIN TEKSHIR", block)
        self.assertIn("Bo'lsa yozadigan raqam kelishilgan narx, katalog narxi EMAS", block)

    def test_every_colloquial_form_from_the_real_chat_is_covered(self):
        block = self.migration.INSERT
        for verb in ["qberasla", "qberas", "bolishi", "qo'yib berasiz"]:
            self.assertIn(f'"{verb}"', block, f"fe'l ro'yxatda yo'q: {verb}")

    def test_a_second_question_does_not_cancel_the_check(self):
        block = self.migration.INSERT
        self.assertIn("xabarda boshqa savol ham turgani", block)
        self.assertIn("mijoz ikkita xabar ketma-ket yozgani", block)
        self.assertIn("savdolashuv savoli ikkinchi bo'lib kelgani", block)

    def test_it_inserts_once_right_after_the_verb_rule(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "oxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count("NARX YOZISHDAN OLDIN TEKSHIR"), 1)
        self.assertLess(row.system_prompt.index(self.migration.ANCHOR), row.system_prompt.index(self.migration.INSERT))

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "oxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class AHaggleReadsTheNoteBeforeTheGenericAnswerTests(TestCase):
    """Savdolashuvni tanigach izohdagi kelishilgan narx aytiladi, umumiy javob emas.

    0143 dan keyin model savdolashuvni tanidi, lekin izohda 800 000 turgan bo'lsa
    ham "gullarimizning yangiligi... budjetingiz qancha" degan 10C javobini yozdi.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0144_ai_prompt_haggle_reads_the_note_first")

    def test_it_names_the_two_steps(self):
        block = self.migration.INSERT
        self.assertIn("get_catalog chaqirib mijoz gapirayotgan gulni top", block)
        self.assertIn("Uning izohidagi kelishilgan narxni bitta qatorda ayt", block)

    def test_it_does_not_ask_which_flower_again(self):
        self.assertIn('"Qaysi gulni nazarda tutyapsiz" deb qayta SO\'RAMA', self.migration.INSERT)

    def test_the_generic_discount_answer_is_ruled_out_when_a_note_price_exists(self):
        block = self.migration.INSERT
        self.assertIn("10C dagi umumiy javobni YOZMA", block)
        self.assertIn("budjetingiz qancha", block)
        self.assertIn("faqat izohda\nkelishilgan narx bo'lmaganda ishlatiladi", block)

    def test_it_inserts_once_after_the_check(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "oxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count("Fe'l topilgach shu ikki qadamni bajar"), 1)

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "oxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class PhoneNumberIsTakenInEveryShapeCustomersWriteTests(TestCase):
    """Mijoz raqamni qanday yozsa ham qabul qilinadi, kam bo'lsa to'ldirilmaydi."""

    def test_the_nine_digit_form_is_complete(self):
        from .services import normalize_phone
        self.assertEqual(normalize_phone("901234567"), "+998901234567")
        self.assertEqual(normalize_phone("90 123 45 67"), "+998901234567")

    def test_the_shapes_customers_actually_use(self):
        from .services import normalize_phone
        for written, expected in [
            ("+998901234567", "+998901234567"),
            ("998901234567", "+998901234567"),
            ("+998 90 240 95 15", "+998902409515"),
            ("8 998 90 123 45 67", "+998901234567"),
            ("0901234567", "+998901234567"),
            ("telefonim 93 555 66 77", "+998935556677"),
        ]:
            self.assertEqual(normalize_phone(written), expected, f"noto'g'ri o'qildi: {written}")

    def test_a_short_number_is_never_guessed(self):
        from .services import normalize_phone
        for written in ["9012345", "90123456", "123", "", "**** 4567"]:
            self.assertEqual(normalize_phone(written), "", f"to'ldirib taxmin qilindi: {written}")


class NoDataMeansOperatorNotAGuessTests(TestCase):
    """Javob senda bo'lmasa taxmin qilinmaydi — operator Telegramiga yo'naltiriladi."""

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0145_ai_prompt_no_data_no_guess")

    def test_it_searches_before_it_redirects(self):
        block = self.migration.BLOCK
        self.assertIn("AVVAL YAXSHILAB QIDIR, KEYINGINA OPERATORGA YO'NALTIR", block)
        self.assertIn("Suhbat tarixi", block)
        self.assertIn("get_catalog", block)
        self.assertIn("Qidirmasdan turib\noperatorga yo'naltirish ham xato", block)

    def test_every_missing_fact_from_the_real_chats_is_listed(self):
        block = self.migration.BLOCK
        for missing in ["karta raqami", "zaklad", "zapiska narxi", "harf yoki yozuv narxi",
                        "aksiya qachongacha", "kelin buket", "stol bezagi", "diametri"]:
            self.assertIn(missing, block, f"ro'yxatda yo'q: {missing}")
        self.assertIn("business.operator_telegram", block)
        self.assertIn("telefon raqami SO'RAMA va lead YARATMA", block)

    def test_it_refuses_to_swap_one_flower_for_another(self):
        block = self.migration.BLOCK
        self.assertIn("FAQAT KATALOGDA BOR GUL BILAN ISHLA", block)
        self.assertIn("Katalina gortenziya emas", block)
        for absent in ["gortenziya", "pion", "orxideya", "ramashka", "gerbera", "lola", "krizantema"]:
            self.assertIn(absent, block)

    def test_a_voice_message_is_answered_by_asking_for_text(self):
        block = self.migration.BLOCK
        self.assertIn("Ovozli xabarni tinglay olmadim, yozib yuborsangiz", block)
        self.assertIn("Ovozli xabar uchun telefon so'rama va operatorga topshirma", block)

    def test_the_greeting_happens_once(self):
        block = self.migration.BLOCK
        self.assertIn("SALOMLASHISH BIR MARTA", block)
        self.assertIn("ENG BIRINCHI", block)

    def test_a_catalog_number_still_reads_as_a_haggle(self):
        block = self.migration.BLOCK
        self.assertIn("2 chisi nechpul qberas", block)
        self.assertIn("Katalog narxini qaytarish XATO", block)
        self.assertIn("mijozni haydash bo'ladi", block)

    def test_the_short_answers_are_spelled_out(self):
        block = self.migration.BLOCK
        self.assertIn("Nalichida bormi", block)
        self.assertIn("hammasi tabiiy tirik gul", block)
        self.assertIn("BUKETGA YOZUV YOZILMAYDI", block)
        self.assertIn("Do'kon 24/7 ochiq VA administratorlar", block)

    def test_the_phone_rule_names_the_nine_digit_form(self):
        block = self.migration.BLOCK
        self.assertIn('"901234567" ko\'rinishida bersa ham qabul qilasan', block)
        self.assertIn("to'qqiz raqamdan kam", block)
        self.assertIn("to'liq yozib yuborasizmi", block)

    def test_it_inserts_once_before_section_00(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count("00D. JAVOBNI QAYERDAN OLASAN"), 1)
        self.assertLess(row.system_prompt.index("00D."), row.system_prompt.index(self.migration.ANCHOR))

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class SaleGroupMessageCarriesTheSizeTests(TestCase):
    """Guruhdagi florist qaysi gul sotilganini hajmidan tanib oladi."""

    def setUp(self):
        self.user = User.objects.create_user("sale-size", password="password", first_name="Sotuvchi", last_name="A")
        flower = Flower.objects.create(name_uz="Atirgul hajm", slug="rose-size")
        variant = FlowerVariant.objects.create(flower=flower, name_uz="Freedom", color_uz="Qizil")
        self.batch = StockBatch.objects.create(variant=variant, batch_number="SZ-1", height_cm=60, stems_per_bunch=20,
                                               received_stems=100, remaining_stems=100, cost_per_stem=1000,
                                               sale_price_per_stem=5000, sale_price_per_bunch=100000)

    def _caption(self, **overrides):
        from .inventory_services import sale_group_caption
        data = {"name_uz": "Alfalob buket", "arrangement_type": "bouquet", "catalog_kind": "standard",
                "price": Decimal("1000000"), "quantity_total": 1, "status": "available"}
        data.update(overrides)
        item = CatalogItem.objects.create(**data)
        history = CatalogHistory.objects.create(
            catalog_item=item, action="sold", created_by=self.user, quantity=1,
            listed_unit_price=Decimal("1000000"), sold_unit_price=Decimal("850000"),
            snapshot={"delivery_amount": "50000"},
        )
        return sale_group_caption(item, history, "mixed")

    def test_the_size_line_is_there(self):
        caption = self._caption(volume="katta", height_cm=60, diameter_cm=45)
        self.assertIn("Hajmi: katta · bo‘yi 60 sm · diametri 45 sm", caption)

    def test_the_english_volume_key_is_written_in_uzbek(self):
        for stored, shown in [("large", "katta"), ("medium", "o‘rtacha"),
                              ("small", "kichik"), ("extra_large", "juda katta")]:
            self.assertIn(f"Hajmi: {shown}", self._caption(volume=stored), f"o'girilmadi: {stored}")

    def test_an_unknown_volume_is_shown_as_written(self):
        self.assertIn("Hajmi: pastak savat", self._caption(volume="pastak savat"))

    def test_it_is_skipped_when_nothing_is_known(self):
        self.assertNotIn("Hajmi", self._caption())

    def test_partial_size_still_shows(self):
        self.assertIn("Hajmi: bo‘yi 55 sm", self._caption(height_cm=55))

    def test_the_money_lines_stay_intact(self):
        caption = self._caption(volume="o'rtacha")
        self.assertIn("Savdo: *800 000 so‘m*", caption)
        self.assertIn("Dastafka: 50 000 so‘m", caption)
        self.assertIn("Jami olingan: *850 000 so‘m*", caption)
        self.assertIn("To‘lov: *Aralash*", caption)
        self.assertIn("Sotdi: Sotuvchi A", caption)


class TheSecondRoundOfPromptFixesTests(TestCase):
    """Birinchi test to'plamidan keyin qolgan kamchiliklar.

    Ish vaqti yarim aytildi, kelin buketi oddiy yasatma deb qabul qilindi, pion
    o'rniga pionavidniy ko'rsatildi, salomlashish ikkinchi javobda ham qaytdi,
    yozuv narxi so'ralganda katalog yuborildi.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0146_ai_prompt_hours_bride_and_greeting")

    def test_working_hours_answer_carries_both_halves(self):
        self.assertIn("IKKALASI ham aytiladi", self.migration.NEW_HOURS)
        self.assertIn("Do'kon 24/7 ochiq. Administratorlarimiz esa har kuni operator_hours", self.migration.NEW_HOURS)
        self.assertIn("administratorlar vaqtini ber", self.migration.NEW_ONLY_HOURS)

    def test_a_bridal_bouquet_is_not_an_ordinary_custom_order(self):
        block = self.migration.NEW_CUSTOM
        self.assertIn("KELIN BUKETI BU BO'LIMGA KIRMAYDI", block)
        self.assertIn("business.operator_telegram", block)
        self.assertIn("Telefon so'rama, lead yaratma", block)

    def test_a_peony_is_not_a_peony_shaped_rose(self):
        block = self.migration.NEW_CATALOG
        self.assertIn("PION va PIONAVIDNIY boshqa-boshqa gul", block)
        self.assertIn("Pionaviy gullardan tayyor", block)
        self.assertIn("zakazga olinadi", block)
        self.assertIn("Yo'qligini aytmasdan katalogni yuborish yarim javob", block)

    def test_the_greeting_does_not_come_back_on_the_second_reply(self):
        block = self.migration.NEW_GREETING
        self.assertIn("IKKINCHI va undan keyingi javoblarida", block)
        self.assertIn("Buket Bambastic — 900 000 so'm", block)

    def test_a_phone_number_is_never_ignored(self):
        block = self.migration.NEW_GREETING
        self.assertIn("MIJOZ RAQAM YUBORSA E'TIBORSIZ QOLDIRMA", block)
        self.assertIn("raqam yo'qoladi", block)

    def test_the_lettering_question_is_not_answered_with_an_album(self):
        block = self.migration.NEW_LETTER
        self.assertIn("yozuv va harfning narxi katalogda yo'q", block)
        self.assertIn("telefon bloki birga yozilmaydi", block)

    def test_every_patch_applies_once_and_only_when_its_anchor_is_there(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "\n".join(old for old, _ in self.migration.PATCHES)
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        for marker in self.migration.MARKERS:
            self.assertEqual(row.system_prompt.count(marker), 1, f"takrorlandi yoki tushmadi: {marker[:40]}")

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "\n".join(old for old, _ in self.migration.PATCHES)
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)

    def test_it_does_nothing_when_the_prompt_has_no_anchors(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "boshqa prompt"
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, "boshqa prompt")


class EveryMessageToTheTestAccountIsATestChatTests(TestCase):
    """Test akkauntga kim yozsa ham AI javob beradi.

    Ilgari faqat yozgan mijozning username i tekshirilardi. Boshqa odam test
    akkauntga yozganda AI jim qolardi — serverda saidbek_ab extra_teest ga
    yozdi va javob olmadi.
    """

    def setUp(self):
        AISettings.objects.update_or_create(pk=1, defaults={"is_active": False})
        self.test_account = "17841476392326035"
        patcher = override_settings(
            AI_TEST_INSTAGRAM_ACCOUNT_IDS=[self.test_account],
            AI_TEST_INSTAGRAM_USERNAMES=["extra_teest"],
            AI_TEST_INSTAGRAM_USER_IDS=[],
        )
        patcher.enable()
        self.addCleanup(patcher.disable)

    def _conversation(self, username, account_id):
        customer = Customer.objects.create(instagram_username=username, instagram_user_id=f"ig-{username}")
        conversation = Conversation.objects.create(customer=customer)
        Message.objects.create(
            conversation=conversation, sender="customer", text="Assalomu alaykum",
            metadata={"instagram_account_id": account_id, "instagram_recipient_id": account_id},
        )
        return conversation

    def test_a_stranger_writing_to_the_test_account_gets_the_ai(self):
        self.assertTrue(ai_allowed_for_conversation(self._conversation("saidbek_ab", self.test_account)))

    def test_the_same_stranger_on_the_live_account_does_not(self):
        self.assertFalse(ai_allowed_for_conversation(self._conversation("saidbek_ab", "17841460916008920")))

    def test_the_recipient_id_is_enough_when_the_account_id_is_missing(self):
        customer = Customer.objects.create(instagram_username="kimdir", instagram_user_id="ig-kimdir")
        conversation = Conversation.objects.create(customer=customer)
        Message.objects.create(conversation=conversation, sender="customer", text="salom",
                               metadata={"instagram_recipient_id": self.test_account})
        self.assertTrue(ai_allowed_for_conversation(conversation))

    def test_the_old_username_allowlist_still_works(self):
        customer = Customer.objects.create(instagram_username="extra_teest", instagram_user_id="ig-extra")
        self.assertTrue(ai_allowed_for_conversation(Conversation.objects.create(customer=customer)))

    def test_a_conversation_with_no_metadata_stays_off(self):
        customer = Customer.objects.create(instagram_username="oddiy", instagram_user_id="ig-oddiy")
        conversation = Conversation.objects.create(customer=customer)
        Message.objects.create(conversation=conversation, sender="customer", text="salom", metadata={})
        self.assertFalse(ai_allowed_for_conversation(conversation))

    def test_the_latest_account_decides(self):
        conversation = self._conversation("kochma", "17841460916008920")
        self.assertFalse(ai_allowed_for_conversation(conversation))
        Message.objects.create(conversation=conversation, sender="customer", text="yana",
                               metadata={"instagram_account_id": self.test_account})
        self.assertTrue(ai_allowed_for_conversation(conversation))

    def test_the_reply_task_keeps_the_account_when_it_reschedules(self):
        source = Path(__file__).with_name("tasks.py").read_text(encoding="utf-8")
        reschedule = [line for line in source.splitlines() if "apply_async(args=[conversation_id, expected_message_id, recipient_id" in line]
        self.assertEqual(len(reschedule), 1)
        self.assertIn("account_id]", reschedule[0], "qayta rejalashtirishda akkaunt tushib qolgan")


class OneAmbiguousWordDoesNotSwitchTheLanguageTests(TestCase):
    """Javob tili oxirgi xabardan emas, javob berilayotgan hammasidan aniqlanadi.

    Real suhbatda mijoz "жойила катта", "вилоятга борми", "доставка" deb yozdi.
    Oxirgisi ruscha ham, o'zbekcha ham bir xil yoziladigan so'z — AI shuni yolg'iz
    o'qib ruscha javob berdi va shundan keyin suhbat ruschaga o'tib ketdi.
    """

    def test_a_lone_loanword_reads_as_russian(self):
        from .services import detect_text_script
        self.assertEqual(detect_text_script("доставка"), "ru")

    def test_but_the_batch_it_arrived_in_is_uzbek(self):
        from .services import conversation_script
        self.assertEqual(conversation_script(["жойила катта", "вилоятга борми", "доставка"]), "uz_cyril")

    def test_uzbek_wins_wherever_it_sits_in_the_batch(self):
        from .services import conversation_script
        self.assertEqual(conversation_script(["доставка", "вилоятга борми"]), "uz_cyril")
        self.assertEqual(conversation_script(["вилоятга борми", "доставка"]), "uz_cyril")

    def test_a_genuinely_russian_batch_stays_russian(self):
        from .services import conversation_script
        self.assertEqual(conversation_script(["Здравствуйте", "доставка есть?"]), "ru")

    def test_latin_and_empty_batches(self):
        from .services import conversation_script
        self.assertEqual(conversation_script(["Assalomu alaykum", "narxi qancha"]), "latin")
        self.assertEqual(conversation_script([]), "latin")
        self.assertEqual(conversation_script(["", "   "]), "latin")

    def test_the_reply_builder_decides_from_the_pending_batch(self):
        source = Path(__file__).with_name("services.py").read_text(encoding="utf-8")
        self.assertIn("customer_script = conversation_script(pending_customer_messages or [latest_customer_text])", source)
        self.assertIn('cyrillic_mode = customer_script == "uz_cyril"', source)
        # pending ro'yxati yozuv aniqlanishidan OLDIN hisoblanishi kerak
        self.assertLess(source.index("pending_customer_messages = [message.text"),
                        source.index("customer_script = conversation_script("))


class EveryPendingMessageGetsAnAnswerTests(TestCase):
    """Kirillcha yozilgan savol va ketma-ket kelgan xabarlar."""

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0147_ai_prompt_cyrillic_and_pending")

    def test_the_prompt_finally_uses_pending_customer_messages(self):
        block = self.migration.BLOCK
        self.assertIn("pending_customer_messages", block)
        self.assertIn("har biriga\njavob ber", block)
        self.assertIn("Javob baribir bitta xabar bo'ladi", block)

    def test_the_real_three_message_mistake_is_written_out(self):
        block = self.migration.BLOCK
        for part in ["жойила катта", "вилоятга борми", "доставка"]:
            self.assertIn(part, block)

    def test_the_k_and_q_confusion_is_explained(self):
        block = self.migration.BLOCK
        self.assertIn('"k" harfi "q" ni ham bildirishi mumkin', block)
        for pair in ["komaganmi   = qolmaganmi", "kanaka      = qanaqa", "arzonrok    = arzonroq"]:
            self.assertIn(pair, block)

    def test_the_freshness_question_is_listed_in_both_scripts(self):
        block = self.migration.BLOCK
        self.assertIn("solib qolmaganmi", block)
        self.assertIn("солиб комаганми", block)
        self.assertIn("гул солиб комаганми", block)

    def test_a_freshness_question_never_moves_the_price(self):
        block = self.migration.BLOCK
        self.assertIn("Kelishilgan narxni bu savolga AYTMA", block)
        self.assertIn('"950 000 so\'m qilib beramiz" deb narx tushirding', block)
        self.assertIn("yetkazib berish savoli deb o'qish ham XATO", block)

    def test_it_inserts_once_before_section_00(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count(self.migration.MARKER), 1)
        self.assertLess(row.system_prompt.index("00E."), row.system_prompt.index(self.migration.ANCHOR))

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class TheReplyLanguageComesFromTheScriptFieldTests(TestCase):
    """«доставка» ruscha belgi ro'yxatida turgani uchun butun javob ruschaga o'tardi."""

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0148_ai_prompt_language_from_script")

    def test_the_shared_loanwords_left_the_russian_marker_list(self):
        for word in ["доставка", "адрес"]:
            self.assertIn(word, self.migration.OLD_MARKERS)
            self.assertNotIn(word, self.migration.NEW_MARKERS.split("Quyidagi so'zlar")[0])

    def test_the_reply_language_is_read_from_the_context(self):
        block = self.migration.NEW_MARKERS
        self.assertIn("conversation.customer_script", block)
        for value in ['"latin"', '"uz_cyril"', '"ru"']:
            self.assertIn(value, block)

    def test_the_shared_words_are_named_as_not_russian(self):
        block = self.migration.NEW_MARKERS
        self.assertIn("rus tilining belgisi EMAS", block)
        for word in ["доставка", "дастафка", "адрес", "локация", "наличия"]:
            self.assertIn(word, block)

    def test_the_real_mistake_is_written_out(self):
        block = self.migration.NEW_MARKERS
        self.assertIn("вилоятга борми", block)
        self.assertIn("butun javobni rus tilida berding", block)

    def test_genuinely_russian_markers_stay(self):
        for word in ["здравствуйте", "сколько", "букет из"]:
            self.assertIn(word, self.migration.NEW_MARKERS)

    def test_the_context_carries_the_script(self):
        source = Path(__file__).with_name("services.py").read_text(encoding="utf-8")
        self.assertIn('"customer_script": customer_script,', source)
        self.assertIn("customer_script = conversation_script(", source)

    def test_it_applies_once_and_reverts(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.OLD_MARKERS + "\noxir"
        row.system_prompt = original
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count(self.migration.MARKER), 1)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class WiltedAndNaturalAreTwoDifferentQuestionsTests(TestCase):
    """0147 da ikkita savolni bitta ro'yxatga qo'shib qo'ygandim.

    "Табиийми" ga so'lish haqidagi javob qaytdi — mijoz esa gul tirikmi deb
    so'ragan edi.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0149_ai_prompt_natural_is_not_wilted")

    def test_natural_left_the_wilting_list(self):
        listed = [line for line in self.migration.NEW_BLOCK.splitlines()
                  if line.strip().startswith(("lotin:", "kirill:")) or line.startswith("          \"")]
        wilting = "\n".join(listed)
        self.assertIn("solib qolmaganmi", wilting, "ro'yxat topilmadi")
        for word in ["tabiiymi", "jivoymi", "табиийми", "живойми"]:
            self.assertNotIn(word, wilting, f"so'lish ro'yxatida qolib ketgan: {word}")

    def test_the_wilting_phrasings_are_still_all_there(self):
        block = self.migration.NEW_BLOCK
        for word in ["solib qolmaganmi", "solib komaganmi", "солиб комаганми",
                     "гул солиб комаганми", "svejiymi", "свежийми"]:
            self.assertIn(word, block)

    def test_natural_has_its_own_one_line_answer(self):
        block = self.migration.NEW_BLOCK
        self.assertIn("BOSHQA savol", block)
        self.assertIn("ha, hammasi tabiiy tirik gul", block)
        self.assertIn("Unga so'lish javobini berma", block)

    def test_it_applies_once_and_reverts(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.OLD_BLOCK + "\noxir"
        row.system_prompt = original
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count(self.migration.MARKER), 1)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class TelegramHandleSurvivesTheCyrillicConversionTests(TestCase):
    """Kirill javobda Telegram username o'girilib ketmasligi kerak.

    "@euroflowerspremium" da "EuroFlowers" brend sifatida saqlanardi, qolgan
    "premium" esa kirillga o'girilib "@euroflowersпремиум" bo'lib chiqardi —
    mijoz bunday akkauntni topolmaydi.
    """

    def test_the_handle_stays_latin(self):
        from .services import latin_to_cyrillic
        reply = latin_to_cyrillic("Aniq javob uchun @euroflowerspremium ga yozing")
        self.assertIn("@euroflowerspremium", reply)
        self.assertNotIn("премиум", reply)

    def test_the_bride_channel_link_stays_latin(self):
        from .services import latin_to_cyrillic
        self.assertIn("t.me/euroflowers_kelinbuket", latin_to_cyrillic("kelin buket: t.me/euroflowers_kelinbuket"))

    def test_links_emails_and_brands_still_survive(self):
        from .services import latin_to_cyrillic
        reply = latin_to_cyrillic("https://yandex.uz/maps/-/CTfQ6TMD va info@euroflowers.uz, Next Mall")
        self.assertIn("https://yandex.uz/maps/-/CTfQ6TMD", reply)
        self.assertIn("info@euroflowers.uz", reply)
        self.assertIn("Next Mall", reply)

    def test_ordinary_words_are_still_converted(self):
        from .services import latin_to_cyrillic
        self.assertEqual(latin_to_cyrillic("Narxi 800 000 so‘m"), "Нархи 800 000 сўм")


class TheAssistantDoesNotRepeatItselfTests(TestCase):
    """Story javobidan keyin yangi savolga o'sha nom va narx qaytardi."""

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0150_ai_prompt_no_repeat_no_greeting")

    def test_the_greeting_is_tied_to_a_context_field(self):
        block = self.migration.BLOCK
        self.assertIn("conversation.has_ai_reply_in_session", block)
        self.assertIn("media\nkelgani suhbatni boshidan boshlamaydi", block)
        self.assertIn("Ассалому алайкум! Ҳозирда бизда бор гуллар шулар", block)

    def test_the_repeat_mistake_is_written_out(self):
        block = self.migration.BLOCK
        self.assertIn("O'ZINGNI TAKRORLAMA", block)
        self.assertIn('mijoz: "solib qomaganmi"', block)
        self.assertIn("narxni ikkinchi marta yozding", block)

    def test_more_options_means_the_whole_album(self):
        block = self.migration.BLOCK
        self.assertIn("catalog_ids BO'SH massiv", block)
        for phrase in ["yana shunaqa variantla bormi", "yokida faqat shumi",
                       "boya ko'proq gullar tashuvdingku", "яна шунака вариантла борми"]:
            self.assertIn(phrase, block)
        self.assertIn('"Katalogimiz shu" deb yozib, albomni yubormaslik ham XATO', block)

    def test_the_wilting_answer_needs_a_wilting_question(self):
        block = self.migration.BLOCK
        self.assertIn("SO'LISH JAVOBINI FAQAT SO'RALGANDA BER", block)
        self.assertIn("boya koproq gullar tashudinku katalogda", block)

    def test_it_inserts_once_before_the_previous_block(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count(self.migration.MARKER), 1)
        self.assertLess(row.system_prompt.index("00F."), row.system_prompt.index(self.migration.ANCHOR))

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class AQualityQuestionDescribesTheFlowerTests(TestCase):
    """«Sifati qanaqa» ga so'lish javobi qaytardi — mijoz gulni bilmoqchi edi."""

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0151_ai_prompt_quality_from_the_note")

    def test_the_answer_comes_from_the_note(self):
        block = self.migration.BLOCK
        self.assertIn("Javobni izohdan ol", block)
        self.assertIn("nechta guli borligi", block)
        self.assertIn("get_catalog", block)

    def test_both_scripts_are_listed(self):
        block = self.migration.BLOCK
        self.assertIn("Sifati qanaqa", block)
        self.assertIn("сифати канака", block)

    def test_it_rules_out_the_wilting_answer(self):
        self.assertIn("so'lish haqidagi javobni berma", self.migration.BLOCK)

    def test_it_applies_once_and_reverts(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.system_prompt = original
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count(self.migration.MARKER), 1)
        self.assertLess(row.system_prompt.index(self.migration.MARKER), row.system_prompt.index(self.migration.ANCHOR))
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class TheCustomerCanAskForTheCatalogAgainTests(TestCase):
    """Albom bir marta ketgach mijoz uni boshqa ko'ra olmasdi.

    Real suhbatda mijoz "Korsat katalogni", keyin "Korsatmading yubor katalog"
    deb yozdi — backend to'sig'i har uchalasini ham bloklab turdi va AI
    "katalogimiz shu" deb yozib hech narsa yubormadi.
    """

    def setUp(self):
        self.customer = Customer.objects.create(instagram_username="katalog_probe", instagram_user_id="ig-katalog")
        self.conversation = Conversation.objects.create(customer=self.customer)

    def _album_sent(self):
        return Message.objects.create(
            conversation=self.conversation, sender="system", text="",
            metadata={"catalog_album_result": {"whole_catalog": True, "items": [{"delivered": True}]}},
        )

    def test_nothing_sent_yet_means_not_blocked(self):
        from .services import whole_catalog_already_sent
        self.assertFalse(whole_catalog_already_sent(self.conversation))

    def test_it_blocks_a_second_send_in_the_same_turn(self):
        from .services import whole_catalog_already_sent
        self._album_sent()
        self.assertTrue(whole_catalog_already_sent(self.conversation))

    def test_the_customer_asking_again_unblocks_it(self):
        from .services import whole_catalog_already_sent
        self._album_sent()
        Message.objects.create(conversation=self.conversation, sender="customer", text="Korsat katalogni")
        self.assertFalse(whole_catalog_already_sent(self.conversation))

    def test_it_blocks_again_after_the_new_album_goes_out(self):
        from .services import whole_catalog_already_sent
        self._album_sent()
        Message.objects.create(conversation=self.conversation, sender="customer", text="Korsat katalogni")
        self._album_sent()
        self.assertTrue(whole_catalog_already_sent(self.conversation))

    def test_a_partial_album_never_blocks(self):
        from .services import whole_catalog_already_sent
        Message.objects.create(
            conversation=self.conversation, sender="system", text="",
            metadata={"catalog_album_result": {"whole_catalog": False, "items": [{"delivered": True}]}},
        )
        self.assertFalse(whole_catalog_already_sent(self.conversation))

    def test_an_undelivered_album_never_blocks(self):
        from .services import whole_catalog_already_sent
        Message.objects.create(
            conversation=self.conversation, sender="system", text="",
            metadata={"catalog_album_result": {"whole_catalog": True, "items": [{"delivered": False}]}},
        )
        self.assertFalse(whole_catalog_already_sent(self.conversation))


class APhotoIsAlwaysAnalysedNotAnsweredFromAnOldReelTests(TestCase):
    """Mijoz reel yuborib, keyin boshqa gulning rasmini tashladi.

    Javob "siz yuborgan reeldan borlari shular" bo'lib chiqdi — rasm umuman
    tahlil qilinmadi, suhbatdagi eski reel havolasi ishlatildi.
    """

    def setUp(self):
        self.customer = Customer.objects.create(instagram_username="rasm_probe", instagram_user_id="ig-rasm")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.item = AICatalogItem.objects.create(
            name="Reeldagi gul", arrangement_type="bouquet", price=Decimal("500000"),
            instagram_link="https://www.instagram.com/reel/AAA111/", is_active=True,
        )
        Message.objects.create(
            conversation=self.conversation, sender="customer", text="shundan bormi",
            metadata={"attachments": [{"url": "https://www.instagram.com/reel/AAA111/", "kind": "reel"}]},
        )

    def test_a_shared_reel_still_matches_from_the_conversation(self):
        from .services import direct_ai_catalog_link_matches
        matched = direct_ai_catalog_link_matches(
            [self.item], "https://lookaside.fbsbx.com/x?asset_id=1",
            attachment={"url": "https://lookaside.fbsbx.com/x?asset_id=1", "kind": "reel"},
            conversation=self.conversation,
        )
        self.assertEqual([row.id for row in matched], [self.item.id])

    def test_a_screenshot_of_the_reel_just_shared_still_matches(self):
        """Reel yuborib darhol skrinshot tashlash — bitta savol."""
        from .services import direct_ai_catalog_link_matches
        matched = direct_ai_catalog_link_matches(
            [self.item], "https://lookaside.fbsbx.com/x?asset_id=2",
            attachment={"url": "https://lookaside.fbsbx.com/x?asset_id=2", "kind": "photo"},
            conversation=self.conversation,
        )
        self.assertEqual([row.id for row in matched], [self.item.id])

    def test_a_photo_sent_after_we_answered_the_reel_does_not(self):
        """Reel haqida javob berilgach kelgan rasm — yangi savol."""
        from .services import direct_ai_catalog_link_matches
        Message.objects.create(conversation=self.conversation, sender="ai",
                               text="Siz yuborgan reeldan hozir bizda borlari shular.")
        matched = direct_ai_catalog_link_matches(
            [self.item], "https://lookaside.fbsbx.com/x?asset_id=2",
            attachment={"url": "https://lookaside.fbsbx.com/x?asset_id=2", "kind": "photo"},
            conversation=self.conversation,
        )
        self.assertEqual(matched, [], "rasm eski reel havolasi bilan javob oldi")

    def test_a_shared_reel_still_matches_after_an_ai_reply(self):
        """Havolaning o'zi kelsa butun suhbat bo'yicha qidiriladi."""
        from .services import direct_ai_catalog_link_matches
        Message.objects.create(conversation=self.conversation, sender="ai", text="javob")
        matched = direct_ai_catalog_link_matches(
            [self.item], "https://lookaside.fbsbx.com/x?asset_id=3",
            attachment={"url": "https://lookaside.fbsbx.com/x?asset_id=3", "kind": "reel"},
            conversation=self.conversation,
        )
        self.assertEqual([row.id for row in matched], [self.item.id])

    def test_a_photo_whose_own_link_matches_still_works(self):
        from .services import direct_ai_catalog_link_matches
        matched = direct_ai_catalog_link_matches(
            [self.item], "https://www.instagram.com/reel/AAA111/",
            attachment={"url": "https://www.instagram.com/reel/AAA111/", "kind": "photo"},
            conversation=self.conversation,
        )
        self.assertEqual([row.id for row in matched], [self.item.id])


class TheAssistantNeverAsksForAPhoneTests(TestCase):
    """Raqam so'rash promptdan butunlay olib tashlandi.

    extra_teest suhbatida mijoz "Yasab berolislami" deb so'raganda AI
    "telefon raqamingizni yozib qoldiring" deb javob berdi. Endi mijoz nima
    so'raganini takrorlab, Telegram akkauntga yo'naltiriladi.
    """

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0152_ai_prompt_never_ask_for_a_phone")

    def test_the_top_rule_forbids_it_outright(self):
        block = self.migration.NEW_TOP
        self.assertIn("ISM VA TELEFON HECH QACHON SO'RALMAYDI", block)
        self.assertIn("client_lead_create ni ham\nCHAQIRMAYSAN", block)
        self.assertIn("shu qoida ulardan ustun", block)

    def test_the_forbidden_sentences_are_named(self):
        block = self.migration.NEW_TOP
        for line in ['"Telefon raqamingizni qoldiring"',
                     '"Ism va telefon raqamingizni yozib yuboring"',
                     '"Operatorlarimiz aloqaga chiqib aytishadi"']:
            self.assertIn(line, block)

    def test_a_custom_order_repeats_the_request_and_redirects(self):
        block = self.migration.NEW_CUSTOM
        self.assertIn("O'Z SO'ZI bilan qisqa takrorla", block)
        self.assertIn("business.operator_telegram", block)
        self.assertIn("yuborgan rasmingizdagi guldan", block)
        self.assertIn("Telefon raqami SO'RAMA va client_lead_create CHAQIRMA.", block)

    def test_the_custom_order_no_longer_gathers_details(self):
        block = self.migration.NEW_CUSTOM
        self.assertIn("Ma'lumot yig'ma, ketma-ket savol berma", block)
        self.assertNotIn("buketmi yoki savatmi", block)
        self.assertNotIn("topic       custom_order", block)

    def test_the_when_section_says_never(self):
        self.assertIn("Hech qachon va hech qanday holatda.", self.migration.NEW_WHEN)

    def test_every_patch_applies_once(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "\n\n".join(old for old, _ in self.migration.PATCHES)
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        for marker in self.migration.MARKERS:
            self.assertEqual(row.system_prompt.count(marker), 1, f"takrorlandi yoki tushmadi: {marker[:40]}")
        self.assertNotIn("ISM VA TELEFON FAQAT BUYURTMA UCHUN", row.system_prompt)

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "\n\n".join(old for old, _ in self.migration.PATCHES)
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class APhotoWithMakeItRequestIsACustomOrderTests(TestCase):
    """Rasm + "shu guldan yasab berolislami" katalog qidiruvi emas."""

    def setUp(self):
        self.migration = importlib.import_module("core.migrations.0153_ai_prompt_photo_plus_make_it_is_custom")

    def test_the_phrasings_are_listed(self):
        block = self.migration.INSERT
        for phrase in ["shu guldan yasab berolislami", "shunaqasini yasang", "shu guldan buket qb bering"]:
            self.assertIn(phrase, block)

    def test_the_album_answer_is_named_as_wrong(self):
        block = self.migration.INSERT
        self.assertIn("savolga javob emas", block)
        self.assertIn("mijoz gul so'ramadi, yasab\nberishni so'radi", block)

    def test_the_right_answer_redirects_with_the_photo_named(self):
        block = self.migration.INSERT
        self.assertIn("Yuborgan rasmingizdagi guldan buket", block)
        self.assertIn("@euroflowerspremium", block)

    def test_it_appends_once_after_the_custom_order_block(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        row.system_prompt = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.save()
        for _ in range(2):
            self.migration.apply_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt.count(self.migration.MARKER), 1)
        self.assertLess(row.system_prompt.index(self.migration.ANCHOR), row.system_prompt.index(self.migration.MARKER))

    def test_it_can_be_reverted(self):
        from django.apps import apps as installed_apps
        row = AISettings.objects.get_or_create(pk=1)[0]
        original = "bosh\n" + self.migration.ANCHOR + "\noxir"
        row.system_prompt = original
        row.save()
        self.migration.apply_prompt(installed_apps, None)
        self.migration.revert_prompt(installed_apps, None)
        row.refresh_from_db()
        self.assertEqual(row.system_prompt, original)


class TheMediaInstructionLeavesRoomForACustomOrderTests(TestCase):
    """Tool ko'rsatmasi promptdan ustun keladi, shuning uchun izn o'sha yerda yozildi."""

    def test_both_not_found_instructions_carry_the_note(self):
        from .services import MEDIA_MATCH_NOT_FOUND_INSTRUCTION, MEDIA_MATCH_SIMILAR_INSTRUCTION, MEDIA_MATCH_CUSTOM_ORDER_NOTE
        self.assertIn(MEDIA_MATCH_CUSTOM_ORDER_NOTE, MEDIA_MATCH_NOT_FOUND_INSTRUCTION)
        self.assertIn(MEDIA_MATCH_CUSTOM_ORDER_NOTE, MEDIA_MATCH_SIMILAR_INSTRUCTION)

    def test_the_note_names_the_phrasings_and_the_answer(self):
        from .services import MEDIA_MATCH_CUSTOM_ORDER_NOTE as note
        self.assertIn("yasab berolislami", note)
        self.assertIn("albom YUBORMA", note)
        self.assertIn("yuborgan rasmingizdagi guldan", note)
        self.assertIn("business.operator_telegram", note)

    def test_the_original_instructions_still_say_what_they_said(self):
        from .services import MEDIA_MATCH_NOT_FOUND_INSTRUCTION, MEDIA_MATCH_SIMILAR_INSTRUCTION
        self.assertIn("butun katalogni yubor", MEDIA_MATCH_NOT_FOUND_INSTRUCTION)
        self.assertIn("Telefon raqami SO'RAMA", MEDIA_MATCH_SIMILAR_INSTRUCTION)


class TheGroupMessageCarriesNoMediaLinksTests(TestCase):
    """Rasmlar slideshow bo'lib ketadi, havolalar ro'yxati yozilmaydi.

    Uzun signed CDN havolalari xabarni o'qishga xalaqit qilardi va bir necha
    soatdan keyin baribir ochilmaydi.
    """

    def setUp(self):
        self.customer = Customer.objects.create(name="Ahmad", phone="+998901234567",
                                                instagram_username="link_probe", instagram_user_id="ig-link")
        self.conversation = Conversation.objects.create(customer=self.customer)
        Message.objects.create(
            conversation=self.conversation, sender="customer", text="shu nechpul",
            metadata={"attachments": [{"url": "https://lookaside.fbsbx.com/ig_messaging_cdn/?asset_id=999&signature=Ab1xyz", "kind": "photo"}]},
        )
        self.lead = Lead.objects.create(conversation=self.conversation, customer=self.customer,
                                        request_uz="Mijoz rasm yubordi", source="ai")

    def _html(self):
        from .services import operator_lead_rich_message
        return operator_lead_rich_message(self.lead, self.conversation)

    def test_no_link_list_and_no_anchor_tags(self):
        message = self._html()
        self.assertNotIn("Media havolalar", message["html"])
        self.assertNotIn("<a href=", message["html"])
        self.assertNotIn("lookaside.fbsbx.com", message["html"])

    def test_the_photo_still_travels_as_media(self):
        message = self._html()
        self.assertTrue(message["media"], "rasm slideshow'dan ham tushib qolgan")
        self.assertIn("lookaside.fbsbx.com", message["media"][0]["media"]["media"])
        self.assertIn("tg://photo", message["html"])

    def test_the_rest_of_the_message_is_untouched(self):
        html = self._html()["html"]
        self.assertIn(f"Yangi lead #{self.lead.id}", html)
        self.assertIn("Ahmad", html)
        self.assertIn("+998901234567", html)
        self.assertIn("Mijoz rasm yubordi", html)


class TheGroupMessageCarriesNoCatalogNoteTests(TestCase):
    """Katalog izohi ichki yozuv — guruhdagi xabarga tushmaydi."""

    def setUp(self):
        self.customer = Customer.objects.create(name="Sardor", phone="+998935556677",
                                                instagram_username="izoh_probe", instagram_user_id="ig-izoh")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.item = AICatalogItem.objects.create(
            name="Alfalob Kompazitsia", arrangement_type="bouquet", price=Decimal("1000000"),
            note="ALfalob gulidan yasalgan, 100 ta guli boladi, narxi:1000000 kelishtirilgan narxi 800000",
            image_url="https://cdn.example.com/alfalob.jpg", is_active=True,
        )
        self.lead = Lead.objects.create(
            conversation=self.conversation, customer=self.customer, source="ai",
            request_uz="Mijoz katalogdan tanladi",
            details={"catalog_items": [{"ai_catalog_item": self.item.id, "catalog_name": self.item.name,
                                        "price": "1000000", "quantity": 1}]},
        )

    def test_the_note_is_not_in_the_message(self):
        from .services import operator_lead_rich_message
        html = operator_lead_rich_message(self.lead, self.conversation)["html"]
        self.assertNotIn("kelishtirilgan narxi", html)
        self.assertNotIn("100 ta guli boladi", html)
        self.assertNotIn("<i>", html)

    def test_only_the_catalog_name_stays(self):
        from .services import operator_lead_rich_message
        html = operator_lead_rich_message(self.lead, self.conversation)["html"]
        self.assertIn("Alfalob Kompazitsia", html)
        self.assertNotIn("1 000 000 so&#x27;m", html)
        self.assertIn("Tanlagan mahsuloti", html)

    def test_the_catalog_photo_still_travels(self):
        from .services import operator_lead_rich_message
        message = operator_lead_rich_message(self.lead, self.conversation)
        self.assertIn("https://cdn.example.com/alfalob.jpg",
                      [row["media"]["media"] for row in message["media"]])

    def test_the_helper_no_longer_returns_a_note(self):
        from .services import lead_catalog_lines
        rows = lead_catalog_lines(self.lead)
        self.assertEqual(len(rows), 1)
        self.assertNotIn("note", rows[0])


class PaymentFlowTests(TestCase):
    """To'lov turi, chek va operatorning qarori."""

    def setUp(self):
        from unittest.mock import patch
        BusinessSettings.objects.update_or_create(pk=1, defaults={
            "payment_card_number": "5614 6821 2301 7099", "payment_card_holder": "Toxtasinov Boxodir"})
        self.customer = Customer.objects.create(name="Ahmad", phone="+998901112233",
                                                instagram_username="pay_probe", instagram_user_id="ig-pay")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.lead = Lead.objects.create(conversation=self.conversation, customer=self.customer,
                                        source="ai", request_uz="Buket tanladi")
        self.calls = []
        patcher = patch("core.payment_services.telegram_api_with_token",
                        side_effect=lambda token, method, payload: self.calls.append((method, payload)) or {"ok": True, "result": {"message_id": 500}})
        patcher.start()
        self.addCleanup(patcher.stop)
        settings_patch = override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
        settings_patch.enable()
        self.addCleanup(settings_patch.disable)

    def _methods(self):
        return [method for method, _ in self.calls]

    def test_cash_needs_no_receipt(self):
        from .payment_services import set_payment_type
        result = set_payment_type(self.lead, "cash")
        self.assertTrue(result["ok"])
        self.assertNotIn("card", result)
        self.assertIn("Chek so'rama", result["instruction_uz"])

    def test_card_returns_the_stored_requisites(self):
        from .payment_services import set_payment_type
        result = set_payment_type(self.lead, "card")
        self.assertEqual(result["card"]["number"], "5614 6821 2301 7099")
        self.assertEqual(result["card"]["holder"], "Toxtasinov Boxodir")
        self.assertIn("chekining rasmini", result["instruction_uz"])

    def test_an_unset_card_never_invents_a_number(self):
        from .payment_services import set_payment_type
        BusinessSettings.objects.filter(pk=1).update(payment_card_number="")
        result = set_payment_type(self.lead, "card")
        self.assertEqual(result["card"], {})
        self.assertIn("o'zingdan yozma", result["instruction_uz"])

    def test_a_later_order_never_shows_the_card_number(self):
        """Buyurtma keyingi kunga — model karta raqamini ko'rmaydi ham."""
        from .payment_services import set_payment_type
        self.lead.desired_date = timezone.localdate() + timedelta(days=3)
        self.lead.save(update_fields=["desired_date"])
        result = set_payment_type(self.lead, "card")
        self.assertTrue(result["future_order"])
        self.assertNotIn("card", result)
        self.assertNotIn("5614", json.dumps(result))
        self.assertIn("Karta raqamini BERMA", result["instruction_uz"])
        self.assertIn("chek ham SO'RAMA", result["instruction_uz"].replace("Chek", "chek"))

    def test_a_later_order_asks_the_customer_to_write_again(self):
        from .payment_services import set_payment_type
        self.lead.desired_date = timezone.localdate() + timedelta(days=1)
        self.lead.save(update_fields=["desired_date"])
        for payment_type in ("card", "cash"):
            result = set_payment_type(self.lead, payment_type)
            self.assertIn("yana bir marta", result["instruction_uz"])
            self.assertEqual(result["desired_date"], self.lead.desired_date.isoformat())

    def test_todays_order_still_gets_the_card_and_the_receipt(self):
        """Bugunga bo'lsa oqim o'zgarmaydi."""
        from .payment_services import set_payment_type
        self.lead.desired_date = timezone.localdate()
        self.lead.save(update_fields=["desired_date"])
        result = set_payment_type(self.lead, "card")
        self.assertNotIn("future_order", result)
        self.assertEqual(result["card"]["number"], "5614 6821 2301 7099")
        self.assertIn("chekining rasmini", result["instruction_uz"])

    def test_the_payment_type_is_added_to_the_group_message(self):
        from .payment_services import set_payment_type, save_payment_state
        save_payment_state(self.lead, operator_message_id=77)
        set_payment_type(self.lead, "card")
        method, payload = self.calls[-1]
        self.assertIn(method, {"editMessageCaption", "editMessageText"})
        self.assertEqual(payload["message_id"], 77)
        self.assertIn("💳 Karta", payload.get("caption") or payload.get("text"))

    def test_the_lead_details_survive_the_payment_update(self):
        """Avval tahrir lead matnini o'chirib, o'rniga faqat to'lov qatorini yozardi."""
        from .payment_services import set_payment_type, save_payment_state
        body = "🌸 Yangi lead #1\n\n👤 Ahmad\n📞 +998901112233\n🛍 Luchiana Gulidan Buket — 900 000 so'm"
        save_payment_state(self.lead, operator_message_id=77, operator_body=body,
                           operator_keyboard={"inline_keyboard": [[{"text": "CRM chatni ochish", "url": "https://crm/x"}]]})
        set_payment_type(self.lead, "card")
        _, payload = self.calls[-1]
        written = payload.get("caption") or payload.get("text")
        self.assertIn("👤 Ahmad", written)
        self.assertIn("📞 +998901112233", written)
        self.assertIn("Luchiana Gulidan Buket", written)
        self.assertIn("💳 Karta", written)
        # "CRM chatni ochish" tugmasi ham joyida qoladi.
        self.assertEqual(payload["reply_markup"]["inline_keyboard"][0][0]["text"], "CRM chatni ochish")

    def test_the_lead_message_never_carries_the_payment_buttons(self):
        from .payment_services import register_receipt, save_payment_state
        save_payment_state(self.lead, operator_message_id=77, type="card",
                           operator_body="🌸 Yangi lead #1\n👤 Ahmad",
                           operator_keyboard={"inline_keyboard": [[{"text": "CRM chatni ochish", "url": "https://crm/x"}]]})
        register_receipt(self.lead, "https://cdn.example.com/chek.jpg")
        edits = [p for m, p in self.calls if m in {"editMessageCaption", "editMessageText"}]
        self.assertTrue(edits)
        for payload in edits:
            labels = [b.get("text") for row in (payload.get("reply_markup") or {}).get("inline_keyboard", []) for b in row]
            self.assertNotIn("✅ To'lovni tasdiqlash", labels)
            self.assertNotIn("❌ To'lovni rad etish", labels)
        # Tugmalar faqat alohida yuborilgan chek xabarida.
        photo = [p for m, p in self.calls if m == "sendPhoto"][0]
        photo_labels = [b["text"] for row in photo["reply_markup"]["inline_keyboard"] for b in row]
        self.assertEqual(photo_labels, ["✅ To'lovni tasdiqlash", "❌ To'lovni rad etish"])

    def test_a_photo_is_only_classified_once_an_order_exists(self):
        """Buyurtma bo'lmagan suhbatda chek ham bo'lmaydi — bekorga so'rov ketmaydi."""
        source = Path(__file__).with_name("services.py").read_text(encoding="utf-8")
        self.assertIn('if attachment.get("kind") == "photo" and conversation.leads.exists():', source)

    def test_a_receipt_goes_to_the_group_with_two_buttons(self):
        from .payment_services import register_receipt, save_payment_state
        save_payment_state(self.lead, operator_message_id=77, type="card")
        result = register_receipt(self.lead, "https://cdn.example.com/chek.jpg")
        self.assertTrue(result["ok"])
        self.assertFalse(result["repeated"])
        photo = [p for m, p in self.calls if m == "sendPhoto"][0]
        self.assertEqual(photo["photo"], "https://cdn.example.com/chek.jpg")
        self.assertEqual(photo["reply_to_message_id"], 77)
        self.assertIn("To'landi ✅ chekni tekshirish kerak", photo["caption"])
        labels = [b["text"] for row in photo["reply_markup"]["inline_keyboard"] for b in row]
        self.assertEqual(labels, ["✅ To'lovni tasdiqlash", "❌ To'lovni rad etish"])
        data = [b["callback_data"] for row in photo["reply_markup"]["inline_keyboard"] for b in row]
        self.assertEqual(data, [f"pay:ok:{self.lead.id}", f"pay:no:{self.lead.id}"])

    def test_a_repeated_receipt_is_marked_as_such(self):
        from .payment_services import register_receipt, save_payment_state, RECEIPT_REJECTED
        save_payment_state(self.lead, operator_message_id=77, receipt_status=RECEIPT_REJECTED)
        result = register_receipt(self.lead, "https://cdn.example.com/chek2.jpg")
        self.assertTrue(result["repeated"])
        photo = [p for m, p in self.calls if m == "sendPhoto"][0]
        self.assertIn("Chek qayta yuborildi", photo["caption"])
        self.assertEqual(photo["reply_to_message_id"], 77)

    def test_confirming_tells_the_customer(self):
        from unittest.mock import patch
        from .payment_services import handle_callback, payment_state, RECEIPT_CONFIRMED
        with patch("core.platform_services.instagram_send", return_value={"ok": True}) as send:
            result = handle_callback({"callback_query": {"id": "cb1", "data": f"pay:ok:{self.lead.id}",
                                                         "message": {"message_id": 90, "chat": {"id": -100}}}})
        self.assertTrue(result["approved"])
        self.lead.refresh_from_db()
        self.assertEqual(payment_state(self.lead)["receipt_status"], RECEIPT_CONFIRMED)
        self.assertIn("To'lovingiz tasdiqlandi", send.call_args.args[1])

    def test_rejecting_asks_for_a_real_receipt(self):
        from unittest.mock import patch
        from .payment_services import handle_callback, payment_state, RECEIPT_REJECTED
        with patch("core.platform_services.instagram_send", return_value={"ok": True}) as send:
            handle_callback({"callback_query": {"id": "cb2", "data": f"pay:no:{self.lead.id}",
                                                "message": {"message_id": 90, "chat": {"id": -100}}}})
        self.lead.refresh_from_db()
        self.assertEqual(payment_state(self.lead)["receipt_status"], RECEIPT_REJECTED)
        self.assertIn("qaytadan haqiqiy to'lov chekini yuboring", send.call_args.args[1])

    def test_the_buttons_are_removed_after_a_decision(self):
        from unittest.mock import patch
        from .payment_services import handle_callback
        with patch("core.platform_services.instagram_send", return_value={"ok": True}):
            handle_callback({"callback_query": {"id": "cb3", "data": f"pay:ok:{self.lead.id}",
                                                "message": {"message_id": 90, "chat": {"id": -100}}}})
        markup = [p for m, p in self.calls if m == "editMessageReplyMarkup"]
        self.assertTrue(markup)
        self.assertEqual(markup[-1]["reply_markup"], {"inline_keyboard": []})

    def test_a_stray_callback_is_ignored(self):
        from .payment_services import handle_callback
        self.assertFalse(handle_callback({"callback_query": {"data": "something:else"}})["ok"])
        self.assertFalse(handle_callback({"callback_query": {"data": "pay:ok:not-a-number"}})["ok"])

    def test_the_tool_refuses_when_there_is_no_lead(self):
        empty = Conversation.objects.create(customer=Customer.objects.create(
            instagram_username="no_lead", instagram_user_id="ig-nolead"))
        result = execute_ai_tool("client_payment_update", {"payment_type": "card", "receipt_url": None}, empty)
        self.assertEqual(result["detail"], "no_lead_yet")

    def test_the_context_says_whether_the_payment_type_is_already_known(self):
        from .services import lead_payment_type
        from .payment_services import set_payment_type
        self.assertEqual(lead_payment_type(self.lead), "")
        self.assertEqual(lead_payment_type(None), "")
        set_payment_type(self.lead, "card")
        self.assertEqual(lead_payment_type(self.lead), "card")


class ReplyToOurPhotoTests(TestCase):
    """Mijoz yuborgan albomdagi rasmga reply qilib savol bersa."""

    def setUp(self):
        self.customer = Customer.objects.create(instagram_user_id="ig-reply", instagram_username="shukhr")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.savat = AICatalogItem.objects.create(
            name="Katalina Gulidan Savat Kompazitsia", arrangement_type="basket", price=800000,
            quantity=1, image_url="https://cdn.example.com/katalina.jpg",
            note="narxi 800000 kelishtirilgan narxi 700000")
        self.buket = AICatalogItem.objects.create(
            name="Luchiana Gulidan Buket", arrangement_type="bouquet", price=1000000,
            quantity=1, image_url="https://cdn.example.com/luchiana.jpg")

    def _album(self, message_id, items):
        self.conversation.messages.create(sender="system", text="", metadata={"catalog_album_result": {
            "ok": True,
            "items": [{"catalog_id": row.id, "name": row.name, "price": str(row.price)} for row in items],
            "sent_message_ids": [message_id],
            "sent_groups": [{"message_id": message_id, "catalog_ids": [row.id for row in items]}],
        }})

    def test_a_reply_to_one_photo_names_that_product(self):
        from .services import replied_to_note
        self._album("mid-album-1", [self.savat])
        note = replied_to_note(self.conversation, "mid-album-1")
        self.assertIn("Katalina Gulidan Savat Kompazitsia", note)
        self.assertIn("800 000", note)

    def test_a_reply_to_a_whole_album_asks_which_one(self):
        from .services import replied_to_note
        self._album("mid-album-2", [self.savat, self.buket])
        note = replied_to_note(self.conversation, "mid-album-2")
        self.assertIn("Katalina Gulidan Savat Kompazitsia", note)
        self.assertIn("Luchiana Gulidan Buket", note)
        self.assertIn("Qaysi biri", note)

    def test_a_reply_to_a_single_catalog_image_is_resolved(self):
        from .services import replied_to_note
        self.conversation.messages.create(sender="system", text="", metadata={"image_tool_result": {
            "catalog_id": self.savat.id, "catalog_name": self.savat.name,
            "sent": {"message_id": "mid-single-1"}}})
        self.assertIn("Katalina", replied_to_note(self.conversation, "mid-single-1"))

    def test_a_reply_to_something_we_never_sent_adds_nothing(self):
        from .services import replied_to_note
        self._album("mid-album-3", [self.savat])
        self.assertEqual(replied_to_note(self.conversation, "mid-somebody-else"), "")
        self.assertEqual(replied_to_note(self.conversation, ""), "")

    def test_the_webhook_puts_the_product_into_the_message(self):
        self._album("mid-album-4", [self.savat])
        payload = {"entry": [{"messaging": [{
            "sender": {"id": "ig-reply"}, "recipient": {"id": "ig-business"},
            "message": {"mid": "mid-customer-1", "text": "неч пул кберасила клиентлага",
                        "reply_to": {"mid": "mid-album-4", "is_self_reply": False}},
        }]}]}
        resolve_instagram_event(payload)
        saved = self.conversation.messages.filter(sender="customer").order_by("-id").first()
        self.assertIn("неч пул кберасила", saved.text)
        self.assertIn("Katalina Gulidan Savat Kompazitsia", saved.text)
        self.assertEqual(saved.metadata["instagram_reply_to_mid"], "mid-album-4")

    def test_a_story_reply_is_left_to_the_story_path(self):
        payload = {"entry": [{"messaging": [{
            "sender": {"id": "ig-reply"}, "recipient": {"id": "ig-business"},
            "message": {"mid": "mid-customer-2", "text": "shu nechpul",
                        "reply_to": {"story": {"url": "https://lookaside.fbsbx.com/x?asset_id=1"}}},
        }]}]}
        from unittest.mock import patch
        with patch("core.webhook_services.find_active_story_by_media_url", return_value={}):
            resolve_instagram_event(payload)
        saved = self.conversation.messages.filter(sender="customer").order_by("-id").first()
        self.assertNotIn("javob qildi", saved.text)


class OurOwnEchoIsNotAnOperatorTests(TestCase):
    """Biz yuborgan xabar echo bo'lib qaytganda suhbat operatorga o'tmasligi kerak."""

    def setUp(self):
        self.customer = Customer.objects.create(instagram_user_id="ig-echo-loc", instagram_username="ahmad")
        self.conversation = Conversation.objects.create(customer=self.customer)

    def test_the_location_reply_records_its_message_id(self):
        from unittest.mock import patch
        from .tasks import deliver_ai_reply
        reply = self.conversation.messages.create(sender="ai", text="Manzilingizni oldik.")
        with patch("core.tasks.instagram_send", return_value={"message_id": "mid-loc-echo"}):
            deliver_ai_reply(self.conversation, reply)
        reply.refresh_from_db()
        self.assertEqual(reply.instagram_message_id, "mid-loc-echo")

    def test_that_echo_is_recognised_and_the_ai_keeps_the_chat(self):
        from unittest.mock import patch
        from .tasks import deliver_ai_reply
        reply = self.conversation.messages.create(sender="ai", text="Manzilingizni oldik.")
        with patch("core.tasks.instagram_send", return_value={"message_id": "mid-loc-echo-2"}):
            deliver_ai_reply(self.conversation, reply)
        payload = {"entry": [{"messaging": [{
            "sender": {"id": "ig-business"}, "recipient": {"id": "ig-echo-loc"},
            "message": {"mid": "mid-loc-echo-2", "text": "Manzilingizni oldik.", "is_echo": True},
        }]}]}
        with patch("core.webhook_services.instagram_account_token_pairs", return_value={"ig-business": "tok"}):
            resolve_instagram_event(payload)
        self.conversation.refresh_from_db()
        self.assertEqual(self.conversation.status, "ai")
        self.assertIsNone(self.conversation.ai_paused_until)
        self.assertFalse(self.conversation.messages.filter(sender="operator").exists())

    def test_the_payment_decision_message_records_its_id_too(self):
        from unittest.mock import patch
        from .payment_services import notify_customer
        lead = Lead.objects.create(conversation=self.conversation, customer=self.customer,
                                   source="ai", request_uz="Buket")
        with patch("core.platform_services.instagram_send", return_value={"message_id": "mid-pay-echo"}):
            self.assertTrue(notify_customer(lead, "To'lovingiz tasdiqlandi."))
        saved = self.conversation.messages.filter(sender="ai").order_by("-id").first()
        self.assertEqual(saved.instagram_message_id, "mid-pay-echo")


class NoFloristFeeForTheCustomerTests(TestCase):
    """Katalog narxiga florist haqi qo'shilmaydi."""

    @override_settings(OPENAI_API_KEY="test-key")
    def test_the_context_hands_the_ai_no_florist_fee(self):
        from unittest.mock import patch
        BusinessSettings.objects.update_or_create(pk=1, defaults={"default_florist_fee": Decimal("50000")})
        customer = Customer.objects.create(instagram_user_id="ig-nofee", name="Ahmad", phone="+998901112233")
        conversation = Conversation.objects.create(customer=customer)
        conversation.messages.create(sender="customer", text="nechpul to'lashim kerak")
        payload = {"reply": "199 000 so'm", "detected_language": "uz", "customer_name": None, "phone": None,
                   "lead_ready": False, "lead_request": None, "arrangement_type": None, "estimated_price": None,
                   "handoff": False, "catalog_items": [], "stock_items": []}
        with patch("core.services.OpenAI") as openai_class:
            client = openai_class.return_value
            client.responses.create.return_value = SimpleNamespace(output_text=json.dumps(payload), output=[], id="r1")
            ai_reply(conversation)
        sent = client.responses.create.call_args.kwargs["input"][0]["content"]
        context = json.loads(sent.split("REAL_CONTEXT_JSON:\n", 1)[1])
        self.assertNotIn("florist_fee", context["business"])
        # Yetkazib berish narxi qoladi — u mijozga aytiladigan haqiqiy summa.
        self.assertEqual(context["business"]["delivery_fee"], "50000.00")
        self.assertNotIn("florist", json.dumps(context))

    def test_the_lead_tool_no_longer_takes_a_florist_fee(self):
        from .services import ai_tool_definitions
        for tool in ai_tool_definitions():
            name = tool.get("name") or (tool.get("function") or {}).get("name")
            if name not in {"client_lead_create", "client_lead_edit"}:
                continue
            schema = tool.get("parameters") or (tool.get("function") or {}).get("parameters") or {}
            self.assertNotIn("florist_fee", schema.get("properties") or {})
            self.assertNotIn("florist_fee", schema.get("required") or [])

    def test_the_prompt_forbids_adding_it(self):
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("FLORIST HAQI HECH QAYERGA QO'SHILMAYDI", prompt)
        self.assertIn("QAT'IY TAQIQLANADI", prompt)
        self.assertNotIn("Kontekstdagi florist_fee ni ham", prompt)

    def test_the_prompt_takes_the_contact_before_the_delivery_choice(self):
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("mahsulot → sana → ISM VA TELEFON → yetkazib berish/kelib olish", prompt)
        self.assertIn("Manzilingizni xaritada belgilang", prompt)


class FreshnessAnswerPromptTests(TestCase):
    """«качон ясалган» savoli operatorga emas, javobga boradi."""

    def test_the_prompt_answers_when_it_was_made(self):
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("Gullarimiz har doim yangi", prompt)
        self.assertIn("качон ясалган", prompt)
        self.assertNotIn('"Qachon yasalgan", "necha kun turadi", "diametri qancha"', prompt)

    def test_the_reply_block_is_in_the_prompt(self):
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("00K. MIJOZ RASMGA JAVOB QILSA", prompt)
        self.assertIn("mijoz shu mahsulot rasmiga javob qildi", prompt)


class CaptionPriceTests(TestCase):
    """Reel izohidagi narx bo'yicha katalog."""

    def test_prices_are_read_in_every_shape_operators_write(self):
        from .services import prices_from_caption
        self.assertEqual(prices_from_caption("Narxi 199 000 so'm"), [Decimal("199000")])
        self.assertEqual(prices_from_caption("1.600.000 сум"), [Decimal("1600000")])
        self.assertEqual(prices_from_caption("800 ming"), [Decimal("800000")])
        self.assertEqual(prices_from_caption("Buket 250000"), [Decimal("250000")])

    def test_small_and_silly_numbers_are_skipped(self):
        from .services import prices_from_caption
        self.assertEqual(prices_from_caption("100 ta gul, 60 sm"), [])
        self.assertEqual(prices_from_caption("2026-yil 25-avgust"), [])

    def test_the_catalog_is_filtered_near_that_price(self):
        from .services import catalog_items_near_price
        cheap = AICatalogItem.objects.create(name="Arzon", arrangement_type="bouquet", price=Decimal("199000"))
        close = AICatalogItem.objects.create(name="Yaqin", arrangement_type="bouquet", price=Decimal("210000"))
        far = AICatalogItem.objects.create(name="Uzoq", arrangement_type="bouquet", price=Decimal("900000"))
        rows = catalog_items_near_price([cheap, close, far], Decimal("200000"))
        self.assertEqual([row.id for row in rows], [cheap.id, close.id])

    def test_a_caption_we_cannot_read_changes_nothing(self):
        from unittest.mock import patch
        from .services import caption_price_matches
        with patch("core.services.media_caption_for_attachment", return_value=""):
            rows, price = caption_price_matches([], {"url": "https://www.instagram.com/reel/AAA/", "kind": "reel"})
        self.assertEqual((rows, price), ([], None))


class RecallOnTheRequestedDateTests(TestCase):
    """Mijoz sanani aytsa o'sha kun ertalab 9:00 ga eslatma qo'yiladi."""

    def setUp(self):
        self.customer = Customer.objects.create(name="Ahmad", phone="+998901112233",
                                                instagram_username="recall_probe", instagram_user_id="ig-recall")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.item = AICatalogItem.objects.create(name="Alfalob Buket", arrangement_type="bouquet",
                                                 price=Decimal("800000"), is_active=True)

    def _lead(self, **extra):
        data = {"conversation": self.conversation, "customer": self.customer, "source": "ai",
                "request_uz": "Mijoz buket tanladi"}
        data.update(extra)
        return Lead.objects.create(**data)

    def test_the_recall_lands_at_nine_in_the_morning(self):
        from datetime import date
        from .recall_services import schedule_from_desired_date
        lead = self._lead(desired_date=date(2026, 9, 3))
        moment = schedule_from_desired_date(lead)
        local = timezone.localtime(moment)
        self.assertEqual((local.year, local.month, local.day), (2026, 9, 3))
        self.assertEqual((local.hour, local.minute), (9, 0))
        lead.refresh_from_db()
        self.assertEqual(lead.recall_at, moment)

    def test_no_date_means_no_recall(self):
        from .recall_services import schedule_from_desired_date
        self.assertIsNone(schedule_from_desired_date(self._lead()))

    def test_an_operator_time_is_never_overwritten(self):
        from datetime import date
        from .recall_services import schedule_from_desired_date
        chosen = timezone.now()
        lead = self._lead(desired_date=date(2026, 9, 3), recall_at=chosen)
        self.assertIsNone(schedule_from_desired_date(lead))
        lead.refresh_from_db()
        self.assertEqual(lead.recall_at, chosen)

    def test_the_card_carries_everything_the_operator_needs(self):
        from datetime import date
        from .recall_services import recall_card
        self.conversation.messages.create(sender="customer", text="Ertaga kerak edi")
        lead = self._lead(desired_date=date(2026, 9, 3), desired_time="14:00",
                          details={"catalog_items": [{"ai_catalog_item": self.item.id,
                                                      "catalog_name": self.item.name,
                                                      "price": "800000", "quantity": 1}]})
        card = recall_card(lead)
        self.assertIn(f"Lead #{lead.id}", card)
        self.assertIn("Ahmad", card)
        self.assertIn("+998901112233", card)
        self.assertIn("@recall_probe", card)
        self.assertIn("Alfalob Buket", card)
        self.assertIn("03.09.2026", card)
        self.assertIn("14:00", card)
        self.assertIn("Yozgan:", card)

    def test_a_custom_order_shows_the_customers_own_words(self):
        from .recall_services import recall_card
        lead = self._lead(details={"flowers_text": "Jumila pushti atirgul", "size_text": "51 dona"})
        self.assertIn("Jumila pushti atirgul · 51 dona", recall_card(lead))


class RecallGroupSurvivesASupergroupUpgradeTests(TestCase):
    """Oddiy guruh superguruhga o'tsa id o'zgaradi — yangi id eslab qolinadi."""

    def setUp(self):
        patcher = override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_RECALL_GROUP_ID="-5385608916")
        patcher.enable()
        self.addCleanup(patcher.disable)

    def test_the_configured_id_is_used_first(self):
        from .recall_services import recall_group_id
        self.assertEqual(recall_group_id(), "-5385608916")

    def test_a_remembered_id_wins(self):
        from .recall_services import recall_group_id, remember_group_id
        remember_group_id("-1005385608916")
        self.assertEqual(recall_group_id(), "-1005385608916")

    def test_a_migration_error_is_retried_with_the_new_id(self):
        from unittest.mock import patch
        from types import SimpleNamespace
        from .recall_services import send_to_group, recall_group_id

        class Boom(Exception):
            def __init__(self):
                self.response = SimpleNamespace(json=lambda: {"parameters": {"migrate_to_chat_id": -1009999}})

        calls = []

        def fake(token, method, payload):
            calls.append(payload["chat_id"])
            if len(calls) == 1:
                raise Boom()
            return {"ok": True}

        with patch("core.recall_services.telegram_api_with_token", fake):
            result = send_to_group("tok", "-5385608916", {"text": "salom"})
        self.assertTrue(result["ok"])
        self.assertEqual(calls, ["-5385608916", -1009999])
        self.assertEqual(recall_group_id(), "-1009999")

    def test_an_ordinary_failure_is_not_retried(self):
        from unittest.mock import patch
        from .recall_services import send_to_group
        with patch("core.recall_services.telegram_api_with_token", side_effect=RuntimeError("tarmoq")):
            self.assertEqual(send_to_group("tok", "-1", {"text": "x"})["detail"], "send_failed")

    def test_nothing_is_sent_without_a_group(self):
        from .recall_services import send_to_group
        self.assertEqual(send_to_group("tok", "", {"text": "x"})["detail"], "recall_group_not_configured")


class TheCustomerNeverGetsATelegramHandleTests(TestCase):
    """Username butunlay olib tashlandi, o'rniga operator chaqiriladi."""

    def setUp(self):
        self.customer = Customer.objects.create(name="Sardor", phone="+998935556677",
                                                instagram_username="op_probe", instagram_user_id="ig-op")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.conversation.messages.create(sender="customer", text="Kelin buket kerak edi")

    def test_the_ready_phrase_has_no_handle(self):
        from .services import operator_telegram_text
        for handle in ["@euroflowerspremium", "", None]:
            text = operator_telegram_text(handle)
            self.assertEqual(text, "Operatorlarimiz sizga tez orada yozib yuborishadi")
            self.assertNotIn("@", text)

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_calling_an_operator_notifies_the_group_with_a_chat_button(self):
        from unittest.mock import patch
        with patch("core.services.telegram_send_with", return_value={"ok": True}) as send:
            result = execute_ai_tool("call_operator", {"reason": "Kelin buketi so'raldi"}, self.conversation)
        self.assertTrue(result["ok"])
        body = send.call_args.args[2]
        self.assertIn("🙋 Operator kerak", body)
        self.assertIn("Sardor", body)
        self.assertIn("+998935556677", body)
        self.assertIn("@op_probe", body)
        self.assertIn("Kelin buket kerak edi", body)
        self.assertIn("Kelin buketi so'raldi", body)
        keyboard = send.call_args.kwargs["reply_markup"]["inline_keyboard"][0][0]
        self.assertEqual(keyboard["text"], "CRM chatni ochish")
        self.assertIn(str(self.conversation.id), keyboard["url"])

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_the_instruction_forbids_a_handle_and_a_phone(self):
        from unittest.mock import patch
        with patch("core.services.telegram_send_with", return_value={"ok": True}):
            result = execute_ai_tool("call_operator", {"reason": "x"}, self.conversation)
        self.assertIn("operatorlarimiz sizga tez orada yozib yuborishadi", result["instruction_uz"])
        self.assertIn("Telegram username BERMA", result["instruction_uz"])

    def test_the_tool_is_in_the_list(self):
        self.assertIn("call_operator", [tool["name"] for tool in ai_tool_definitions()])

    def test_the_prompt_block_spells_the_rule_out(self):
        migration = importlib.import_module("core.migrations.0157_ai_prompt_call_operator")
        block = migration.BLOCK
        self.assertIn("MIJOZGA TELEGRAM USERNAME BERILMAYDI", block)
        self.assertIn("Operatorlarimiz sizga tez orada yozib yuborishadi", block)
        self.assertIn("call_operator NI CHAQIRASAN", block)
        self.assertIn("Gap yozib, tool chaqirmaslik XATO", block)


class DeliveryLocationTests(TestCase):
    """Mijoz xaritada belgilagan manzil."""

    def setUp(self):
        self.customer = Customer.objects.create(name="Ahmad", phone="+998901112233",
                                                instagram_username="loc_probe", instagram_user_id="ig-loc")
        self.conversation = Conversation.objects.create(customer=self.customer)
        self.lead = Lead.objects.create(conversation=self.conversation, customer=self.customer,
                                        source="ai", request_uz="Buket tanladi")
        self.client = APIClient()

    def _post(self, **overrides):
        from .location_services import ensure_token
        body = {"lead_id": self.lead.id, "token": ensure_token(self.lead),
                "latitude": "41.2995000000", "longitude": "69.2401000000"}
        body.update(overrides)
        return self.client.post("/api/delivery-location/", body, format="json")

    # --- havola ---

    def test_the_token_is_made_once_and_kept(self):
        from .location_services import ensure_token
        first = ensure_token(self.lead)
        self.lead.refresh_from_db()
        self.assertEqual(ensure_token(self.lead), first)
        self.assertTrue(len(first) >= 8)

    @override_settings(DELIVERY_LOCATION_URL="https://front.uz/loc/{lead_id}?t={token}")
    def test_the_link_carries_the_lead_and_the_token(self):
        from .location_services import location_link, ensure_token
        link = location_link(self.lead)
        self.assertIn(f"/loc/{self.lead.id}?t=", link)
        self.assertIn(ensure_token(self.lead), link)

    @override_settings(DELIVERY_LOCATION_URL="")
    def test_an_unset_url_gives_no_link(self):
        from .location_services import location_link
        self.assertEqual(location_link(self.lead), "")

    @override_settings(DELIVERY_LOCATION_URL="https://front.uz/loc/{lead_id}?t={token}")
    def test_a_token_with_url_characters_is_escaped(self):
        from .location_services import location_link, save_location_state
        save_location_state(self.lead, token="A b-_=+/x")
        self.assertEqual(location_link(self.lead),
                         "https://front.uz/loc/%d?t=A%%20b-_%%3D%%2B%%2Fx" % self.lead.id)

    def test_a_nudged_pin_is_not_a_new_address(self):
        from .location_services import point_moved
        here = {"latitude": "41.2995", "longitude": "69.2401"}
        self.assertFalse(point_moved(here, "41.29952", "69.24012"))
        self.assertTrue(point_moved(here, "41.3100", "69.2401"))
        self.assertTrue(point_moved({}, "41.2995", "69.2401"))

    @override_settings(DELIVERY_LOCATION_URL="https://front.uz/loc/{lead_id}?t={token}")
    def test_the_tool_hands_the_link_to_the_ai(self):
        result = execute_ai_tool("delivery_location_link", {}, self.conversation)
        self.assertTrue(result["ok"])
        self.assertIn("front.uz/loc/", result["link"])
        self.assertIn("aynan shu ko'rinishda", result["instruction_uz"])

    def test_the_tool_refuses_without_a_lead(self):
        empty = Conversation.objects.create(customer=Customer.objects.create(instagram_user_id="ig-empty"))
        self.assertEqual(execute_ai_tool("delivery_location_link", {}, empty)["detail"], "no_lead_yet")

    @override_settings(DELIVERY_LOCATION_URL="")
    def test_an_unset_url_tells_the_ai_to_ask_in_text(self):
        result = execute_ai_tool("delivery_location_link", {}, self.conversation)
        self.assertFalse(result["ok"])
        self.assertIn("matn bilan", result["instruction_uz"])

    # --- API ---

    def test_a_good_request_is_accepted(self):
        from unittest.mock import patch
        with patch("core.location_services.send_location_to_group", return_value={"ok": True}), \
             patch("core.tasks.process_location_reply.delay") as later:
            response = self._post(address="Chilonzor 5, 12-uy")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "OK")
        self.lead.refresh_from_db()
        state = (self.lead.details or {})["location"]
        self.assertEqual(state["latitude"], "41.2995000000")
        self.assertEqual(state["longitude"], "69.2401000000")
        self.assertEqual(self.lead.delivery_address, "Chilonzor 5, 12-uy")
        self.assertTrue(later.called)

    def test_a_wrong_token_is_refused(self):
        response = self._post(token="notatoken")
        self.assertEqual(response.status_code, 403)
        self.lead.refresh_from_db()
        self.assertNotIn("latitude", (self.lead.details or {}).get("location", {}))

    def test_an_unknown_lead_is_skipped(self):
        response = self._post(lead_id=99999999, token="whatever")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "SKIPPED")

    def test_broken_coordinates_are_refused(self):
        self.assertEqual(self._post(latitude="200").status_code, 400)
        self.assertEqual(self._post(longitude="abc").status_code, 400)

    def test_pressing_select_again_on_the_same_spot_stays_quiet(self):
        from unittest.mock import patch
        with patch("core.location_services.send_location_to_group", return_value={"ok": True}) as to_group, \
             patch("core.tasks.process_location_reply.delay") as later:
            self.assertEqual(self._post().json()["status"], "OK")
            self.assertEqual(self._post().json()["status"], "OK")
            self.assertEqual(self._post(latitude="41.2995200000").json()["status"], "OK")
        self.assertEqual(to_group.call_count, 1)
        self.assertEqual(later.call_count, 1)

    def test_a_moved_pin_corrects_the_group_without_a_second_ai_reply(self):
        from unittest.mock import patch
        from .location_services import accept_location, ensure_token
        token = ensure_token(self.lead)
        with patch("core.location_services.send_location_to_group", return_value={"ok": True}) as to_group, \
             patch("core.tasks.process_location_reply.delay") as later:
            accept_location(self.lead.id, token, "41.2995000000", "69.2401000000")
            moved = accept_location(self.lead.id, token, "41.3200000000", "69.2401000000")
        self.assertEqual(moved["detail"], "updated")
        self.assertEqual(to_group.call_count, 2)
        self.assertTrue(to_group.call_args.kwargs["updated"])
        self.assertEqual(later.call_count, 1)
        self.lead.refresh_from_db()
        self.assertEqual((self.lead.details or {})["location"]["latitude"], "41.3200000000")

    def test_a_corrected_location_is_labelled_in_the_group(self):
        from .location_services import location_caption, save_location_state
        save_location_state(self.lead, latitude="41.2995", longitude="69.2401")
        self.assertIn("Yetkazib berish manzili", location_caption(self.lead))
        self.assertIn("Manzil yangilandi", location_caption(self.lead, updated=True))

    def test_a_typed_address_is_not_overwritten(self):
        from unittest.mock import patch
        self.lead.delivery_address = "Mijoz o'zi yozgan manzil"
        self.lead.save(update_fields=["delivery_address"])
        with patch("core.location_services.send_location_to_group", return_value={"ok": True}), \
             patch("core.tasks.process_location_reply.delay"):
            self._post(address="Xaritadan kelgan manzil")
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.delivery_address, "Mijoz o'zi yozgan manzil")
        self.assertEqual((self.lead.details or {})["location"]["address"], "Xaritadan kelgan manzil")

    # --- guruh va suhbat ---

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_the_location_replies_to_the_lead_message(self):
        from unittest.mock import patch
        from .location_services import save_location_state, send_location_to_group
        from .payment_services import save_payment_state
        save_payment_state(self.lead, operator_message_id=321)
        save_location_state(self.lead, latitude="41.2995", longitude="69.2401", address="Chilonzor 5")
        calls = []
        with patch("core.location_services.telegram_api_with_token_and_chat_fallback",
                   side_effect=lambda t, m, p: calls.append((m, p)) or {"ok": True, "result": {"message_id": len(calls)}}):
            self.assertTrue(send_location_to_group(self.lead)["ok"])
        method, payload = calls[0]
        self.assertEqual(method, "sendLocation")
        self.assertEqual(payload["reply_to_message_id"], 321)
        self.assertAlmostEqual(payload["latitude"], 41.2995, places=4)
        self.assertAlmostEqual(payload["longitude"], 69.2401, places=4)
        note = calls[1][1]
        self.assertIn(f"Lead #{self.lead.id}", note["text"])
        self.assertIn("Chilonzor 5", note["text"])
        self.assertEqual(note["reply_to_message_id"], 321)

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_location_send_retries_then_records_message_id(self):
        from unittest.mock import patch
        from .location_services import save_location_state, send_location_to_group
        save_location_state(self.lead, latitude="41.2995", longitude="69.2401")
        calls = []
        def sender(token, method, payload):
            calls.append(method)
            if len(calls) == 1:
                raise ConnectionError("reset")
            return {"ok": True, "result": {"message_id": 777 + len(calls)}}
        with patch("core.location_services.time.sleep"), \
                patch("core.location_services.telegram_api_with_token_and_chat_fallback", side_effect=sender):
            self.assertTrue(send_location_to_group(self.lead)["ok"])
        self.lead.refresh_from_db()
        state = (self.lead.details or {})["location"]
        self.assertEqual(state["operator_group_location"]["message_id"], 779)
        self.assertGreaterEqual(calls.count("sendLocation"), 2)

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_same_location_is_not_sent_twice_to_group(self):
        from unittest.mock import patch
        from .location_services import save_location_state, send_location_to_group
        save_location_state(
            self.lead,
            latitude="41.2995000000",
            longitude="69.2401000000",
            operator_group_location={"message_id": 555, "latitude": "41.2995", "longitude": "69.2401"},
        )
        with patch("core.location_services.telegram_api_with_token_and_chat_fallback") as sender:
            result = send_location_to_group(self.lead)
        self.assertEqual(result["detail"], "already_sent")
        self.assertEqual(sender.call_count, 0)

    @override_settings(AI_OPERATOR_HANDOFF_BOT_TOKEN="tok", AI_OPERATOR_HANDOFF_GROUP_ID="-100")
    def test_without_coordinates_nothing_is_sent(self):
        from .location_services import send_location_to_group
        self.assertEqual(send_location_to_group(self.lead)["detail"], "no_coordinates")

    def test_the_location_lands_in_the_chat_as_a_customer_message(self):
        from .location_services import save_location_state, record_customer_location_message
        save_location_state(self.lead, latitude="41.2995", longitude="69.2401", token="secret")
        message = record_customer_location_message(self.lead, "Chilonzor 5")
        self.assertEqual(message.sender, "customer")
        self.assertIn("xaritada belgiladi", message.text)
        self.assertIn("Chilonzor 5", message.text)
        # Maxfiy kod suhbatga yozilmaydi.
        self.assertNotIn("secret", json.dumps(message.metadata))

    def test_the_prompt_block_covers_the_flow(self):
        migration = importlib.import_module("core.migrations.0158_ai_prompt_delivery_location")
        block = migration.BLOCK
        self.assertIn("delivery_location_link", block)
        self.assertIn("AYNAN o'sha", block)
        self.assertIn("MATN MANZILNI HAM QABUL QIL", block)
        self.assertIn("Manzilingizni oldik", block)
        self.assertIn("Koordinatani mijozga o'qib berma", block)


class RussianReplyLanguageTests(TestCase):
    """Ruscha yozgan mijozga o'zbekcha so'z ketmasin."""

    def setUp(self):
        self.item = AICatalogItem.objects.create(
            name="London Gulidan Savat Kompazitsia", arrangement_type="basket",
            price=1500000, quantity=1, image_url="https://cdn.example.com/london.jpg")

    def test_the_name_is_rebuilt_in_russian(self):
        from .services import catalog_name_ru
        self.assertEqual(catalog_name_ru("London Gulidan Savat Kompazitsia"),
                         "Корзина-композиция из London")
        self.assertEqual(catalog_name_ru("Luchiana Gulidan Buket"), "Букет из Luchiana")
        self.assertEqual(catalog_name_ru("Qizil Atirgul Gulidan Karobka"),
                         "Коробка из красных роз")

    def test_the_messy_real_names_come_out_in_russian(self):
        """Katalogdagi nomlar erkin yozilgan — hammasi ruscha chiqishi kerak."""
        from .services import catalog_name_ru
        self.assertEqual(catalog_name_ru("Oq Atir Guldan Kompazitsia"),
                         "Композиция из белых роз")
        self.assertEqual(catalog_name_ru("Qizil Va Oq Atir Guldan Kompazitsia"),
                         "Композиция из красных и белых роз")
        self.assertEqual(catalog_name_ru("Aziza Va Luchiana Gulidan Buket"),
                         "Букет из Aziza и Luchiana")
        self.assertEqual(catalog_name_ru("Buket Kotta Shoxli Bambastic Gulidan Yasalgan"),
                         "Большой ветвистый Букет из Bambastic")
        self.assertEqual(catalog_name_ru("Oq Jumila Atir Gulidan Yasalgan Kompazitsia 100 Tali"),
                         "Композиция из 100 шт белых роз Jumila")
        self.assertEqual(catalog_name_ru("Hermossodan Kompazitsia"), "Композиция из Hermosso")
        self.assertEqual(catalog_name_ru("Savat Kompazitsa"), "Корзина-композиция")

    def test_no_uzbek_word_survives_any_real_catalog_name(self):
        import re
        from .services import catalog_name_ru
        forbidden = {"gul", "gullar", "gulli", "gulidan", "guldan", "gulimizdan",
                     "gullarimiz", "yasalgan", "savat", "savatli", "buket", "quti",
                     "kompazitsia", "kompozitsiyasi", "kompazitsa", "oq", "qizil",
                     "katta", "kotta", "shoxli", "atir", "atirguldan", "ta", "tali", "va"}
        names = [
            "London Gulidan Savat Kompazitsia", "All For Love Gulidan Kompazitsia",
            "Oq Atir Guldan Kompazitsia", "Hermossodan Kompazitsia",
            "Sendi Avalanch Gulidan Buket", "Aziza Va Luchiana Gulidan Buket",
            "Jumila Va Qizil Atir Guldan Kompazitsia", "London Va Oq Atirguldan Kompazitsia",
            "Jumilia Kompozitsiyasi", "Katalina/bables Kompazitsia", "Buket Bambastic",
            "Katalina Bables Gulidan Yasalgan Kompazitsia", "Savat Bables Gulidan",
            "Bables Gulidan Savatli Kompazitsia", "London Gulidan Kompazitsia Savat",
            "Savat Jumila Oq Atir Guldan Yasalgan Kompazitsia", "Shoxli Bambastic",
            "Buket Aziza Gulidan Yasalgan Kompazitsia 100 Ta Gulli",
            "London Gulimizdan Kompazitsia 100 Ta Gulli",
            "Qizil Atir Guldan Kompazitsia 100 Tali Gullarimiz",
            "Alfalob Gulidan Katta Kompazitsia 100 Tali Gul",
            "Buket Jumila Va Oq Atir Guldan Yasalgan Kompazitsia",
        ]
        for name in names:
            russian = catalog_name_ru(name)
            self.assertTrue(russian, name)
            left = {word.lower() for word in re.findall(r"[A-Za-z']+", russian)} & forbidden
            self.assertEqual(left, set(), f"{name} -> {russian}")

    def test_uzbek_cyrillic_never_reaches_a_russian_reply(self):
        """Real suhbatda "Оқ Жумила" ruscha javob ichida qolib ketgan."""
        from .services import catalog_name_ru
        name = catalog_name_ru("Оқ Жумила")
        self.assertEqual(name, "Белая Jumila")
        self.assertFalse(set(name) & set("ўқғҳЎҚҒҲ"))

    def test_the_currency_word_follows_the_language(self):
        from .services import money_text
        self.assertEqual(money_text(1500000, "ru"), "1 500 000 сум")
        self.assertEqual(money_text(1500000, "latin"), "1 500 000 so'm")
        self.assertEqual(money_text(1500000, "uz_cyril"), "1 500 000 so'm")

    def test_the_catalog_rows_carry_the_russian_name_and_price(self):
        from .services import ai_catalog_rows
        row = next(row for row in ai_catalog_rows("") if row["catalog_id"] == self.item.id)
        self.assertEqual(row["name_ru"], "Корзина-композиция из London")
        self.assertEqual(row["price_text_ru"], "1 500 000 сум")

    def _album_captions(self, texts):
        from unittest.mock import patch
        customer = Customer.objects.create(instagram_user_id="ig-ru-album", instagram_username="asadnabiev")
        conversation = Conversation.objects.create(customer=customer)
        for text in texts:
            conversation.messages.create(sender="customer", text=text)
        captions = []
        with patch("core.services.send_catalog_album_chunk",
                   side_effect=lambda *args, **kwargs: (captions.extend(row["caption"] for row in args[3]), (True, "mocked", None))[1]):
            services.send_catalog_album(conversation, [self.item])
        return captions

    def test_the_album_caption_is_russian_for_a_russian_customer(self):
        captions = self._album_captions(["Можно заказать?", "Сколько стоит"])
        self.assertEqual(captions, ["1. Корзина-композиция из London — 1 500 000 сум"])

    def test_the_album_caption_is_untouched_for_an_uzbek_customer(self):
        captions = self._album_captions(["Nechpul", "qanaqa gullar bor"])
        self.assertEqual(captions, ["1. London Gulidan Savat Kompazitsia — 1 500 000 so'm"])

    def test_a_short_russian_message_does_not_flip_the_album_to_uzbek(self):
        """Real suhbat: "На сегодня", "А какая длина" — belgisiz, lekin ruscha."""
        captions = self._album_captions(["Можно заказать?", "На сегодня", "А какая длина"])
        self.assertEqual(captions, ["1. Корзина-композиция из London — 1 500 000 сум"])

    def test_one_uzbek_message_keeps_the_album_uzbek(self):
        """"доставка" deb yozgan o'zbek mijoz ruschaga o'tib ketmasin."""
        captions = self._album_captions(["Каерда жойлашгансиз", "доставка"])
        self.assertEqual(captions, ["1. London Gulidan Savat Kompazitsia — 1 500 000 so'm"])

    def test_the_prompt_carries_the_russian_block(self):
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("00L. RUS TILIDA JAVOB", prompt)
        self.assertIn("Здравствуйте! Магазин премиум-цветов EuroFlowers", prompt)
        self.assertIn("price_text emas → price_text_ru", prompt)
        self.assertIn('"so\'m" so\'zi HECH QACHON yozilmaydi', prompt)

    def test_the_prompt_carries_the_later_order_block(self):
        prompt = AISettings.objects.get(pk=1).system_prompt
        self.assertIn("BUYURTMA BUGUNGA EMAS BO'LSA", prompt)
        self.assertIn("future_order = true", prompt)
        self.assertIn("Chek ham SO'RAMA", prompt)
