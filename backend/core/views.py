from django.conf import settings
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from django.core.serializers.json import DjangoJSONEncoder
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Count, DecimalField, F, IntegerField, Max, OuterRef, Q, Subquery, Sum, Value
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce, TruncDate
from django.utils.dateparse import parse_date, parse_datetime
from django.utils import timezone
from datetime import datetime, time, timedelta
from decimal import Decimal
import hashlib
import hmac
import json
import requests
import django_filters
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from urllib.parse import parse_qsl
from drf_spectacular.utils import OpenApiParameter, OpenApiResponse, extend_schema, inline_serializer
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, parser_classes, permission_classes
from rest_framework import serializers
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from .models import MaterialDelivery, StockDelivery, AISettings, AuditLog, Branch, BusinessSettings, CatalogTransfer, CatalogComposition, CatalogHistory, CatalogItem, Conversation, Customer, FloristAttendance, FloristProfile, FloristSalaryEntry, FloristVolumeRate, Flower, FlowerVariant, InstagramSettings, InstagramWebhookEvent, IntegrationSettings, Lead, LeadCatalogUsage, LeadStatus, LeadStockUsage, Notification, Packaging, PackagingMovement, PagePermission, Reservation, ReservationPayment, FloristDayOff, FloristFaceSample, FloristStockBalance, FloristStockIssue, SocialPost, StockBatch, StockMovement, Supplier, SupplierPayment
from .permissions import RolePermission, has_page_permission
from .serializers import backdate_record, FloristCloseIssueSerializer, FloristStockIssueBulkRequestSerializer, FloristStockIssueEditSerializer, MaterialDeliverySerializer, MaterialReceiveSerializer, StockDeliverySerializer, AISettingsSerializer, BranchSerializer, CatalogRestoreFlowersSerializer, CatalogTransferRequestSerializer, CatalogTransferSerializer, AIPauseRequestSerializer, AuditLogSerializer, BusinessSettingsSerializer, CatalogItemSerializer, CatalogSellRequestSerializer, ChangePasswordSerializer, ConversationSerializer, CustomerSerializer, EuroFlowersTokenObtainPairSerializer, FloristAttendanceSerializer, FloristProfileSerializer, FloristDayOffSerializer, FloristFaceSampleSerializer, FloristSalaryEntrySerializer, FloristStockBalanceSerializer, FloristLeftoverRequestSerializer, FloristStockIssueRequestSerializer, FloristStockIssueSerializer, FloristStockReturnRequestSerializer, FloristVolumeRateSerializer, FlowerSerializer, FlowerVariantSerializer, InstagramSettingsSerializer, InstagramWebhookEventSerializer, IntegrationSettingsSerializer, LeadColumnReorderSerializer, LeadMoveSerializer, LeadSerializer, LeadStatusSerializer, MiniAppInitSerializer, MiniAppLeadSerializer, MiniAppQuoteSerializer, MovementRequestSerializer, NotificationSerializer, PackagingMovementRequestSerializer, PackagingMovementSerializer, PackagingSerializer, PagePermissionSerializer, ReservationPaymentRequestSerializer, ReservationPaymentSerializer, ReservationSerializer, SendResponseSerializer, SimulateResponseSerializer, SocialPostSerializer, StockBatchSerializer, StockMovementSerializer, SupplierPaymentSerializer, SupplierSerializer, TextRequestSerializer, UploadResponseSerializer, UploadSerializer, UserSerializer, UserWriteSerializer
from . import face_services
from .inventory_services import edit_florist_stock_issue, delete_florist_stock_issue, receive_material_into_delivery, catalog_cost_breakdown, adjust_florist_stems, close_all_florist_issues, close_florist_issue, florist_close_plan, florist_stem_plan, transfer_catalog_to_branch, issue_multiple_stock_to_florist, issue_stock_to_florist, return_stock_from_florist, apply_packaging_movement, apply_stock_movement, deduct_catalog_stock, deduct_lead_stock, mark_catalog_sold, restore_catalog_flowers, restore_catalog_inventory, restore_lead_stock, sync_reservation_payment_status
from .platform_services import instagram_send, telegram_send
from .services import mini_app_custom_quote_ai, normalize_phone, process_customer_message


class CreatedAtRangeFilter(django_filters.FilterSet):
    created_at = django_filters.DateTimeFromToRangeFilter()


class LeadFilter(CreatedAtRangeFilter):
    class Meta:
        model = Lead
        fields = ["status", "arrangement_type", "assigned_to", "social_post", "created_at"]


def schedule_lead_recall(lead):
    if not lead.recall_at or lead.recall_sent_at or lead.status == "lost":
        return
    from .tasks import process_lead_recall
    eta = lead.recall_at if lead.recall_at > timezone.now() else None
    if eta:
        process_lead_recall.apply_async(args=[lead.id], eta=eta)
    else:
        process_lead_recall.delay(lead.id)


def next_lead_sort_order(status_value):
    current = Lead.objects.filter(status=status_value).aggregate(value=Max("sort_order"))["value"] or Decimal("0")
    return current + Decimal("1000")


def lead_sort_order_between(before, after, status_value):
    if before and before.status != status_value:
        raise serializers.ValidationError({"before": "Lead statusi yangi column bilan mos emas"})
    if after and after.status != status_value:
        raise serializers.ValidationError({"after": "Lead statusi yangi column bilan mos emas"})
    if before and after:
        return (before.sort_order + after.sort_order) / Decimal("2")
    if before:
        return before.sort_order + Decimal("1000")
    if after:
        return after.sort_order - Decimal("1000")
    return next_lead_sort_order(status_value)


def create_user_notification(user, notification_type, title, body, reference_type="", reference_id=None):
    if not user:
        return None
    if getattr(getattr(user, "profile", None), "role", None) == "developer":
        return None
    return Notification.objects.create(
        target_user=user,
        notification_type=notification_type,
        title_uz=title,
        title_ru=title,
        body_uz=body,
        body_ru=body,
        reference_type=reference_type,
        reference_id=reference_id,
    )


def normalize_coordinate(value):
    if value in [None, ""]:
        return value
    return Decimal(str(value)).quantize(Decimal("0.0000000001"))


def notification_references_developer(queryset):
    developer_user_ids = list(User.objects.filter(profile__role="developer").values_list("id", flat=True))
    developer_florist_ids = list(FloristProfile.objects.filter(user_id__in=developer_user_ids).values_list("id", flat=True))
    developer_attendance_ids = list(FloristAttendance.objects.filter(florist_id__in=developer_florist_ids).values_list("id", flat=True))
    developer_salary_ids = list(FloristSalaryEntry.objects.filter(florist_id__in=developer_florist_ids).values_list("id", flat=True))
    developer_filters = Q(target_user__profile__role="developer")
    if developer_attendance_ids:
        developer_filters |= Q(reference_type="attendance", reference_id__in=developer_attendance_ids)
    if developer_salary_ids:
        developer_filters |= Q(reference_type="florist_salary", reference_id__in=developer_salary_ids)
    return queryset.exclude(developer_filters)


def notification_queryset_for_user(user):
    queryset = Notification.objects.all()
    role = getattr(getattr(user, "profile", None), "role", None)
    if role != "developer":
        queryset = notification_references_developer(queryset)
    if role in ["florist", "apprentice"]:
        return queryset.filter(target_user=user)
    return queryset.filter(Q(target_user__isnull=True) | Q(target_user=user))


class StockMovementFilter(CreatedAtRangeFilter):
    supplier = django_filters.ModelChoiceFilter(field_name="batch__supplier", queryset=Supplier.objects.all())

    class Meta:
        model = StockMovement
        fields = ["batch", "supplier", "movement_type", "created_at"]


class PackagingMovementFilter(CreatedAtRangeFilter):
    class Meta:
        model = PackagingMovement
        fields = ["packaging", "movement_type", "created_at"]


class StockBatchFilter(CreatedAtRangeFilter):
    class Meta:
        model = StockBatch
        fields = ["variant", "supplier", "delivery", "is_free", "height_cm", "height_from_cm", "height_to_cm", "is_active", "created_at"]


class FloristAttendanceFilter(CreatedAtRangeFilter):
    class Meta:
        model = FloristAttendance
        fields = ["florist", "work_date", "source", "created_at"]


class FloristSalaryEntryFilter(CreatedAtRangeFilter):
    class Meta:
        model = FloristSalaryEntry
        fields = ["florist", "source", "work_date", "created_at"]


class ConversationFilter(CreatedAtRangeFilter):
    class Meta:
        model = Conversation
        fields = ["status", "assigned_to", "created_at"]


class AuditLogFilter(CreatedAtRangeFilter):
    user = django_filters.ModelChoiceFilter(queryset=User.objects.all())
    user_id = django_filters.NumberFilter(field_name="user_id")

    class Meta:
        model = AuditLog
        fields = ["action", "entity_type", "user", "user_id", "created_at"]


class PagePermissionFilter(django_filters.FilterSet):
    permission_page = django_filters.ChoiceFilter(field_name="page", choices=PagePermission.PAGE_CHOICES)

    class Meta:
        model = PagePermission
        fields = ["user", "permission_page", "can_view", "can_control"]


def parse_date_range_params(request):
    date_from = parse_date(request.query_params.get("date_from", "") or "")
    date_to = parse_date(request.query_params.get("date_to", "") or "")
    return date_from, date_to


def filter_date_field(queryset, field_name, date_from=None, date_to=None):
    if date_from:
        queryset = queryset.filter(**{f"{field_name}__date__gte": date_from})
    if date_to:
        queryset = queryset.filter(**{f"{field_name}__date__lte": date_to})
    return queryset


def local_datetime_label(value):
    if not value:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M")


def money_label(value):
    return float(value or 0)


def styled_workbook(title, headers):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.freeze_panes = "A2"
    return workbook, sheet


def autosize_sheet(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def excel_response(workbook, filename):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def catalog_payment_type(item, history=None):
    sources = []
    if history:
        sources.append(history.snapshot or {})
    sources.append({})
    for source in sources:
        value = source.get("payment_type") or source.get("payment_method")
        if value:
            return str(value)
    return "Aniqlanmagan"


def user_full_name(user):
    if not user:
        return ""
    return user.get_full_name() or user.username


def payment_label(value):
    if value in ["cash", "naqd"]:
        return "Naqd"
    if value in ["card", "karta"]:
        return "Karta"
    if value == "transfer":
        return "O‘tkazma"
    return "Aniqlanmagan"


def catalog_history_sale_total(history):
    return Decimal(history.sold_unit_price or 0) * Decimal(history.quantity or 0)


def catalog_history_listed_total(history):
    return Decimal(history.listed_unit_price or 0) * Decimal(history.quantity or 0)


def catalog_history_cost_total(history):
    item = history.catalog_item
    total = Decimal(item.quantity_total or 1)
    quantity = Decimal(history.quantity or 0)
    if total <= 0:
        return Decimal("0")
    return (Decimal(item.calculated_cost_price or 0) / total * quantity).quantize(Decimal("0.01"))


def catalog_history_cost_breakdown(history):
    """Sotuv tannarxini gul, material va florist haqiga ajratadi."""
    item = history.catalog_item
    total = Decimal(item.quantity_total or 1)
    quantity = Decimal(history.quantity or 0)
    empty = {"flower_cost": Decimal("0"), "material_cost": Decimal("0"), "florist_fee_cost": Decimal("0")}
    if total <= 0 or quantity <= 0:
        return empty
    breakdown = catalog_cost_breakdown(item)
    share = quantity / total
    rows = {key: (breakdown[key] * share).quantize(Decimal("0.01")) for key in empty}
    # yaxlitlash farqini gul tannarxiga qo'shib, yig'indi cost_total bilan mos bo'lishini ta'minlaymiz
    diff = catalog_history_cost_total(history) - sum(rows.values())
    rows["flower_cost"] += diff
    return rows


def catalog_history_flower_stems(history):
    """Sotuvga ketgan gul donasi. Filialga o'tkazilgan katalogda tarkib nusxalanadi,
    shuning uchun filial sotuvida ham gul soni to'g'ri chiqadi."""
    item = history.catalog_item
    per_item = sum(int(row.quantity_stems or 0) for row in item.composition.all())
    return per_item * int(history.quantity or 0)


def accounting_branch_selection(request):
    """Hisob-kitob qaysi filiallarni qamrashini aniqlaydi.

    Filial foydalanuvchisi faqat o'z filialini ko'radi va buni kengaytira olmaydi.
    Asosiy filial foydalanuvchisi sukut bo'yicha hamma filialni ko'radi:
    ?branch=main faqat asosiy filial, ?branch=<id> esa bitta filial.
    """
    branch = user_branch(getattr(request, "user", None))
    if branch:
        return {"mode": "branch", "branch": branch}
    raw = str((request.query_params.get("branch") if hasattr(request, "query_params") else "") or "").strip().lower()
    if raw in ["", "all"]:
        return {"mode": "all", "branch": None}
    if raw == "main":
        return {"mode": "main", "branch": None}
    selected = Branch.objects.filter(pk=raw).first() if raw.isdigit() else Branch.objects.filter(name__iexact=raw).first()
    if not selected:
        return {"mode": "all", "branch": None}
    if selected.is_main:
        return {"mode": "main", "branch": None}
    return {"mode": "branch", "branch": selected}


def accounting_includes_main(selection):
    """Asosiy filial hisobga kiradimi. Chiqit faqat asosiy skladda bo'lgani uchun kerak."""
    return selection["mode"] in ["all", "main"]


def sold_catalog_history_queryset(request):
    date_from, date_to = parse_date_range_params(request)
    rows = (
        CatalogHistory.objects.filter(action="sold")
        .select_related("catalog_item__florist__user", "catalog_item__branch", "created_by")
        .prefetch_related("catalog_item__composition")
    )
    selection = accounting_branch_selection(request)
    if selection["mode"] == "branch":
        rows = rows.filter(catalog_item__branch=selection["branch"])
    elif selection["mode"] == "main":
        rows = rows.filter(catalog_item__branch__isnull=True)
    if date_from:
        rows = rows.filter(created_at__date__gte=date_from)
    if date_to:
        rows = rows.filter(created_at__date__lte=date_to)
    return rows, date_from, date_to


MAIN_BRANCH_LABEL = "Toshkent (asosiy filial)"


def blank_accounting_bucket(branch_id=None, branch_name=MAIN_BRANCH_LABEL, is_main=True):
    """Bitta filial uchun bo'sh hisob-kitob qutisi. Umumiy yig'indi ham shu shaklda."""
    return {
        "branch_id": branch_id,
        "branch_name": branch_name,
        "is_main": is_main,
        "sales_count": 0,
        "total_quantity": 0,
        "flower_stems": 0,
        "standard_quantity": 0,
        "custom_quantity": 0,
        "total_sales": Decimal("0"),
        "cash_total": Decimal("0"),
        "card_total": Decimal("0"),
        "unknown_total": Decimal("0"),
        "cash_quantity": 0,
        "card_quantity": 0,
        "unknown_quantity": 0,
        "cash_count": 0,
        "card_count": 0,
        "unknown_count": 0,
        "discount_total": Decimal("0"),
        "discounted_sales_count": 0,
        "discounted_quantity": 0,
        "cost_total": Decimal("0"),
        "flower_cost_total": Decimal("0"),
        "material_cost_total": Decimal("0"),
        "florist_fee_cost_total": Decimal("0"),
        "waste_cost_total": Decimal("0"),
        "waste_stems": 0,
        "net_profit": Decimal("0"),
    }


def add_sale_to_bucket(bucket, payment, quantity, stems, sale_total, discount, cost_total, cost_breakdown, kind):
    """Bitta sotuvni qutiga qo'shadi. Umumiy va filial qutilariga bir xil qo'llanadi."""
    bucket["sales_count"] += 1
    bucket["total_quantity"] += quantity
    bucket["flower_stems"] += stems
    bucket["total_sales"] += sale_total
    bucket[f"{payment}_total"] += sale_total
    bucket[f"{payment}_quantity"] += quantity
    bucket[f"{payment}_count"] += 1
    bucket["discount_total"] += discount
    bucket["cost_total"] += cost_total
    bucket["flower_cost_total"] += cost_breakdown["flower_cost"]
    bucket["material_cost_total"] += cost_breakdown["material_cost"]
    bucket["florist_fee_cost_total"] += cost_breakdown["florist_fee_cost"]
    if discount > 0:
        bucket["discounted_sales_count"] += 1
        bucket["discounted_quantity"] += quantity
    if kind == "custom":
        bucket["custom_quantity"] += quantity
    else:
        bucket["standard_quantity"] += quantity


def accounting_report_data(request):
    histories, date_from, date_to = sold_catalog_history_queryset(request)
    selection = accounting_branch_selection(request)
    summary = blank_accounting_bucket()
    summary["branch_id"] = None
    summary["branch_name"] = "Umumiy"
    summary["is_main"] = False
    branch_buckets = {}
    if accounting_includes_main(selection):
        branch_buckets[None] = blank_accounting_bucket()
    if selection["mode"] == "all":
        for row in Branch.objects.filter(is_main=False, is_active=True):
            branch_buckets[row.id] = blank_accounting_bucket(row.id, row.name, False)
    elif selection["mode"] == "branch":
        branch = selection["branch"]
        branch_buckets[branch.id] = blank_accounting_bucket(branch.id, branch.name, False)
    by_kind = {
        "standard": {"quantity": 0, "sales": Decimal("0"), "discount": Decimal("0")},
        "custom": {"quantity": 0, "sales": Decimal("0"), "discount": Decimal("0")},
    }
    by_payment = {
        "cash": {"label": "Naqd", "quantity": 0, "sales": Decimal("0")},
        "card": {"label": "Karta", "quantity": 0, "sales": Decimal("0")},
        "unknown": {"label": "Aniqlanmagan", "quantity": 0, "sales": Decimal("0")},
    }
    by_volume = {}
    discount_rows = []
    history_rows = []
    reservation_payment_rows = []
    reservation_payment_summary = {
        "count": 0,
        "total": Decimal("0"),
        "cash_total": Decimal("0"),
        "card_total": Decimal("0"),
        "transfer_total": Decimal("0"),
    }
    for history in histories.order_by("-created_at", "-id"):
        item = history.catalog_item
        kind = item.catalog_kind or "standard"
        payment = catalog_payment_type(item, history).lower()
        if payment == "naqd":
            payment = "cash"
        elif payment == "karta":
            payment = "card"
        if payment not in ["cash", "card"]:
            payment = "unknown"
        quantity = int(history.quantity or 0)
        stems = catalog_history_flower_stems(history)
        sale_total = catalog_history_sale_total(history)
        cost_total = catalog_history_cost_total(history)
        cost_breakdown = catalog_history_cost_breakdown(history)
        discount = Decimal(history.discount_amount or 0)
        add_sale_to_bucket(summary, payment, quantity, stems, sale_total, discount, cost_total, cost_breakdown, kind)
        if item.branch_id not in branch_buckets:
            name = item.branch.name if item.branch_id else MAIN_BRANCH_LABEL
            branch_buckets[item.branch_id] = blank_accounting_bucket(item.branch_id, name, item.branch_id is None)
        add_sale_to_bucket(branch_buckets[item.branch_id], payment, quantity, stems, sale_total, discount, cost_total, cost_breakdown, kind)
        by_kind.setdefault(kind, {"quantity": 0, "sales": Decimal("0"), "discount": Decimal("0")})
        by_kind[kind]["quantity"] += quantity
        by_kind[kind]["sales"] += sale_total
        by_kind[kind]["discount"] += discount
        by_payment[payment]["quantity"] += quantity
        by_payment[payment]["sales"] += sale_total
        volume_key = item.volume or "Belgilanmagan"
        volume_row = by_volume.setdefault((kind, volume_key), {"catalog_kind": kind, "volume": volume_key, "quantity": 0, "sales": Decimal("0"), "discount": Decimal("0")})
        volume_row["quantity"] += quantity
        volume_row["sales"] += sale_total
        volume_row["discount"] += discount
        row = {
            "history_id": history.id,
            "catalog_id": item.id,
            "catalog_name": item.name_uz,
            "catalog_kind": kind,
            "arrangement_type": item.arrangement_type,
            "volume": item.volume,
            "quantity": quantity,
            "flower_stems": stems,
            "branch_id": item.branch_id,
            "branch_name": item.branch.name if item.branch_id else MAIN_BRANCH_LABEL,
            "is_main_branch": item.branch_id is None,
            "created_at": history.created_at,
            "catalog_created_at": item.created_at,
            "sold_at": history.created_at,
            "florist_id": item.florist_id,
            "florist_name": str(item.florist) if item.florist_id else "",
            "listed_unit_price": history.listed_unit_price,
            "sold_unit_price": history.sold_unit_price,
            "listed_total": catalog_history_listed_total(history),
            "sale_total": sale_total,
            "cost_total": cost_total,
            "flower_cost": cost_breakdown["flower_cost"],
            "material_cost": cost_breakdown["material_cost"],
            "florist_fee_cost": cost_breakdown["florist_fee_cost"],
            "net_profit": sale_total - cost_total,
            "payment_type": payment,
            "payment_label": payment_label(payment),
            "discount_amount": discount,
            "discount_percent": history.discount_percent,
            "discount_reason": history.discount_reason or history.note,
            "sold_by": user_full_name(history.created_by),
        }
        history_rows.append(row)
        if discount > 0:
            discount_rows.append(row)
    if accounting_includes_main(selection):
        reservation_payments = ReservationPayment.objects.select_related("reservation__customer", "reservation__catalog_item", "created_by")
        if date_from:
            reservation_payments = reservation_payments.filter(paid_at__date__gte=date_from)
        if date_to:
            reservation_payments = reservation_payments.filter(paid_at__date__lte=date_to)
        for payment in reservation_payments.order_by("-paid_at", "-id"):
            method = payment.method if payment.method in ["cash", "card", "transfer"] else "unknown"
            amount = Decimal(payment.amount or 0)
            reservation_payment_summary["count"] += 1
            reservation_payment_summary["total"] += amount
            if method in ["cash", "card", "transfer"]:
                reservation_payment_summary[f"{method}_total"] += amount
            reservation_payment_rows.append({
                "id": payment.id,
                "reservation_id": payment.reservation_id,
                "customer_id": payment.reservation.customer_id,
                "customer_name": payment.reservation.customer.name,
                "catalog_id": payment.reservation.catalog_item_id,
                "catalog_name": payment.reservation.catalog_item.name_uz if payment.reservation.catalog_item_id else "",
                "amount": amount,
                "method": method,
                "method_label": payment_label(method),
                "paid_at": payment.paid_at,
                "note": payment.note,
                "created_by": user_full_name(payment.created_by),
            })
    # Chiqit faqat asosiy skladda bo'ladi, filiallarda gul saqlanmaydi.
    if accounting_includes_main(selection):
        waste_movements = StockMovement.objects.filter(movement_type="waste").select_related("batch")
        if date_from:
            waste_movements = waste_movements.filter(created_at__date__gte=date_from)
        if date_to:
            waste_movements = waste_movements.filter(created_at__date__lte=date_to)
        main_bucket = branch_buckets.setdefault(None, blank_accounting_bucket())
        for movement in waste_movements:
            waste = abs(int(movement.quantity_stems or 0))
            cost = (Decimal(waste) * Decimal(movement.batch.cost_per_stem or 0)).quantize(Decimal("0.01"))
            summary["waste_stems"] += waste
            summary["waste_cost_total"] += cost
            main_bucket["waste_stems"] += waste
            main_bucket["waste_cost_total"] += cost
    summary["net_profit"] = summary["total_sales"] - summary["cost_total"]
    for bucket in branch_buckets.values():
        bucket["net_profit"] = bucket["total_sales"] - bucket["cost_total"]
        bucket["share_percent"] = (
            (bucket["total_sales"] / summary["total_sales"] * 100).quantize(Decimal("0.01"))
            if summary["total_sales"] else Decimal("0")
        )
    by_branch = sorted(branch_buckets.values(), key=lambda row: (not row["is_main"], row["branch_name"]))
    return {
        "period": {"date_from": date_from.isoformat() if date_from else None, "date_to": date_to.isoformat() if date_to else None},
        "branch_filter": {
            "mode": selection["mode"],
            "branch_id": selection["branch"].id if selection["branch"] else None,
            "branch_name": selection["branch"].name if selection["branch"] else None,
        },
        "summary": summary,
        "by_branch": by_branch,
        "by_kind": [{"catalog_kind": key, **value} for key, value in by_kind.items()],
        "by_payment": [{"payment_type": key, **value} for key, value in by_payment.items()],
        "by_volume": sorted(by_volume.values(), key=lambda row: (row["catalog_kind"], row["volume"])),
        "discounted_sales": discount_rows,
        "history": history_rows,
        "reservation_payments_summary": reservation_payment_summary,
        "reservation_payments": reservation_payment_rows,
    }


class EuroFlowersTokenObtainPairView(TokenObtainPairView):
    serializer_class = EuroFlowersTokenObtainPairSerializer


def forbidden():
    return Response({"detail": "Sizda bu sahifa uchun ruxsat yo‘q."}, status=status.HTTP_403_FORBIDDEN)


def json_safe(value):
    return json.loads(json.dumps(value, cls=DjangoJSONEncoder))


def instance_snapshot(instance):
    data = {}
    for field in instance._meta.concrete_fields:
        key = field.name
        value = getattr(instance, field.attname)
        data[key] = value
    if isinstance(instance, Lead):
        data["catalog_usage"] = [{"catalog_item": row.catalog_item_id, "catalog_name": row.catalog_item.name_uz, "quantity": row.quantity} for row in instance.catalog_usage.select_related("catalog_item")]
        data["stock_usage"] = [{"stock_batch": row.stock_batch_id, "batch_number": row.stock_batch.batch_number, "quantity_stems": row.quantity_stems, "quantity_bunches": row.quantity_bunches} for row in instance.stock_usage.select_related("stock_batch")]
        data["packaging_usage"] = [{"packaging": row.packaging_id, "packaging_name": row.packaging.name_uz, "quantity": row.quantity} for row in instance.packaging_usage.select_related("packaging")]
    if isinstance(instance, User):
        profile = getattr(instance, "profile", None)
        data["profile"] = {"role": profile.role, "language": profile.language} if profile else None
        data["permissions"] = [{"page": row.page, "can_view": row.can_view, "can_control": row.can_control} for row in instance.page_permissions.order_by("page")]
    if isinstance(instance, StockBatch):
        data["batch_number"] = instance.batch_number
        data["flower"] = str(instance.variant)
        data["supplier"] = instance.supplier.name if instance.supplier_id else ""
        data["remaining_bunches"] = instance.remaining_bunches
        data["stock_value"] = instance.stock_value
    if isinstance(instance, Packaging):
        data["material"] = instance.name_uz
        data["type_label"] = instance.get_packaging_type_display()
        data["quantity_label"] = instance.quantity_label
    if isinstance(instance, FloristProfile):
        data["full_name"] = str(instance)
        data["volume_rates"] = [{"arrangement_type": row.arrangement_type, "volume": row.volume, "default_stems": row.default_stems, "florist_fee": row.florist_fee, "is_active": row.is_active} for row in instance.volume_rates.order_by("arrangement_type", "volume")]
    return json_safe(data)


def changed_snapshot(before, after):
    keys = sorted(set(before) | set(after))
    before_changed = {}
    after_changed = {}
    for key in keys:
        if before.get(key) != after.get(key):
            before_changed[key] = before.get(key)
            after_changed[key] = after.get(key)
    return before_changed, after_changed


def request_ip(request):
    if not request:
        return None
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def write_audit(user, action, instance, before=None, after=None, request=None, summary=""):
    if isinstance(instance, AuditLog):
        return None
    return AuditLog.objects.create(
        user=user if getattr(user, "is_authenticated", False) else None,
        action=action,
        summary=summary or f"{instance.__class__.__name__}: {action}",
        entity_type=instance.__class__.__name__,
        entity_id=str(getattr(instance, "id", "")),
        before=json_safe(before or {}),
        after=json_safe(after or {}),
        ip_address=request_ip(request),
        request_method=getattr(request, "method", "") if request else "",
        request_path=getattr(request, "path", "")[:255] if request else "",
    )


PROTECTED_LABELS = {
    "supplierpayment": "To‘lovlar",
    "stockbatch": "Sklad partiyalari",
    "stockmovement": "Sklad harakatlari",
    "catalogitem": "Katalog mahsulotlari",
    "catalogcomposition": "Katalog tarkibi",
    "catalogmaterialusage": "Katalog materiallari",
    "leadstockusage": "Lead gullari",
    "leadpackagingusage": "Lead qadoqlari",
    "leadcatalogusage": "Lead katalogi",
    "lead": "Leadlar",
    "packagingmovement": "Qadoq harakatlari",
    "floristsalaryentry": "Florist ish haqi",
    "floristattendance": "Keldi-ketdi",
    "flowervariant": "Gul navlari",
    "conversation": "Suhbatlar",
}


def protected_blockers(error):
    """ProtectedError ichidagi bog'liq yozuvlarni model bo'yicha sanaydi."""
    counts = {}
    for obj in getattr(error, "protected_objects", []) or []:
        key = obj._meta.model_name
        counts[key] = counts.get(key, 0) + 1
    return [
        {"model": key, "label": PROTECTED_LABELS.get(key, key), "count": value}
        for key, value in sorted(counts.items(), key=lambda row: -row[1])
    ]


class ScopedViewSet(viewsets.ModelViewSet):
    permission_classes = [RolePermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        if not queryset.query.order_by and not queryset.model._meta.ordering:
            queryset = queryset.order_by("id")
        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        write_audit(self.request.user, f"{instance.__class__.__name__.lower()}_created", instance, before={}, after=instance_snapshot(instance), request=self.request)

    def perform_update(self, serializer):
        before = instance_snapshot(serializer.instance)
        instance = serializer.save()
        after = instance_snapshot(instance)
        before_changed, after_changed = changed_snapshot(before, after)
        if before_changed or after_changed:
            write_audit(self.request.user, f"{instance.__class__.__name__.lower()}_updated", instance, before=before_changed, after=after_changed, request=self.request)

    def perform_destroy(self, instance):
        self._archive_blockers = None
        before = instance_snapshot(instance)
        try:
            instance.delete()
            write_audit(self.request.user, f"{instance.__class__.__name__.lower()}_deleted", instance, before=before, after={}, request=self.request)
        except ProtectedError as error:
            field_names = {field.name for field in instance._meta.fields}
            if "is_active" not in field_names:
                raise
            instance.is_active = False
            instance.save(update_fields=["is_active", "updated_at"])
            before_changed, after_changed = changed_snapshot(before, instance_snapshot(instance))
            write_audit(self.request.user, f"{instance.__class__.__name__.lower()}_archived", instance, before=before_changed, after=after_changed, request=self.request)
            self._archive_blockers = protected_blockers(error)

    def destroy(self, request, *args, **kwargs):
        """Bog'liq yozuvlari bor obyekt o'chirilmaydi, arxivga olinadi.
        O'chirilganda 204, arxivlanganda 200 va tushuntirish qaytadi."""
        instance = self.get_object()
        label = str(instance)
        self.perform_destroy(instance)
        blockers = getattr(self, "_archive_blockers", None)
        if blockers is None:
            return Response(status=status.HTTP_204_NO_CONTENT)
        reason = ", ".join(f"{row['label']} ({row['count']} ta)" for row in blockers)
        return Response(
            {
                "detail": f"{label} bog'liq yozuvlari borligi uchun o'chirilmadi, arxivga olindi. Sabab: {reason}.",
                "archived": True,
                "deleted": False,
                "id": instance.pk,
                "is_active": False,
                "blocked_by": blockers,
            },
            status=status.HTTP_200_OK,
        )


class UserViewSet(viewsets.ModelViewSet):
    permission_classes = [RolePermission]
    permission_page = "users"
    queryset = User.objects.select_related("profile").prefetch_related("page_permissions").order_by("id")
    serializer_class = UserSerializer
    search_fields = ["username", "first_name", "last_name", "email"]
    filterset_fields = ["is_active", "profile__role"]

    def get_serializer_class(self):
        if self.request.method in ["POST", "PUT", "PATCH"]:
            return UserWriteSerializer
        return UserSerializer

    def get_developer_role(self):
        return getattr(getattr(self.request.user, "profile", None), "role", None) == "developer"

    def create(self, request, *args, **kwargs):
        if request.data.get("role") == "developer" and not self.get_developer_role():
            return forbidden()
        response = super().create(request, *args, **kwargs)
        if response.status_code < 400:
            user = User.objects.filter(id=response.data.get("id")).first()
            if user:
                write_audit(request.user, "user_created", user, before={}, after=instance_snapshot(user), request=request)
        return response

    def update(self, request, *args, **kwargs):
        if request.data.get("role") == "developer" and not self.get_developer_role():
            return forbidden()
        user = self.get_object()
        before = instance_snapshot(user)
        response = super().update(request, *args, **kwargs)
        if response.status_code < 400:
            user.refresh_from_db()
            before_changed, after_changed = changed_snapshot(before, instance_snapshot(user))
            if before_changed or after_changed:
                write_audit(request.user, "user_updated", user, before=before_changed, after=after_changed, request=request)
        return response

    def partial_update(self, request, *args, **kwargs):
        if request.data.get("role") == "developer" and not self.get_developer_role():
            return forbidden()
        user = self.get_object()
        before = instance_snapshot(user)
        response = super().partial_update(request, *args, **kwargs)
        if response.status_code < 400:
            user.refresh_from_db()
            before_changed, after_changed = changed_snapshot(before, instance_snapshot(user))
            if before_changed or after_changed:
                write_audit(request.user, "user_updated", user, before=before_changed, after=after_changed, request=request)
        return response

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(getattr(self.request.user, "profile", None), "role", None)
        if role != "developer":
            queryset = queryset.exclude(profile__role="developer")
        return queryset

    @action(detail=True, methods=["post"])
    def deactivate(self, request, pk=None):
        user = self.get_object()
        before = instance_snapshot(user)
        user.is_active = False
        user.save(update_fields=["is_active"])
        before_changed, after_changed = changed_snapshot(before, instance_snapshot(user))
        write_audit(request.user, "user_deactivated", user, before=before_changed, after=after_changed, request=request)
        return Response(UserSerializer(user).data)


class FlowerViewSet(ScopedViewSet):
    permission_page = "inventory"
    write_roles = ["admin", "warehouse"]
    queryset = Flower.objects.prefetch_related("variants").all()
    serializer_class = FlowerSerializer
    search_fields = ["name_uz"]

    def perform_destroy(self, instance):
        before = instance_snapshot(instance)
        try:
            instance.delete()
            write_audit(self.request.user, "flower_deleted", instance, before=before, after={}, request=self.request)
        except ProtectedError:
            instance.is_active = False
            instance.save(update_fields=["is_active", "updated_at"])
            archived_variants = list(instance.variants.filter(is_active=True).values_list("id", flat=True))
            instance.variants.filter(is_active=True).update(is_active=False, updated_at=timezone.now())
            after = instance_snapshot(instance)
            after["archived_variants"] = archived_variants
            before_changed, after_changed = changed_snapshot(before, after)
            write_audit(self.request.user, "flower_archived", instance, before=before_changed, after=after_changed, request=self.request)


class FlowerVariantViewSet(ScopedViewSet):
    permission_page = "inventory"
    write_roles = ["admin", "warehouse"]
    queryset = FlowerVariant.objects.select_related("flower").all()
    serializer_class = FlowerVariantSerializer
    filterset_fields = ["flower", "is_active"]
    search_fields = ["name_uz", "color_uz", "flower__name_uz"]

    def perform_destroy(self, instance):
        before = instance_snapshot(instance)
        try:
            instance.delete()
            write_audit(self.request.user, "flowervariant_deleted", instance, before=before, after={}, request=self.request)
        except ProtectedError:
            instance.is_active = False
            instance.save(update_fields=["is_active", "updated_at"])
            before_changed, after_changed = changed_snapshot(before, instance_snapshot(instance))
            write_audit(self.request.user, "flowervariant_archived", instance, before=before_changed, after=after_changed, request=self.request)


def supplier_rollup_queryset():
    money = DecimalField(max_digits=16, decimal_places=2)
    purchase = Subquery(
        StockBatch.objects.filter(supplier=OuterRef("pk"))
        .values("supplier")
        .annotate(total=Coalesce(Sum(F("received_stems") * F("cost_per_stem"), output_field=money), Value(Decimal("0"), output_field=money)))
        .values("total")[:1],
        output_field=money,
    )
    paid = Subquery(
        SupplierPayment.objects.filter(supplier=OuterRef("pk"))
        .values("supplier")
        .annotate(total=Coalesce(Sum("amount", output_field=money), Value(Decimal("0"), output_field=money)))
        .values("total")[:1],
        output_field=money,
    )
    last_paid = Subquery(SupplierPayment.objects.filter(supplier=OuterRef("pk")).order_by("-paid_at", "-id").values("paid_at")[:1])
    zero = Value(Decimal("0"), output_field=money)
    # Postavshikdan har safar to'liq to'lab olinadi, qarz tushunchasi yo'q.
    # Shuning uchun faqat umumiy sotib olingan summa hisoblanadi.
    return Supplier.objects.annotate(
        batches_count=Count("stock_batches", distinct=True),
        total_received_stems=Coalesce(Sum("stock_batches__received_stems"), 0),
        purchase_total=Coalesce(purchase, zero),
        paid_total=Coalesce(paid, zero),
        last_payment_at=last_paid,
    )


class SupplierViewSet(ScopedViewSet):
    permission_page = "suppliers"
    write_roles = ["admin", "warehouse"]
    queryset = supplier_rollup_queryset()
    serializer_class = SupplierSerializer
    filterset_fields = ["is_active", "supplier_type"]
    search_fields = ["name", "phone", "notes"]
    ordering_fields = ["name", "purchase_total", "paid_total", "last_payment_at", "created_at"]


class SupplierPaymentViewSet(ScopedViewSet):
    permission_page = "suppliers"
    write_roles = ["admin", "warehouse"]
    queryset = SupplierPayment.objects.select_related("supplier", "created_by").all()
    serializer_class = SupplierPaymentSerializer
    filterset_fields = ["supplier", "method", "paid_at"]
    search_fields = ["note", "supplier__name"]
    ordering_fields = ["paid_at", "amount", "created_at"]

    def perform_create(self, serializer):
        payment = serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)
        write_audit(self.request.user, "supplierpayment_created", payment, before={}, after=instance_snapshot(payment), request=self.request)


class FloristVolumeRateViewSet(ScopedViewSet):
    permission_page = "florists"
    write_roles = ["admin"]
    queryset = FloristVolumeRate.objects.all()
    serializer_class = FloristVolumeRateSerializer
    filterset_fields = ["arrangement_type", "volume", "is_active"]


ARRANGEMENT_LABELS = {"bouquet": "Buket", "basket": "Savat", "box": "Quti"}
CATALOG_KIND_LABELS = {"standard": "Standart", "custom": "Custom"}


def florist_salary_queryset(profile, date_from, date_to):
    rows = FloristSalaryEntry.objects.filter(florist=profile).select_related("catalog_item__florist", "created_by", "attendance")
    if date_from:
        rows = rows.filter(work_date__gte=date_from)
    if date_to:
        rows = rows.filter(work_date__lte=date_to)
    return rows.order_by("-work_date", "-id")


def florist_item_revenue_map(item_ids):
    """Har bir katalog mahsuloti bo'yicha sotuvdan tushgan real summa va sotilgan soni."""
    revenue = {}
    if not item_ids:
        return revenue
    for history in CatalogHistory.objects.filter(action="sold", catalog_item_id__in=item_ids):
        row = revenue.setdefault(history.catalog_item_id, {"revenue": Decimal("0"), "sold_quantity": 0, "last_sold_at": None})
        row["revenue"] += catalog_history_sale_total(history)
        row["sold_quantity"] += int(history.quantity or 0)
        if not row["last_sold_at"] or history.created_at > row["last_sold_at"]:
            row["last_sold_at"] = history.created_at
    return revenue


def florist_stats_data(profile, request, include_sales=True):
    """Florist bo'yicha to'liq statistika. Detail API, florist dashboard va Excel eksport shundan foydalanadi.
    include_sales=False bo'lsa sotuv narxi, tushum va foyda chiqarilmaydi — florist o'zi ko'rgan holat."""
    date_from, date_to = parse_date_range_params(request)
    salary = list(florist_salary_queryset(profile, date_from, date_to))
    item_ids = [row.catalog_item_id for row in salary if row.catalog_item_id]
    revenue_map = florist_item_revenue_map(item_ids)

    summary = {
        "salary_total": Decimal("0"),
        "salary_entries_count": len(salary),
        "catalog_salary_total": Decimal("0"),
        "decoration_salary_total": Decimal("0"),
        "daily_salary_total": Decimal("0"),
        "manual_salary_total": Decimal("0"),
        "catalog_count": 0,
        "bouquet_count": 0,
        "basket_count": 0,
        "standard_count": 0,
        "custom_count": 0,
        "sold_quantity": 0,
        "sale_revenue": Decimal("0"),
        "avg_fee_per_item": Decimal("0"),
    }
    by_source, by_arrangement, by_volume, by_day, entries = {}, {}, {}, {}, []

    for row in salary:
        item = row.catalog_item
        amount = Decimal(row.amount or 0)
        arrangement = item.arrangement_type if item else ""
        volume = (item.volume if item else "") or "Belgilanmagan"
        kind = (item.catalog_kind if item else "") or ""
        sold = revenue_map.get(row.catalog_item_id, {}) if row.catalog_item_id else {}
        sold_quantity = int(sold.get("sold_quantity") or 0)
        sale_revenue = Decimal(sold.get("revenue") or 0)
        is_production = bool(item and row.source in ["catalog", "custom_catalog"])
        produced_quantity = int(item.quantity_total or 1) if is_production else 0

        summary["salary_total"] += amount
        if row.source == "daily":
            summary["daily_salary_total"] += amount
        elif row.source == "manual":
            summary["manual_salary_total"] += amount
        elif row.source in ["decoration", "sale_decoration"]:
            summary["decoration_salary_total"] += amount
        else:
            summary["catalog_salary_total"] += amount
        if is_production:
            summary["sold_quantity"] += sold_quantity
            summary["sale_revenue"] += sale_revenue
            summary["catalog_count"] += produced_quantity
            if arrangement == "bouquet":
                summary["bouquet_count"] += produced_quantity
            elif arrangement == "basket":
                summary["basket_count"] += produced_quantity
            if kind == "custom":
                summary["custom_count"] += produced_quantity
            else:
                summary["standard_count"] += produced_quantity

        source_row = by_source.setdefault(row.source, {"source": row.source, "source_label": row.get_source_display(), "count": 0, "amount": Decimal("0")})
        source_row["count"] += 1
        source_row["amount"] += amount

        if arrangement and is_production:
            arr_row = by_arrangement.setdefault(arrangement, {"arrangement_type": arrangement, "arrangement_label": ARRANGEMENT_LABELS.get(arrangement, arrangement), "count": 0, "amount": Decimal("0"), "sold_quantity": 0, "sale_revenue": Decimal("0")})
            arr_row["count"] += produced_quantity
            arr_row["amount"] += amount
            arr_row["sold_quantity"] += sold_quantity
            arr_row["sale_revenue"] += sale_revenue

            vol_key = (arrangement, volume)
            vol_row = by_volume.setdefault(vol_key, {"arrangement_type": arrangement, "arrangement_label": ARRANGEMENT_LABELS.get(arrangement, arrangement), "volume": volume, "count": 0, "amount": Decimal("0"), "sold_quantity": 0, "sale_revenue": Decimal("0")})
            vol_row["count"] += produced_quantity
            vol_row["amount"] += amount
            vol_row["sold_quantity"] += sold_quantity
            vol_row["sale_revenue"] += sale_revenue

        day_row = by_day.setdefault(row.work_date, {"work_date": row.work_date, "count": 0, "amount": Decimal("0"), "bouquets": 0, "baskets": 0, "sold_quantity": 0, "sale_revenue": Decimal("0")})
        day_row["count"] += produced_quantity
        day_row["amount"] += amount
        if is_production:
            day_row["sold_quantity"] += sold_quantity
            day_row["sale_revenue"] += sale_revenue
            if arrangement == "bouquet":
                day_row["bouquets"] += produced_quantity
            elif arrangement == "basket":
                day_row["baskets"] += produced_quantity

        entries.append({
            "id": row.id,
            "work_date": row.work_date,
            "created_at": row.created_at,
            "source": row.source,
            "source_label": row.get_source_display(),
            "amount": amount,
            "note": row.note,
            "added_by": user_full_name(row.created_by),
            "catalog_item_id": row.catalog_item_id,
            "catalog_name": item.name_uz if item else "",
            "catalog_kind": kind,
            "catalog_kind_label": CATALOG_KIND_LABELS.get(kind, kind),
            "arrangement_type": arrangement,
            "arrangement_label": ARRANGEMENT_LABELS.get(arrangement, ""),
            "volume": item.volume if item else "",
            "quantity_total": int(item.quantity_total or 0) if item else 0,
            "quantity_sold": int(item.quantity_sold or 0) if item else 0,
            "listed_price": Decimal(item.price or 0) if item else Decimal("0"),
            "sold_quantity": sold_quantity,
            "sale_revenue": sale_revenue,
            "last_sold_at": sold.get("last_sold_at"),
            "is_sold": bool(sold_quantity),
        })

    if summary["catalog_count"]:
        summary["avg_fee_per_item"] = (summary["catalog_salary_total"] / Decimal(summary["catalog_count"])).quantize(Decimal("0.01"))

    if not include_sales:
        for key in ["sold_quantity", "sale_revenue", "unsold_quantity"]:
            summary.pop(key, None)
        for row in entries:
            for key in ["listed_price", "sold_quantity", "sale_revenue", "last_sold_at", "is_sold", "quantity_sold"]:
                row.pop(key, None)
        for bucket in (by_arrangement, by_volume, by_day):
            for row in bucket.values():
                for key in ["sold_quantity", "sale_revenue"]:
                    row.pop(key, None)

    attendance = FloristAttendance.objects.filter(florist=profile)
    if date_from:
        attendance = attendance.filter(work_date__gte=date_from)
    if date_to:
        attendance = attendance.filter(work_date__lte=date_to)
    attendance = list(attendance.order_by("-work_date", "-id"))
    summary["attendance_days"] = len(attendance)
    if include_sales:
        summary["unsold_quantity"] = max(summary["catalog_count"] - summary["sold_quantity"], 0)

    return {
        "florist": {
            "id": profile.id,
            "name": profile.user.get_full_name() or profile.user.username,
            "username": profile.user.username,
            "staff_type": profile.staff_type,
            "staff_type_label": profile.get_staff_type_display(),
            "phone": profile.phone,
            "daily_pay": Decimal(profile.daily_pay or 0),
            "is_active": profile.is_active,
        },
        "period": {"date_from": date_from.isoformat() if date_from else None, "date_to": date_to.isoformat() if date_to else None},
        "summary": summary,
        "by_source": sorted(by_source.values(), key=lambda row: -row["amount"]),
        "by_arrangement": sorted(by_arrangement.values(), key=lambda row: -row["count"]),
        "by_volume": sorted(by_volume.values(), key=lambda row: (row["arrangement_type"], row["volume"])),
        "by_day": sorted(by_day.values(), key=lambda row: row["work_date"], reverse=True),
        "salary_entries": entries,
        "attendance": [
            {
                "id": row.id,
                "work_date": row.work_date,
                "check_in_at": row.check_in_at,
                "check_out_at": row.check_out_at,
                "source": row.source,
                "source_label": row.get_source_display(),
                "note": row.note,
            }
            for row in attendance
        ],
    }


class FloristProfileViewSet(ScopedViewSet):
    permission_page = "florists"
    write_roles = ["admin", "supervisor"]
    queryset = FloristProfile.objects.select_related("user").all()
    serializer_class = FloristProfileSerializer
    filterset_fields = ["staff_type", "is_active"]
    search_fields = ["user__first_name", "user__last_name", "user__username", "phone"]

    def get_queryset(self):
        catalog_quantity = CatalogItem.objects.filter(florist=OuterRef("pk")).values("florist").annotate(total=Coalesce(Sum("quantity_total"), 0)).values("total")[:1]
        queryset = super().get_queryset().annotate(
            salary_total=Coalesce(Sum("salary_entries__amount"), Decimal("0")),
            catalog_count=Coalesce(Subquery(catalog_quantity, output_field=IntegerField()), 0),
        )
        role = getattr(getattr(self.request.user, "profile", None), "role", None)
        if role in ["florist", "apprentice"]:
            return queryset.filter(user=self.request.user)
        return queryset

    @action(detail=False, methods=["get"], url_path="me")
    def me_profile(self, request):
        profile = FloristProfile.objects.select_related("user").filter(user=request.user).first()
        if not profile:
            return Response({"detail": "Florist profili topilmadi"}, status=status.HTTP_404_NOT_FOUND)
        return Response(self.get_serializer(profile).data)

    @action(detail=True, methods=["get"], url_path="stats")
    def florist_stats(self, request, pk=None):
        profile = self.get_object()
        return Response(json_safe(florist_stats_data(profile, request)))

    @action(detail=False, methods=["get"], url_path="me/dashboard")
    def me_dashboard(self, request):
        profile = FloristProfile.objects.select_related("user").filter(user=request.user).first()
        if not profile:
            return Response({"detail": "Florist profili topilmadi"}, status=status.HTTP_404_NOT_FOUND)
        # Florist o'z sahifasida sotuv narxi, tushum va foydani ko'rmaydi.
        # Faqat nechta yasagani va ish haqiga qancha qo'shilgani ko'rinadi.
        return Response(json_safe(florist_stats_data(profile, request, include_sales=False)))


class FloristStockIssueViewSet(viewsets.ReadOnlyModelViewSet):
    """Skladdan floristga chiqarilgan va qaytarilgan gullar tarixi."""

    permission_classes = [RolePermission]
    permission_page = "inventory"
    queryset = FloristStockIssue.objects.select_related("florist__user", "batch__variant__flower", "performed_by").all()
    serializer_class = FloristStockIssueSerializer
    filterset_fields = ["florist", "batch", "kind"]
    ordering_fields = ["created_at", "quantity_stems"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if has_page_permission(self.request.user, "inventory", False):
            return queryset
        profile = FloristProfile.objects.filter(user=self.request.user).first()
        return queryset.filter(florist=profile) if profile else queryset.none()

    @extend_schema(request=FloristStockIssueEditSerializer, responses=FloristStockIssueSerializer)
    @action(detail=True, methods=["patch"], url_path="edit")
    def edit(self, request, pk=None):
        """Chiqarilgan gul soni noto'g'ri yozilgan bo'lsa to'g'rilaydi."""
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        serializer = FloristStockIssueEditSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            issue = edit_florist_stock_issue(
                self.get_object(),
                serializer.validated_data["quantity_stems"],
                serializer.validated_data.get("reason"),
                request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(FloristStockIssueSerializer(issue).data)

    @extend_schema(request=None, responses=OpenApiResponse(description="Bekor qilindi"))
    @action(detail=True, methods=["delete"], url_path="cancel")
    def cancel(self, request, pk=None):
        """Chiqim yozuvini butunlay bekor qiladi va qoldiqlarni tiklaydi."""
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        try:
            delete_florist_stock_issue(self.get_object(), request.user)
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @extend_schema(request=FloristStockIssueRequestSerializer, responses=FloristStockIssueSerializer)
    @action(detail=False, methods=["post"], url_path="issue")
    def issue(self, request):
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        serializer = FloristStockIssueRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            issue = issue_stock_to_florist(
                serializer.validated_data["florist"], serializer.validated_data["batch"],
                serializer.validated_data["quantity_stems"], serializer.validated_data.get("reason", ""), request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        write_audit(request.user, "floriststockissue_created", issue, before={}, after=instance_snapshot(issue), request=request)
        return Response(FloristStockIssueSerializer(issue).data, status=status.HTTP_201_CREATED)

    @extend_schema(request=FloristStockIssueBulkRequestSerializer, responses=FloristStockIssueSerializer(many=True))
    @action(detail=False, methods=["post"], url_path="bulk-issue")
    def bulk_issue(self, request):
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        serializer = FloristStockIssueBulkRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            issues = issue_multiple_stock_to_florist(
                serializer.validated_data["florist"],
                serializer.validated_data["items"],
                serializer.validated_data.get("reason", ""),
                request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(FloristStockIssueSerializer(issues, many=True).data, status=status.HTTP_201_CREATED)

    @extend_schema(request=FloristStockReturnRequestSerializer, responses=FloristStockIssueSerializer)
    @action(detail=False, methods=["post"], url_path="return")
    def return_stock(self, request):
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        serializer = FloristStockReturnRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            issue = return_stock_from_florist(
                serializer.validated_data["florist"], serializer.validated_data["batch"],
                serializer.validated_data["quantity_stems"], serializer.validated_data.get("reason", ""),
                request.user, serializer.validated_data.get("kind", "return"),
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        write_audit(request.user, "floriststockissue_created", issue, before={}, after=instance_snapshot(issue), request=request)
        return Response(FloristStockIssueSerializer(issue).data, status=status.HTTP_201_CREATED)


class FloristStockBalanceViewSet(viewsets.ReadOnlyModelViewSet):
    """Floristlarning qo'lidagi gul qoldig'i. Katalog qo'shishda shu ro'yxatdan tanlanadi."""

    permission_classes = [RolePermission]
    permission_page = "inventory"
    queryset = FloristStockBalance.objects.select_related("florist__user", "batch__variant__flower").all()
    serializer_class = FloristStockBalanceSerializer
    filterset_fields = ["florist", "batch"]
    ordering_fields = ["remaining_stems"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.query_params.get("only_available", "true").lower() != "false":
            queryset = queryset.filter(remaining_stems__gt=0)
        if has_page_permission(self.request.user, "inventory", False):
            return queryset
        profile = FloristProfile.objects.filter(user=self.request.user).first()
        return queryset.filter(florist=profile) if profile else queryset.none()

    @extend_schema(
        parameters=[
            OpenApiParameter("florist", int, description="Florist id — majburiy"),
            OpenApiParameter("batch", int, description="Bitta partiya bo‘yicha. to_florist da majburiy."),
            OpenApiParameter("direction", str, description="to_catalog (sukut) yoki to_florist"),
            OpenApiParameter("quantity_stems", int, description="Faqat to_florist uchun: qaytariladigan gul soni"),
        ],
        responses=OpenApiResponse(description="Qaysi katalogga qancha o‘zgarishi. Hech narsa o‘zgarmaydi."),
    )
    @action(detail=False, methods=["get"], url_path="adjust-preview")
    def adjust_preview(self, request):
        """Hisobni to'g'rilashdan oldin nima bo'lishini ko'rsatadi."""
        if not has_page_permission(request.user, "inventory", False):
            return forbidden()
        florist = FloristProfile.objects.filter(pk=request.query_params.get("florist")).first()
        if not florist:
            return Response({"detail": "Florist tanlanmadi."}, status=status.HTTP_400_BAD_REQUEST)
        batch = StockBatch.objects.filter(pk=request.query_params.get("batch")).first() if request.query_params.get("batch") else None
        direction = request.query_params.get("direction") or "to_catalog"
        if direction not in ["to_catalog", "to_florist"]:
            return Response({"detail": "Yo‘nalish noto‘g‘ri."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            rows = florist_stem_plan(florist, batch, direction, request.query_params.get("quantity_stems"))
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({
            "florist": florist.id,
            "florist_name": str(florist),
            "direction": direction,
            "batches": rows,
            "total_florist_stems": sum(row["florist_stems_now"] for row in rows),
            "blocked_count": sum(1 for row in rows if row["blocked"]),
        })

    @extend_schema(
        parameters=[
            OpenApiParameter("florist", int, description="Florist id — majburiy"),
            OpenApiParameter("batch", int, description="Partiyadagi gul id — majburiy"),
            OpenApiParameter("return_stems", int, description="Skladga qaytariladigan ortiqcha son"),
        ],
        responses=OpenApiResponse(description="Chiqim yopilganda nima bo‘lishi. Hech narsa o‘zgarmaydi."),
    )
    @action(detail=False, methods=["get"], url_path="close-issue-preview")
    def close_issue_preview(self, request):
        """Chiqimni yopishdan oldin taqsimotni ko'rsatadi."""
        if not has_page_permission(request.user, "inventory", False):
            return forbidden()
        florist = FloristProfile.objects.filter(pk=request.query_params.get("florist")).first()
        batch = StockBatch.objects.filter(pk=request.query_params.get("batch")).first()
        if not florist or not batch:
            return Response({"detail": "Florist va gul tanlanishi kerak."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            return_stems = int(request.query_params.get("return_stems") or 0)
        except (TypeError, ValueError):
            return Response({"detail": "Qaytariladigan son noto‘g‘ri."}, status=status.HTTP_400_BAD_REQUEST)
        plan = florist_close_plan(florist, batch, return_stems, True)
        return Response({"florist": florist.id, "florist_name": str(florist), **plan})

    @extend_schema(request=FloristCloseIssueSerializer, responses=OpenApiResponse(description="Yopish natijasi"))
    @action(detail=False, methods=["post"], url_path="close-issue")
    def close_issue(self, request):
        """Chiqarilgan gul tugadi: ortig'i skladga qaytadi, qolgani kataloglarga bo'linadi."""
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        serializer = FloristCloseIssueSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            if serializer.validated_data.get("close_all"):
                result = close_all_florist_issues(
                    serializer.validated_data["florist"],
                    request.user,
                    True,
                )
            else:
                result = close_florist_issue(
                    serializer.validated_data["florist"],
                    serializer.validated_data["batch"],
                    serializer.validated_data.get("return_stems") or 0,
                    request.user,
                    serializer.validated_data.get("absorb_remainder", False),
                )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result)

    @extend_schema(request=FloristLeftoverRequestSerializer, responses=OpenApiResponse(description="To‘g‘rilash natijasi"))
    @action(detail=False, methods=["post"], url_path="adjust")
    def adjust(self, request):
        """Florist standartdan farqli gul ishlatganda hisobni to'g'rilaydi."""
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        serializer = FloristLeftoverRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            result = adjust_florist_stems(
                serializer.validated_data["florist"],
                serializer.validated_data.get("batch"),
                serializer.validated_data.get("direction", "to_catalog"),
                serializer.validated_data.get("quantity_stems"),
                request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result)


class FloristDayOffViewSet(ScopedViewSet):
    permission_page = "attendance"
    write_roles = ["admin", "supervisor"]
    queryset = FloristDayOff.objects.select_related("florist__user", "created_by").all()
    serializer_class = FloristDayOffSerializer
    filterset_fields = ["florist", "kind", "work_date", "is_paid"]
    ordering_fields = ["work_date", "created_at"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if has_page_permission(self.request.user, "attendance", True):
            return queryset
        profile = FloristProfile.objects.filter(user=self.request.user).first()
        return queryset.filter(florist=profile) if profile else queryset.none()

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)
        write_audit(self.request.user, "floristdayoff_created", instance, before={}, after=instance_snapshot(instance), request=self.request)


class FloristFaceSampleViewSet(viewsets.ModelViewSet):
    """Floristning yuz namunalari. Ro'yxatdan o'tkazish uchun."""

    permission_classes = [RolePermission]
    permission_page = "attendance"
    queryset = FloristFaceSample.objects.select_related("florist__user").all()
    serializer_class = FloristFaceSampleSerializer
    filterset_fields = ["florist", "is_active"]

    def create(self, request, *args, **kwargs):
        if not has_page_permission(request.user, "attendance", True):
            return forbidden()
        florist_id = request.data.get("florist")
        profile = FloristProfile.objects.filter(id=florist_id).first()
        if not profile:
            return Response({"florist": ["Florist topilmadi"]}, status=status.HTTP_400_BAD_REQUEST)
        source = request.FILES.get("image") or request.data.get("image_base64")
        if not source:
            return Response({"image": ["Rasm yuboring: image fayli yoki image_base64"]}, status=status.HTTP_400_BAD_REQUEST)
        try:
            face = face_services.face_from_source(source)
        except face_services.FaceError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        image_url = ""
        uploaded = request.FILES.get("image")
        if uploaded:
            uploaded.seek(0)
            _, media_url = face_services.save_face_image(uploaded, profile.id)
            image_url = f"{settings.PUBLIC_BASE_URL}{media_url}" if settings.PUBLIC_BASE_URL else request.build_absolute_uri(media_url)
        sample = FloristFaceSample.objects.create(
            florist=profile, image_url=image_url,
            descriptor=face_services.encode_face(face),
            created_by=request.user if request.user.is_authenticated else None,
        )
        face_services.invalidate_cache()
        write_audit(request.user, "floristfacesample_created", sample, before={}, after={"florist": str(profile)}, request=request)
        return Response(self.get_serializer(sample).data, status=status.HTTP_201_CREATED)

    def perform_destroy(self, instance):
        instance.delete()
        face_services.invalidate_cache()


class AttendanceDeviceView(APIView):
    """Do'kondagi qurilma uchun. Yuz rasmidan floristni tanib keldi-ketdi qiladi.
    Lokatsiya so'ralmaydi."""

    permission_classes = [RolePermission]
    permission_page = "attendance"

    @extend_schema(
        request=inline_serializer(name="AttendanceDeviceRequest", fields={
            "image_base64": serializers.CharField(required=False),
            "action": serializers.ChoiceField(choices=["auto", "check_in", "check_out"], required=False),
        }),
        responses=OpenApiResponse(description="Tanilgan florist va keldi-ketdi holati"),
    )
    def post(self, request):
        if not has_page_permission(request.user, "attendance", True):
            return forbidden()
        source = request.FILES.get("image") or request.data.get("image_base64")
        if not source:
            return Response({"detail": "Rasm yuboring: image fayli yoki image_base64"}, status=status.HTTP_400_BAD_REQUEST)
        samples = list(FloristFaceSample.objects.filter(is_active=True))
        try:
            florist_id, confidence = face_services.recognize(source, samples)
        except face_services.FaceError as exc:
            return Response({"detail": str(exc), "recognized": False}, status=status.HTTP_400_BAD_REQUEST)
        profile = FloristProfile.objects.select_related("user").filter(id=florist_id, is_active=True).first()
        if not profile:
            return Response({"detail": "Florist topilmadi yoki faol emas", "recognized": False}, status=status.HTTP_404_NOT_FOUND)
        now = timezone.now()
        work_date = timezone.localtime(now).date()
        if FloristDayOff.objects.filter(florist=profile, work_date=work_date).exists():
            return Response({"detail": f"{profile} bugun dam olish kunida", "recognized": True, "florist": {"id": profile.id, "name": str(profile)}}, status=status.HTTP_400_BAD_REQUEST)
        row, _ = FloristAttendance.objects.get_or_create(florist=profile, work_date=work_date, defaults={"source": "device"})
        requested = (request.data.get("action") or "auto").lower()
        if requested == "auto":
            requested = "check_out" if row.check_in_at else "check_in"
        if requested == "check_in":
            if row.check_in_at:
                return Response({"detail": f"{profile} bugun allaqachon ishga kelgan", "recognized": True, "action": "check_in", "already": True, "florist": {"id": profile.id, "name": str(profile)}, "check_in_at": row.check_in_at}, status=status.HTTP_200_OK)
            row.check_in_at = now
        else:
            if not row.check_in_at:
                return Response({"detail": f"{profile} bugun hali ishga kelmagan", "recognized": True, "action": "check_out"}, status=status.HTTP_400_BAD_REQUEST)
            row.check_out_at = now
        row.source = "device"
        row.save(update_fields=["check_in_at", "check_out_at", "source", "updated_at"])
        write_audit(request.user, f"attendance_{requested}", row, before={}, after=instance_snapshot(row), request=request, summary=f"{profile} qurilmadan {requested}")
        create_user_notification(profile.user, "attendance", "Ish vaqti belgilandi", f"{timezone.localtime(now).strftime('%H:%M')} da {'ishga kelish' if requested == 'check_in' else 'ishdan ketish'} belgilandi.", "attendance", row.id)
        return Response({
            "recognized": True,
            "confidence": confidence,
            "action": requested,
            "florist": {"id": profile.id, "name": str(profile), "staff_type": profile.staff_type},
            "work_date": work_date,
            "check_in_at": row.check_in_at,
            "check_out_at": row.check_out_at,
        })


class FloristAttendanceViewSet(ScopedViewSet):
    permission_page = "attendance"
    write_roles = ["admin", "supervisor", "florist", "apprentice"]
    queryset = FloristAttendance.objects.select_related("florist__user").all()
    serializer_class = FloristAttendanceSerializer
    filterset_class = FloristAttendanceFilter

    def get_queryset(self):
        queryset = super().get_queryset()
        if has_page_permission(self.request.user, "attendance", True):
            return queryset
        profile = FloristProfile.objects.filter(user=self.request.user).first()
        return queryset.filter(florist=profile) if profile else queryset.none()

    def _profile_from_request(self, request):
        florist_id = request.data.get("florist")
        if florist_id and has_page_permission(request.user, "attendance", True):
            profile = FloristProfile.objects.filter(id=florist_id).first()
        else:
            profile = FloristProfile.objects.filter(user=request.user).first()
        if not profile:
            raise serializers.ValidationError({"florist": "Florist profili topilmadi"})
        return profile

    @action(detail=False, methods=["post"], url_path="check-in")
    def check_in(self, request):
        profile = self._profile_from_request(request)
        checked_at = parse_datetime(request.data.get("checked_at") or "") or timezone.now()
        if timezone.is_naive(checked_at):
            checked_at = timezone.make_aware(checked_at)
        work_date = timezone.localtime(checked_at).date()
        row, _ = FloristAttendance.objects.get_or_create(florist=profile, work_date=work_date, defaults={"source": request.data.get("source") or "mobile"})
        first_check_in = not row.check_in_at
        row.check_in_at = row.check_in_at or checked_at
        row.check_in_latitude = normalize_coordinate(request.data.get("latitude")) or row.check_in_latitude
        row.check_in_longitude = normalize_coordinate(request.data.get("longitude")) or row.check_in_longitude
        row.source = request.data.get("source") or row.source
        row.note = request.data.get("note") or row.note
        row.save(update_fields=["check_in_at", "check_in_latitude", "check_in_longitude", "source", "note", "updated_at"])
        write_audit(request.user, "attendance_check_in", row, before={}, after=instance_snapshot(row), request=request, summary=f"{profile} ishga keldi")
        if first_check_in:
            create_user_notification(profile.user, "attendance", "Ishga keldingiz", f"{timezone.localtime(row.check_in_at).strftime('%Y-%m-%d %H:%M')} da ishga kelganingiz belgilandi.", "attendance", row.id)
            if getattr(getattr(profile.user, "profile", None), "role", None) != "developer":
                Notification.objects.create(notification_type="attendance", title_uz=f"{profile} ishga keldi", title_ru=f"{profile} ishga keldi", body_uz=f"{profile} {timezone.localtime(row.check_in_at).strftime('%Y-%m-%d %H:%M')} da ishga keldi.", body_ru=f"{profile} {timezone.localtime(row.check_in_at).strftime('%Y-%m-%d %H:%M')} da ishga keldi.", reference_type="attendance", reference_id=row.id)
        return Response(self.get_serializer(row).data)

    @action(detail=False, methods=["post"], url_path="check-out")
    def check_out(self, request):
        profile = self._profile_from_request(request)
        checked_at = parse_datetime(request.data.get("checked_at") or "") or timezone.now()
        if timezone.is_naive(checked_at):
            checked_at = timezone.make_aware(checked_at)
        work_date = timezone.localtime(checked_at).date()
        row, _ = FloristAttendance.objects.get_or_create(florist=profile, work_date=work_date, defaults={"source": request.data.get("source") or "mobile"})
        first_check_out = not row.check_out_at
        row.check_out_at = checked_at
        row.check_out_latitude = normalize_coordinate(request.data.get("latitude")) or row.check_out_latitude
        row.check_out_longitude = normalize_coordinate(request.data.get("longitude")) or row.check_out_longitude
        row.source = request.data.get("source") or row.source
        row.note = request.data.get("note") or row.note
        row.save(update_fields=["check_out_at", "check_out_latitude", "check_out_longitude", "source", "note", "updated_at"])
        if profile.staff_type == "apprentice" and profile.daily_pay:
            existing_salary = FloristSalaryEntry.objects.filter(florist=profile, source="daily", attendance=row).first()
            old_amount = existing_salary.amount if existing_salary else None
            salary, _ = FloristSalaryEntry.objects.update_or_create(florist=profile, source="daily", attendance=row, defaults={"amount": profile.daily_pay, "work_date": work_date, "note": "Shogird kunlik ish haqi", "created_by": request.user})
            write_audit(request.user, "apprentice_daily_salary_recorded", salary, before={}, after=instance_snapshot(salary), request=request, summary=f"{profile} uchun kunlik ish haqi yozildi")
            if old_amount != salary.amount:
                create_user_notification(profile.user, "florist_salary", "Kunlik ish haqi yozildi", f"{work_date} uchun {salary.amount} so‘m kunlik ish haqi yozildi.", "florist_salary", salary.id)
        write_audit(request.user, "attendance_check_out", row, before={}, after=instance_snapshot(row), request=request, summary=f"{profile} ishdan ketdi")
        if first_check_out:
            create_user_notification(profile.user, "attendance", "Ishdan ketdingiz", f"{timezone.localtime(row.check_out_at).strftime('%Y-%m-%d %H:%M')} da ishdan ketganingiz belgilandi.", "attendance", row.id)
        return Response(self.get_serializer(row).data)


class FloristSalaryEntryViewSet(ScopedViewSet):
    permission_page = "florists"
    write_roles = ["admin", "supervisor"]
    queryset = FloristSalaryEntry.objects.select_related("florist__user", "catalog_item", "attendance", "created_by").all()
    serializer_class = FloristSalaryEntrySerializer
    filterset_class = FloristSalaryEntryFilter

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(getattr(self.request.user, "profile", None), "role", None)
        if role in ["florist", "apprentice"]:
            profile = FloristProfile.objects.filter(user=self.request.user).first()
            return queryset.filter(florist=profile) if profile else queryset.none()
        return queryset

    def perform_create(self, serializer):
        entry = serializer.save(created_by=self.request.user)
        write_audit(self.request.user, "florist_salary_created", entry, before={}, after=instance_snapshot(entry), request=self.request, summary=f"{entry.florist} uchun ish haqi qo‘shildi")
        create_user_notification(entry.florist.user, "florist_salary", "Ish haqi qo‘shildi", f"{entry.work_date} uchun {entry.amount} so‘m ish haqi qo‘shildi.", "florist_salary", entry.id)

    def perform_update(self, serializer):
        before = instance_snapshot(serializer.instance)
        entry = serializer.save()
        before_changed, after_changed = changed_snapshot(before, instance_snapshot(entry))
        if before_changed or after_changed:
            write_audit(self.request.user, "florist_salary_updated", entry, before=before_changed, after=after_changed, request=self.request, summary=f"{entry.florist} ish haqi o‘zgartirildi")
            create_user_notification(entry.florist.user, "florist_salary", "Ish haqi o‘zgartirildi", f"{entry.work_date} uchun ish haqi {entry.amount} so‘m qilib yangilandi.", "florist_salary", entry.id)


class StockDeliveryViewSet(ScopedViewSet):
    """Partiya. Avval ochiladi, keyin ichiga gullar qo'shiladi."""

    permission_page = "inventory"
    write_roles = ["admin", "warehouse"]
    queryset = StockDelivery.objects.select_related("supplier", "created_by").prefetch_related("batches").all()
    serializer_class = StockDeliverySerializer
    filterset_fields = ["supplier", "is_active", "received_at"]
    search_fields = ["number", "note", "supplier__name"]
    ordering_fields = ["received_at", "number", "created_at"]

    def perform_create(self, serializer):
        delivery = serializer.save(created_by=self.request.user)
        write_audit(self.request.user, "stock_delivery_created", delivery, before={}, after=instance_snapshot(delivery), request=self.request, summary=f"{delivery.number} partiya ochildi")

    @extend_schema(responses=StockBatchSerializer(many=True))
    @action(detail=True, methods=["get"])
    def batches(self, request, pk=None):
        """Partiya ichidagi gullar."""
        rows = StockBatch.objects.filter(delivery_id=pk).select_related("variant__flower", "supplier", "delivery")
        return Response(StockBatchSerializer(rows, many=True).data)


class StockBatchViewSet(ScopedViewSet):
    permission_page = "inventory"
    write_roles = ["admin", "warehouse"]
    queryset = StockBatch.objects.select_related("variant__flower", "supplier", "delivery").all()
    serializer_class = StockBatchSerializer
    filterset_class = StockBatchFilter
    search_fields = ["batch_number", "variant__flower__name_uz", "variant__name_uz", "variant__color_uz", "supplier__name", "supplier__phone"]
    ordering_fields = ["received_at", "remaining_stems", "sale_price_per_stem", "height_cm", "height_from_cm", "height_to_cm"]

    def perform_create(self, serializer):
        batch = serializer.save()
        StockMovement.objects.create(batch=batch, movement_type="in", quantity_stems=batch.received_stems, quantity_bunches=batch.received_stems / batch.stems_per_bunch, reason="Partiya kirimi", performed_by=self.request.user)
        write_audit(self.request.user, "stock_received", batch, before={}, after=instance_snapshot(batch), request=self.request, summary=f"{batch.batch_number} partiya kirim qilindi")
        if batch.supplier_id:
            title = "Yangi gul kirimi"
            body = f"{batch.supplier.name} postavshikdan {batch.variant.flower.name_uz} {batch.variant.name_uz} {batch.variant.color_uz} keldi. Partiya: {batch.batch_number}. Miqdor: {batch.received_stems} dona."
            Notification.objects.create(notification_type="supplier_stock", title_uz=title, title_ru=title, body_uz=body, body_ru=body, reference_type="stock_batch", reference_id=batch.id)
            integration, _ = IntegrationSettings.objects.get_or_create(pk=1)
            group_chat_id = integration.telegram_group_chat_id or settings.TELEGRAM_GROUP_CHAT_ID
            if group_chat_id:
                try:
                    telegram_send(group_chat_id, f"{title}\n{body}")
                except Exception as exc:
                    print(f"SUPPLIER_STOCK_TELEGRAM_FAILED batch={batch.id} error={exc}", flush=True)

    def perform_update(self, serializer):
        """Kelgan son tahrirlanganda qoldiq va kirim yozuvi ham to'g'rilanadi.

        Xato kiritilgan son tuzatilganda allaqachon ishlatilgan gul unutilib
        qolmasligi kerak: qoldiq farq qancha bo'lsa o'shancha siljiydi.
        """
        instance = serializer.instance
        old_received = instance.received_stems
        old_remaining = instance.remaining_stems
        used = max(old_received - old_remaining, 0)
        new_received = serializer.validated_data.get("received_stems", old_received)
        explicit_remaining = "remaining_stems" in serializer.validated_data
        if new_received != old_received and not explicit_remaining and new_received < used:
            raise serializers.ValidationError({
                "received_stems": f"Bu partiyadan allaqachon {used} dona ishlatilgan. "
                                  f"Kelgan sonni undan kam qilib bo‘lmaydi.",
            })
        batch = serializer.save()
        if new_received != old_received and not explicit_remaining:
            batch.remaining_stems = max(new_received - used, 0)
            batch.save(update_fields=["remaining_stems", "updated_at"])
            # kirim yozuvi ham yangi songa moslanadi, aks holda jurnal to'g'ri kelmaydi
            movement = StockMovement.objects.filter(batch=batch, movement_type="in").order_by("created_at", "id").first()
            if movement:
                movement.quantity_stems = new_received
                movement.quantity_bunches = Decimal(new_received) / Decimal(batch.stems_per_bunch or 1)
                movement.save(update_fields=["quantity_stems", "quantity_bunches", "updated_at"])
            write_audit(
                self.request.user, "stock_batch_quantity_edited", batch,
                before={"received_stems": old_received, "remaining_stems": old_remaining},
                after={"received_stems": batch.received_stems, "remaining_stems": batch.remaining_stems, "used": used},
                request=self.request,
                summary=f"{batch.batch_number} partiyada kelgan son {old_received} dan {new_received} ga o‘zgartirildi",
            )

    def destroy(self, request, *args, **kwargs):
        batch = self.get_object()
        try:
            return super().destroy(request, *args, **kwargs)
        except ProtectedError:
            batch.is_active = False
            batch.save(update_fields=["is_active", "updated_at"])
            return Response({"detail": "Bu partiyada sklad tarixi bor. Partiya o‘chirilmadi, is_active=false qilib arxivlandi.", "id": batch.id, "is_active": batch.is_active})

    @extend_schema(request=MovementRequestSerializer, responses=StockMovementSerializer)
    @action(detail=True, methods=["post"])
    def movement(self, request, pk=None):
        batch = self.get_object()
        serializer = MovementRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            movement = apply_stock_movement(batch, serializer.validated_data["movement_type"], serializer.validated_data.get("quantity_stems"), serializer.validated_data.get("reason", ""), request.user, serializer.validated_data.get("quantity_bunches"))
            backdate_record(movement, serializer.validated_data.get("created_at"))
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(StockMovementSerializer(movement).data)


class StockMovementViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [RolePermission]
    permission_page = "inventory"
    queryset = StockMovement.objects.select_related("batch__variant__flower", "performed_by").all()
    serializer_class = StockMovementSerializer
    filterset_class = StockMovementFilter


class PackagingViewSet(ScopedViewSet):
    permission_page = "inventory"
    write_roles = ["admin", "warehouse"]
    queryset = Packaging.objects.all()
    serializer_class = PackagingSerializer
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    filterset_fields = ["packaging_type", "is_active", "unit", "basket_material", "size"]
    search_fields = ["name_uz", "size"]

    def perform_create(self, serializer):
        packaging = serializer.save()
        # yukka bog'lab qo'shilgan bo'lsa kirim yozuvi allaqachon yaratilgan
        if packaging.quantity and not getattr(packaging, "received_via_delivery", False):
            movement = PackagingMovement.objects.create(packaging=packaging, movement_type="in", quantity=packaging.quantity, reason="Qadoq/savat kirimi", performed_by=self.request.user)
            write_audit(self.request.user, "packaging_received", packaging, before={}, after={**instance_snapshot(packaging), "movement": movement.id}, request=self.request, summary=f"{packaging.name_uz} material kirim qilindi")

    def perform_update(self, serializer):
        before = serializer.instance.quantity
        packaging = serializer.save()
        if "quantity" in serializer.validated_data and packaging.quantity != before:
            delta = packaging.quantity - before
            movement = PackagingMovement.objects.create(packaging=packaging, movement_type="adjustment", quantity=delta, reason="Qadoq/savat qoldig‘i tahrirlandi", performed_by=self.request.user)
            write_audit(self.request.user, "packaging_adjusted", packaging, before={"quantity": before}, after={**instance_snapshot(packaging), "movement": movement.id}, request=self.request, summary=f"{packaging.name_uz} material qoldig‘i o‘zgartirildi")

    @extend_schema(request=PackagingMovementRequestSerializer, responses=PackagingMovementSerializer)
    @action(detail=True, methods=["post"])
    def movement(self, request, pk=None):
        packaging = self.get_object()
        serializer = PackagingMovementRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            movement = apply_packaging_movement(packaging, serializer.validated_data["movement_type"], serializer.validated_data["quantity"], serializer.validated_data.get("reason", ""), request.user)
            backdate_record(movement, serializer.validated_data.get("created_at"))
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(PackagingMovementSerializer(movement).data)


class PackagingMovementViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [RolePermission]
    permission_page = "inventory"
    queryset = PackagingMovement.objects.select_related("packaging", "performed_by", "delivery__supplier").all()
    serializer_class = PackagingMovementSerializer
    filterset_class = PackagingMovementFilter


class MaterialDeliveryViewSet(ScopedViewSet):
    """Material partiyasi. Avval ochiladi, keyin ichiga materiallar kiritiladi."""

    permission_page = "inventory"
    write_roles = ["admin", "warehouse"]
    queryset = MaterialDelivery.objects.select_related("supplier", "created_by").prefetch_related("movements__packaging").all()
    serializer_class = MaterialDeliverySerializer
    filterset_fields = ["supplier", "is_active", "received_at"]
    search_fields = ["number", "note", "supplier__name"]
    ordering_fields = ["received_at", "number", "created_at"]

    def perform_create(self, serializer):
        delivery = serializer.save(created_by=self.request.user)
        write_audit(self.request.user, "material_delivery_created", delivery, before={}, after=instance_snapshot(delivery), request=self.request, summary=f"{delivery.number} material partiyasi ochildi")

    @extend_schema(responses=PackagingMovementSerializer(many=True))
    @action(detail=True, methods=["get"])
    def items(self, request, pk=None):
        """Partiya ichiga kiritilgan materiallar."""
        rows = PackagingMovement.objects.filter(delivery_id=pk, movement_type="in").select_related("packaging", "performed_by", "delivery__supplier")
        return Response(PackagingMovementSerializer(rows, many=True).data)

    @extend_schema(request=MaterialReceiveSerializer, responses=PackagingMovementSerializer)
    @action(detail=True, methods=["post"])
    def receive(self, request, pk=None):
        """Partiyaga material kiritadi: soni oshadi, tannarxi yangilanadi."""
        if not has_page_permission(request.user, "inventory", True):
            return forbidden()
        delivery = self.get_object()
        serializer = MaterialReceiveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            movement = receive_material_into_delivery(
                delivery,
                serializer.validated_data["packaging"],
                serializer.validated_data["quantity"],
                serializer.validated_data.get("cost_price"),
                serializer.validated_data.get("reason", ""),
                request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(PackagingMovementSerializer(movement).data, status=status.HTTP_201_CREATED)


class InventoryMovementJournalView(APIView):
    permission_classes = [RolePermission]
    permission_page = "inventory"

    @extend_schema(
        parameters=[
            OpenApiParameter("stock_type", str, OpenApiParameter.QUERY, description="all, flower yoki material"),
            OpenApiParameter("movement_type", str, OpenApiParameter.QUERY, description="in, out, adjustment, waste"),
            OpenApiParameter("date_from", str, OpenApiParameter.QUERY),
            OpenApiParameter("date_to", str, OpenApiParameter.QUERY),
        ],
        responses=inline_serializer(name="InventoryMovementJournal", fields={"results": serializers.ListField(child=serializers.DictField())}),
    )
    def get(self, request):
        stock_type = request.query_params.get("stock_type", "all")
        movement_type = request.query_params.get("movement_type", "")
        date_from, date_to = parse_date_range_params(request)
        rows = []
        if stock_type in ["all", "flower"]:
            flower_rows = StockMovement.objects.select_related("batch__variant__flower", "batch__supplier", "performed_by").all()
            if movement_type:
                flower_rows = flower_rows.filter(movement_type=movement_type)
            flower_rows = filter_date_field(flower_rows, "created_at", date_from, date_to)
            for row in flower_rows[:500]:
                rows.append({
                    "id": f"flower-{row.id}",
                    "stock_type": "flower",
                    "stock_type_label": "Gul sklad",
                    "created_at": row.created_at,
                    "movement_type": row.movement_type,
                    "item_name": str(row.batch.variant),
                    "batch_number": row.batch.batch_number,
                    "quantity": row.quantity_stems,
                    "quantity_label": f"{row.quantity_stems} dona / {row.quantity_bunches} pochka",
                    "reason": row.reason,
                    "performed_by": row.performed_by_id,
                    "performed_by_name": user_full_name(row.performed_by),
                    "reference_type": row.reference_type,
                    "reference_id": row.reference_id,
                })
        if stock_type in ["all", "material"]:
            material_rows = PackagingMovement.objects.select_related("packaging", "performed_by").all()
            if movement_type:
                material_rows = material_rows.filter(movement_type=movement_type)
            material_rows = filter_date_field(material_rows, "created_at", date_from, date_to)
            for row in material_rows[:500]:
                rows.append({
                    "id": f"material-{row.id}",
                    "stock_type": "material",
                    "stock_type_label": "Material sklad",
                    "created_at": row.created_at,
                    "movement_type": row.movement_type,
                    "item_name": row.packaging.name_uz,
                    "batch_number": "",
                    "quantity": row.quantity,
                    "quantity_label": f"{row.quantity} dona",
                    "reason": row.reason,
                    "performed_by": row.performed_by_id,
                    "performed_by_name": user_full_name(row.performed_by),
                    "reference_type": row.reference_type,
                    "reference_id": row.reference_id,
                })
        rows = sorted(rows, key=lambda row: row["created_at"], reverse=True)[:500]
        for row in rows:
            row["created_at"] = local_datetime_label(row["created_at"])
        return Response({"results": rows})


class FloristSelfExcelExportView(APIView):
    """Florist o'z hisobotini yuklaydi. florists sahifasiga ruxsati bor foydalanuvchi
    ?florist=<id> bilan boshqa floristning hisobotini ham yuklashi mumkin."""

    permission_classes = [RolePermission]
    permission_page = None

    def get(self, request):
        requested_id = request.query_params.get("florist")
        if requested_id and has_page_permission(request.user, "florists", False):
            profile = FloristProfile.objects.select_related("user").filter(id=requested_id).first()
            if not profile:
                return Response({"detail": "Florist topilmadi"}, status=status.HTTP_404_NOT_FOUND)
            return export_florist_workbook(profile, request)
        profile = FloristProfile.objects.select_related("user").filter(user=request.user).first()
        if not profile:
            return Response({"detail": "Florist profile topilmadi"}, status=status.HTTP_404_NOT_FOUND)
        return export_florist_workbook(profile, request, include_sales=False)


class AdminFloristsExcelExportView(APIView):
    permission_classes = [RolePermission]
    permission_page = "florists"

    def get(self, request):
        if not has_page_permission(request.user, "florists", True):
            return forbidden()
        return export_all_florists_workbook(request)


class AdminProfitExcelExportView(APIView):
    permission_classes = [RolePermission]
    permission_page = "dashboard"

    def get(self, request):
        if not has_page_permission(request.user, "dashboard", False):
            return forbidden()
        return export_profit_workbook(request)


class AccountingReportView(APIView):
    permission_classes = [RolePermission]
    permission_page = "dashboard"

    @extend_schema(parameters=[
        OpenApiParameter("branch", str, description="all — hamma filial (sukut), main — faqat asosiy filial, <id> — bitta filial. Filial foydalanuvchisiga ta'sir qilmaydi."),
    ])
    def get(self, request):
        if not has_page_permission(request.user, "dashboard", False):
            return forbidden()
        return Response(json_safe(accounting_report_data(request)))


def user_branch(user):
    """Foydalanuvchi qaysi filialga tegishli. None bo'lsa asosiy filial."""
    return getattr(getattr(user, "profile", None), "branch", None)


def scope_catalog_to_branch(queryset, user):
    """Filiallar katalogi aralashmaydi. Filial foydalanuvchisi faqat o'z filialini,
    asosiy filial foydalanuvchisi faqat asosiy katalogni ko'radi."""
    branch = user_branch(user)
    return queryset.filter(branch=branch) if branch else queryset.filter(branch__isnull=True)


def branch_report_data(request):
    branch_id = request.query_params.get("branch")
    date_from, date_to = parse_date_range_params(request)
    branches = Branch.objects.filter(is_main=False, is_active=True)
    if branch_id:
        branches = branches.filter(id=branch_id)
    rows = []
    for branch in branches:
        transfers = CatalogTransfer.objects.filter(branch=branch)
        items = CatalogItem.objects.filter(branch=branch)
        history = CatalogHistory.objects.filter(action="sold", catalog_item__branch=branch).select_related("catalog_item")
        if date_from:
            transfers = transfers.filter(created_at__date__gte=date_from)
            history = history.filter(created_at__date__gte=date_from)
        if date_to:
            transfers = transfers.filter(created_at__date__lte=date_to)
            history = history.filter(created_at__date__lte=date_to)
        received_quantity = transfers.aggregate(value=Coalesce(Sum("quantity"), 0))["value"]
        # Filialga katalog ikki yo'l bilan tushadi: asosiy filialdan yuborilgan (transfer)
        # yoki o'sha zahoti filial uchun qo'shilgan. Ikkinchisi transferda ko'rinmaydi.
        direct_items = items.filter(source_item__isnull=True)
        if date_from:
            direct_items = direct_items.filter(created_at__date__gte=date_from)
        if date_to:
            direct_items = direct_items.filter(created_at__date__lte=date_to)
        direct_quantity = direct_items.aggregate(value=Coalesce(Sum("quantity_total"), 0))["value"]
        sold_quantity = 0
        sold_revenue = Decimal("0")
        source_value = Decimal("0")
        discounted_sales = 0
        discounted_quantity = 0
        discount_total = Decimal("0")
        for row in history:
            quantity = int(row.quantity or 0)
            sold_quantity += quantity
            sold_revenue += catalog_history_sale_total(row)
            source_value += Decimal(row.catalog_item.source_price or 0) * Decimal(quantity)
            discount = Decimal(row.discount_amount or 0)
            if discount > 0:
                discounted_sales += 1
                discounted_quantity += quantity
                discount_total += discount
        rows.append({
            "branch_id": branch.id,
            "branch_name": branch.name,
            "received_transfers": transfers.count(),
            "received_quantity": received_quantity,
            "direct_quantity": direct_quantity,
            "incoming_quantity": received_quantity + direct_quantity,
            "catalog_items": items.count(),
            "available_quantity": max(items.aggregate(value=Coalesce(Sum("quantity_total"), 0))["value"] - items.aggregate(value=Coalesce(Sum("quantity_sold"), 0))["value"], 0),
            "sold_quantity": sold_quantity,
            "sold_revenue": sold_revenue,
            "source_value": source_value,
            "markup_total": sold_revenue - source_value,
            "discounted_sales_count": discounted_sales,
            "discounted_quantity": discounted_quantity,
            "discount_total": discount_total,
        })
    totals = {
        "received_quantity": sum(row["received_quantity"] for row in rows),
        "direct_quantity": sum(row["direct_quantity"] for row in rows),
        "incoming_quantity": sum(row["incoming_quantity"] for row in rows),
        "sold_quantity": sum(row["sold_quantity"] for row in rows),
        "sold_revenue": sum((row["sold_revenue"] for row in rows), Decimal("0")),
        "discounted_quantity": sum(row["discounted_quantity"] for row in rows),
        "discount_total": sum((row["discount_total"] for row in rows), Decimal("0")),
    }
    return {
        "period": {"date_from": date_from.isoformat() if date_from else None, "date_to": date_to.isoformat() if date_to else None},
        "branches": rows,
        "totals": totals,
    }


class BranchViewSet(ScopedViewSet):
    permission_page = "settings"
    write_roles = ["admin"]
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    filterset_fields = ["is_active", "is_main"]

    def get_permissions(self):
        return super().get_permissions()


class CatalogTransferViewSet(viewsets.ReadOnlyModelViewSet):
    """Filialga yuborilgan katalog tarixi."""

    permission_classes = [RolePermission]
    permission_page = "catalog"
    queryset = CatalogTransfer.objects.select_related("branch", "target_item", "source_item", "created_by").all()
    serializer_class = CatalogTransferSerializer
    filterset_fields = ["branch", "source_item", "target_item"]
    ordering_fields = ["created_at", "quantity"]

    def get_queryset(self):
        # filial faqat o'ziga kelgan yuborishlarni ko'radi
        branch = user_branch(self.request.user)
        queryset = super().get_queryset()
        return queryset.filter(branch=branch) if branch else queryset


class BranchReportView(APIView):
    """Admin uchun: filialga qancha katalog yuborilgan, qanchasi sotilgan,
    qanchasi chegirma bilan sotilgan."""

    permission_classes = [RolePermission]
    permission_page = "dashboard"

    def get(self, request):
        if not has_page_permission(request.user, "dashboard", False):
            return forbidden()
        return Response(json_safe(branch_report_data(request)))


class CatalogItemViewSet(ScopedViewSet):
    permission_page = "catalog"
    write_roles = ["admin", "florist", "content", "warehouse"]
    queryset = CatalogItem.objects.select_related("social_post", "florist__user", "decoration_florist__user", "customer").prefetch_related("composition__stock_batch__variant__flower", "materials__packaging").all()
    serializer_class = CatalogItemSerializer
    filterset_fields = ["status", "arrangement_type", "catalog_kind", "florist", "customer"]
    search_fields = ["name_uz", "description_uz", "description_ru", "customer__name", "customer__phone"]

    def get_queryset(self):
        return scope_catalog_to_branch(super().get_queryset(), self.request.user)

    def perform_create(self, serializer):
        # Filial foydalanuvchisi yangi katalog yaratmaydi, unga faqat yuboriladi.
        branch = user_branch(self.request.user)
        if branch:
            raise serializers.ValidationError({"detail": "Filialda yangi katalog yaratilmaydi. Asosiy filialdan yuboriladi."})
        serializer.save(created_by=self.request.user)

    @extend_schema(request=CatalogTransferRequestSerializer, responses=CatalogTransferSerializer)
    @action(detail=True, methods=["post"])
    def transfer(self, request, pk=None):
        if user_branch(request.user):
            return Response({"detail": "Faqat asosiy filial katalogni boshqa filialga yuboradi"}, status=status.HTTP_403_FORBIDDEN)
        serializer = CatalogTransferRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            transfer = transfer_catalog_to_branch(
                self.get_object(), serializer.validated_data["branch"], serializer.validated_data["quantity"],
                serializer.validated_data.get("price"), serializer.validated_data.get("note", ""), request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CatalogTransferSerializer(transfer).data, status=status.HTTP_201_CREATED)

    @extend_schema(request=CatalogSellRequestSerializer, responses=CatalogItemSerializer)
    @action(detail=True, methods=["post"])
    def sell(self, request, pk=None):
        serializer = CatalogSellRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            item = mark_catalog_sold(
                self.get_object(),
                request.user,
                serializer.validated_data.get("quantity", 1),
                serializer.validated_data.get("sale_price"),
                serializer.validated_data.get("discount_reason", ""),
                serializer.validated_data.get("payment_type", ""),
                serializer.validated_data.get("sold_at"),
                serializer.validated_data.get("reservation"),
                serializer.validated_data.get("materials", []),
                serializer.validated_data.get("decoration_florist"),
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(item).data)

    @extend_schema(request=CatalogRestoreFlowersSerializer, responses=CatalogItemSerializer)
    @action(detail=True, methods=["post"], url_path="restore-flowers")
    def restore_flowers(self, request, pk=None):
        if not has_page_permission(request.user, "catalog", True):
            return forbidden()
        serializer = CatalogRestoreFlowersSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            item = restore_catalog_flowers(
                self.get_object(),
                serializer.validated_data["florist"],
                serializer.validated_data["old_batch"],
                serializer.validated_data["new_batch"],
                serializer.validated_data["quantity_stems"],
                serializer.validated_data.get("reason", ""),
                request.user,
            )
        except (ValueError, TypeError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(item).data)

    @extend_schema(request=None, responses=CatalogItemSerializer)
    @action(detail=True, methods=["post"])
    def deduct_stock(self, request, pk=None):
        try:
            quantity = request.data.get("quantity")
            item = deduct_catalog_stock(self.get_object(), request.user, quantity)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(item).data)

    def perform_destroy(self, instance):
        with transaction.atomic():
            item = CatalogItem.objects.select_for_update().get(pk=instance.pk)
            before = instance_snapshot(item)
            restore_catalog_inventory(item, self.request.user)
            item.refresh_from_db()
            try:
                item.delete()
                write_audit(self.request.user, "catalog_deleted", item, before=before, after={}, request=self.request)
            except ProtectedError:
                item.status = "archived"
                item.save(update_fields=["status", "updated_at"])
                write_audit(self.request.user, "catalog_archived", item, before=before, after=instance_snapshot(item), request=self.request)


def _style_header(sheet):
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def export_florist_workbook(profile, request, include_sales=True):
    data = florist_stats_data(profile, request, include_sales=include_sales)
    summary = data["summary"]
    period = data["period"]
    florist = data["florist"]

    workbook, sheet = styled_workbook("Ish haqi tarixi", [
        "Sana", "Qo‘shilgan vaqt", "Manba", "Katalog mahsuloti", "Turi",
        "Buket yoki savat", "Hajm", "Soni", "Sotilgan", "Narxi",
        "Sotuvdan tushgan", "Floristga qo‘shilgan", "Sotilgan vaqti", "Izoh", "Kim qo‘shdi",
    ])
    for row in data["salary_entries"]:
        sheet.append([
            row["work_date"].isoformat() if row["work_date"] else "",
            local_datetime_label(row["created_at"]),
            row["source_label"],
            row["catalog_name"],
            row["catalog_kind_label"],
            row["arrangement_label"],
            row["volume"],
            row["quantity_total"] or "",
            row.get("sold_quantity") or "",
            money_label(row["listed_price"]) if row.get("listed_price") else "",
            money_label(row["sale_revenue"]) if row.get("sale_revenue") else "",
            money_label(row["amount"]),
            local_datetime_label(row.get("last_sold_at")),
            row["note"],
            row["added_by"],
        ])
    total_row = sheet.max_row + 2
    sheet.cell(total_row, 11, "Jami")
    sheet.cell(total_row, 11).font = Font(bold=True)
    sheet.cell(total_row, 12, money_label(summary["salary_total"]))
    sheet.cell(total_row, 12).font = Font(bold=True)

    info = workbook.create_sheet("Umumiy")
    info.append(["Ko‘rsatkich", "Qiymat"])
    _style_header(info)
    for label, value in [
        ("Florist", florist["name"]),
        ("Lavozim", florist["staff_type_label"]),
        ("Telefon", florist["phone"]),
        ("Davr boshi", period["date_from"] or "Boshidan"),
        ("Davr oxiri", period["date_to"] or "Bugungacha"),
        ("Jami ish haqi", money_label(summary["salary_total"])),
        ("Katalog uchun", money_label(summary["catalog_salary_total"])),
        ("Kunlik ish haqi", money_label(summary["daily_salary_total"])),
        ("Qo‘lda qo‘shilgan", money_label(summary["manual_salary_total"])),
        ("Yozuvlar soni", summary["salary_entries_count"]),
        ("Yasagan mahsulot", summary["catalog_count"]),
        ("Buket", summary["bouquet_count"]),
        ("Savat", summary["basket_count"]),
        ("Standart", summary["standard_count"]),
        ("Custom", summary["custom_count"]),
        ("Sotilgan dona", summary.get("sold_quantity", "")),
        ("Sotilmagan", summary.get("unsold_quantity", "")),
        ("Sotuvdan tushgan", money_label(summary["sale_revenue"]) if "sale_revenue" in summary else ""),
        ("Bitta mahsulotga o‘rtacha haq", money_label(summary["avg_fee_per_item"])),
        ("Ishlagan kunlar", summary["attendance_days"]),
    ]:
        info.append([label, value])

    daily_sheet = workbook.create_sheet("Kunlar bo‘yicha")
    daily_sheet.append(["Sana", "Yozuvlar", "Buket", "Savat", "Sotilgan dona", "Sotuvdan tushgan", "Floristga qo‘shilgan"])
    _style_header(daily_sheet)
    for row in data["by_day"]:
        daily_sheet.append([
            row["work_date"].isoformat() if row["work_date"] else "",
            row["count"], row["bouquets"], row["baskets"], row.get("sold_quantity", ""),
            money_label(row["sale_revenue"]) if "sale_revenue" in row else "", money_label(row["amount"]),
        ])

    volume_sheet = workbook.create_sheet("Hajm bo‘yicha")
    volume_sheet.append(["Buket yoki savat", "Hajm", "Soni", "Sotilgan dona", "Sotuvdan tushgan", "Floristga qo‘shilgan"])
    _style_header(volume_sheet)
    for row in data["by_volume"]:
        volume_sheet.append([
            row["arrangement_label"], row["volume"], row["count"], row.get("sold_quantity", ""),
            money_label(row["sale_revenue"]) if "sale_revenue" in row else "", money_label(row["amount"]),
        ])

    source_sheet = workbook.create_sheet("Manba bo‘yicha")
    source_sheet.append(["Manba", "Yozuvlar", "Summa"])
    _style_header(source_sheet)
    for row in data["by_source"]:
        source_sheet.append([row["source_label"], row["count"], money_label(row["amount"])])

    attendance_sheet = workbook.create_sheet("Keldi-ketdi")
    attendance_sheet.append(["Sana", "Keldi", "Ketdi", "Manba", "Izoh"])
    _style_header(attendance_sheet)
    for row in data["attendance"]:
        attendance_sheet.append([
            row["work_date"].isoformat() if row["work_date"] else "",
            local_datetime_label(row["check_in_at"]),
            local_datetime_label(row["check_out_at"]),
            row["source_label"], row["note"],
        ])

    for current in [sheet, info, daily_sheet, volume_sheet, source_sheet, attendance_sheet]:
        autosize_sheet(current)
    suffix = f"_{period['date_from'] or 'boshidan'}_{period['date_to'] or 'bugun'}"
    return excel_response(workbook, f"florist_{profile.id}{suffix}.xlsx")


def export_all_florists_workbook(request):
    date_from, date_to = parse_date_range_params(request)
    workbook, sheet = styled_workbook("Hamma floristlar", ["Florist", "Turi", "Katalog soni", "Custom katalog", "Standart katalog", "Jami ish haqi", "Keldi kunlari"])
    profiles = FloristProfile.objects.select_related("user").filter(is_active=True).order_by("user__first_name", "user__username")
    detail_sheet = workbook.create_sheet("Kunlik hajm")
    detail_sheet.append(["Florist", "Sana", "Katalog turi", "Arrangement", "Hajm", "Yasalgan soni", "Floristga qo‘shilgan summa"])
    for cell in detail_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for profile in profiles:
        salary = FloristSalaryEntry.objects.filter(florist=profile)
        catalog = CatalogItem.objects.filter(florist=profile)
        attendance = FloristAttendance.objects.filter(florist=profile)
        if date_from:
            salary = salary.filter(work_date__gte=date_from)
            catalog = catalog.filter(created_at__date__gte=date_from)
            attendance = attendance.filter(work_date__gte=date_from)
        if date_to:
            salary = salary.filter(work_date__lte=date_to)
            catalog = catalog.filter(created_at__date__lte=date_to)
            attendance = attendance.filter(work_date__lte=date_to)
        sheet.append([
            str(profile),
            profile.get_staff_type_display(),
            sum(int(item.quantity_total or 0) for item in catalog),
            sum(int(item.quantity_total or 0) for item in catalog.filter(catalog_kind="custom")),
            sum(int(item.quantity_total or 0) for item in catalog.filter(catalog_kind="standard")),
            money_label(salary.aggregate(value=Coalesce(Sum("amount"), Decimal("0")))["value"]),
            attendance.filter(check_in_at__isnull=False).count(),
        ])
        daily_rows = {}
        for row in salary.select_related("catalog_item"):
            item = row.catalog_item
            if not item or row.source not in ["catalog", "custom_catalog"]:
                continue
            key = (row.work_date, item.catalog_kind if item else "", item.arrangement_type if item else "", item.volume if item else "")
            current = daily_rows.setdefault(key, {"count": 0, "amount": Decimal("0")})
            current["count"] += int(item.quantity_total or 1)
            current["amount"] += Decimal(row.amount or 0)
        for key, value in sorted(daily_rows.items(), key=lambda row: row[0], reverse=True):
            work_date, catalog_kind, arrangement_type, volume = key
            detail_sheet.append([str(profile), work_date.isoformat(), catalog_kind, arrangement_type, volume or "Belgilanmagan", value["count"], money_label(value["amount"])])
    autosize_sheet(sheet)
    autosize_sheet(detail_sheet)
    return excel_response(workbook, "all_florists_export.xlsx")


def export_profit_workbook(request):
    data = accounting_report_data(request)
    summary = data["summary"]
    workbook, sheet = styled_workbook("Hisob-kitob", ["Ko‘rsatkich", "Qiymat"])
    summary_rows = [
        ("Umumiy savdo", summary["total_sales"]),
        ("Naqd", summary["cash_total"]),
        ("Karta", summary["card_total"]),
        ("Aniqlanmagan to‘lov", summary["unknown_total"]),
        ("Bron/zaklad to‘lovlari", data["reservation_payments_summary"]["total"]),
        ("Sotuvlar soni", summary["sales_count"]),
        ("Sotilgan son", summary["total_quantity"]),
        ("Sotilgan gul donasi", summary["flower_stems"]),
        ("Standart sotilgan", summary["standard_quantity"]),
        ("Custom sotilgan", summary["custom_quantity"]),
        ("Sof foyda", summary["net_profit"]),
        ("Umumiy skidka", summary["discount_total"]),
        ("Skidka bilan sotuvlar", summary["discounted_sales_count"]),
    ]
    for label, value in summary_rows:
        sheet.append([label, money_label(value) if isinstance(value, Decimal) else value])
    branch_sheet = workbook.create_sheet("Filiallar")
    branch_sheet.append([
        "Filial", "Sotuvlar soni", "Sotilgan son", "Sotilgan gul donasi", "Savdo",
        "Naqd", "Naqd soni", "Karta", "Karta soni", "Aniqlanmagan", "Skidka",
        "Skidkali sotuv", "Tannarx", "Sof foyda", "Ulush %",
    ])
    _style_header(branch_sheet)
    for row in data["by_branch"]:
        branch_sheet.append([
            row["branch_name"], row["sales_count"], row["total_quantity"], row["flower_stems"],
            money_label(row["total_sales"]), money_label(row["cash_total"]), row["cash_count"],
            money_label(row["card_total"]), row["card_count"], money_label(row["unknown_total"]),
            money_label(row["discount_total"]), row["discounted_sales_count"],
            money_label(row["cost_total"]), money_label(row["net_profit"]), str(row["share_percent"]),
        ])
    volume_sheet = workbook.create_sheet("Hajmlar")
    volume_sheet.append(["Katalog turi", "Hajm", "Soni", "Savdo", "Skidka"])
    for cell in volume_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in data["by_volume"]:
        volume_sheet.append([row["catalog_kind"], row["volume"], row["quantity"], money_label(row["sales"]), money_label(row["discount"])])
    history_sheet = workbook.create_sheet("Sotuv history")
    history_sheet.append(["Sotilgan vaqt", "Katalogga qo‘shilgan vaqt", "Filial", "Katalog", "Turi", "Arrangement", "Hajm", "Florist", "Soni", "Gul donasi", "Asl jami", "Sotuv jami", "Cost", "Sof foyda", "To‘lov", "Skidka", "Skidka foiz", "Kim sotdi", "Izoh"])
    for cell in history_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in data["history"]:
        history_sheet.append([
            local_datetime_label(row["sold_at"]),
            local_datetime_label(row["catalog_created_at"]),
            row["branch_name"],
            row["catalog_name"],
            row["catalog_kind"],
            row["arrangement_type"],
            row["volume"] or "Belgilanmagan",
            row["florist_name"],
            row["quantity"],
            row["flower_stems"],
            money_label(row["listed_total"]),
            money_label(row["sale_total"]),
            money_label(row["cost_total"]),
            money_label(row["net_profit"]),
            row["payment_label"],
            money_label(row["discount_amount"]),
            row["discount_percent"],
            row["sold_by"],
            row["discount_reason"],
        ])
    discount_sheet = workbook.create_sheet("Skidkalar")
    discount_sheet.append(["Sana", "Filial", "Katalog", "Turi", "Hajm", "Soni", "Skidka summa", "Skidka foiz", "Izoh", "Kim sotdi"])
    for cell in discount_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in data["discounted_sales"]:
        discount_sheet.append([
            local_datetime_label(row["sold_at"]),
            row["branch_name"],
            row["catalog_name"],
            row["catalog_kind"],
            row["volume"] or "Belgilanmagan",
            row["quantity"],
            money_label(row["discount_amount"]),
            row["discount_percent"],
            row["discount_reason"],
            row["sold_by"],
        ])
    reservation_sheet = workbook.create_sheet("Bron to‘lovlari")
    reservation_sheet.append(["To‘langan vaqt", "Bron ID", "Mijoz", "Katalog", "Summa", "To‘lov turi", "Kim qo‘shdi", "Izoh"])
    for cell in reservation_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in data["reservation_payments"]:
        reservation_sheet.append([
            local_datetime_label(row["paid_at"]),
            row["reservation_id"],
            row["customer_name"],
            row["catalog_name"],
            money_label(row["amount"]),
            row["method_label"],
            row["created_by"],
            row["note"],
        ])
    autosize_sheet(sheet)
    autosize_sheet(branch_sheet)
    autosize_sheet(volume_sheet)
    autosize_sheet(history_sheet)
    autosize_sheet(discount_sheet)
    autosize_sheet(reservation_sheet)
    return excel_response(workbook, "profit_export.xlsx")


class CustomerViewSet(ScopedViewSet):
    permission_page = "customers"
    write_roles = ["admin", "operator"]
    queryset = Customer.objects.annotate(purchases_count=Count("leads", filter=Q(leads__status="won")), total_spent=Coalesce(Sum("leads__estimated_price", filter=Q(leads__status="won")), Decimal("0"))).all()
    serializer_class = CustomerSerializer
    filterset_fields = ["language", "is_blocked"]
    search_fields = ["name", "phone", "instagram_username"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.query_params.get("include_incomplete") == "true":
            return queryset
        return queryset.exclude(Q(name="") | Q(phone=""))

    def destroy(self, request, *args, **kwargs):
        customer = self.get_object()
        try:
            return super().destroy(request, *args, **kwargs)
        except ProtectedError:
            before = instance_snapshot(customer)
            customer.name = ""
            customer.phone = ""
            customer.instagram_username = ""
            customer.instagram_user_id = f"deleted:{customer.id}"
            customer.notes = (customer.notes + "\n" if customer.notes else "") + "Client arxivlandi. Lead tarixi saqlandi."
            customer.is_blocked = True
            customer.save(update_fields=["name", "phone", "instagram_username", "instagram_user_id", "notes", "is_blocked", "updated_at"])
            before_changed, after_changed = changed_snapshot(before, instance_snapshot(customer))
            write_audit(self.request.user, "customer_archived", customer, before=before_changed, after=after_changed, request=self.request)
            return Response({"detail": "Client arxivlandi. Lead tarixi saqlandi.", "id": customer.id, "archived": True})


class ReservationViewSet(ScopedViewSet):
    permission_page = "crm"
    queryset = Reservation.objects.select_related("customer", "catalog_item", "created_by").prefetch_related("payments").all()
    serializer_class = ReservationSerializer
    filterset_fields = ["status", "payment_status", "customer", "catalog_item", "desired_date"]
    search_fields = ["customer__name", "customer__phone", "request_uz", "note"]

    def perform_create(self, serializer):
        reservation = serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)
        write_audit(self.request.user, "reservation_created", reservation, before={}, after=instance_snapshot(reservation), request=self.request)

    def perform_update(self, serializer):
        before = instance_snapshot(self.get_object())
        reservation = serializer.save()
        reservation = sync_reservation_payment_status(reservation)
        write_audit(self.request.user, "reservation_updated", reservation, before=before, after=instance_snapshot(reservation), request=self.request)

    @extend_schema(request=ReservationPaymentRequestSerializer, responses=ReservationPaymentSerializer)
    @action(detail=True, methods=["post"], url_path="add-payment")
    def add_payment(self, request, pk=None):
        serializer = ReservationPaymentRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            reservation = Reservation.objects.select_for_update().get(pk=self.get_object().pk)
            payment = ReservationPayment.objects.create(
                reservation=reservation,
                amount=serializer.validated_data["amount"],
                method=serializer.validated_data.get("method", "cash"),
                paid_at=serializer.validated_data.get("paid_at") or timezone.now(),
                note=serializer.validated_data.get("note", ""),
                created_by=request.user if request.user.is_authenticated else None,
            )
            reservation = sync_reservation_payment_status(reservation)
            write_audit(request.user, "reservation_payment_created", payment, before={}, after=instance_snapshot(payment), request=request)
        return Response(ReservationPaymentSerializer(payment).data, status=status.HTTP_201_CREATED)

    @extend_schema(request=None, responses=ReservationSerializer)
    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        with transaction.atomic():
            reservation = Reservation.objects.select_for_update().get(pk=self.get_object().pk)
            before = instance_snapshot(reservation)
            reservation.status = "cancelled"
            reservation.save(update_fields=["status", "updated_at"])
            write_audit(request.user, "reservation_cancelled", reservation, before=before, after=instance_snapshot(reservation), request=request)
        return Response(self.get_serializer(reservation).data)


class LeadStatusViewSet(ScopedViewSet):
    permission_page = "crm"
    write_roles = ["admin", "operator"]
    queryset = LeadStatus.objects.all()
    serializer_class = LeadStatusSerializer
    filterset_fields = ["is_active"]
    search_fields = ["key", "name_uz"]
    ordering_fields = ["order", "created_at"]


class LeadViewSet(ScopedViewSet):
    permission_page = "crm"
    write_roles = ["admin", "operator"]
    queryset = Lead.objects.select_related("customer", "assigned_to", "social_post").all()
    serializer_class = LeadSerializer
    filterset_class = LeadFilter
    search_fields = ["customer__name", "customer__phone", "request_uz", "request_ru"]
    ordering_fields = ["sort_order", "created_at", "estimated_price"]
    ordering = ["status", "sort_order", "-created_at", "id"]

    def perform_create(self, serializer):
        with transaction.atomic():
            extra = {}
            if "sort_order" not in serializer.validated_data:
                status_value = serializer.validated_data.get("status", "new")
                extra["sort_order"] = next_lead_sort_order(status_value)
            lead = serializer.save(**extra)
            write_audit(self.request.user, "lead_created", lead, before={}, after=instance_snapshot(lead), request=self.request)
            if lead.status == "won":
                try:
                    deduct_lead_stock(lead, self.request.user)
                    serializer.instance.refresh_from_db()
                except ValueError as exc:
                    raise serializers.ValidationError({"detail": str(exc)})
            transaction.on_commit(lambda lead_id=lead.id: schedule_lead_recall(Lead.objects.get(id=lead_id)))

    @extend_schema(request=LeadMoveSerializer, responses=LeadSerializer)
    @action(detail=True, methods=["post"])
    def move(self, request, pk=None):
        lead = self.get_object()
        serializer = LeadMoveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        status_value = serializer.validated_data.get("status") or lead.status
        if not LeadStatus.objects.filter(key=status_value).exists():
            return Response({"status": "Bunday lead statusi mavjud emas"}, status=status.HTTP_400_BAD_REQUEST)
        before = serializer.validated_data.get("before")
        after = serializer.validated_data.get("after")
        sort_order = serializer.validated_data.get("sort_order")
        with transaction.atomic():
            lead = Lead.objects.select_for_update().get(pk=lead.pk)
            before_snapshot = instance_snapshot(lead)
            before_status = lead.status
            if sort_order is None:
                sort_order = lead_sort_order_between(before, after, status_value)
            lead.status = status_value
            lead.sort_order = sort_order
            lead.save(update_fields=["status", "sort_order", "updated_at"])
            if lead.status == "won" and before_status != "won":
                try:
                    deduct_lead_stock(lead, self.request.user)
                    lead.refresh_from_db()
                except ValueError as exc:
                    raise serializers.ValidationError({"detail": str(exc)})
            elif before_status == "won" and lead.status != "won":
                restore_lead_stock(lead, self.request.user)
                lead.refresh_from_db()
            after_snapshot = instance_snapshot(lead)
            before_changed, after_changed = changed_snapshot(before_snapshot, after_snapshot)
            if before_changed or after_changed:
                write_audit(self.request.user, "lead_moved", lead, before=before_changed, after=after_changed, request=self.request)
        return Response(LeadSerializer(lead, context={"request": request}).data)

    @extend_schema(request=LeadColumnReorderSerializer, responses=inline_serializer(name="LeadColumnReorderResponse", fields={"updated": serializers.IntegerField()}))
    @action(detail=False, methods=["post"], url_path="reorder-column")
    def reorder_column(self, request):
        serializer = LeadColumnReorderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        status_value = serializer.validated_data["status"]
        if not LeadStatus.objects.filter(key=status_value).exists():
            return Response({"status": "Bunday lead statusi mavjud emas"}, status=status.HTTP_400_BAD_REQUEST)
        lead_ids = serializer.validated_data["lead_ids"]
        if len(lead_ids) != len(set(lead_ids)):
            return Response({"lead_ids": "Lead id takrorlanmasligi kerak"}, status=status.HTTP_400_BAD_REQUEST)
        scoped_queryset = self.get_queryset()
        leads = list(scoped_queryset.filter(id__in=lead_ids))
        if len(leads) != len(set(lead_ids)):
            return Response({"lead_ids": "Lead topilmadi yoki sizda ruxsat yo‘q"}, status=status.HTTP_400_BAD_REQUEST)
        target_existing_ids = set(scoped_queryset.filter(status=status_value).values_list("id", flat=True))
        incoming_ids = set(lead_ids)
        missing_ids = target_existing_ids - incoming_ids
        if missing_ids:
            return Response({"lead_ids": "Target column lead_ids to‘liq yuborilishi kerak", "missing_ids": sorted(missing_ids)}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            for index, lead_id in enumerate(lead_ids, start=1):
                lead = Lead.objects.select_for_update().get(id=lead_id)
                before_snapshot = instance_snapshot(lead)
                before_status = lead.status
                lead.status = status_value
                lead.sort_order = Decimal(index * 1000)
                lead.save(update_fields=["status", "sort_order", "updated_at"])
                if lead.status == "won" and before_status != "won":
                    try:
                        deduct_lead_stock(lead, self.request.user)
                    except ValueError as exc:
                        raise serializers.ValidationError({"detail": str(exc)})
                elif before_status == "won" and lead.status != "won":
                    restore_lead_stock(lead, self.request.user)
                lead.refresh_from_db()
                after_snapshot = instance_snapshot(lead)
                before_changed, after_changed = changed_snapshot(before_snapshot, after_snapshot)
                if before_changed or after_changed:
                    write_audit(self.request.user, "lead_reordered", lead, before=before_changed, after=after_changed, request=self.request)
        return Response({"updated": len(lead_ids)})

    def perform_update(self, serializer):
        with transaction.atomic():
            before_snapshot = instance_snapshot(serializer.instance)
            before_status = serializer.instance.status
            lead = serializer.save()
            if lead.status == "won" and before_status != "won":
                try:
                    deduct_lead_stock(lead, self.request.user)
                    serializer.instance.refresh_from_db()
                except ValueError as exc:
                    raise serializers.ValidationError({"detail": str(exc)})
            elif before_status == "won" and lead.status != "won":
                restore_lead_stock(lead, self.request.user)
                serializer.instance.refresh_from_db()
            lead.refresh_from_db()
            after_snapshot = instance_snapshot(lead)
            before_changed, after_changed = changed_snapshot(before_snapshot, after_snapshot)
            if before_changed or after_changed:
                write_audit(self.request.user, "lead_updated", lead, before=before_changed, after=after_changed, request=self.request)
            transaction.on_commit(lambda lead_id=lead.id: schedule_lead_recall(Lead.objects.get(id=lead_id)))

    def perform_destroy(self, instance):
        with transaction.atomic():
            lead = Lead.objects.select_for_update().get(pk=instance.pk)
            before_snapshot = instance_snapshot(lead)
            if lead.stock_deducted_at:
                restore_lead_stock(lead, self.request.user)
                lead.refresh_from_db()
            after_restore_snapshot = instance_snapshot(lead)
            write_audit(self.request.user, "lead_deleted", lead, before=before_snapshot, after={"restored_before_delete": after_restore_snapshot}, request=self.request)
            lead.delete()


class SocialPostViewSet(ScopedViewSet):
    permission_page = "social_posts"
    write_roles = ["admin", "content"]
    queryset = SocialPost.objects.prefetch_related("catalog_items__composition__stock_batch__variant__flower", "leads__customer", "leads__catalog_usage__catalog_item").annotate(reply_count=Count("conversations", distinct=True), lead_count=Count("leads", distinct=True)).all()
    serializer_class = SocialPostSerializer
    filterset_fields = ["post_type", "is_targeted", "is_active"]
    search_fields = ["title_uz", "title_ru", "media_id", "permalink"]


class ConversationViewSet(ScopedViewSet):
    permission_page = "conversations"
    write_roles = ["admin", "operator"]
    queryset = Conversation.objects.select_related("customer", "social_post", "assigned_to").prefetch_related("messages").all()
    serializer_class = ConversationSerializer
    filterset_class = ConversationFilter
    search_fields = ["customer__name", "customer__instagram_username", "messages__text"]

    @extend_schema(request=TextRequestSerializer, responses=SendResponseSerializer)
    @action(detail=True, methods=["post"])
    def send(self, request, pk=None):
        conversation = self.get_object()
        text = request.data.get("text", "").strip()
        if not text:
            return Response({"detail": "Xabar bo‘sh"}, status=status.HTTP_400_BAD_REQUEST)
        external_id = conversation.customer.instagram_user_id
        delivery_status = "sent"
        platform_status = None
        platform_response = ""
        try:
            if external_id.startswith("telegram:"):
                telegram_message = conversation.messages.filter(instagram_message_id__startswith="telegram:").order_by("-created_at", "-id").first()
                parts = telegram_message.instagram_message_id.split(":") if telegram_message else []
                chat_id = parts[1] if len(parts) >= 3 else external_id.removeprefix("telegram:")
                telegram_send(chat_id, text)
            else:
                instagram_send(external_id, text)
        except requests.HTTPError as exc:
            response = getattr(exc, "response", None)
            platform_status = getattr(response, "status_code", None)
            platform_body = getattr(response, "text", "")
            platform_response = platform_body[:500]
            delivery_status = "failed"
        except requests.RequestException as exc:
            platform_response = str(exc)
            delivery_status = "failed"
        metadata = {}
        if delivery_status != "sent":
            metadata = {
                "delivery_status": delivery_status,
                "platform_status": platform_status,
                "platform_response": platform_response,
            }
        message = conversation.messages.create(sender="operator", text=text, metadata=metadata)
        conversation.last_message_at = timezone.now()
        conversation.ai_paused_until = timezone.now() + timedelta(minutes=15)
        conversation.ai_pause_reason = "operator_message"
        conversation.assigned_to = request.user
        conversation.save(update_fields=["last_message_at", "ai_paused_until", "ai_pause_reason", "assigned_to", "updated_at"])
        return Response({"id": message.id, "text": message.text, "delivery_status": delivery_status, "platform_status": platform_status, "platform_response": platform_response})

    @extend_schema(request=None, responses=ConversationSerializer)
    @action(detail=True, methods=["post"])
    def handoff(self, request, pk=None):
        conversation = self.get_object()
        conversation.status = "operator"
        conversation.assigned_to = request.user
        conversation.save(update_fields=["status", "assigned_to", "updated_at"])
        return Response(self.get_serializer(conversation).data)

    @extend_schema(request=None, responses=ConversationSerializer)
    @action(detail=True, methods=["post"])
    def resume_ai(self, request, pk=None):
        conversation = self.get_object()
        conversation.status = "ai"
        conversation.assigned_to = None
        conversation.ai_paused_until = None
        conversation.ai_pause_reason = ""
        conversation.save(update_fields=["status", "assigned_to", "ai_paused_until", "ai_pause_reason", "updated_at"])
        return Response(self.get_serializer(conversation).data)

    @extend_schema(request=AIPauseRequestSerializer, responses=ConversationSerializer)
    @action(detail=True, methods=["post"])
    def pause_ai(self, request, pk=None):
        conversation = self.get_object()
        serializer = AIPauseRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        if serializer.validated_data.get("paused_until"):
            paused_until = serializer.validated_data["paused_until"]
        else:
            paused_until = timezone.now() + timedelta(minutes=serializer.validated_data["minutes"])
        conversation.ai_paused_until = paused_until
        conversation.ai_pause_reason = serializer.validated_data.get("reason", "manual")
        conversation.assigned_to = request.user
        conversation.save(update_fields=["ai_paused_until", "ai_pause_reason", "assigned_to", "updated_at"])
        return Response(self.get_serializer(conversation).data)

    @extend_schema(request=TextRequestSerializer, responses=SimulateResponseSerializer)
    @action(detail=True, methods=["post"])
    def simulate(self, request, pk=None):
        serializer = TextRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reply = process_customer_message(self.get_object(), serializer.validated_data["text"])
        return Response({"reply": reply.text if reply else None})


class NotificationViewSet(ScopedViewSet):
    permission_page = "notifications"
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    filterset_fields = ["notification_type", "is_read"]
    write_roles = ["admin", "operator", "florist", "warehouse", "content"]
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        scoped_ids = notification_queryset_for_user(self.request.user).values("id")
        return super().get_queryset().filter(id__in=scoped_ids)

    def _mark_read_response(self):
        notification = self.get_object()
        notification.is_read = True
        notification.save(update_fields=["is_read", "updated_at"])
        return Response(self.get_serializer(notification).data)

    @action(detail=True, methods=["post"])
    def read(self, request, pk=None):
        return self._mark_read_response()

    @action(detail=True, methods=["post"], url_path="mark-read")
    def mark_read(self, request, pk=None):
        return self._mark_read_response()

    @action(detail=False, methods=["post"])
    def read_all(self, request):
        queryset = self.filter_queryset(self.get_queryset()).filter(is_read=False)
        count = queryset.update(is_read=True)
        return Response({"updated": count})


class AuditLogViewSet(ScopedViewSet):
    permission_page = "audit"
    queryset = AuditLog.objects.select_related("user").all()
    serializer_class = AuditLogSerializer
    filterset_class = AuditLogFilter
    write_roles = []
    http_method_names = ["get", "head", "options"]

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(getattr(self.request.user, "profile", None), "role", None)
        if role != "developer":
            queryset = queryset.exclude(user__profile__role="developer")
        return queryset

    @extend_schema(
        parameters=[
            OpenApiParameter("user", int, OpenApiParameter.QUERY, description="Audit loglarni bitta user bo‘yicha filterlash"),
            OpenApiParameter("user_id", int, OpenApiParameter.QUERY, description="Audit loglarni bitta user ID bo‘yicha filterlash"),
            OpenApiParameter("action", str, OpenApiParameter.QUERY, description="Action kodi bo‘yicha filter"),
            OpenApiParameter("entity_type", str, OpenApiParameter.QUERY, description="Entity turi bo‘yicha filter"),
            OpenApiParameter("created_at_after", str, OpenApiParameter.QUERY, description="Boshlanish sanasi yoki vaqti"),
            OpenApiParameter("created_at_before", str, OpenApiParameter.QUERY, description="Tugash sanasi yoki vaqti"),
        ]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)


class PagePermissionViewSet(viewsets.ModelViewSet):
    permission_classes = [RolePermission]
    permission_page = "users"
    queryset = PagePermission.objects.select_related("user").all()
    serializer_class = PagePermissionSerializer
    filterset_class = PagePermissionFilter

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(getattr(self.request.user, "profile", None), "role", None)
        if role != "developer":
            queryset = queryset.exclude(page__in=PagePermission.DEVELOPER_ONLY_PAGES).exclude(user__profile__role="developer")
        return queryset

    def perform_create(self, serializer):
        permission = serializer.save()
        write_audit(self.request.user, "pagepermission_created", permission, before={}, after=instance_snapshot(permission), request=self.request)

    def perform_update(self, serializer):
        before = instance_snapshot(serializer.instance)
        permission = serializer.save()
        before_changed, after_changed = changed_snapshot(before, instance_snapshot(permission))
        if before_changed or after_changed:
            write_audit(self.request.user, "pagepermission_updated", permission, before=before_changed, after=after_changed, request=self.request)

    def perform_destroy(self, instance):
        before = instance_snapshot(instance)
        write_audit(self.request.user, "pagepermission_deleted", instance, before=before, after={}, request=self.request)
        instance.delete()


class InstagramWebhookEventViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [RolePermission]
    permission_page = "integrations"
    queryset = InstagramWebhookEvent.objects.all()
    serializer_class = InstagramWebhookEventSerializer
    filterset_fields = ["event_type", "sender_id", "message_id", "media_id", "story_id"]
    search_fields = ["text", "story_url", "media_id", "story_id", "sender_id", "message_id"]


@extend_schema(responses=UserSerializer)
@api_view(["GET"])
def me(request):
    return Response(UserSerializer(request.user).data)


@extend_schema(request=ChangePasswordSerializer, responses=inline_serializer(name="ChangePasswordResponse", fields={"detail": serializers.CharField()}))
@api_view(["POST"])
def change_password(request):
    serializer = ChangePasswordSerializer(data=request.data, context={"request": request})
    serializer.is_valid(raise_exception=True)
    request.user.set_password(serializer.validated_data["new_password"])
    request.user.save(update_fields=["password"])
    write_audit(request.user, "password_changed", request.user, before={}, after={"password": "changed"}, request=request)
    return Response({"detail": "Password o‘zgartirildi"})


dashboard_date_parameters = [
    OpenApiParameter("date_from", str, OpenApiParameter.QUERY, required=False, description="Boshlanish sanasi: YYYY-MM-DD yoki ISO datetime"),
    OpenApiParameter("date_to", str, OpenApiParameter.QUERY, required=False, description="Tugash sanasi: YYYY-MM-DD yoki ISO datetime"),
    OpenApiParameter("from", str, OpenApiParameter.QUERY, required=False, description="date_from alias"),
    OpenApiParameter("to", str, OpenApiParameter.QUERY, required=False, description="date_to alias"),
]


@extend_schema(parameters=dashboard_date_parameters, responses=inline_serializer(
    name="Dashboard",
    fields={
        "active_leads": serializers.IntegerField(),
        "new_leads_today": serializers.IntegerField(),
        "available_catalog": serializers.IntegerField(),
        "pending_deductions": serializers.IntegerField(),
        "unread_notifications": serializers.IntegerField(),
        "ai_conversations": serializers.IntegerField(),
        "operator_conversations": serializers.IntegerField(),
        "stock_stems": serializers.IntegerField(),
        "low_stock": serializers.IntegerField(),
        "lead_pipeline": serializers.ListField(child=serializers.DictField()),
        "recent_leads": LeadSerializer(many=True),
        "recent_notifications": NotificationSerializer(many=True),
        "revenue_today": serializers.DecimalField(max_digits=14, decimal_places=2),
        "orders_today": serializers.IntegerField(),
        "revenue_7d": serializers.DecimalField(max_digits=14, decimal_places=2),
        "conversion_rate": serializers.FloatField(),
        "period": serializers.DictField(),
        "period_revenue": serializers.DecimalField(max_digits=14, decimal_places=2),
        "period_orders": serializers.IntegerField(),
        "period_leads": serializers.IntegerField(),
        "period_customers": serializers.IntegerField(),
        "period_conversations": serializers.IntegerField(),
        "daily_stats": serializers.ListField(child=serializers.DictField()),
        "top_selling_flowers": serializers.ListField(child=serializers.DictField()),
        "florist_revenue": serializers.DecimalField(max_digits=14, decimal_places=2),
        "florist_salary_total": serializers.DecimalField(max_digits=14, decimal_places=2),
        "flowers_sold_stems": serializers.IntegerField(),
        "catalog_revenue": serializers.DecimalField(max_digits=14, decimal_places=2),
        "catalog_cost": serializers.DecimalField(max_digits=14, decimal_places=2),
        "catalog_discount": serializers.DecimalField(max_digits=14, decimal_places=2),
        "discounted_catalog_sales_count": serializers.IntegerField(),
        "discounted_catalog_quantity": serializers.IntegerField(),
        "discounted_catalog_amount": serializers.DecimalField(max_digits=14, decimal_places=2),
        "net_profit": serializers.DecimalField(max_digits=14, decimal_places=2),
        "batch_inventory_stats": serializers.ListField(child=serializers.DictField()),
        "florist_production_stats": serializers.ListField(child=serializers.DictField()),
    },
))
@api_view(["GET"])
def dashboard(request):
    if not has_page_permission(request.user, "dashboard", False):
        return forbidden()
    today = timezone.localdate()
    week_start = today - timedelta(days=6)
    period_start, period_end = dashboard_period(request)
    stock = StockBatch.objects.filter(is_active=True)
    stock_movements = StockMovement.objects.all()
    leads = Lead.objects.all()
    customers = Customer.objects.all()
    branch = user_branch(request.user)
    catalog = CatalogItem.objects.filter(branch=branch) if branch else CatalogItem.objects.filter(branch__isnull=True)
    notifications = notification_queryset_for_user(request.user)
    conversations = Conversation.objects.all()
    won_leads = leads.filter(status="won")
    period_leads = apply_created_range(leads, period_start, period_end)
    period_customers = apply_created_range(customers, period_start, period_end)
    period_conversations = apply_created_range(conversations, period_start, period_end)
    period_won_leads = apply_updated_range(won_leads, period_start, period_end)
    period_stock_out = apply_created_range(stock_movements.filter(movement_type="out", quantity_stems__lt=0), period_start, period_end)
    flowers_sold = period_stock_out.aggregate(value=Coalesce(Sum("quantity_stems"), 0))["value"] or 0
    period_catalog_sold = apply_updated_range(catalog.filter(quantity_sold__gt=0), period_start, period_end)
    catalog_financials = catalog_sale_financials(period_catalog_sold)
    catalog_discount_stats = catalog_sale_discount_stats(period_start, period_end, branch)
    today_start = timezone.make_aware(datetime.combine(today, time.min))
    today_end = timezone.make_aware(datetime.combine(today, time.max))
    week_start_dt = timezone.make_aware(datetime.combine(week_start, time.min))
    catalog_today = catalog_sales_totals(catalog_sales_queryset(today_start, today_end, branch))
    catalog_week = catalog_sales_totals(catalog_sales_queryset(week_start_dt, today_end, branch))
    catalog_period_rows = list(catalog_sales_queryset(period_start, period_end, branch))
    catalog_period = catalog_sales_totals(catalog_period_rows)
    lead_revenue_today = won_leads.filter(updated_at__date=today).aggregate(value=Coalesce(Sum("estimated_price"), Decimal("0")))["value"]
    lead_revenue_7d = won_leads.filter(updated_at__date__gte=week_start).aggregate(value=Coalesce(Sum("estimated_price"), Decimal("0")))["value"]
    lead_revenue_period = period_won_leads.aggregate(value=Coalesce(Sum("estimated_price"), Decimal("0")))["value"]
    leads_total = leads.count()
    conversations_total = conversations.count()
    conversion_base = conversations_total or leads_total
    data = {
        "active_leads": leads.exclude(status__in=["won", "lost"]).count(),
        "new_leads_today": leads.filter(created_at__date=today).count(),
        "orders_today": won_leads.filter(updated_at__date=today).count() + catalog_today["orders"],
        "revenue_today": lead_revenue_today + catalog_today["revenue"],
        "revenue_7d": lead_revenue_7d + catalog_week["revenue"],
        "lead_revenue_today": lead_revenue_today,
        "lead_revenue_7d": lead_revenue_7d,
        "catalog_sales_revenue_today": catalog_today["revenue"],
        "catalog_sales_revenue_7d": catalog_week["revenue"],
        "catalog_sales_orders_today": catalog_today["orders"],
        "catalog_sales_quantity_today": catalog_today["quantity"],
        "period": {"from": period_start, "to": period_end},
        "period_orders": period_won_leads.count() + catalog_period["orders"],
        "period_revenue": lead_revenue_period + catalog_period["revenue"],
        "period_lead_revenue": lead_revenue_period,
        "period_catalog_sales_revenue": catalog_period["revenue"],
        "period_catalog_sales_orders": catalog_period["orders"],
        "period_catalog_sales_quantity": catalog_period["quantity"],
        "period_leads": period_leads.count(),
        "period_customers": period_customers.count(),
        "period_conversations": period_conversations.count(),
        "daily_stats": dashboard_daily_stats(period_leads, period_conversations, period_start, period_end, catalog_period_rows),
        "top_selling_flowers": top_selling_flowers(period_won_leads)[:5],
        "florist_revenue": period_won_leads.aggregate(value=Coalesce(Sum("florist_fee"), Decimal("0")))["value"],
        "florist_salary_total": apply_created_range(FloristSalaryEntry.objects.all(), period_start, period_end).aggregate(value=Coalesce(Sum("amount"), Decimal("0")))["value"],
        "flowers_sold_stems": abs(int(flowers_sold)),
        "catalog_revenue": catalog_financials["revenue"],
        "catalog_cost": catalog_financials["cost"],
        "catalog_discount": catalog_financials["discount"],
        "discounted_catalog_sales_count": catalog_discount_stats["discounted_sales_count"],
        "discounted_catalog_quantity": catalog_discount_stats["discounted_quantity"],
        "discounted_catalog_amount": catalog_discount_stats["discounted_amount"],
        "net_profit": catalog_financials["profit"],
        "batch_inventory_stats": batch_inventory_stats(period_start, period_end)[:10],
        "florist_production_stats": florist_production_stats(period_start, period_end)[:10],
        "conversion_rate": round((won_leads.count() / conversion_base) * 100, 2) if conversion_base else 0,
        "available_catalog": catalog.filter(status="available").count(),
        "pending_deductions": catalog.filter(quantity_sold__gt=F("quantity_stock_deducted")).count(),
        "unread_notifications": notifications.filter(is_read=False).count(),
        "ai_conversations": conversations.filter(status="ai").count(),
        "operator_conversations": conversations.filter(status="operator").count(),
        "stock_stems": stock.aggregate(value=Coalesce(Sum("remaining_stems"), 0))["value"],
        "low_stock": stock.filter(remaining_stems__lte=F("minimum_sale_stems")).count(),
        "lead_pipeline": list(leads.values("status").annotate(count=Count("id")).order_by("status")),
        "recent_leads": LeadSerializer(leads.select_related("customer")[:6], many=True).data,
        "recent_notifications": NotificationSerializer(notifications.filter(is_read=False)[:6], many=True).data,
    }
    return Response(data)


@extend_schema(parameters=dashboard_date_parameters, responses=inline_serializer(
    name="Analytics",
    fields={
        "period": serializers.DictField(),
        "summary": serializers.DictField(),
        "daily_stats": serializers.ListField(child=serializers.DictField()),
        "top_selling_flowers": serializers.ListField(child=serializers.DictField()),
        "top_catalog_items": serializers.ListField(child=serializers.DictField()),
        "recent_top_catalog_items": serializers.ListField(child=serializers.DictField()),
        "lead_statuses": serializers.ListField(child=serializers.DictField()),
        "arrangement_types": serializers.ListField(child=serializers.DictField()),
        "conversation_sources": serializers.ListField(child=serializers.DictField()),
        "revenue_by_source": serializers.ListField(child=serializers.DictField()),
    },
))
@api_view(["GET"])
def analytics(request):
    if not has_page_permission(request.user, "dashboard", False):
        return forbidden()
    period_start, period_end = dashboard_period(request)
    leads = Lead.objects.select_related("customer").all()
    conversations = Conversation.objects.select_related("customer").all()
    customers = Customer.objects.all()
    stock_movements = StockMovement.objects.all()
    period_leads = apply_created_range(leads, period_start, period_end)
    period_conversations = apply_created_range(conversations, period_start, period_end)
    period_customers = apply_created_range(customers, period_start, period_end)
    won_leads = leads.filter(status="won")
    period_won_leads = apply_updated_range(won_leads, period_start, period_end)
    period_stock_out = apply_created_range(stock_movements.filter(movement_type="out", quantity_stems__lt=0), period_start, period_end)
    flowers_sold = period_stock_out.aggregate(value=Coalesce(Sum("quantity_stems"), 0))["value"] or 0
    branch = user_branch(request.user)
    branch_catalog = CatalogItem.objects.filter(branch=branch) if branch else CatalogItem.objects.filter(branch__isnull=True)
    period_catalog_sold = apply_updated_range(branch_catalog.filter(quantity_sold__gt=0), period_start, period_end)
    catalog_financials = catalog_sale_financials(period_catalog_sold)
    catalog_discount_stats = catalog_sale_discount_stats(period_start, period_end, branch)
    catalog_rows = list(catalog_sales_queryset(period_start, period_end, branch))
    catalog_sales = catalog_sales_totals(catalog_rows)
    lead_revenue = period_won_leads.aggregate(value=Coalesce(Sum("estimated_price"), Decimal("0")))["value"]
    total_orders = period_won_leads.count() + catalog_sales["orders"]
    data = {
        "period": {"from": period_start, "to": period_end},
        "summary": {
            "leads": period_leads.count(),
            "customers": period_customers.count(),
            "conversations": period_conversations.count(),
            "orders": total_orders,
            "revenue": lead_revenue + catalog_sales["revenue"],
            "lead_orders": period_won_leads.count(),
            "lead_revenue": lead_revenue,
            "catalog_sales_orders": catalog_sales["orders"],
            "catalog_sales_revenue": catalog_sales["revenue"],
            "catalog_sales_quantity": catalog_sales["quantity"],
            "florist_revenue": period_won_leads.aggregate(value=Coalesce(Sum("florist_fee"), Decimal("0")))["value"],
            "florist_salary_total": apply_created_range(FloristSalaryEntry.objects.all(), period_start, period_end).aggregate(value=Coalesce(Sum("amount"), Decimal("0")))["value"],
            "flowers_sold_stems": abs(int(flowers_sold)),
            "catalog_revenue": catalog_financials["revenue"],
            "catalog_cost": catalog_financials["cost"],
            "catalog_discount": catalog_financials["discount"],
            "discounted_catalog_sales_count": catalog_discount_stats["discounted_sales_count"],
            "discounted_catalog_quantity": catalog_discount_stats["discounted_quantity"],
            "discounted_catalog_amount": catalog_discount_stats["discounted_amount"],
            "net_profit": catalog_financials["profit"],
            "conversion_rate": round((period_won_leads.count() / (period_conversations.count() or period_leads.count())) * 100, 2) if (period_conversations.count() or period_leads.count()) else 0,
        },
        "daily_stats": analytics_daily_stats(period_leads, period_conversations, period_won_leads, period_start, period_end, catalog_rows),
        "top_selling_flowers": top_selling_flowers(period_won_leads),
        "top_catalog_items": top_catalog_items(period_won_leads, catalog_rows),
        "recent_top_catalog_items": recent_top_catalog_items(period_won_leads, catalog_rows),
        "batch_inventory_stats": batch_inventory_stats(period_start, period_end),
        "florist_production_stats": florist_production_stats(period_start, period_end),
        "lead_statuses": list(period_leads.values("status").annotate(count=Count("id")).order_by("status")),
        "arrangement_types": list(period_leads.values("arrangement_type").annotate(count=Count("id")).order_by("arrangement_type")),
        "conversation_sources": conversation_source_breakdown(period_conversations),
        "revenue_by_source": revenue_by_source(period_won_leads, catalog_rows),
    }
    return Response(data)


@extend_schema(request=BusinessSettingsSerializer, responses=BusinessSettingsSerializer)
@api_view(["GET", "PATCH"])
@permission_classes([AllowAny])
def business_settings(request):
    obj, _ = BusinessSettings.objects.get_or_create(pk=1)
    if request.method == "PATCH" and (not request.user or not request.user.is_authenticated):
        return Response({"detail": "Authentication credentials were not provided."}, status=status.HTTP_401_UNAUTHORIZED)
    if request.method == "PATCH" and not has_page_permission(request.user, "settings", True):
        return forbidden()
    if request.method == "PATCH":
        serializer = BusinessSettingsSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    return Response(BusinessSettingsSerializer(obj).data)


@extend_schema(request=InstagramSettingsSerializer, responses=InstagramSettingsSerializer)
@api_view(["GET", "PATCH"])
@permission_classes([AllowAny])
def instagram_status(request):
    if not request.user or not request.user.is_authenticated:
        return Response({"detail": "Authentication credentials were not provided."}, status=status.HTTP_401_UNAUTHORIZED)
    obj, _ = InstagramSettings.objects.get_or_create(pk=1)
    integration, _ = IntegrationSettings.objects.get_or_create(pk=1)
    context = {"instagram_access_token": integration.instagram_access_token or settings.INSTAGRAM_ACCESS_TOKEN, "instagram_account_id": integration.instagram_account_id or settings.INSTAGRAM_ACCOUNT_ID}
    if request.method == "PATCH" and not has_page_permission(request.user, "settings", True):
        return forbidden()
    if request.method == "PATCH":
        serializer = InstagramSettingsSerializer(obj, data=request.data, partial=True, context=context)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    return Response(InstagramSettingsSerializer(obj, context=context).data)


def is_developer(user):
    return bool(user and user.is_authenticated and getattr(getattr(user, "profile", None), "role", None) == "developer")


def dashboard_period(request):
    start_value = request.query_params.get("from") or request.query_params.get("date_from")
    end_value = request.query_params.get("to") or request.query_params.get("date_to")
    start = parse_datetime(start_value) if start_value else None
    end = parse_datetime(end_value) if end_value else None
    if start_value and not start:
        parsed = parse_date(start_value)
        if parsed:
            start = timezone.make_aware(datetime.combine(parsed, time.min))
    if end_value and not end:
        parsed = parse_date(end_value)
        if parsed:
            end = timezone.make_aware(datetime.combine(parsed, time.max))
    if start and timezone.is_naive(start):
        start = timezone.make_aware(start)
    if end and timezone.is_naive(end):
        end = timezone.make_aware(end)
    today = timezone.localdate()
    if not end:
        end = timezone.make_aware(datetime.combine(today, time.max))
    if not start:
        end_date = timezone.localtime(end).date()
        start = timezone.make_aware(datetime.combine(end_date - timedelta(days=29), time.min))
    return start, end


def dashboard_daily_stats(leads, conversations, start, end, catalog_rows=None):
    start_date = timezone.localtime(start).date()
    end_date = timezone.localtime(end).date()
    lead_counts = {
        row["day"]: row["count"]
        for row in leads.annotate(day=TruncDate("created_at", tzinfo=timezone.get_current_timezone())).values("day").annotate(count=Count("id"))
    }
    conversation_counts = {
        row["day"]: row["count"]
        for row in conversations.annotate(day=TruncDate("created_at", tzinfo=timezone.get_current_timezone())).values("day").annotate(count=Count("id"))
    }
    catalog_days = catalog_sales_by_day(catalog_rows or [])
    days = []
    current = start_date
    while current <= end_date:
        catalog_day = catalog_days.get(current, {})
        days.append({
            "date": current.isoformat(),
            "leads": lead_counts.get(current, 0),
            "conversations": conversation_counts.get(current, 0),
            "catalog_revenue": catalog_day.get("revenue", Decimal("0")),
            "catalog_orders": catalog_day.get("orders", 0),
            "catalog_quantity": catalog_day.get("quantity", 0),
        })
        current += timedelta(days=1)
    return days


def analytics_daily_stats(leads, conversations, won_leads, start, end, catalog_rows=None):
    start_date = timezone.localtime(start).date()
    end_date = timezone.localtime(end).date()
    lead_counts = {
        row["day"]: row["count"]
        for row in leads.annotate(day=TruncDate("created_at", tzinfo=timezone.get_current_timezone())).values("day").annotate(count=Count("id"))
    }
    conversation_counts = {
        row["day"]: row["count"]
        for row in conversations.annotate(day=TruncDate("created_at", tzinfo=timezone.get_current_timezone())).values("day").annotate(count=Count("id"))
    }
    order_rows = won_leads.annotate(day=TruncDate("updated_at", tzinfo=timezone.get_current_timezone())).values("day").annotate(count=Count("id"), revenue=Coalesce(Sum("estimated_price"), Decimal("0")))
    order_counts = {row["day"]: row["count"] for row in order_rows}
    revenue_counts = {row["day"]: row["revenue"] for row in order_rows}
    catalog_days = catalog_sales_by_day(catalog_rows or [])
    days = []
    current = start_date
    while current <= end_date:
        catalog_day = catalog_days.get(current, {})
        lead_revenue = revenue_counts.get(current, Decimal("0"))
        catalog_revenue = catalog_day.get("revenue", Decimal("0"))
        days.append({
            "date": current.isoformat(),
            "leads": lead_counts.get(current, 0),
            "conversations": conversation_counts.get(current, 0),
            "orders": order_counts.get(current, 0) + catalog_day.get("orders", 0),
            "revenue": lead_revenue + catalog_revenue,
            "lead_orders": order_counts.get(current, 0),
            "lead_revenue": lead_revenue,
            "catalog_revenue": catalog_revenue,
            "catalog_orders": catalog_day.get("orders", 0),
            "catalog_quantity": catalog_day.get("quantity", 0),
        })
        current += timedelta(days=1)
    return days


def catalog_sales_queryset(start=None, end=None, branch=None):
    """Katalogdan haqiqiy sotuvlar. /api/accounting/ bilan bir xil manba.
    Filiallar aralashmasligi uchun har doim filial bo'yicha ajratiladi."""
    rows = CatalogHistory.objects.filter(action="sold").select_related("catalog_item")
    rows = rows.filter(catalog_item__branch=branch) if branch else rows.filter(catalog_item__branch__isnull=True)
    return apply_created_range(rows, start, end)


def catalog_sales_totals(rows):
    revenue = Decimal("0")
    quantity = 0
    orders = 0
    for history in rows:
        revenue += catalog_history_sale_total(history)
        quantity += int(history.quantity or 0)
        orders += 1
    return {"revenue": revenue, "orders": orders, "quantity": quantity}


def catalog_sales_by_day(rows):
    days = {}
    for history in rows:
        day = timezone.localtime(history.created_at).date()
        row = days.setdefault(day, {"revenue": Decimal("0"), "orders": 0, "quantity": 0})
        row["revenue"] += catalog_history_sale_total(history)
        row["orders"] += 1
        row["quantity"] += int(history.quantity or 0)
    return days


def catalog_sales_top_items(rows, limit=20):
    items = {}
    for history in rows:
        item = history.catalog_item
        if not item:
            continue
        row = items.setdefault(item.id, {
            "catalog_item_id": item.id,
            "catalog_item__name_uz": item.name_uz,
            "catalog_item__arrangement_type": item.arrangement_type,
            "catalog_item__image_url": item.image_url,
            "catalog_kind": item.catalog_kind,
            "quantity": 0,
            "orders": 0,
            "revenue": Decimal("0"),
            "last_sold_at": None,
        })
        row["quantity"] += int(history.quantity or 0)
        row["orders"] += 1
        row["revenue"] += catalog_history_sale_total(history)
        if not row["last_sold_at"] or history.created_at > row["last_sold_at"]:
            row["last_sold_at"] = history.created_at
    return sorted(items.values(), key=lambda row: (-row["quantity"], row["last_sold_at"] or timezone.now()))[:limit]


def top_selling_flowers(won_leads):
    lead_ids = list(won_leads.values_list("id", flat=True))
    rows = {}
    stock_rows = LeadStockUsage.objects.filter(lead_id__in=lead_ids).select_related("stock_batch__variant__flower").values(
        "stock_batch__variant__flower_id",
        "stock_batch__variant__flower__name_uz",
        "stock_batch__variant__color_uz",
    ).annotate(stems=Coalesce(Sum("quantity_stems"), 0), bunches=Coalesce(Sum("quantity_bunches"), Decimal("0")))
    for row in stock_rows:
        key = (row["stock_batch__variant__flower_id"], row["stock_batch__variant__color_uz"] or "")
        rows.setdefault(key, {"flower_id": row["stock_batch__variant__flower_id"], "name_uz": row["stock_batch__variant__flower__name_uz"], "color_uz": row["stock_batch__variant__color_uz"], "stems": 0, "bunches": Decimal("0")})
        rows[key]["stems"] += int(row["stems"] or 0)
        rows[key]["bunches"] += row["bunches"] or Decimal("0")
    catalog_usage = LeadCatalogUsage.objects.filter(lead_id__in=lead_ids).select_related("catalog_item").values("catalog_item_id").annotate(quantity=Coalesce(Sum("quantity"), 0))
    catalog_quantities = {row["catalog_item_id"]: row["quantity"] for row in catalog_usage}
    composition_rows = CatalogComposition.objects.filter(catalog_item_id__in=catalog_quantities.keys()).select_related("stock_batch__variant__flower")
    for composition in composition_rows:
        quantity = catalog_quantities.get(composition.catalog_item_id, 0)
        variant = composition.stock_batch.variant
        flower = variant.flower
        key = (flower.id, variant.color_uz or "")
        rows.setdefault(key, {"flower_id": flower.id, "name_uz": flower.name_uz, "color_uz": variant.color_uz, "stems": 0, "bunches": Decimal("0")})
        rows[key]["stems"] += int(composition.quantity_stems * quantity)
        rows[key]["bunches"] += composition.quantity_bunches * quantity
    return sorted([dict(row, bunches=str(row["bunches"])) for row in rows.values()], key=lambda row: row["stems"], reverse=True)[:20]


def top_catalog_items(won_leads, catalog_rows=None):
    """Katalogdan sotilgan mahsulotlar reytingi. Lead orqali sotilganlar ham qo'shiladi."""
    rows = catalog_sales_top_items(catalog_rows if catalog_rows is not None else [])
    merged = {row["catalog_item_id"]: row for row in rows}
    for lead_row in LeadCatalogUsage.objects.filter(lead__in=won_leads).select_related("catalog_item").values("catalog_item_id", "catalog_item__name_uz", "catalog_item__arrangement_type", "catalog_item__image_url").annotate(quantity=Coalesce(Sum("quantity"), 0), orders=Count("lead", distinct=True), revenue=Coalesce(Sum("lead__estimated_price"), Decimal("0")), last_sold_at=Max("lead__updated_at")):
        key = lead_row["catalog_item_id"]
        if key in merged:
            merged[key]["quantity"] += lead_row["quantity"]
            merged[key]["orders"] += lead_row["orders"]
            merged[key]["revenue"] += lead_row["revenue"]
            if lead_row["last_sold_at"] and (not merged[key]["last_sold_at"] or lead_row["last_sold_at"] > merged[key]["last_sold_at"]):
                merged[key]["last_sold_at"] = lead_row["last_sold_at"]
        else:
            merged[key] = {**lead_row, "catalog_kind": ""}
    return sorted(merged.values(), key=lambda row: (-row["quantity"], -(row["revenue"] or Decimal("0"))))[:20]


def recent_top_catalog_items(won_leads, catalog_rows=None):
    rows = top_catalog_items(won_leads, catalog_rows)
    return sorted(rows, key=lambda row: (row["last_sold_at"] or timezone.now(), row["quantity"]), reverse=True)[:20]


def conversation_source_breakdown(conversations):
    rows = {"instagram": 0, "telegram": 0, "mini_app": 0}
    for conversation in conversations.select_related("customer"):
        external_id = conversation.customer.instagram_user_id if conversation.customer_id else ""
        if external_id.startswith("telegram:"):
            rows["telegram"] += 1
        elif external_id.startswith("miniapp:"):
            rows["mini_app"] += 1
        else:
            rows["instagram"] += 1
    return [{"source": key, "count": value} for key, value in rows.items()]


def revenue_by_source(won_leads, catalog_rows=None):
    rows = list(won_leads.values("source").annotate(orders=Count("id"), revenue=Coalesce(Sum("estimated_price"), Decimal("0"))).order_by("source"))
    result = [{"source": row["source"] or "unknown", "source_label": SOURCE_LABELS.get(row["source"] or "unknown", row["source"] or "unknown"), "orders": row["orders"], "revenue": row["revenue"]} for row in rows]
    if catalog_rows is not None:
        totals = catalog_sales_totals(catalog_rows)
        if totals["orders"]:
            result.append({"source": "catalog", "source_label": SOURCE_LABELS["catalog"], "orders": totals["orders"], "revenue": totals["revenue"]})
    return sorted(result, key=lambda row: -row["revenue"])


SOURCE_LABELS = {
    "instagram": "Instagram",
    "telegram": "Telegram",
    "mini_app": "Mini app",
    "catalog": "Katalogdan sotuv",
    "unknown": "Aniqlanmagan",
}


def catalog_sale_financials(queryset):
    revenue = Decimal("0")
    cost = Decimal("0")
    florist_salary = Decimal("0")
    discount = Decimal("0")
    for item in queryset:
        sold = Decimal(item.quantity_sold or 0)
        total = Decimal(item.quantity_total or 1)
        ratio = sold / total if total else Decimal("0")
        sale_rows = list(item.history.filter(action="sold"))
        if sale_rows:
            revenue += sum(Decimal(row.sold_unit_price or 0) * Decimal(row.quantity or 0) for row in sale_rows)
            discount += sum(Decimal(row.discount_amount or 0) for row in sale_rows)
        else:
            revenue += Decimal(item.price or 0) * sold
            discount += Decimal(item.discount_amount or 0) * ratio
        cost += Decimal(item.calculated_cost_price or 0) * ratio
        florist_salary += Decimal(item.florist_fee or 0) * sold
    return {"revenue": revenue, "cost": cost, "florist_salary": florist_salary, "discount": discount, "profit": revenue - cost}


def catalog_sale_discount_stats(start, end, branch=None):
    rows = CatalogHistory.objects.filter(action="sold", discount_amount__gt=0)
    rows = rows.filter(catalog_item__branch=branch) if branch else rows.filter(catalog_item__branch__isnull=True)
    rows = apply_created_range(rows, start, end)
    totals = rows.aggregate(
        discounted_sales_count=Count("id"),
        discounted_quantity=Coalesce(Sum("quantity"), 0),
        discounted_amount=Coalesce(Sum("discount_amount"), Decimal("0")),
    )
    return totals


def batch_inventory_stats(start, end):
    movements = apply_created_range(StockMovement.objects.select_related("batch__variant__flower", "batch__supplier").filter(Q(reference_type="catalog_item") | Q(movement_type="waste")), start, end)
    catalog_ids = [row.reference_id for row in movements if row.reference_type == "catalog_item" and row.reference_id]
    catalog_kinds = {item.id: item.catalog_kind for item in CatalogItem.objects.filter(id__in=catalog_ids)}
    rows = {}
    for movement in movements:
        batch = movement.batch
        key = batch.id
        variant = batch.variant
        row = rows.setdefault(key, {
            "batch_id": batch.id,
            "batch_number": batch.batch_number,
            "supplier_id": batch.supplier_id,
            "supplier_name": batch.supplier.name if batch.supplier_id else "",
            "flower": variant.flower.name_uz,
            "variant": variant.name_uz,
            "color": variant.color_uz,
            "standard_catalog_stems": 0,
            "custom_catalog_stems": 0,
            "waste_stems": 0,
            "waste_cost_value": Decimal("0"),
            "total_out_stems": 0,
            "total_out_cost_value": Decimal("0"),
        })
        stems = abs(int(movement.quantity_stems or 0))
        cost_value = (Decimal(stems) * Decimal(batch.cost_per_stem or 0)).quantize(Decimal("0.01"))
        if movement.movement_type == "waste":
            row["waste_stems"] += stems
            row["waste_cost_value"] += cost_value
        elif movement.reference_type == "catalog_item":
            if catalog_kinds.get(movement.reference_id) == "custom":
                row["custom_catalog_stems"] += stems
            else:
                row["standard_catalog_stems"] += stems
        if movement.quantity_stems < 0:
            row["total_out_stems"] += stems
            row["total_out_cost_value"] += cost_value
    return sorted(rows.values(), key=lambda row: row["total_out_stems"], reverse=True)


def florist_production_stats(start, end):
    catalog = apply_created_range(CatalogItem.objects.select_related("florist__user").filter(florist__isnull=False), start, end)
    salary = apply_created_range(FloristSalaryEntry.objects.select_related("florist__user"), start, end)
    salary_by_florist = {row["florist_id"]: row["amount"] for row in salary.values("florist_id").annotate(amount=Coalesce(Sum("amount"), Decimal("0")))}
    rows = {}
    for item in catalog:
        profile = item.florist
        row = rows.setdefault(profile.id, {
            "florist_id": profile.id,
            "name": profile.user.get_full_name() or profile.user.username,
            "staff_type": profile.staff_type,
            "standard_bouquets": 0,
            "standard_baskets": 0,
            "custom_bouquets": 0,
            "custom_baskets": 0,
            "catalog_total": 0,
            "salary_total": salary_by_florist.get(profile.id, Decimal("0")),
        })
        row["catalog_total"] += 1
        key = f"{item.catalog_kind}_{'baskets' if item.arrangement_type == 'basket' else 'bouquets'}"
        if key in row:
            row[key] += 1
    return sorted(rows.values(), key=lambda row: row["catalog_total"], reverse=True)


def apply_created_range(queryset, start, end):
    if start:
        queryset = queryset.filter(created_at__gte=start)
    if end:
        queryset = queryset.filter(created_at__lte=end)
    return queryset


def apply_updated_range(queryset, start, end):
    if start:
        queryset = queryset.filter(updated_at__gte=start)
    if end:
        queryset = queryset.filter(updated_at__lte=end)
    return queryset


@extend_schema(request=AISettingsSerializer, responses=AISettingsSerializer)
@api_view(["GET", "PATCH"])
def ai_settings(request):
    if not is_developer(request.user):
        return forbidden()
    obj, _ = AISettings.objects.get_or_create(pk=1)
    if request.method == "PATCH":
        serializer = AISettingsSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    return Response(AISettingsSerializer(obj).data)


@extend_schema(request=IntegrationSettingsSerializer, responses=IntegrationSettingsSerializer)
@api_view(["GET", "PATCH"])
def integrations_settings(request):
    if not is_developer(request.user):
        return forbidden()
    obj, _ = IntegrationSettings.objects.get_or_create(pk=1)
    if request.method == "PATCH":
        serializer = IntegrationSettingsSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    return Response(IntegrationSettingsSerializer(obj).data)


@extend_schema(request=UploadSerializer, responses=UploadResponseSerializer)
@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_file(request):
    if not has_page_permission(request.user, "inventory", True):
        return forbidden()
    serializer = UploadSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    uploaded = serializer.validated_data["file"]
    path = default_storage.save(f"uploads/{uploaded.name}", uploaded)
    media_url = default_storage.url(path)
    url = f"{settings.PUBLIC_BASE_URL}{media_url}" if settings.PUBLIC_BASE_URL else request.build_absolute_uri(media_url)
    return Response({"url": url, "path": path}, status=status.HTTP_201_CREATED)


def mini_app_identity(init_data, require_user=False):
    integration, _ = IntegrationSettings.objects.get_or_create(pk=1)
    bot_token = integration.telegram_bot_token
    values = dict(parse_qsl(init_data or "", keep_blank_values=True))
    if bot_token and init_data and not values.get("hash"):
        raise ValueError("Mini app init data noto‘g‘ri")
    if bot_token and values.get("hash"):
        received_hash = values.pop("hash")
        data_check = "\n".join(f"{key}={values[key]}" for key in sorted(values))
        secret = hmac.new(b"WebAppData", bot_token.encode(), hashlib.sha256).digest()
        expected = hmac.new(secret, data_check.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(received_hash, expected):
            raise ValueError("Mini app init data noto‘g‘ri")
    user_payload = {}
    user_id = values.get("user", "")
    if user_id:
        try:
            user_payload = json.loads(user_id)
            user_id = str(user_payload.get("id", user_id))
        except (TypeError, ValueError):
            pass
    if require_user and not user_id:
        raise ValueError("Telegram init data kerak")
    return {"user_id": (user_id[:100] or "guest"), "user": user_payload, "auth_date": values.get("auth_date", ""), "query_id": values.get("query_id", "")}


def mini_app_user(init_data):
    return mini_app_identity(init_data, require_user=True)["user_id"]


def mini_app_customer(identity):
    return Customer.objects.filter(instagram_user_id=f"miniapp:{identity['user_id']}").first()


def mini_app_order_rows(customer):
    if not customer:
        return []
    rows = Lead.objects.filter(customer=customer).order_by("-created_at")[:30]
    statuses = {row.key: row.name_uz for row in LeadStatus.objects.filter(key__in=[lead.status for lead in rows])}
    return [{
        "id": row.id,
        "created_at": row.created_at,
        "updated_at": row.updated_at,
        "status": row.status,
        "status_label": statuses.get(row.status, row.status),
        "source": row.source,
        "arrangement_type": row.arrangement_type,
        "request": row.request_uz or row.request_ru,
        "estimated_price": row.estimated_price,
        "details": row.details or {},
    } for row in rows]


def mini_app_quote_payload(data):
    if data["arrangement_type"] in ["bouquet", "basket"]:
        return mini_app_custom_quote_ai(data["request_text"], data["arrangement_type"])
    total = Decimal("0")
    lines = []
    for item in data["items"]:
        if item.get("catalog_item"):
            catalog = CatalogItem.objects.filter(id=item["catalog_item"], status="available").first()
            if not catalog:
                raise ValueError("Katalog guli topilmadi")
            quantity = item.get("quantity") or 1
            line_total = catalog.price * quantity
            total += line_total
            lines.append({"type": "catalog", "id": catalog.id, "name_uz": catalog.name_uz, "quantity": quantity, "unit_price": str(catalog.price), "total": str(line_total)})
            continue
        raise ValueError("Mini app katalogda faqat catalog_item ishlatiladi")
    return {"lines": lines, "packaging": None, "florist_fee": str(Decimal("0")), "estimated_price": str(total), "price_is_estimate": False, "ai_note": ""}


def mini_app_request_text(arrangement_type, quote, note=""):
    labels = {"bouquet": "Buket", "basket": "Savat", "stems": "Donalab gul", "catalog": "Tayyor katalog"}
    lines = [f"Mini app buyurtma: {labels.get(arrangement_type, arrangement_type)}"]
    for row in quote["lines"]:
        if row["type"] == "catalog":
            lines.append(f"- {row['name_uz']}: {row['quantity']} ta, {row['total']} so‘m")
        elif row["type"] == "stock":
            name = f"{row['flower_uz']} {row['variant_uz']} {row['color_uz']}".strip()
            lines.append(f"- {name}: {row['quantity_stems']} dona, {row['total']} so‘m")
        elif row["type"] == "custom_text":
            lines.append(f"- Mijoz matni: {row['request_text']}")
    if quote.get("packaging"):
        packaging = quote["packaging"]
        lines.append(f"- Savat/qadoq: {packaging['name_uz']} ({packaging.get('size', '')}), {packaging['sale_price']} so‘m")
    if Decimal(str(quote["florist_fee"])) > 0:
        lines.append(f"- Florist xizmati: {quote['florist_fee']} so‘m")
    total_label = "Jami taxminan" if quote.get("price_is_estimate") else "Jami"
    lines.append(f"{total_label}: {quote['estimated_price']} so‘m")
    if note:
        lines.append(f"Izoh: {note}")
    return "\n".join(lines)


@extend_schema(parameters=[MiniAppInitSerializer], responses=inline_serializer(name="MiniAppCustomer", fields={"customer": CustomerSerializer(allow_null=True), "orders": serializers.ListField(child=serializers.DictField())}))
@api_view(["GET"])
@permission_classes([AllowAny])
def mini_app_me(request):
    try:
        identity = mini_app_identity(request.query_params.get("init_data", ""), require_user=True)
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
    customer = mini_app_customer(identity)
    return Response({"customer": CustomerSerializer(customer).data if customer else None, "orders": mini_app_order_rows(customer)})


@extend_schema(parameters=[MiniAppInitSerializer], responses=inline_serializer(name="MiniAppCatalog", fields={"catalog": CatalogItemSerializer(many=True), "customer": CustomerSerializer(allow_null=True), "orders": serializers.ListField(child=serializers.DictField())}))
@api_view(["GET"])
@permission_classes([AllowAny])
def mini_app_catalog(request):
    try:
        identity = mini_app_identity(request.query_params.get("init_data", ""))
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
    catalog = CatalogItem.objects.filter(status="available").select_related("social_post")[:50]
    customer = mini_app_customer(identity)
    return Response({"catalog": CatalogItemSerializer(catalog, many=True).data, "customer": CustomerSerializer(customer).data if customer else None, "orders": mini_app_order_rows(customer)})


@extend_schema(request=MiniAppQuoteSerializer, responses=inline_serializer(name="MiniAppQuoteResponse", fields={"lines": serializers.ListField(child=serializers.DictField()), "packaging": serializers.DictField(allow_null=True), "florist_fee": serializers.CharField(), "estimated_price": serializers.CharField(), "price_is_estimate": serializers.BooleanField(), "ai_note": serializers.CharField(allow_blank=True)}))
@api_view(["POST"])
@permission_classes([AllowAny])
def mini_app_quote(request):
    serializer = MiniAppQuoteSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        mini_app_user(serializer.validated_data.get("init_data", ""))
        quote = mini_app_quote_payload(serializer.validated_data)
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(quote)


@extend_schema(methods=["GET"], parameters=[MiniAppInitSerializer], responses=inline_serializer(name="MiniAppOrders", fields={"orders": serializers.ListField(child=serializers.DictField()), "customer": CustomerSerializer(allow_null=True)}))
@extend_schema(methods=["POST"], request=MiniAppLeadSerializer, responses=LeadSerializer)
@api_view(["GET", "POST"])
@permission_classes([AllowAny])
def mini_app_lead(request):
    if request.method == "GET":
        try:
            identity = mini_app_identity(request.query_params.get("init_data", ""), require_user=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        customer = mini_app_customer(identity)
        return Response({"customer": CustomerSerializer(customer).data if customer else None, "orders": mini_app_order_rows(customer)})
    serializer = MiniAppLeadSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        mini_user = mini_app_user(serializer.validated_data.get("init_data", ""))
        quote = mini_app_quote_payload(serializer.validated_data)
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    phone = normalize_phone(serializer.validated_data["phone"]) or serializer.validated_data["phone"]
    customer, _ = Customer.objects.update_or_create(instagram_user_id=f"miniapp:{mini_user}", defaults={"name": serializer.validated_data["name"], "phone": phone, "language": "uz"})
    request_text = mini_app_request_text(serializer.validated_data["arrangement_type"], quote, serializer.validated_data.get("note", ""))
    details = {"lines": quote["lines"], "packaging": quote["packaging"], "florist_fee": quote["florist_fee"], "estimated_price": quote["estimated_price"], "price_is_estimate": quote["price_is_estimate"], "ai_note": quote.get("ai_note", ""), "note": serializer.validated_data.get("note", "")}
    lead = Lead.objects.create(customer=customer, status="new", request_uz=request_text, arrangement_type=serializer.validated_data["arrangement_type"], estimated_price=quote["estimated_price"], source="mini_app", details=details)
    for row in quote["lines"]:
        if row["type"] == "catalog":
            catalog_item = CatalogItem.objects.filter(id=row["id"]).first()
            if catalog_item:
                LeadCatalogUsage.objects.create(lead=lead, catalog_item=catalog_item, quantity=row["quantity"])
        elif row["type"] == "stock":
            batch = StockBatch.objects.filter(id=row["id"]).first()
            if batch:
                LeadStockUsage.objects.create(lead=lead, stock_batch=batch, quantity_stems=row["quantity_stems"], quantity_bunches=Decimal(row["quantity_stems"]) / Decimal(batch.stems_per_bunch))
    Notification.objects.create(notification_type="lead", title_uz=f"Mini app lead: {customer}", title_ru=f"Mini app лид: {customer}", body_uz=request_text, body_ru=request_text, reference_type="lead", reference_id=lead.id)
    return Response(LeadSerializer(lead).data, status=status.HTTP_201_CREATED)


@extend_schema(
    request=inline_serializer(name="InstagramWebhookPayload", fields={}),
    responses={
        200: OpenApiResponse(description="Webhook verified or event received"),
        403: OpenApiResponse(description="Invalid verify token"),
    },
)
@api_view(["GET", "POST"])
@permission_classes([AllowAny])
def instagram_webhook(request):
    if request.method == "GET":
        integration, _ = IntegrationSettings.objects.get_or_create(pk=1)
        verify_token = integration.instagram_verify_token or settings.INSTAGRAM_VERIFY_TOKEN
        if request.query_params.get("hub.verify_token") == verify_token:
            return Response(int(request.query_params.get("hub.challenge", "0")))
        return Response({"detail": "Invalid verify token"}, status=status.HTTP_403_FORBIDDEN)
    from .tasks import process_instagram_webhook
    process_instagram_webhook.delay(request.data)
    return Response({"status": "EVENT_RECEIVED"})


@extend_schema(request=inline_serializer(name="TelegramWebhookPayload", fields={}), responses={200: OpenApiResponse(description="Telegram event received")})
@api_view(["POST"])
@permission_classes([AllowAny])
def telegram_webhook(request):
    from .tasks import process_telegram_webhook
    process_telegram_webhook.delay(request.data)
    return Response({"status": "EVENT_RECEIVED"})


@extend_schema(request=inline_serializer(name="BackupTelegramWebhookPayload", fields={}), responses={200: OpenApiResponse(description="Backup bot event received")})
@api_view(["POST"])
@permission_classes([AllowAny])
def backup_telegram_webhook(request):
    from .backup_services import backup_command_matches
    if backup_command_matches(request.data):
        from .tasks import send_telegram_backup
        message = request.data.get("message") or {}
        sender = message.get("from") or {}
        triggered_by = sender.get("username") or sender.get("first_name") or "telegram_command"
        send_telegram_backup.delay(f"manual:{triggered_by}")
        return Response({"status": "BACKUP_QUEUED"})
    return Response({"status": "IGNORED"})
